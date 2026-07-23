import json
import io
import os
from datetime import datetime, date, timedelta
from decimal import Decimal
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app, send_from_directory, make_response
from flask_login import login_required, current_user
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import (
    db, Plant, Size, Field, StockBalance, Document, DocumentRow,
    AppSetting, PriceHistory, PriceChangeLog, Order, OrderItem, Client, ActionLog,
    ProjectPottingRecountLine, SeedlingEventLog, StockPurchaseLot,
)
from app.utils import log_action, get_or_create_stock, msk_now, natural_key, apply_excel_styles
from app.services import get_unit_cost_for_year
from app.stock_helpers import get_reserved_map, aggregate_stock_movements
from app.shop_prices import get_shop_price_map, resolve_shop_price, transform_stock_report_price_mode

bp = Blueprint('stock', __name__)

RECOUNT_DOC_TYPES = ('field_recount', 'potting_recount')


def build_stock_row_price_map(sorted_groups, price_overrides=None):
    """{plant_id_size_id: {wholesale, retail}} для выгрузок остатков."""
    ov = price_overrides if price_overrides is not None else get_shop_price_map()
    result = {}
    for group in sorted_groups:
        if group.get('is_section'):
            continue
        for row in group.get('data', {}).get('rows', []):
            key = f"{row['plant_id']}_{row['size_id']}"
            if key in result:
                continue
            wholesale = float(row.get('price') or 0)
            result[key] = {
                'wholesale': wholesale,
                'retail': resolve_shop_price(
                    row['plant_id'], row['size_id'], wholesale, ov,
                ),
            }
    return result


def _is_recount_doc(doc_type):
    return doc_type in RECOUNT_DOC_TYPES


def _group_recount_rows(doc):
    """Разделяет строки пересчета на приход/списание для карточки."""
    rows_in = []
    rows_out = []
    total_in = 0
    total_out = 0
    field_name = ''

    sorted_rows = sorted(
        doc.rows,
        key=lambda r: (
            r.plant.name if r.plant else '',
            r.size.name if r.size else '',
            r.year or 0
        )
    )
    for row in sorted_rows:
        qty = int(row.quantity or 0)
        if not field_name and row.field_to:
            field_name = row.field_to.name
        item = {
            'plant_name': row.plant.name if row.plant else '-',
            'size_name': row.size.name if row.size else '-',
            'year': row.year,
            'qty': abs(qty),
            'delta': qty,
            'row': row
        }
        if qty >= 0:
            rows_in.append(item)
            total_in += abs(qty)
        else:
            rows_out.append(item)
            total_out += abs(qty)

    return {
        'doc': doc,
        'field_name': field_name or '-',
        'rows_in': rows_in,
        'rows_out': rows_out,
        'total_in': total_in,
        'total_out': total_out
    }


def resolve_stock_report_filters(report_mode, args=None):
    """Разбор фильтров остатков: как на экране, так и в Excel.

    Возвращает dict:
      selected_sizes, query_sizes, size_filter_manual,
      selected_fields, selected_plants, selected_years
    """
    src = args if args is not None else request.args
    selected_sizes = [int(x) for x in src.getlist('filter_size')]
    size_filter_manual = bool(selected_sizes)
    query_sizes = list(selected_sizes)
    if not selected_sizes and report_mode == 'product':
        from app.seedlings import is_excluded_from_product_stock, is_seedling_size_name
        all_s = Size.query.all()
        selected_sizes = [s.id for s in all_s if not is_excluded_from_product_stock(s.name)]
        raw_ids = [
            s.id for s in all_s
            if is_seedling_size_name(s.name) and is_excluded_from_product_stock(s.name)
        ]
        query_sizes = (selected_sizes + raw_ids) if selected_sizes else [s.id for s in all_s]
        if not selected_sizes:
            selected_sizes = [s.id for s in all_s]
            query_sizes = selected_sizes
    if not query_sizes:
        query_sizes = selected_sizes or None

    return {
        'selected_sizes': selected_sizes,
        'query_sizes': query_sizes,
        'size_filter_manual': size_filter_manual,
        'selected_fields': [int(x) for x in src.getlist('filter_field')],
        'selected_plants': [int(x) for x in src.getlist('filter_plant')],
        'selected_years': [int(x) for x in src.getlist('filter_year')],
    }


def grand_total_from_stock_groups(sorted_groups, exclude_section_keys=None):
    """Суммы по группам отчёта; секции из exclude_section_keys не входят в итог."""
    exclude = set(exclude_section_keys or ())
    total = {
        'income': 0, 'reserved': 0, 'free': 0, 'shipped': 0,
        'qty': 0, 'sum': 0, 'free_sum': 0,
    }
    skip = False
    for g in sorted_groups:
        if g.get('is_section'):
            skip = g.get('section_key') in exclude
            continue
        if skip:
            continue
        t = (g.get('data') or {}).get('totals') or {}
        total['income'] += t.get('income') or 0
        total['reserved'] += t.get('reserved') or 0
        total['free'] += t.get('free') or 0
        total['shipped'] += t.get('shipped') or 0
        total['qty'] += t.get('qty') or 0
        total['sum'] += t.get('sum') or 0
        total['free_sum'] += t.get('free_sum') or 0
    return total


def build_stock_report_data(report_mode, end_date, selected_fields=None, selected_plants=None, selected_sizes=None, selected_years=None, all_plants=None, all_sizes=None, all_fields=None):
    """Собирает данные для отчета по остаткам (та же логика, что и в stock_report).

    Возвращает кортеж (sorted_groups, grand_total).
    """
    # --- Загрузочные справочники ---
    if all_plants is None:
        all_plants = sorted(Plant.query.all(), key=lambda x: x.name)
    if all_sizes is None:
        all_sizes = sorted(Size.query.all(), key=natural_key)
    if all_fields is None:
        all_fields = sorted(Field.query.all(), key=natural_key)

    if selected_fields is None or not selected_fields:
        selected_fields = [f.id for f in all_fields]

    fact_map, income_map, shipped_map = aggregate_stock_movements(end_date, selected_fields)

    reserve_map = {}
    for (pid, sid, fid, yr), res_q in get_reserved_map().items():
        if fid in selected_fields and res_q > 0:
            reserve_map[(pid, sid, fid, yr)] = res_q

    # --- Собираем уникальные ключи ---
    all_keys = set(fact_map.keys()) | set(income_map.keys()) | set(reserve_map.keys())

    plants_dict = {p.id: p for p in all_plants}
    sizes_dict = {s.id: s.name for s in all_sizes}
    fields_dict = {f.id: f.name for f in all_fields}

    # --- Логика цен и себестоимости ---
    target_price_year = end_date.year
    try:
        current_unit_cost = float(get_unit_cost_for_year(target_price_year))
    except Exception:
        current_unit_cost = 0.0

    price_years = {target_price_year}
    for (_pid, _sid, _fid, byear) in all_keys:
        if byear is not None:
            price_years.add(byear)

    hist_rows = PriceHistory.query.filter(PriceHistory.year.in_(price_years)).all()
    hist_prices_map = {}
    hist_plant_size_map = {}
    for h in hist_rows:
        hist_prices_map[(h.plant_id, h.size_id, h.field_id, h.year)] = h.price
        hist_plant_size_map[(h.plant_id, h.size_id, h.year)] = h.price
    stock_prices_map = {
        (sb.plant_id, sb.size_id, sb.field_id, sb.year): sb.price
        for sb in StockBalance.query.with_entities(
            StockBalance.plant_id,
            StockBalance.size_id,
            StockBalance.field_id,
            StockBalance.year,
            StockBalance.price,
        ).all()
    }

    aggregated = {}
    grand_total = {'income': 0, 'reserved': 0, 'free': 0, 'shipped': 0, 'qty': 0, 'sum': 0, 'free_sum': 0}

    for (pid, sid, fid, byear) in all_keys:
        if selected_plants and pid not in selected_plants:
            continue
        if selected_sizes and sid not in selected_sizes:
            continue
        if selected_years and byear not in selected_years:
            continue

        fact = fact_map.get((pid, sid, fid, byear), 0)
        inc = income_map.get((pid, sid, fid, byear), 0)
        res = reserve_map.get((pid, sid, fid, byear), 0)
        shp = shipped_map.get((pid, sid, fid, byear), 0)
        if fact == 0:
            # Не показываем нулевые остатки, пока факт снова не станет ≠ 0
            continue

        size_name = sizes_dict.get(sid, '')
        is_netov = 'нетов' in size_name.lower()

        if is_netov:
            calc_price = current_unit_cost
            display_price = current_unit_cost
        else:
            actual_price = lookup_sale_price(
                pid, sid, fid, byear, target_price_year,
                hist_prices_map, stock_prices_map, hist_plant_size_map,
            )
            calc_price = float(actual_price)
            display_price = calc_price

        batch_sum = calc_price * fact

        if report_mode == 'inventory':
            k_agg = (pid, fid, byear)
            size_val = None
        else:
            k_agg = (pid, sid)
            size_val = size_name

        pl_obj = plants_dict.get(pid)

        if k_agg not in aggregated:
            aggregated[k_agg] = {
                'plant_id': pid,
                'size_id': sid if report_mode != 'inventory' else 0,
                'plant': pl_obj.name if pl_obj else 'Unknown',
                'latin_name': pl_obj.latin_name if pl_obj else '',
                'characteristic': pl_obj.characteristic if pl_obj else '',
                'size': size_val,
                'fields': set(),
                'field_id': fid if report_mode == 'inventory' else None,
                'year': byear if report_mode == 'inventory' else None,
                'price': display_price,
                'income': 0,
                'reserved': 0,
                'shipped': 0,
                'quantity': 0,
                'total_val': 0
            }

        aggregated[k_agg]['fields'].add(fields_dict.get(fid, 'Unknown'))
        aggregated[k_agg]['income'] += inc
        aggregated[k_agg]['reserved'] += res
        aggregated[k_agg]['shipped'] += shp
        aggregated[k_agg]['quantity'] += fact
        aggregated[k_agg]['total_val'] += batch_sum

        if not is_netov and display_price > aggregated[k_agg]['price']:
            aggregated[k_agg]['price'] = display_price

    grouped_final = {}
    for k, v in aggregated.items():
        free = v['quantity'] - v['reserved']
        sm = v['total_val']
        free_sm = (v['price'] or 0) * free
        row = {
            'plant_id': v['plant_id'],
            'size_id': v['size_id'],
            'field_id': v.get('field_id'),
            'size': v['size'],
            'fields_str': ", ".join(sorted(list(v['fields']))),
            'characteristic': v['characteristic'],
            'latin_name': v['latin_name'],
            'year': v.get('year'),
            'price': v['price'],
            'sum': sm,
            'free_sum': free_sm,
            'income': v['income'],
            'reserved': v['reserved'],
            'free': free,
            'shipped': v['shipped'],
            'quantity': v['quantity']
        }

        p_name = v['plant']
        p_latin = v['latin_name']
        if p_name not in grouped_final:
            grouped_final[p_name] = {'latin_name': p_latin, 'rows': [], 'totals': {'sum': 0, 'free_sum': 0, 'income': 0, 'reserved': 0, 'free': 0, 'shipped': 0, 'qty': 0}}

        grouped_final[p_name]['rows'].append(row)
        t = grouped_final[p_name]['totals']
        t['sum'] += sm
        t['free_sum'] += free_sm
        t['income'] += v['income']
        t['reserved'] += v['reserved']
        t['free'] += free
        t['shipped'] += v['shipped']
        t['qty'] += v['quantity']
        grand_total['sum'] += sm
        grand_total['free_sum'] += free_sm
        grand_total['income'] += v['income']
        grand_total['reserved'] += v['reserved']
        grand_total['free'] += free
        grand_total['shipped'] += v['shipped']
        grand_total['qty'] += v['quantity']

    sorted_groups = []
    for p_name in sorted(grouped_final.keys()):
        group = grouped_final[p_name]
        if report_mode == 'inventory':
            group['rows'].sort(key=lambda x: (x.get('year') or 0, x['fields_str']))
        else:
            group['rows'].sort(key=lambda x: (natural_key(x['size']), x['fields_str']))
        sorted_groups.append({'name': p_name, 'data': group})

    # Товарный режим: сначала грунт, затем товарные саженцы (Саженцы + контейнер).
    if report_mode == 'product':
        from app.seedlings import is_product_seedling_size_name, is_seedling_size_name

        def _empty_totals():
            return {'sum': 0, 'free_sum': 0, 'income': 0, 'reserved': 0, 'free': 0, 'shipped': 0, 'qty': 0}

        def _recalc_group(name, latin_name, rows):
            totals = _empty_totals()
            for r in rows:
                totals['sum'] += r.get('sum') or 0
                totals['free_sum'] += r.get('free_sum') or 0
                totals['income'] += r.get('income') or 0
                totals['reserved'] += r.get('reserved') or 0
                totals['free'] += r.get('free') or 0
                totals['shipped'] += r.get('shipped') or 0
                totals['qty'] += r.get('quantity') or 0
            return {
                'name': name,
                'data': {'latin_name': latin_name or '', 'rows': rows, 'totals': totals},
            }

        def _section_marker(key, title, print_include=True):
            return {
                'is_section': True,
                'section_key': key,
                'print_include': print_include,
                'name': title,
                'data': {'latin_name': '', 'rows': [], 'totals': _empty_totals()},
            }

        ground_groups = []
        container_groups = []
        raw_groups = []
        for g in sorted_groups:
            ground_rows = []
            container_rows = []
            raw_rows = []
            for row in g['data']['rows']:
                sn = row.get('size') or ''
                if is_product_seedling_size_name(sn):
                    container_rows.append(row)
                elif is_seedling_size_name(sn):
                    raw_rows.append(row)
                else:
                    ground_rows.append(row)
            latin = g['data'].get('latin_name') or ''
            if ground_rows:
                ground_groups.append(_recalc_group(g['name'], latin, ground_rows))
            if raw_rows:
                raw_groups.append(_recalc_group(g['name'], latin, raw_rows))
            if container_rows:
                container_groups.append(_recalc_group(g['name'], latin, container_rows))

        sorted_groups = []
        if ground_groups:
            sorted_groups.append(_section_marker('ground', 'Растения в грунте'))
            sorted_groups.extend(ground_groups)
        if container_groups:
            sorted_groups.append(_section_marker('containers', 'Товарные саженцы'))
            sorted_groups.extend(container_groups)
        # Нетоварные (нетов / без контейнера) — на экране, но не в КП/PDF автофильтра
        if raw_groups:
            sorted_groups.append(_section_marker(
                'raw_seedlings',
                'Саженцы (не товарные)',
                print_include=False,
            ))
            sorted_groups.extend(raw_groups)

    return sorted_groups, grand_total


def lookup_sale_price(plant_id, size_id, field_id, batch_year, target_price_year,
                      hist_prices_map, stock_prices_map, hist_plant_size_map=None):
    """Цена продажи партии: точная запись в истории → позиция+год → остаток."""
    for y in (target_price_year, batch_year):
        p = hist_prices_map.get((plant_id, size_id, field_id, y))
        if p is not None:
            return float(p)
    if hist_plant_size_map:
        for y in (target_price_year, batch_year):
            p = hist_plant_size_map.get((plant_id, size_id, y))
            if p is not None:
                return float(p)
    p = stock_prices_map.get((plant_id, size_id, field_id, batch_year))
    if p is not None and float(p) > 0:
        return float(p)
    return 0.0


def _append_price_change_log(
    *,
    plant_id: int,
    size_id: int,
    new_price,
    old_price=None,
    field_id=None,
    year=None,
    source: str = 'ui',
    user_id=None,
):
    """Пишет строку в журнал; текущая цена в PriceHistory при этом может перезаписываться отдельно."""
    try:
        new_val = Decimal(str(new_price or 0))
    except Exception:
        new_val = Decimal('0')
    old_val = None
    if old_price is not None:
        try:
            old_val = Decimal(str(old_price))
        except Exception:
            old_val = None
    if old_val is not None and old_val == new_val:
        return
    if user_id is None:
        try:
            if current_user and getattr(current_user, 'is_authenticated', False):
                user_id = current_user.id
        except Exception:
            user_id = None
    db.session.add(PriceChangeLog(
        plant_id=plant_id,
        size_id=size_id,
        field_id=field_id,
        year=year,
        old_price=old_val,
        new_price=new_val,
        changed_at=msk_now(),
        user_id=user_id,
        source=source or 'ui',
    ))


def apply_price_for_plant_size(plant_id: int, size_id: int, price, source: str = 'ui') -> dict:
    """Перезаписывает текущую цену продажи во всех партиях и дописывает журнал изменений."""
    price_val = Decimal(str(price))

    stocks = StockBalance.query.filter_by(plant_id=plant_id, size_id=size_id).all()
    if not stocks:
        orphan_hist = PriceHistory.query.filter_by(plant_id=plant_id, size_id=size_id).all()
        if orphan_hist:
            for h in orphan_hist:
                _append_price_change_log(
                    plant_id=plant_id, size_id=size_id,
                    old_price=h.price, new_price=price_val,
                    field_id=h.field_id, year=h.year, source=source,
                )
                h.price = price_val
            db.session.commit()
            return {
                'ok': True,
                'stock_rows': 0,
                'hist_created': 0,
                'hist_updated': len(orphan_hist),
                'batches': len(orphan_hist),
                'message': f'Обновлено {len(orphan_hist)} записей истории (на складе партий нет).',
            }
        return {
            'ok': False,
            'message': 'Нет партий на складе для этой позиции — некуда применить цену.',
        }

    hist_created = hist_updated = stock_rows = 0
    seen_keys = set()

    for sb in stocks:
        key = (sb.field_id, sb.year)
        seen_keys.add(key)
        old_stock = sb.price
        sb.price = price_val
        stock_rows += 1

        ex = PriceHistory.query.filter_by(
            plant_id=plant_id, size_id=size_id,
            field_id=sb.field_id, year=sb.year,
        ).first()
        old_hist = ex.price if ex else old_stock
        _append_price_change_log(
            plant_id=plant_id, size_id=size_id,
            old_price=old_hist, new_price=price_val,
            field_id=sb.field_id, year=sb.year, source=source,
        )
        if ex:
            ex.price = price_val
            hist_updated += 1
        else:
            db.session.add(PriceHistory(
                plant_id=plant_id, size_id=size_id,
                field_id=sb.field_id, year=sb.year, price=price_val,
            ))
            hist_created += 1

    for h in PriceHistory.query.filter_by(plant_id=plant_id, size_id=size_id).all():
        if (h.field_id, h.year) in seen_keys:
            continue
        _append_price_change_log(
            plant_id=plant_id, size_id=size_id,
            old_price=h.price, new_price=price_val,
            field_id=h.field_id, year=h.year, source=source,
        )
        h.price = price_val
        hist_updated += 1

    db.session.commit()
    batches = len(seen_keys)
    return {
        'ok': True,
        'stock_rows': stock_rows,
        'hist_created': hist_created,
        'hist_updated': hist_updated,
        'batches': batches,
        'message': (
            f'Цена {float(price_val):,.2f} ₽ применена к {batches} '
            f'{"партии" if batches == 1 else "партиям"} '
            f'({stock_rows} строк остатков перезаписано).'
        ).replace(',', ' '),
    }


def _price_history_position_options(plant_id: int, size_id: int) -> dict:
    """Собирает партии (год+поле) из остатков и истории цен для подсказок в форме."""
    fields_dict = {f.id: f.name for f in Field.query.all()}
    batches = []
    seen = set()

    stocks = (
        StockBalance.query.filter_by(plant_id=plant_id, size_id=size_id)
        .order_by(StockBalance.year.desc(), StockBalance.field_id.asc())
        .all()
    )
    for sb in stocks:
        key = (sb.year, sb.field_id)
        seen.add(key)
        hist = PriceHistory.query.filter_by(
            plant_id=plant_id, size_id=size_id, field_id=sb.field_id, year=sb.year
        ).first()
        batches.append({
            'year': sb.year,
            'field_id': sb.field_id,
            'field_name': fields_dict.get(sb.field_id, '—'),
            'quantity': int(sb.quantity or 0),
            'stock_price': float(sb.price or 0),
            'history_price': float(hist.price) if hist else None,
            'has_history': hist is not None,
            'source': 'stock',
        })

    history_rows = (
        PriceHistory.query.filter_by(plant_id=plant_id, size_id=size_id)
        .order_by(PriceHistory.year.desc(), PriceHistory.field_id.asc())
        .all()
    )
    for h in history_rows:
        key = (h.year, h.field_id)
        if key in seen:
            continue
        seen.add(key)
        batches.append({
            'year': h.year,
            'field_id': h.field_id,
            'field_name': h.field.name if h.field else fields_dict.get(h.field_id, '—'),
            'quantity': 0,
            'stock_price': 0.0,
            'history_price': float(h.price or 0),
            'has_history': True,
            'source': 'history_only',
        })

    batches.sort(
        key=lambda b: (
            0 if b['quantity'] > 0 else 1,
            0 if b['has_history'] else 1,
            -b['year'],
            natural_key(b['field_name']),
        )
    )

    hist_items = [
        {
            'year': h.year,
            'field_id': h.field_id,
            'field_name': h.field.name if h.field else fields_dict.get(h.field_id, '—'),
            'price': float(h.price or 0),
        }
        for h in history_rows
    ]

    suggested = batches[0] if batches else None
    unique_prices = sorted({b['history_price'] for b in batches if b['history_price'] is not None}
                           | {b['stock_price'] for b in batches if b['stock_price'] > 0})
    total_qty = sum(b['quantity'] for b in batches)
    return {
        'batches': batches,
        'history': hist_items,
        'suggested': suggested,
        'batch_count': len(batches),
        'total_qty': total_qty,
        'unique_prices': unique_prices,
        'suggested_price': (
            float(suggested['history_price']) if suggested and suggested['history_price'] is not None
            else (float(suggested['stock_price']) if suggested and suggested['stock_price'] > 0 else None)
        ),
        'current_year': msk_now().year,
    }


def build_price_edit_positions(product_only: bool = True):
    """Позиции для массового редактирования цен: группировка plant+size.

    Поля с остатками склеиваются через запятую. Цену берём из StockBalance;
    если она разная по партиям — old_price=None и mixed_prices=True.
    """
    from collections import defaultdict
    from app.seedlings import is_excluded_from_product_stock

    fields_map = {f.id: f.name for f in Field.query.all()}
    plants_map = {p.id: p.name for p in Plant.query.all()}
    sizes_map = {s.id: s.name for s in Size.query.all()}

    excluded_size_ids = set()
    if product_only:
        for sid, sname in sizes_map.items():
            if is_excluded_from_product_stock(sname):
                excluded_size_ids.add(sid)

    groups = defaultdict(lambda: {
        'qty': 0,
        'fields': set(),
        'prices': set(),
    })
    for sb in StockBalance.query.filter(StockBalance.quantity > 0).all():
        if sb.size_id in excluded_size_ids:
            continue
        key = (sb.plant_id, sb.size_id)
        g = groups[key]
        g['qty'] += int(sb.quantity or 0)
        fname = fields_map.get(sb.field_id)
        if fname:
            g['fields'].add(fname)
        try:
            pp = float(sb.price or 0)
        except (TypeError, ValueError):
            pp = 0.0
        if pp > 0:
            g['prices'].add(round(pp, 2))

    rows = []
    for (plant_id, size_id), g in groups.items():
        prices = sorted(g['prices'])
        mixed = len(prices) > 1
        old_price = prices[0] if len(prices) == 1 else None
        if not prices:
            hist = (
                PriceHistory.query.filter_by(plant_id=plant_id, size_id=size_id)
                .order_by(PriceHistory.year.desc())
                .first()
            )
            if hist and hist.price is not None:
                try:
                    old_price = float(hist.price)
                except (TypeError, ValueError):
                    old_price = None
        rows.append({
            'plant_id': plant_id,
            'size_id': size_id,
            'plant_name': plants_map.get(plant_id, f'#{plant_id}'),
            'size_name': sizes_map.get(size_id, f'#{size_id}'),
            'fields_str': ', '.join(sorted(g['fields'], key=natural_key)),
            'quantity': g['qty'],
            'old_price': old_price,
            'mixed_prices': mixed,
            'unique_prices': prices,
        })

    rows.sort(key=lambda r: (natural_key(r['plant_name']), natural_key(r['size_name'])))
    return rows


@bp.route('/stock/api/price_history_options')
@login_required
def price_history_options():
    if current_user.role != 'admin':
        return jsonify({'error': 'forbidden'}), 403
    plant_id = request.args.get('plant_id', type=int)
    size_id = request.args.get('size_id', type=int)
    if not plant_id or not size_id:
        return jsonify({
            'batches': [], 'history': [], 'suggested': None,
            'batch_count': 0, 'total_qty': 0, 'unique_prices': [],
            'suggested_price': None, 'current_year': msk_now().year,
        })
    return jsonify(_price_history_position_options(plant_id, size_id))


@bp.route('/price_history', methods=['GET', 'POST'])
@login_required
def price_history():
    if current_user.role != 'admin':
        return redirect(url_for('directory.directory'))

    if request.method == 'POST':
        payload = request.get_json(silent=True) if request.is_json else None
        act = (payload or {}).get('action') if payload else request.form.get('action')

        if act == 'bulk_add':
            items = (payload or {}).get('items') if payload else None
            if items is None:
                plants = request.form.getlist('ph_plant')
                sizes = request.form.getlist('ph_size')
                prices = request.form.getlist('ph_price')
                items = []
                for i in range(min(len(plants), len(sizes), len(prices))):
                    items.append({
                        'plant_id': plants[i],
                        'size_id': sizes[i],
                        'price': prices[i],
                    })
            ok_n = 0
            err_n = 0
            messages = []
            for it in items or []:
                try:
                    p = int(it.get('plant_id'))
                    s = int(it.get('size_id'))
                    raw = str(it.get('price') or '').replace(',', '.').replace(' ', '').replace('\xa0', '').replace('₽', '').strip()
                    if not raw:
                        continue
                    pr = float(raw)
                    result = apply_price_for_plant_size(p, s, pr, source='bulk')
                    if result.get('ok'):
                        ok_n += 1
                        log_action(
                            f"Цена {pr} для plant={p} size={s}: "
                            f"{result.get('batches', 0)} партий, stock={result.get('stock_rows', 0)}"
                        )
                    else:
                        err_n += 1
                        messages.append(result.get('message') or 'ошибка')
                except (TypeError, ValueError) as e:
                    err_n += 1
                    messages.append(str(e))
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                return jsonify({
                    'status': 'ok' if ok_n else 'error',
                    'updated': ok_n,
                    'errors': err_n,
                    'messages': messages[:5],
                })
            if ok_n:
                flash(f'Обновлено позиций: {ok_n}' + (f', ошибок: {err_n}' if err_n else ''), 'success')
            elif err_n:
                flash('Не удалось сохранить цены', 'warning')
            return redirect(url_for('stock.price_history'))

        if act == 'add':
            p = int(request.form.get('plant_id'))
            s = int(request.form.get('size_id'))
            pr = float(request.form.get('price'))
            result = apply_price_for_plant_size(p, s, pr, source='ui')
            if result['ok']:
                flash(result['message'], 'success')
                log_action(
                    f"Цена {pr} для plant={p} size={s}: "
                    f"{result.get('batches', 0)} партий, stock={result.get('stock_rows', 0)}"
                )
            else:
                flash(result['message'], 'warning')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'status': 'ok' if result.get('ok') else 'error', **result})
        elif act == 'delete':
            # Удаление только из журнала логов (append-only журнал редкий, но админ может почистить)
            PriceChangeLog.query.filter_by(id=request.form.get('id')).delete()
            db.session.commit()
        elif act == 'delete_position':
            p = int(request.form.get('plant_id'))
            s = int(request.form.get('size_id'))
            n = PriceChangeLog.query.filter_by(plant_id=p, size_id=s).delete()
            db.session.commit()
            flash(f'Удалено записей журнала: {n}', 'success')
            log_action(f"Удалил журнал цен plant={p} size={s} ({n} записей)")
        return redirect(url_for('stock.price_history'))

    change_log = (
        PriceChangeLog.query
        .order_by(PriceChangeLog.changed_at.desc(), PriceChangeLog.id.desc())
        .limit(500)
        .all()
    )
    return render_template(
        'stock/price_history.html',
        change_log=change_log,
        stock_positions=build_price_edit_positions(product_only=True),
        plants=sorted(Plant.query.all(), key=lambda x: x.name),
        sizes=sorted(Size.query.all(), key=natural_key),
        fields=sorted(Field.query.all(), key=natural_key),
        current_year=msk_now().year,
    )

@bp.route('/stock', methods=['GET', 'POST'])
@login_required
def stock_report():
    report_mode = request.args.get('mode', 'product')
    end_date_str = request.args.get('end_date')
    # Если дата не выбрана, берем текущий момент
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59) if end_date_str else msk_now()

    filters = resolve_stock_report_filters(report_mode)
    selected_sizes = filters['selected_sizes']
    query_sizes = filters['query_sizes']
    size_filter_manual = filters['size_filter_manual']
    selected_fields = filters['selected_fields']
    selected_plants = filters['selected_plants']
    selected_years = filters['selected_years']

    # Обработка POST
    if request.method == 'POST':
        if current_user.role == 'admin':
            # Логика обновления цены
            if 'update_price' in request.form:
                p_id, s_id = request.form.get('plant_id'), request.form.get('size_id')
                try:
                    new_price = float(request.form.get('price').replace(' ', '').replace('₽', ''))
                    apply_price_for_plant_size(int(p_id), int(s_id), new_price, source='stock')
                    log_action(f"Обновил цену в остатках PlantID {p_id}")
                except ValueError: pass
            
            # Логика обновления ИНФО (Латинское название)
            elif 'update_info' in request.form:
                p_id = request.form.get('plant_id')
                latin_val = request.form.get('latin_name')
                plant = Plant.query.get(p_id)
                if plant:
                    plant.latin_name = latin_val
                    db.session.commit()
                    log_action(f"Обновил латинское название для {plant.name}")

        return redirect(url_for('stock.stock_report', end_date=request.args.get('end_date'), mode=report_mode, filter_field=selected_fields, filter_plant=selected_plants, filter_size=selected_sizes, filter_year=selected_years))
    
    # --- Загрузка данных ---
    all_plants = sorted(Plant.query.all(), key=lambda x: x.name)
    all_sizes = sorted(Size.query.all(), key=natural_key)
    all_fields = sorted(Field.query.all(), key=natural_key)
    all_years = sorted([r[0] for r in db.session.query(StockBalance.year).distinct().all() if r[0] is not None])

    sorted_groups, grand_total = build_stock_report_data(
        report_mode,
        end_date,
        selected_fields=selected_fields,
        selected_plants=selected_plants,
        selected_sizes=query_sizes,
        selected_years=selected_years,
        all_plants=all_plants,
        all_sizes=all_sizes,
        all_fields=all_fields,
    )
    # Автофильтр товарных: нетоварные саженцы на экране есть, но в «Итого» не входят
    # (как в Excel / КП). При ручном выборе размеров — считаем всё выбранное.
    if report_mode == 'product' and not size_filter_manual:
        grand_total = grand_total_from_stock_groups(
            sorted_groups, exclude_section_keys={'raw_seedlings'},
        )
    else:
        grand_total = grand_total_from_stock_groups(sorted_groups)
    
    print_settings_obj = AppSetting.query.get('print_header_settings')
    print_settings = json.loads(print_settings_obj.value) if print_settings_obj else {}
    price_overrides = get_shop_price_map()
    stock_row_prices = build_stock_row_price_map(sorted_groups, price_overrides)
    
    return render_template('stock/stock.html', 
                           report_data=sorted_groups, 
                           grand_total=grand_total, 
                           end_date=end_date.strftime('%Y-%m-%d'), 
                           report_mode=report_mode, 
                           all_plants=all_plants, selected_plants=selected_plants, 
                           all_sizes=all_sizes, selected_sizes=selected_sizes, 
                           all_fields=all_fields, selected_fields=selected_fields, 
                           all_years=all_years, selected_years=selected_years, 
                           print_settings=print_settings,
                           price_overrides=price_overrides,
                           resolve_shop_price=resolve_shop_price,
                           stock_row_prices=stock_row_prices,
                           size_filter_manual=size_filter_manual)

@bp.route('/stock/export')
@login_required
def stock_report_export():
    report_mode = request.args.get('mode', 'product')
    end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d').replace(hour=23, minute=59) if request.args.get('end_date') else msk_now()

    filters = resolve_stock_report_filters(report_mode)
    query_sizes = filters['query_sizes']
    selected_fields = filters['selected_fields']
    selected_plants = filters['selected_plants']
    selected_years = filters['selected_years']

    all_plants = sorted(Plant.query.all(), key=lambda x: x.name)
    all_sizes = sorted(Size.query.all(), key=natural_key)
    all_fields = sorted(Field.query.all(), key=natural_key)

    sorted_groups, _ = build_stock_report_data(
        report_mode,
        end_date,
        selected_fields=selected_fields,
        selected_plants=selected_plants,
        selected_sizes=query_sizes,
        selected_years=selected_years,
        all_plants=all_plants,
        all_sizes=all_sizes,
        all_fields=all_fields,
    )

    # В Excel секция «не товарные» не выгружается (как раньше в КП)
    skip_raw_rows = report_mode == 'product'

    price_mode = request.args.get('price_mode', 'wholesale')
    if price_mode not in ('wholesale', 'retail'):
        price_mode = 'wholesale'
    if price_mode == 'retail':
        sorted_groups = transform_stock_report_price_mode(sorted_groups, 'retail')

    if skip_raw_rows:
        grand_total = grand_total_from_stock_groups(
            sorted_groups, exclude_section_keys={'raw_seedlings'},
        )
    else:
        grand_total = grand_total_from_stock_groups(sorted_groups)

    # --- Создание Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"

    # Стили (ближе к экспорту заказов)
    title_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    group_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    border_total = Border(top=Side(style='thick'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Колонки, которые присутствуют в таблице на экране
    columns = ["Растение"]
    include_size = report_mode != 'inventory'
    if include_size:
        columns.append("Размер")

    columns.append("Поле")

    include_price = current_user.role != 'user2'
    if include_price:
        columns.append("Цена (опт.)" if price_mode == 'wholesale' else "Цена (розница)")

    columns.extend(["Резерв", "Отгружено", "Свободно", "Факт"])

    if include_price:
        columns.append("Сумма")

    if report_mode == 'inventory':
        columns.append("Год")

    # Заголовок файла
    title = f"Остатки на {end_date.strftime('%d.%m.%Y')} ({'Товарные' if report_mode=='product' else 'По полям'})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Заголовки столбцов
    header_row = 2
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    row_idx = header_row + 1

    section_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=12)

    # Данные: группы + детали (сырые саженцы в Excel — только если размеры выбраны вручную)
    for group in sorted_groups:
        if group.get('is_section') and group.get('section_key') == 'raw_seedlings' and skip_raw_rows:
            break
        if group.get('is_section'):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(columns))
            scell = ws.cell(row=row_idx, column=1, value=group['name'])
            scell.font = section_font
            scell.fill = section_fill
            scell.alignment = Alignment(horizontal="left", vertical="center")
            row_idx += 1
            continue

        # Группирующая строка
        gcell = ws.cell(row=row_idx, column=1, value=group['name'])
        gcell.font = Font(bold=True)
        gcell.fill = group_fill
        gcell.border = thin_border

        def _set_group_value(col_name, value):
            if col_name in columns:
                idx = columns.index(col_name) + 1
                c = ws.cell(row=row_idx, column=idx, value=value)
                c.font = Font(bold=True)
                c.fill = group_fill
                c.border = thin_border
                c.alignment = align_center

        _set_group_value("Резерв", group['data']['totals']['reserved'])
        _set_group_value("Отгружено", group['data']['totals']['shipped'])
        _set_group_value("Свободно", group['data']['totals']['free'])
        _set_group_value("Факт", group['data']['totals']['qty'])
        if include_price:
            _set_group_value("Сумма", float(group['data']['totals']['sum']))

        row_idx += 1

        # Строки деталей
        for row in group['data']['rows']:
            col = 1
            ws.cell(row=row_idx, column=col, value="").border = thin_border
            col += 1

            if include_size:
                ws.cell(row=row_idx, column=col, value=row.get('size') or '').border = thin_border
                col += 1

            ws.cell(row=row_idx, column=col, value=row.get('fields_str')).border = thin_border
            col += 1

            if include_price:
                c_price = ws.cell(row=row_idx, column=col, value=float(row.get('price') or 0))
                c_price.number_format = '#,##0.00'
                c_price.alignment = align_right
                c_price.border = thin_border
                col += 1

            for field_name in ["reserved", "shipped", "free", "quantity"]:
                c = ws.cell(row=row_idx, column=col, value=row.get(field_name) or 0)
                c.alignment = align_center
                c.border = thin_border
                col += 1

            if include_price:
                c_sum = ws.cell(row=row_idx, column=col, value=float(row.get('sum') or 0))
                c_sum.number_format = '#,##0.00'
                c_sum.font = Font(bold=True)
                c_sum.alignment = align_right
                c_sum.border = thin_border
                col += 1

            if report_mode == 'inventory':
                c_year = ws.cell(row=row_idx, column=col, value=row.get('year'))
                c_year.alignment = align_center
                c_year.border = thin_border

            row_idx += 1

    # Итоговая строка
    total_label = ws.cell(row=row_idx, column=1, value="ИТОГО")
    total_label.font = Font(bold=True)
    total_label.border = border_total

    def _write_total(col_name, value, number_format=None):
        if col_name in columns:
            idx = columns.index(col_name) + 1
            c = ws.cell(row=row_idx, column=idx, value=value)
            c.font = Font(bold=True)
            c.border = border_total
            c.alignment = align_center if number_format is None else align_right
            if number_format:
                c.number_format = number_format

    _write_total("Резерв", grand_total['reserved'])
    _write_total("Отгружено", grand_total['shipped'])
    _write_total("Свободно", grand_total['free'])
    _write_total("Факт", grand_total['qty'])
    if include_price:
        _write_total("Сумма", float(grand_total['sum'] or 0), number_format='#,##0.00 "₽"')

    # Автоширина колонок
    # Берем номера колонок по индексам (чтобы избежать ошибок с MergedCell)
    for idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
        ws.column_dimensions[get_column_letter(idx)].width = max_length + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'stock_export_{end_date.strftime("%Y%m%d")}_{"opt" if price_mode == "wholesale" else "rozn"}.xlsx'
    response = make_response(buf.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


@bp.route('/stock/receipts', methods=['GET', 'POST'])
@login_required
def stock_receipts():
    """Карточки поступления (несколько позиций) — вместо вкладки в журнале документов."""
    from app.models import Supplier
    from app.purchase_lots import apply_purchase_lot

    if current_user.role not in ['admin', 'executive', 'user']:
        flash('Недостаточно прав', 'warning')
        return redirect(url_for('stock.stock_report'))

    if request.method == 'POST':
        if current_user.role not in ['admin', 'executive']:
            flash('Проведение поступления доступно администратору', 'warning')
            return redirect(url_for('stock.stock_receipts'))
        form_date = request.form.get('date')
        doc_date = datetime.strptime(form_date, '%Y-%m-%d') if form_date else msk_now()
        try:
            supplier_id = int(request.form.get('supplier_id') or 0) or None
        except (TypeError, ValueError):
            supplier_id = None
        comment = (request.form.get('comment') or '').strip()
        p_ids = request.form.getlist('plant[]')
        s_ids = request.form.getlist('size[]')
        f_to = request.form.getlist('field_to[]')
        y_ids = request.form.getlist('year[]')
        q_ids = request.form.getlist('quantity[]')
        pp_ids = request.form.getlist('purchase_price[]')

        lines = []
        for i in range(len(p_ids)):
            try:
                pid = int(p_ids[i]); sid = int(s_ids[i]); ft = int(f_to[i])
                year = int(y_ids[i]); qty = int(q_ids[i])
            except (TypeError, ValueError, IndexError):
                continue
            if qty <= 0 or not pid or not sid or not ft or not year:
                continue
            try:
                pp = float((pp_ids[i] if i < len(pp_ids) else '') or 0)
            except (TypeError, ValueError):
                pp = 0
            lines.append({
                'plant_id': pid, 'size_id': sid, 'field_id': ft,
                'year': year, 'quantity': qty, 'purchase_price': pp,
            })
        if not lines:
            flash('Добавьте хотя бы одну позицию поступления', 'warning')
            return redirect(url_for('stock.stock_receipts'))

        try:
            doc = Document(
                doc_type='income',
                user_id=current_user.id,
                date=doc_date,
                comment=comment or 'Поступление',
                supplier_id=supplier_id,
            )
            db.session.add(doc)
            db.session.flush()
            for line in lines:
                st = get_or_create_stock(
                    line['plant_id'], line['size_id'], line['field_id'], line['year'],
                )
                st.quantity = int(st.quantity or 0) + line['quantity']
                row = DocumentRow(
                    document_id=doc.id,
                    plant_id=line['plant_id'],
                    size_id=line['size_id'],
                    field_to_id=line['field_id'],
                    year=line['year'],
                    quantity=line['quantity'],
                    purchase_price=line['purchase_price'] or None,
                )
                db.session.add(row)
                db.session.flush()
                apply_purchase_lot(
                    plant_id=line['plant_id'],
                    size_id=line['size_id'],
                    field_id=line['field_id'],
                    year=line['year'],
                    quantity=line['quantity'],
                    purchase_price=line['purchase_price'],
                    supplier_id=supplier_id,
                    document_id=doc.id,
                    document_row_id=row.id,
                    stock=st,
                )
            db.session.commit()
            log_action(f'Поступление #{doc.id}: {len(lines)} поз.')
            flash(f'Поступление проведено (док. #{doc.id})', 'success')
        except Exception as exc:
            db.session.rollback()
            flash(f'Ошибка поступления: {exc}', 'danger')
        return redirect(url_for('stock.stock_receipts'))

    docs = (
        Document.query
        .options(
            joinedload(Document.user),
            joinedload(Document.supplier),
            joinedload(Document.rows).joinedload(DocumentRow.plant),
            joinedload(Document.rows).joinedload(DocumentRow.size),
            joinedload(Document.rows).joinedload(DocumentRow.field_to),
        )
        .filter_by(doc_type='income')
        .order_by(Document.date.desc(), Document.id.desc())
        .limit(80)
        .all()
    )
    return render_template(
        'stock/receipts.html',
        plants=sorted(Plant.query.all(), key=lambda x: (x.name or '').lower()),
        sizes=sorted(Size.query.all(), key=natural_key),
        fields=sorted(Field.query.all(), key=natural_key),
        suppliers=sorted(Supplier.query.all(), key=lambda s: (s.name or '').lower()),
        docs=docs,
        current_year=msk_now().year,
        today=msk_now().strftime('%Y-%m-%d'),
        can_post=current_user.role in ['admin', 'executive'],
    )


@bp.route('/stock/save_settings', methods=['POST'])
@login_required
def save_stock_settings():
    if current_user.role != 'admin': return jsonify({'status': 'error'})
    data = request.json
    setting = AppSetting.query.get('stock_widths')
    if not setting: 
        setting = AppSetting(key='stock_widths')
        db.session.add(setting)
    setting.value = json.dumps(data)
    db.session.commit()
    return jsonify({'status': 'ok'})

@bp.route('/settings/save_print_header', methods=['POST'])
@login_required
def save_print_header():
    if current_user.role != 'admin': return redirect(url_for('stock.stock_report'))
    
    setting = AppSetting.query.get('print_header_settings')
    data = json.loads(setting.value) if setting and setting.value else {}

    data.update({
        'company_name': request.form.get('company_name'),
        'phone': request.form.get('phone'),
        'email': request.form.get('email'),
        'website': request.form.get('website'),
        'footer_text': request.form.get('footer_text')
    })
    
    def save_file_safely(field_name, target_name):
        file = request.files.get(field_name)
        if file and file.filename and file.filename.strip() != '':
            try:
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
                filename = f"{target_name}.{ext}"
                save_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
                file.save(save_path)
                data[f'{target_name}_url'] = url_for('main.serve_uploaded_file', filename=filename)
            except Exception as e:
                print(f"Error saving {field_name}: {e}")

    save_file_safely('logo', 'company_logo')
    save_file_safely('photo1', 'promo_photo_1')
    save_file_safely('photo2', 'promo_photo_2')
    save_file_safely('photo3', 'promo_photo_3')

    if not setting:
        setting = AppSetting(key='print_header_settings')
        db.session.add(setting)
    
    setting.value = json.dumps(data)
    db.session.commit()
    
    flash('Настройки печати и фото сохранены')
    return redirect(url_for('stock.stock_report'))


@bp.route('/api/field_recount/items')
@login_required
def api_field_recount_items():
    """Возвращает фактические остатки по выбранному полю для формы пересчета."""
    field_id = request.args.get('field_id', type=int)
    if not field_id:
        return jsonify({'items': []})

    # Считаем ФАКТ по документам (аналогично отчету по остаткам), а не из StockBalance,
    # чтобы не зависеть от возможного рассинхрона таблицы остатков.
    fact_map = {}
    movements = db.session.query(
        Document.doc_type,
        DocumentRow.plant_id,
        DocumentRow.size_id,
        DocumentRow.size_to_id,
        DocumentRow.field_from_id,
        DocumentRow.field_to_id,
        DocumentRow.year,
        DocumentRow.quantity
    ).join(Document).all()

    for m in movements:
        dtype = m.doc_type
        pid = m.plant_id
        sid = m.size_id
        yr = m.year
        qty = int(m.quantity or 0)

        if dtype in ['income', 'correction', 'field_recount', 'potting_recount']:
            if m.field_to_id == field_id:
                k = (pid, sid, yr)
                fact_map[k] = fact_map.get(k, 0) + qty
        elif dtype in ['writeoff', 'shipment']:
            if m.field_from_id == field_id:
                k = (pid, sid, yr)
                fact_map[k] = fact_map.get(k, 0) - qty
        elif dtype == 'move':
            if m.field_from_id == field_id:
                k = (pid, sid, yr)
                fact_map[k] = fact_map.get(k, 0) - qty
            if m.field_to_id == field_id:
                k = (pid, sid, yr)
                fact_map[k] = fact_map.get(k, 0) + qty
        elif dtype == 'regrading':
            if m.field_from_id == field_id:
                k_old = (pid, sid, yr)
                fact_map[k_old] = fact_map.get(k_old, 0) - qty
            dst_field = getattr(m, 'field_to_id', None) or m.field_from_id
            if m.size_to_id and dst_field == field_id:
                k_new = (pid, m.size_to_id, yr)
                fact_map[k_new] = fact_map.get(k_new, 0) + qty

    plants_map = {p.id: p.name for p in Plant.query.all()}
    sizes_map = {s.id: s.name for s in Size.query.all()}
    rmap = get_reserved_map()
    items = []
    for (pid, sid, yr), fact_qty in fact_map.items():
        if fact_qty <= 0:
            continue
        reserved = int(rmap.get((pid, sid, field_id, yr), 0) or 0)
        free = max(0, int(fact_qty) - reserved)
        items.append({
            'plant_id': pid,
            'plant_name': plants_map.get(pid, '-'),
            'size_id': sid,
            'size_name': sizes_map.get(sid, '-'),
            'year': yr,
            'current_qty': int(fact_qty),
            'reserved': reserved,
            'free': free,
        })
    items.sort(key=lambda x: (natural_key(x['plant_name']), natural_key(x['size_name']), x['year'] or 0))
    return jsonify({'items': items})

@bp.route('/documents', methods=['GET', 'POST'])
@login_required
def documents():
    if request.method == 'POST':
        dt = request.form.get('doc_type')
        if dt in ['income', 'writeoff', 'shipment'] and current_user.role != 'admin': 
            flash('Только для администратора')
            return redirect(url_for('stock.documents'))
        if dt == 'field_recount' and current_user.role == 'user2':
            flash('Только для менеджера/админа')
            return redirect(url_for('stock.documents', active_tab='regrading'))

        if dt == 'field_recount':
            field_id = request.form.get('recount_field_id', type=int)
            doc_date = msk_now()

            plant_ids = request.form.getlist('recount_plant_id[]')
            size_ids = request.form.getlist('recount_size_id[]')
            years = request.form.getlist('recount_year[]')
            fact_qtys = request.form.getlist('recount_fact_qty[]')

            if not field_id:
                flash('Выберите поле для пересчета')
                return redirect(url_for('stock.documents', active_tab='regrading'))

            try:
                # агрегируем по ключу, чтобы избежать дублей строк
                fact_map = {}
                for i in range(len(plant_ids)):
                    try:
                        pid = int(plant_ids[i])
                        sid = int(size_ids[i])
                        yr = int(years[i])
                        fact_qty = int(fact_qtys[i])
                    except Exception:
                        continue
                    if fact_qty < 0:
                        raise ValueError('Фактическое количество не может быть отрицательным')
                    fact_map[(pid, sid, yr)] = fact_qty

                if not fact_map:
                    flash('Нет данных для пересчета')
                    return redirect(url_for('stock.documents', active_tab='regrading'))

                # Валидация: фактическое кол-во не должно быть меньше зарезервированного
                rmap = get_reserved_map()
                for (pid, sid, yr), fact_qty in fact_map.items():
                    reserved = int(rmap.get((pid, sid, field_id, yr), 0) or 0)
                    if int(fact_qty) < reserved:
                        pl = Plant.query.get(pid); sz = Size.query.get(sid)
                        name = (pl.name if pl else '') + ((' · ' + sz.name) if sz else '')
                        raise ValueError(f'Пересчёт {fact_qty} шт по позиции «{name}» меньше резерва {reserved} шт')

                field_obj = Field.query.get(field_id)
                doc_comment = f'Пересчет по полю: {field_obj.name if field_obj else field_id}'

                doc = Document(
                    doc_type='field_recount',
                    user_id=current_user.id,
                    date=doc_date,
                    comment=doc_comment
                )
                db.session.add(doc)
                db.session.flush()

                changed_rows = 0
                for (pid, sid, yr), fact_qty in fact_map.items():
                    stock = get_or_create_stock(pid, sid, field_id, yr)
                    current_qty = int(stock.quantity or 0)
                    delta = int(fact_qty) - current_qty
                    if delta == 0:
                        continue
                    stock.quantity += delta
                    db.session.add(DocumentRow(
                        document_id=doc.id,
                        plant_id=pid,
                        size_id=sid,
                        field_to_id=field_id,
                        year=yr,
                        quantity=delta
                    ))
                    changed_rows += 1

                if changed_rows == 0:
                    db.session.delete(doc)
                    db.session.commit()
                    flash('Изменений нет: фактические остатки совпадают с базой')
                else:
                    db.session.commit()
                    flash(f'Карточка пересчета сохранена (изменено строк: {changed_rows})')
                    log_action(f"Создал карточку пересчета #{doc.id} по полю {field_id}")
                    # Анализ карточки на расхождения +/- (триггер для админов).
                    # Не ломаем основной поток.
                    try:
                        from app.anomaly_engine import sync_recount_anomaly_for_doc
                        sync_recount_anomaly_for_doc(doc.id)
                    except Exception:
                        try:
                            current_app.logger.exception(
                                'sync_recount_anomaly_for_doc(%s) failed', doc.id
                            )
                        except Exception:
                            pass

            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка пересчета: {e}')

            return redirect(url_for('stock.documents', active_tab='regrading'))
        
        p_ids = request.form.getlist('plant[]')
        s_ids = request.form.getlist('size[]')
        q_ids = request.form.getlist('quantity[]')
        y_ids = request.form.getlist('year[]')
        f_to = request.form.getlist('field_to[]')
        f_from = request.form.getlist('field_from[]')
        
        form_date = request.form.get('date')
        doc_date = datetime.strptime(form_date, '%Y-%m-%d') if form_date else msk_now()
        
        try:
            inc_yr = int(request.form.get('income_year', msk_now().year))
        except (ValueError, TypeError):
            inc_yr = msk_now().year
            
        comm = request.form.get('comment', '')

        # если админ создаёт ручную отгрузку с указанием клиента, заведём "ghost"‑заказ
        ghost_order = None
        if dt == 'shipment':
            client_id = request.form.get('client_id', type=int)
            if client_id:
                ghost_order = Order(client_id=client_id, date=doc_date, status='ghost')
                db.session.add(ghost_order)
                db.session.flush()

        try:
            doc = Document(doc_type=dt, user_id=current_user.id, date=doc_date, comment=comm)
            if ghost_order:
                doc.order_id = ghost_order.id
            db.session.add(doc)
            db.session.flush()

            # Карта резервов нужна только для операций, ограниченных свободным остатком
            rmap_local = get_reserved_map() if dt == 'writeoff' else {}

            for i in range(len(p_ids)):
                try: qty = int(q_ids[i])
                except: continue
                
                if qty <= 0: continue
                pid = int(p_ids[i])
                year = int(y_ids[i])

                if dt == 'income':
                    from app.purchase_lots import apply_purchase_lot
                    sid = int(s_ids[i]); ft = int(f_to[i])
                    # Год партии: из строки или общий income_year документа
                    try:
                        row_year = int(y_ids[i]) if i < len(y_ids) and y_ids[i] else inc_yr
                    except (TypeError, ValueError):
                        row_year = inc_yr
                    st = get_or_create_stock(pid, sid, ft, row_year)
                    st.quantity += qty
                    try:
                        in_purchase = float(request.form.getlist('purchase_price[]')[i] or 0)
                    except (ValueError, TypeError, IndexError):
                        in_purchase = 0
                    try:
                        in_sell = float(request.form.getlist('price[]')[i] or 0)
                    except (ValueError, TypeError, IndexError):
                        in_sell = 0
                    if in_sell > 0:
                        st.price = in_sell
                    row = DocumentRow(
                        document_id=doc.id, plant_id=pid, size_id=sid,
                        field_to_id=ft, year=row_year, quantity=qty,
                        purchase_price=in_purchase or None,
                    )
                    db.session.add(row)
                    db.session.flush()
                    apply_purchase_lot(
                        plant_id=pid, size_id=sid, field_id=ft, year=row_year,
                        quantity=qty, purchase_price=in_purchase,
                        supplier_id=getattr(doc, 'supplier_id', None),
                        document_id=doc.id, document_row_id=row.id, stock=st,
                    )
                
                elif dt == 'writeoff':
                    sid = int(s_ids[i]); ff = int(f_from[i])
                    stock = get_or_create_stock(pid, sid, ff, year)
                    reserved = int(rmap_local.get((pid, sid, ff, year), 0) or 0)
                    free = max(0, int(stock.quantity or 0) - reserved)
                    if qty > free:
                        pl = Plant.query.get(pid); sz = Size.query.get(sid)
                        name = (pl.name if pl else '') + ((' · ' + sz.name) if sz else '')
                        raise ValueError(f'Списание {qty} шт по позиции «{name}» превышает свободный остаток {free} шт (резерв: {reserved})')
                    stock.quantity -= qty
                    db.session.add(DocumentRow(document_id=doc.id, plant_id=pid, size_id=sid, field_from_id=ff, year=year, quantity=qty))
                
                elif dt == 'move':
                    sid = int(s_ids[i]); ff = int(f_from[i]); ft = int(f_to[i])
                    src = get_or_create_stock(pid, sid, ff, year)
                    # Защита: нельзя унести с поля больше, чем там физически
                    # есть свободного с учётом резерва. Так перемещение в
                    # журнале документов получает ту же защиту, что и на
                    # карте (api_visual_move).
                    from app.stock_helpers import compute_reserved
                    reserved_src = compute_reserved(pid, sid, ff, year)
                    src_free = max(0, int(src.quantity or 0) - reserved_src)
                    if qty > src_free:
                        pl = Plant.query.get(pid); sz = Size.query.get(sid)
                        name = (pl.name if pl else '') + ((' · ' + sz.name) if sz else '')
                        raise ValueError(
                            f'Перемещение {qty} шт по «{name}» превышает '
                            f'свободный остаток на исходном поле {src_free} шт '
                            f'(резерв: {reserved_src})'
                        )
                    dest = get_or_create_stock(pid, sid, ft, year)
                    src.quantity -= qty
                    dest.quantity += qty
                    # Переносим закупочную/продажную цену партии на новое
                    # поле, если там пусто. Иначе финансовые отчёты не увидят
                    # себестоимости после move и завысят маржу.
                    try:
                        if (not dest.purchase_price) and src.purchase_price:
                            dest.purchase_price = src.purchase_price
                        if (not dest.price) and src.price:
                            dest.price = src.price
                    except Exception:
                        pass
                    db.session.add(DocumentRow(document_id=doc.id, plant_id=pid, size_id=sid, field_from_id=ff, field_to_id=ft, year=year, quantity=qty))
                
                elif dt == 'shipment':
                    sid = int(s_ids[i]); ff = int(f_from[i])
                    # Защита: ручная отгрузка через журнал не должна уводить
                    # склад в минус и не должна забирать чужой резерв (в т.ч.
                    # активный заказ другого менеджера). Если есть ghost — то
                    # резерв «себе» считаем нормальным; ghost ниже резерва
                    # тоже не учитывается (по правилам get_reserved_map).
                    stock_ship = get_or_create_stock(pid, sid, ff, year)
                    from app.stock_helpers import compute_reserved
                    reserved_ship = compute_reserved(pid, sid, ff, year)
                    free_ship = max(0, int(stock_ship.quantity or 0) - reserved_ship)
                    if qty > free_ship:
                        pl = Plant.query.get(pid); sz = Size.query.get(sid)
                        name = (pl.name if pl else '') + ((' · ' + sz.name) if sz else '')
                        raise ValueError(
                            f'Отгрузка {qty} шт по «{name}» превышает '
                            f'свободный остаток {free_ship} шт '
                            f'(резерв: {reserved_ship}). Если это уже '
                            f'зарезервированный заказ — отгружай через '
                            f'карточку заказа, а не через журнал документов.'
                        )
                    stock_ship.quantity -= qty
                    db.session.add(DocumentRow(document_id=doc.id, plant_id=pid, size_id=sid, field_from_id=ff, year=year, quantity=qty))
                    # если мы создали ghost-заказ, сразу добавим позицию туда тоже
                    if ghost_order:
                        db.session.add(OrderItem(order_id=ghost_order.id, plant_id=pid, size_id=sid, field_id=ff, year=year, quantity=qty, shipped_quantity=qty, price=0))

                elif dt == 'regrading':
                    sid_from = int(s_ids[i])       
                    sid_to = int(f_to[i]) 
                    field_id = int(f_from[i])      
                    get_or_create_stock(pid, sid_from, field_id, year).quantity -= qty
                    get_or_create_stock(pid, sid_to, field_id, year).quantity += qty
                    db.session.add(DocumentRow(document_id=doc.id, plant_id=pid, size_id=sid_from, size_to_id=sid_to, field_from_id=field_id, year=year, quantity=qty))

            db.session.commit()
            flash('Документ проведен')
            log_action(f"Создал документ {dt} #{doc.id}")
        except Exception as e: 
            db.session.rollback()
            flash(f'Error: {e}')
            
        return redirect(url_for('stock.documents'))
    
    # --- ЛОГИКА GET-ЗАПРОСА (С ФИЛЬТРАМИ) ---
    show_all = request.args.get('show_all') == '1'
    f_start = request.args.get('start_date')
    f_end = request.args.get('end_date')
    f_client = request.args.get('client_id', type=int)
    f_doc_type = request.args.get('doc_type')

    list_date_limited = False
    if not show_all and not f_start and not f_end:
        default_start = (msk_now().date() - timedelta(days=183)).isoformat()
        f_start = default_start
        list_date_limited = True

    # дополнительные умные фильтры
    f_plant = request.args.get('plant_id', type=int)
    f_size = request.args.get('size_id', type=int)
    f_field = request.args.get('field_id', type=int)

    # Текущая вкладка (чтобы фильтры не сбрасывали видимую вкладку)
    # Если параметр передан, но пуст, по умолчанию остаемся на "regrading" (Изменение размера).
    active_tab = request.args.get('active_tab') or 'regrading'
    if active_tab == 'income':
        return redirect(url_for('stock.stock_receipts'))

    doc_list_opts = (
        joinedload(Document.rows).joinedload(DocumentRow.plant),
        joinedload(Document.rows).joinedload(DocumentRow.size),
        joinedload(Document.rows).joinedload(DocumentRow.field_from),
        joinedload(Document.order).joinedload(Order.client),
    )

    # 1. Отгрузки (Shipments)
    ship_q = Document.query.options(*doc_list_opts).filter_by(doc_type='shipment')
    # исключаем документы, связанные с ghost‑заказами — они показываются ниже отдельным блоком
    ship_q = ship_q.outerjoin(Order, Document.order_id == Order.id)
    ship_q = ship_q.filter(or_(Order.id == None, Order.status != 'ghost'))
    if f_start: ship_q = ship_q.filter(func.date(Document.date) >= f_start)
    if f_end: ship_q = ship_q.filter(func.date(Document.date) <= f_end)
    if f_client: ship_q = ship_q.join(Order).filter(Order.client_id == f_client)
    # фильтрация по строкам документа
    if f_plant or f_size or f_field:
        ship_q = ship_q.join(DocumentRow)
        if f_plant: ship_q = ship_q.filter(DocumentRow.plant_id == f_plant)
        if f_size: ship_q = ship_q.filter(DocumentRow.size_id == f_size)
        if f_field: ship_q = ship_q.filter(DocumentRow.field_from_id == f_field)
        ship_q = ship_q.distinct()
    if f_doc_type and f_doc_type != 'shipment':
        shipments = []
    else:
        shipments = ship_q.order_by(Document.date.desc()).all()

    # 2. Исторические (Ghost)
    ghost_q = Order.query.options(
        joinedload(Order.client),
        joinedload(Order.items).joinedload(OrderItem.plant),
        joinedload(Order.items).joinedload(OrderItem.size),
        joinedload(Order.items).joinedload(OrderItem.field),
    ).filter_by(status='ghost')
    if f_start: ghost_q = ghost_q.filter(func.date(Order.date) >= f_start)
    if f_end: ghost_q = ghost_q.filter(func.date(Order.date) <= f_end)
    if f_client: ghost_q = ghost_q.filter(Order.client_id == f_client)
    if f_plant or f_size or f_field:
        ghost_q = ghost_q.join(OrderItem)
        if f_plant: ghost_q = ghost_q.filter(OrderItem.plant_id == f_plant)
        if f_size: ghost_q = ghost_q.filter(OrderItem.size_id == f_size)
        if f_field: ghost_q = ghost_q.filter(OrderItem.field_id == f_field)
        ghost_q = ghost_q.distinct()
    if f_doc_type and f_doc_type != 'ghost':
        ghost_orders = []
    else:
        ghost_orders = ghost_q.order_by(Order.date.desc()).all()

    # Фильтруем показываемые строчки внутри ghost-заказов по выбранным фильтрам
    # (если мы отфильтровали по товару/размеру/полю, то показываем только подходящие позиции)
    if f_plant or f_size or f_field:
        filtered_ghost_orders = []
        for o in ghost_orders:
            filtered_items = []
            for item in o.items:
                if f_plant and item.plant_id != f_plant: continue
                if f_size and item.size_id != f_size: continue
                if f_field and item.field_id != f_field: continue
                filtered_items.append(item)
            if filtered_items:
                o.filtered_items = filtered_items
                filtered_ghost_orders.append(o)
        ghost_orders = filtered_ghost_orders
    else:
        for o in ghost_orders:
            o.filtered_items = o.items

    # 3. Черновики инвентаризации (Drafts)
    draft_q = Document.query.filter_by(doc_type='inventory_draft')
    if f_start: draft_q = draft_q.filter(func.date(Document.date) >= f_start)
    if f_end: draft_q = draft_q.filter(func.date(Document.date) <= f_end)
    # Фильтр по клиенту к инвентаризации не применим, поэтому его тут нет
    if f_doc_type and f_doc_type != 'inventory_draft':
        draft_docs = []
    else:
        draft_docs = draft_q.order_by(Document.date.desc()).all()

    # Сбор данных для черновиков (предзагрузка остатков для быстрого поиска)
    stock_lookup = {}
    if draft_docs:
        for sb in StockBalance.query.all():
            stock_lookup[(sb.plant_id, sb.size_id, sb.field_id, sb.year)] = sb.quantity or 0
    drafts_data = []
    for d in draft_docs:
        rows_data = []
        for r in d.rows:
            db_qty = stock_lookup.get((r.plant_id, r.size_id, r.field_to_id, r.year), 0)
            rows_data.append({
                'row_id': r.id,
                'plant_name': r.plant.name if r.plant else '-',
                'size_name': r.size.name if r.size else '-',
                'fact': r.quantity, 
                'db_qty': db_qty,   
                'delta': r.quantity - db_qty
            })
        drafts_data.append({'doc': d, 'rows': rows_data})

    # 4. Карточки пересчета по полю
    recount_q = Document.query.options(
        joinedload(Document.rows).joinedload(DocumentRow.plant),
        joinedload(Document.rows).joinedload(DocumentRow.size),
        joinedload(Document.rows).joinedload(DocumentRow.field_to),
    ).filter(Document.doc_type.in_(RECOUNT_DOC_TYPES))
    if f_start: recount_q = recount_q.filter(func.date(Document.date) >= f_start)
    if f_end: recount_q = recount_q.filter(func.date(Document.date) <= f_end)
    if f_plant or f_size or f_field:
        recount_q = recount_q.join(DocumentRow)
        if f_plant: recount_q = recount_q.filter(DocumentRow.plant_id == f_plant)
        if f_size: recount_q = recount_q.filter(DocumentRow.size_id == f_size)
        if f_field: recount_q = recount_q.filter(DocumentRow.field_to_id == f_field)
        recount_q = recount_q.distinct()
    if f_doc_type and f_doc_type not in RECOUNT_DOC_TYPES:
        recount_docs = []
    else:
        recount_docs = recount_q.order_by(Document.date.desc(), Document.id.desc()).all()
    recount_cards = [_group_recount_rows(d) for d in recount_docs]

    # Документы саженцев (пересадка/товарность/промер/выпад) — для admin edit/delete
    seedling_docs_q = (
        Document.query
        .options(joinedload(Document.user))
        .filter(Document.comment.ilike('Саженцы:%'))
        .order_by(Document.date.desc(), Document.id.desc())
        .limit(80)
    )
    seedling_docs = seedling_docs_q.all()

    filters = {
        'start': f_start,
        'end': f_end,
        'client_id': f_client,
        'doc_type': f_doc_type,
        'plant_id': f_plant,
        'size_id': f_size,
        'field_id': f_field,
        'show_all': show_all,
        'list_date_limited': list_date_limited,
    }

    resp = make_response(render_template('stock/documents.html',
                           plants=Plant.query.all(),
                           sizes=sorted(Size.query.all(), key=natural_key),
                           fields=sorted(Field.query.all(), key=natural_key),
                           clients=Client.query.all(),
                           shipments=shipments,
                           ghost_orders=ghost_orders,
                           drafts_data=drafts_data,
                           recount_cards=recount_cards,
                           seedling_docs=seedling_docs,
                           current_year=msk_now().year,
                           filters=filters,
                           active_tab=active_tab,
                           current_url=request.url))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp

@bp.route('/order/create_ghost', methods=['GET', 'POST'])
@login_required
def order_create_ghost():
    if current_user.role != 'admin': 
        flash('Только для администратора')
        return redirect(url_for('stock.documents'))

    if request.method == 'POST':
        try:
            client_id = int(request.form.get('client'))
            date_str = request.form.get('date')
            order_date = datetime.strptime(date_str, '%Y-%m-%d') if date_str else msk_now()

            order = Order(client_id=client_id, date=order_date, status='ghost', invoice_number='HISTORY')
            db.session.add(order)
            db.session.commit()

            p_ids = request.form.getlist('plant[]')
            s_ids = request.form.getlist('size[]')
            f_ids = request.form.getlist('field[]')
            y_ids = request.form.getlist('year[]')
            q_ids = request.form.getlist('quantity[]')
            prices = request.form.getlist('price[]')

            for i in range(len(p_ids)):
                try:
                    qty = int(q_ids[i])
                    if qty <= 0: continue
                    price = float(prices[i])
                    item = OrderItem(
                        order_id=order.id, plant_id=int(p_ids[i]), size_id=int(s_ids[i]),
                        field_id=int(f_ids[i]), year=int(y_ids[i]), quantity=qty,
                        shipped_quantity=qty, price=price
                    )
                    db.session.add(item)
                except ValueError: continue

            db.session.commit()
            log_action(f"Создал историческую отгрузку (Ghost) #{order.id}")
            flash(f"Историческая отгрузка #{order.id} создана")
            return redirect(url_for('stock.documents'))

        except Exception as e:
            db.session.rollback()
            flash(f"Ошибка: {e}")

    return render_template('stock/ghost_order_form.html', clients=Client.query.all(), plants=Plant.query.all(), sizes=Size.query.all(), fields=Field.query.all())

@bp.route('/order/delete_ghost/<int:order_id>', methods=['POST'])
@login_required
def delete_ghost_order(order_id):
    if current_user.role != 'admin': return redirect(url_for('stock.documents'))
    o = Order.query.get_or_404(order_id)
    if o.status == 'ghost':
        # прежде чем стереть заказ, удалим связанные документы (например, ручную отгрузку)
        docs = Document.query.filter_by(order_id=o.id).all()
        for d in docs:
            # восстанавливаем остаток только для реальных отгрузок
            if d.doc_type == 'shipment':
                for r in d.rows:
                    # если поле указан, возвращаем на склад
                    if r.field_from_id:
                        st = get_or_create_stock(r.plant_id, r.size_id, r.field_from_id, r.year)
                        st.quantity += r.quantity
            db.session.delete(d)
        db.session.delete(o)
        db.session.commit()
        flash('Историческая запись удалена')
        log_action(f"Удалил историческую отгрузку #{order_id}")
    return redirect(url_for('stock.documents'))

@bp.route('/order/mass_edit_ghost', methods=['POST'])
@login_required
def mass_edit_ghost_orders():
    if current_user.role != 'admin': return redirect(url_for('stock.documents'))
    try:
        new_client_id = request.form.get('new_client_id')
        order_ids = request.form.getlist('order_ids[]')
        
        if not new_client_id or not order_ids:
            flash('Не выбраны заказы или новый клиент')
            return redirect(url_for('stock.documents'))

        count = 0
        for oid in order_ids:
            order = Order.query.filter_by(id=int(oid), status='ghost').first()
            if order:
                order.client_id = int(new_client_id)
                count += 1
        
        db.session.commit()
        if count > 0:
            flash(f'Обновлен клиент у {count} исторических отгрузок')
            log_action(f"Массовая смена клиента у Ghost-заказов (кол-во: {count})")
        else:
            flash('Ничего не обновлено')

    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}')
    return redirect(url_for('stock.documents'))

@bp.route('/order/merge_ghost', methods=['POST'])
@login_required
def merge_ghost_orders():
    if current_user.role != 'admin': return redirect(url_for('stock.documents'))
    
    order_ids = request.form.getlist('order_ids[]')
    
    if not order_ids or len(order_ids) < 2:
        flash('Выберите минимум 2 заказа для объединения')
        return redirect(url_for('stock.documents'))
    
    try:
        # Сортируем ID, чтобы первый (самый старый) стал основным
        sorted_ids = sorted([int(x) for x in order_ids])
        main_order_id = sorted_ids[0]
        other_ids = sorted_ids[1:]
        
        main_order = Order.query.get(main_order_id)
        if not main_order or main_order.status != 'ghost':
            flash('Ошибка: Основной заказ не найден или не является Ghost')
            return redirect(url_for('stock.documents'))
            
        items_moved = 0
        merged_orders_count = 0
        
        for oid in other_ids:
            sub_order = Order.query.get(oid)
            if sub_order and sub_order.status == 'ghost':
                # Переносим позиции товаров в основной заказ
                for item in sub_order.items:
                    item.order_id = main_order.id
                    items_moved += 1
                
                # Если у объединяемых заказов есть номер счета, а у основного нет - берем его
                if not main_order.invoice_number and sub_order.invoice_number:
                    main_order.invoice_number = sub_order.invoice_number
                    main_order.invoice_date = sub_order.invoice_date
                
                # Удаляем пустой заказ-донор
                db.session.delete(sub_order)
                merged_orders_count += 1
                
        db.session.commit()
        flash(f'Успешно! Объединено {merged_orders_count + 1} заказов в заказ #{main_order_id}. Перенесено позиций: {items_moved}.')
        log_action(f"Объединение Ghost-заказов {order_ids} в #{main_order_id}")
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка объединения: {e}')
        
    return redirect(url_for('stock.documents'))

@bp.route('/order/download_ghost_template')
@login_required
def download_ghost_template():
    if current_user.role != 'admin': return redirect(url_for('stock.documents'))
    wb = Workbook()
    ws = wb.active
    ws.append(["Клиент", "Дата (YYYY-MM-DD)", "Растение", "Размер", "Поле (Откуда)", "Год партии", "Кол-во", "Цена продажи"])
    ws.append(["Иванов И.И.", "2023-05-20", "Сосна горная", "C3", "Поле 1", 2020, 10, 1500])
    apply_excel_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='ghost_import_template.xlsx', as_attachment=True)

@bp.route('/order/import_ghost', methods=['POST'])
@login_required
def import_ghost_orders():
    if current_user.role != 'admin': return redirect(url_for('stock.documents'))
    file = request.files.get('file')
    if not file: flash('Файл не выбран'); return redirect(url_for('stock.documents'))
    try:
        wb = load_workbook(file)
        ws = wb.active
        plants_map = {p.name.lower().strip(): p.id for p in Plant.query.all()}
        sizes_map = {s.name.lower().strip(): s.id for s in Size.query.all()}
        fields_map = {f.name.lower().strip(): f.id for f in Field.query.all()}
        orders_batch = {} 
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[2]: continue
            c_name = str(row[0]).strip()
            raw_date = row[1]
            if isinstance(raw_date, datetime): d_date = raw_date
            elif isinstance(raw_date, date): d_date = datetime.combine(raw_date, datetime.min.time())
            else:
                try: d_date = datetime.strptime(str(raw_date), '%Y-%m-%d')
                except: d_date = msk_now()
            
            p_id = plants_map.get(str(row[2]).lower().strip())
            s_id = sizes_map.get(str(row[3]).lower().strip())
            f_id = fields_map.get(str(row[4]).lower().strip())
            
            if p_id and s_id and f_id:
                key = (c_name, d_date)
                if key not in orders_batch: orders_batch[key] = []
                orders_batch[key].append({
                    'plant_id': p_id, 'size_id': s_id, 'field_id': f_id,
                    'year': int(row[5]) if row[5] else 2017,
                    'qty': int(row[6] or 0), 'price': float(row[7] or 0)
                })

        created_count = 0
        for (client_name, order_date), items in orders_batch.items():
            if not items: continue
            client = Client.query.filter(func.lower(Client.name) == client_name.lower()).first()
            if not client:
                client = Client(name=client_name)
                db.session.add(client)
                db.session.flush()
            
            order = Order(client_id=client.id, date=order_date, status='ghost', invoice_number='IMPORT')
            db.session.add(order)
            db.session.flush()
            
            for item in items:
                oi = OrderItem(
                    order_id=order.id, plant_id=item['plant_id'], size_id=item['size_id'],
                    field_id=item['field_id'], year=item['year'], quantity=item['qty'],
                    shipped_quantity=item['qty'], price=item['price']
                )
                db.session.add(oi)
            created_count += 1

        db.session.commit()
        flash(f'Успешно импортировано {created_count} исторических заказов')
        log_action(f"Импорт Ghost-заказов ({created_count} шт.)")
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка импорта: {e}')
    return redirect(url_for('stock.documents'))

@bp.route('/document/delete/<int:doc_id>', methods=['POST'])
@login_required
def document_delete(doc_id):
    if current_user.role != 'admin': return redirect(url_for('stock.logs'))
    return_to = request.args.get('return_to') or request.form.get('return_to')
    if return_to == 'None': return_to = None
    doc = Document.query.get_or_404(doc_id)
    try:
        for row in doc.rows:
            if doc.doc_type in ['income', 'correction']:
                if row.field_to_id:
                    get_or_create_stock(row.plant_id, row.size_id, row.field_to_id, row.year).quantity -= row.quantity
            elif _is_recount_doc(doc.doc_type):
                if row.field_to_id:
                    get_or_create_stock(row.plant_id, row.size_id, row.field_to_id, row.year).quantity -= row.quantity
            elif doc.doc_type in ['writeoff', 'shipment']:
                get_or_create_stock(row.plant_id, row.size_id, row.field_from_id, row.year).quantity += row.quantity
            elif doc.doc_type == 'move':
                get_or_create_stock(row.plant_id, row.size_id, row.field_from_id, row.year).quantity += row.quantity
                get_or_create_stock(row.plant_id, row.size_id, row.field_to_id, row.year).quantity -= row.quantity
            elif doc.doc_type == 'regrading':
                dst_field = row.field_to_id or row.field_from_id
                get_or_create_stock(row.plant_id, row.size_id, row.field_from_id, row.year).quantity += row.quantity
                get_or_create_stock(row.plant_id, row.size_to_id, dst_field, row.year).quantity -= row.quantity

            if doc.doc_type == 'shipment' and doc.order_id:
                oi = OrderItem.query.filter_by(order_id=doc.order_id, plant_id=row.plant_id, size_id=row.size_id, field_id=row.field_from_id, year=row.year).first()
                if oi: oi.shipped_quantity = max(0, oi.shipped_quantity - row.quantity)

        if doc.doc_type == 'potting_recount' and doc.project_id:
            ProjectPottingRecountLine.query.filter_by(project_id=doc.project_id).delete()

        # FK: журнал саженцев и партии закупки ссылаются на document
        SeedlingEventLog.query.filter_by(document_id=doc.id).delete(synchronize_session=False)
        StockPurchaseLot.query.filter_by(document_id=doc.id).update(
            {StockPurchaseLot.document_id: None, StockPurchaseLot.document_row_id: None},
            synchronize_session=False,
        )

        db.session.delete(doc)
        db.session.commit()
        flash('Документ удален, остатки пересчитаны')
        log_action(f"Удалил документ #{doc.id}")
    except Exception as e: 
        db.session.rollback()
        flash(f'Ошибка удаления: {e}')

    if return_to:
        return redirect(return_to)
    return redirect(url_for('stock.documents', active_tab='shipments'))

@bp.route('/document/edit/<int:doc_id>', methods=['GET', 'POST'])
@login_required
def document_edit(doc_id):
    if current_user.role != 'admin': 
        flash('Только админ')
        return redirect(url_for('stock.logs'))
    doc = Document.query.get_or_404(doc_id)
    
    if request.method == 'POST':
        return_to = request.form.get('return_to') or request.args.get('return_to')
        if return_to == 'None': return_to = None
        try:
            for row in doc.rows:
                if doc.doc_type in ['income', 'correction']:
                    if row.field_to_id: get_or_create_stock(row.plant_id, row.size_id, row.field_to_id, row.year).quantity -= row.quantity
                elif _is_recount_doc(doc.doc_type):
                    if row.field_to_id: get_or_create_stock(row.plant_id, row.size_id, row.field_to_id, row.year).quantity -= row.quantity
                elif doc.doc_type == 'writeoff':
                    if row.field_from_id: get_or_create_stock(row.plant_id, row.size_id, row.field_from_id, row.year).quantity += row.quantity
                elif doc.doc_type == 'move':
                    if row.field_from_id: get_or_create_stock(row.plant_id, row.size_id, row.field_from_id, row.year).quantity += row.quantity
                    if row.field_to_id: get_or_create_stock(row.plant_id, row.size_id, row.field_to_id, row.year).quantity -= row.quantity
                elif doc.doc_type == 'regrading':
                    dst_field = row.field_to_id or row.field_from_id
                    if row.field_from_id: get_or_create_stock(row.plant_id, row.size_id, row.field_from_id, row.year).quantity += row.quantity
                    if row.field_from_id and row.size_to_id: get_or_create_stock(row.plant_id, row.size_to_id, dst_field, row.year).quantity -= row.quantity
                elif doc.doc_type == 'shipment':
                    if row.field_from_id: get_or_create_stock(row.plant_id, row.size_id, row.field_from_id, row.year).quantity += row.quantity
                    if doc.order_id:
                        oi = OrderItem.query.filter_by(order_id=doc.order_id, plant_id=row.plant_id, size_id=row.size_id, field_id=row.field_from_id, year=row.year).first()
                        if oi: oi.shipped_quantity = max(0, oi.shipped_quantity - row.quantity)
            
            for row in doc.rows: db.session.delete(row)
            
            new_date_str = request.form.get('date')
            if new_date_str: doc.date = datetime.strptime(new_date_str, '%Y-%m-%d')
            
            p_ids = request.form.getlist('plant[]')
            s_ids = request.form.getlist('size[]')
            q_ids = request.form.getlist('quantity[]')
            y_ids = request.form.getlist('year[]')
            
            f_to_ids = request.form.getlist('field_to[]')
            f_from_ids = request.form.getlist('field_from[]')
            op_types = request.form.getlist('operation_type[]')

            for i in range(len(p_ids)):
                try: 
                    if i >= len(s_ids) or i >= len(q_ids) or i >= len(y_ids): continue
                    
                    qty = int(q_ids[i])
                    if not _is_recount_doc(doc.doc_type) and qty <= 0:
                        continue
                    if _is_recount_doc(doc.doc_type) and qty == 0:
                        continue
                    
                    pid = int(p_ids[i])
                    sid = int(s_ids[i])
                    year = int(y_ids[i])
                    
                    new_row = DocumentRow(document_id=doc.id, plant_id=pid, size_id=sid, year=year, quantity=qty)
                    
                    if doc.doc_type in ['income', 'correction']:
                        if i < len(f_to_ids):
                            ft = int(f_to_ids[i])
                            new_row.field_to_id = ft
                            get_or_create_stock(pid, sid, ft, year).quantity += qty
                        else:
                            continue 

                    elif doc.doc_type == 'writeoff':
                        if i < len(f_from_ids):
                            ff = int(f_from_ids[i])
                            new_row.field_from_id = ff
                            get_or_create_stock(pid, sid, ff, year).quantity -= qty
                        else: continue

                    elif doc.doc_type == 'move':
                        if i < len(f_from_ids) and i < len(f_to_ids):
                            ff = int(f_from_ids[i]); ft = int(f_to_ids[i])
                            new_row.field_from_id = ff; new_row.field_to_id = ft
                            get_or_create_stock(pid, sid, ff, year).quantity -= qty
                            get_or_create_stock(pid, sid, ft, year).quantity += qty
                        else: continue

                    elif doc.doc_type == 'regrading':
                        if i < len(f_from_ids) and i < len(f_to_ids):
                            ff = int(f_from_ids[i]); s_to = int(f_to_ids[i])
                            new_row.field_from_id = ff; new_row.size_to_id = s_to
                            get_or_create_stock(pid, sid, ff, year).quantity -= qty
                            get_or_create_stock(pid, s_to, ff, year).quantity += qty
                        else: continue

                    elif doc.doc_type == 'shipment':
                        if i < len(f_from_ids):
                            ff = int(f_from_ids[i])
                            new_row.field_from_id = ff
                            get_or_create_stock(pid, sid, ff, year).quantity -= qty
                            if doc.order_id:
                                oi = OrderItem.query.filter_by(order_id=doc.order_id, plant_id=pid, size_id=sid, field_id=ff, year=year).first()
                                if oi: oi.shipped_quantity += qty
                        else: continue

                    elif _is_recount_doc(doc.doc_type):
                        if i < len(f_to_ids):
                            ft = int(f_to_ids[i])
                            op = op_types[i] if i < len(op_types) else 'income'
                            signed_qty = qty if op == 'income' else -qty
                            if signed_qty == 0:
                                continue
                            new_row.field_to_id = ft
                            new_row.quantity = signed_qty
                            get_or_create_stock(pid, sid, ft, year).quantity += signed_qty
                        else:
                            continue
                        
                    db.session.add(new_row)
                    
                except (ValueError, IndexError) as row_err:
                    print(f"Row error: {row_err}")
                    continue 

            db.session.commit()
            flash('Документ обновлен')
            log_action(f"Отредактировал документ #{doc.id}")
            if return_to:
                return redirect(return_to)
            return redirect(url_for('stock.documents', active_tab='shipments'))
            
        except Exception as e: 
            db.session.rollback()
            flash(f'Ошибка обновления: {e}')
            if return_to:
                return redirect(return_to)
            return redirect(url_for('stock.document_edit', doc_id=doc.id))

    return render_template(
        'stock/document_edit.html',
        doc=doc,
        plants=Plant.query.all(),
        sizes=Size.query.all(),
        fields=Field.query.all(),
        return_to=request.args.get('return_to')
    )

@bp.route('/admin/recalc_stock_balances')
@login_required
def recalc_stock_balances():
    if current_user.role != 'admin': 
        return redirect(url_for('main.index'))
    
    try:
        StockBalance.query.update({StockBalance.quantity: 0})
        all_docs = Document.query.order_by(Document.date.asc(), Document.id.asc()).all()
        doc_count = 0
        
        for doc in all_docs:
            doc_count += 1
            for row in doc.rows:
                pid, sid, yr, qty = row.plant_id, row.size_id, row.year, row.quantity
                
                if doc.doc_type in ['income', 'correction']:
                    if row.field_to_id:
                        get_or_create_stock(pid, sid, row.field_to_id, yr).quantity += qty
                
                elif _is_recount_doc(doc.doc_type):
                    if row.field_to_id:
                        get_or_create_stock(pid, sid, row.field_to_id, yr).quantity += qty
                
                elif doc.doc_type in ['writeoff', 'shipment']:
                    if row.field_from_id:
                        get_or_create_stock(pid, sid, row.field_from_id, yr).quantity -= qty
                
                elif doc.doc_type == 'move':
                    if row.field_from_id and row.field_to_id:
                        get_or_create_stock(pid, sid, row.field_from_id, yr).quantity -= qty
                        get_or_create_stock(pid, sid, row.field_to_id, yr).quantity += qty
                
                elif doc.doc_type == 'regrading':
                    if row.field_from_id and row.size_to_id:
                        dst_field = row.field_to_id or row.field_from_id
                        get_or_create_stock(pid, sid, row.field_from_id, yr).quantity -= qty
                        get_or_create_stock(pid, row.size_to_id, dst_field, yr).quantity += qty

        db.session.commit()
        flash(f'Успешно! Пересчитано {doc_count} документов. Остатки синхронизированы.')
        log_action("Запустил полный пересчет остатков (Repair DB)")
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка пересчета: {e}')
        
    return redirect(url_for('stock.stock_report'))

@bp.route('/logs')
@login_required
def logs():
    resp = make_response(render_template('stock/logs.html',
                           logs=ActionLog.query.order_by(ActionLog.date.desc()).limit(500).all(),
                           documents=Document.query.order_by(Document.date.desc()).limit(500).all()))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp

@bp.route('/changelog', methods=['GET', 'POST'])
@login_required
def changelog():
    # Импорт внутри функции, чтобы избежать циклических ссылок, если они возникнут
    from app.models import ChangeLog 
    
    if request.method == 'POST':
        if current_user.role != 'admin':
            flash('Только админ может создавать релизы')
            return redirect(url_for('stock.changelog'))
            
        version = request.form.get('version')
        changes_text = request.form.get('changes')
        
        if version and changes_text:
            # Сохраняем как есть, переносы строк обработаем в шаблоне
            log = ChangeLog(
                version=version,
                content=changes_text,
                date=msk_now().date() # Автоматическая дата
            )
            db.session.add(log)
            db.session.commit()
            log_action(f"Опубликовал версию системы {version}")
            flash(f'Версия {version} опубликована!')
            
        elif request.form.get('action') == 'delete':
            # Логика удаления (на всякий случай)
            log_id = request.form.get('log_id')
            ChangeLog.query.filter_by(id=log_id).delete()
            db.session.commit()
            flash('Запись удалена')
            
        return redirect(url_for('stock.changelog'))

    logs = ChangeLog.query.order_by(ChangeLog.date.desc(), ChangeLog.id.desc()).all()
    return render_template('stock/changelog.html', logs=logs, today=msk_now().date())

@bp.route('/price_history/download_template')
@login_required
def price_history_download_template():
    if current_user.role != 'admin':
        return redirect(url_for('stock.price_history'))

    include_zero = request.args.get('all') == '1'

    hist_map = {
        (h.plant_id, h.size_id, h.field_id, h.year): float(h.price or 0)
        for h in PriceHistory.query.all()
    }

    q = (
        StockBalance.query
        .options(
            joinedload(StockBalance.plant),
            joinedload(StockBalance.size),
            joinedload(StockBalance.field),
        )
    )
    if not include_zero:
        q = q.filter(StockBalance.quantity > 0)

    stocks = q.all()
    stocks.sort(
        key=lambda sb: (
            (sb.plant.name if sb.plant else '').lower(),
            natural_key(sb.size.name if sb.size else ''),
            -(sb.year or 0),
            natural_key(sb.field.name if sb.field else ''),
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Цены из остатков"
    headers = ["Растение", "Размер", "Поле", "Год (Партия)", "Цена продажи (Руб)", "Остаток (шт)"]
    ws.append(headers)

    for sb in stocks:
        key = (sb.plant_id, sb.size_id, sb.field_id, sb.year)
        price = hist_map.get(key)
        if price is None:
            price = float(sb.price or 0)
        ws.append([
            sb.plant.name if sb.plant else '',
            sb.size.name if sb.size else '',
            sb.field.name if sb.field else '',
            sb.year,
            price,
            int(sb.quantity or 0),
        ])

    if not stocks:
        ws.append([
            "Туя западная Smaragd", "C3", "Поле 1", msk_now().year, 500, 0,
        ])

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14

    try:
        apply_excel_styles(ws)
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = 'all' if include_zero else 'stock'
    fname = f'price_history_{suffix}_{msk_now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, download_name=fname, as_attachment=True)

@bp.route('/price_history/import', methods=['POST'])
@login_required
def price_history_import():
    if current_user.role != 'admin': return redirect(url_for('stock.price_history'))
    file = request.files.get('file')
    if not file: return redirect(url_for('stock.price_history'))
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        plants_map = {p.name.lower().strip(): p.id for p in Plant.query.all()}
        sizes_map = {s.name.lower().strip(): s.id for s in Size.query.all()}
        fields_map = {f.name.lower().strip(): f.id for f in Field.query.all()}
        
        count = 0
        updated = 0
        bulk_count = 0
        bulk_seen = set()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            try:
                price = float(row[4]) if len(row) > 4 and row[4] is not None else None
            except (TypeError, ValueError):
                continue
            if price is None:
                continue

            p_name = str(row[0]).lower().strip()
            s_name = str(row[1]).lower().strip()
            pid = plants_map.get(p_name)
            sid = sizes_map.get(s_name)
            if not pid or not sid:
                continue

            f_cell = row[2] if len(row) > 2 else None
            y_cell = row[3] if len(row) > 3 else None
            f_name = str(f_cell).lower().strip() if f_cell not in (None, '') else ''
            if not f_name:
                bulk_key = (pid, sid)
                if bulk_key in bulk_seen:
                    continue
                bulk_seen.add(bulk_key)
                result = apply_price_for_plant_size(pid, sid, price, source='excel')
                if result.get('ok'):
                    bulk_count += 1
                    updated += result.get('hist_updated', 0)
                    count += result.get('hist_created', 0)
                continue

            fid = fields_map.get(f_name)
            try:
                year = int(y_cell)
            except (TypeError, ValueError):
                year = msk_now().year

            if fid:
                hist = PriceHistory.query.filter_by(
                    plant_id=pid, size_id=sid, field_id=fid, year=year
                ).first()
                old_price = hist.price if hist else None
                if hist:
                    hist.price = price
                    updated += 1
                else:
                    sb0 = StockBalance.query.filter_by(
                        plant_id=pid, size_id=sid, field_id=fid, year=year,
                    ).first()
                    old_price = sb0.price if sb0 else None
                    db.session.add(PriceHistory(
                        plant_id=pid, size_id=sid, field_id=fid, year=year, price=price
                    ))
                    count += 1
                _append_price_change_log(
                    plant_id=pid, size_id=sid,
                    old_price=old_price, new_price=price,
                    field_id=fid, year=year, source='excel',
                )
                sb = StockBalance.query.filter_by(
                    plant_id=pid, size_id=sid, field_id=fid, year=year,
                ).first()
                if sb:
                    sb.price = price

        db.session.commit()
        parts = []
        if bulk_count:
            parts.append(f'позиций целиком: {bulk_count}')
        if count:
            parts.append(f'добавлено: {count}')
        if updated:
            parts.append(f'обновлено: {updated}')
        flash('Импорт: ' + (', '.join(parts) if parts else 'нет изменений'))
        log_action(f"Импорт истории цен: bulk={bulk_count}, +{count}, up={updated}")
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка импорта: {e}')
        
    return redirect(url_for('stock.price_history'))

@bp.route('/document/approve_inventory/<int:doc_id>', methods=['POST'])
@login_required
def approve_inventory(doc_id):
    """Проведение черновика инвентаризации в чистовой документ"""
    # Доступ разрешен admin и user (менеджер). user2 (бригадир) - запрещен.
    if current_user.role == 'user2':
        return redirect(url_for('stock.documents'))
        
    doc = Document.query.get_or_404(doc_id)
    if doc.doc_type != 'inventory_draft':
        return redirect(url_for('stock.documents'))

    try:
        # 1. Обновляем факты из формы (вдруг админ/менеджер поправил руками)
        for row in doc.rows:
            fact_val = request.form.get(f'fact_{row.id}')
            if fact_val is not None:
                row.quantity = int(fact_val)

        # 2. ВАЖНО: проверяем, что факт не меньше уже зарезервированного по
        # активным заказам. Иначе резерв станет «больше факта» — оверсейл,
        # ошибки в свободном остатке, разъезд при отгрузках. Эта же проверка
        # уже стоит в карточке пересчёта (`field_recount`), теперь и тут.
        from app.stock_helpers import compute_reserved
        problems = []
        for row in doc.rows:
            reserved = compute_reserved(
                row.plant_id, row.size_id, row.field_to_id, row.year
            )
            try:
                fact_qty = int(row.quantity or 0)
            except Exception:
                fact_qty = 0
            if fact_qty < reserved:
                pl = Plant.query.get(row.plant_id)
                sz = Size.query.get(row.size_id)
                fld = Field.query.get(row.field_to_id) if row.field_to_id else None
                name_parts = [
                    (pl.name if pl else f'#{row.plant_id}'),
                    (sz.name if sz else f'#{row.size_id}'),
                    (fld.name if fld else ''),
                    str(row.year),
                ]
                name = ' · '.join([p for p in name_parts if p])
                problems.append(
                    f'«{name}» — факт {fact_qty} шт меньше резерва {reserved} шт'
                )
        if problems:
            db.session.rollback()
            flash(
                'Инвентаризация не проведена: введённый факт меньше уже '
                'зарезервированного по активным заказам. Сначала отмените '
                'или скорректируйте заказы.\n• ' + '\n• '.join(problems[:10]),
                'danger'
            )
            return redirect(url_for('stock.documents'))

        # 3. Применяем к складу и превращаем факт в дельту для архива
        for row in doc.rows:
            st = get_or_create_stock(row.plant_id, row.size_id, row.field_to_id, row.year)
            
            delta = row.quantity - st.quantity # Вычисляем излишек/недостачу
            
            st.quantity = row.quantity # Записываем на склад ФАКТ
            row.quantity = delta       # В документе оставляем ДЕЛЬТУ (для истории)

        # 4. Меняем статус документа на "Проведен"
        doc.doc_type = 'correction'
        doc.comment = str(doc.comment).replace("Черновик", "Проведена")
        
        db.session.commit()
        flash('Инвентаризация успешно проведена! Остатки обновлены.')
        log_action(f"Провел инвентаризацию #{doc.id}")
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка проведения: {e}')
        
    return redirect(url_for('stock.documents'))

@bp.route('/api/inventory/get_years')
@login_required
def api_inventory_get_years():
    """Отдает список партий (годов) для конкретного растения на поле"""
    p_id = request.args.get('plant_id', type=int)
    f_id = request.args.get('field_id', type=int)
    
    # Ищем уникальные года, где остаток больше 0
    stocks = StockBalance.query.filter_by(plant_id=p_id, field_id=f_id).filter(StockBalance.quantity > 0).all()
    
    # Собираем уникальные года и сортируем (свежие сверху)
    years = sorted(list(set([st.year for st in stocks])), reverse=True)
    return jsonify({'years': years})

@bp.route('/api/inventory/get_sizes')
@login_required
def api_inventory_get_sizes():
    """Отдает список размеров, которые числятся на поле для конкретного растения"""
    p_id = request.args.get('plant_id', type=int)
    f_id = request.args.get('field_id', type=int)
    y = request.args.get('year', type=int)
    
    # Ищем только те, где остаток больше 0
    stocks = StockBalance.query.filter_by(plant_id=p_id, field_id=f_id, year=y).filter(StockBalance.quantity > 0).all()
    result =[{'id': st.size_id, 'name': st.size.name, 'db_qty': st.quantity} for st in stocks]
    
    return jsonify({'sizes': result})

@bp.route('/inventory/mobile', methods=['GET'])
@login_required
def inventory_mobile():
    plants = Plant.query.order_by(Plant.name).all()
    fields = sorted(Field.query.all(), key=natural_key)
    sizes = sorted(Size.query.all(), key=natural_key)
    return render_template('stock/inventory_mobile.html', plants=plants, fields=fields, sizes=sizes, current_year=msk_now().year)

@bp.route('/api/inventory/save', methods=['POST'])
@login_required
def api_inventory_save():
    """Сохранение инвентаризации как ЧЕРНОВИКА (без изменения склада)"""
    data = request.json
    field_id = data.get('field_id')
    plant_id = data.get('plant_id')
    year = data.get('year')
    counts = data.get('counts') # Словарь: {size_id: quantity}

    try:
        # Создаем документ ЧЕРНОВИКА
        doc = Document(
            doc_type='inventory_draft', 
            user_id=current_user.id, 
            date=msk_now(), 
            comment=f"Черновик инвентаризации: Поле {field_id}, Растение {plant_id}"
        )
        db.session.add(doc)
        db.session.flush()

        for size_id_str, new_qty in counts.items():
            # Записываем в документ то, что накликал менеджер (ФАКТ)
            # ВАЖНО: Мы пока НЕ трогаем склад (get_or_create_stock)
            db.session.add(DocumentRow(
                document_id=doc.id, plant_id=plant_id, size_id=int(size_id_str),
                field_to_id=field_id, year=year, quantity=int(new_qty)
            ))

        db.session.commit()
        return jsonify({'status': 'ok', 'doc_id': doc.id})

    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500