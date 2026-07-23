import io
import re
import os
import calendar
from datetime import datetime, date, timedelta
from decimal import Decimal
from werkzeug.utils import secure_filename
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, current_app, jsonify, send_from_directory
from flask_login import login_required, current_user
from sqlalchemy import func, or_, and_
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side 
from openpyxl.utils import get_column_letter

from app.models import (
    db, Expense, BudgetItem, BudgetPlan, CashflowPlan, Employee, UnitCostOverride, 
    Plant, StockBalance, Order, OrderItem, Field, Payment, Document, DocumentRow, Size,
    Client, PaymentInvoice, Project, ProjectItem, ProjectPottingLog, ProjectPottingDayMeta,
    ProjectPottingRecountLine, ProjectBudget, ChatExpenseMessage, AppSetting,
)
from app.utils import (
    msk_now, msk_today, log_action, natural_key, create_pdf_response, MONTH_NAMES,
    timelog_date_str, timesheet_workers_count_by_date, timesheet_workers_count_on_date,
    get_or_create_stock,
)
from app.services import (
    calculate_cost_data, calculate_container_cost_data, calculate_investor_debt,
    get_detailed_stock_at_year_end, get_bed_field_ids, get_container_field_ids,
    get_cost_container_project_ids, save_cost_container_project_ids,
    build_container_cost_table_rows,
)

bp = Blueprint('finance', __name__)

@bp.route('/expenses', methods=['GET', 'POST'])
@login_required
def expenses():
    if current_user.role not in ['admin', 'executive']: 
        return redirect(url_for('orders.orders_list'))
    
    # Текущий таб (расходы / счета)
    tab = request.args.get('tab', 'expenses')

    # Общие фильтры (для вкладки расходов)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    f_item = request.args.get('filter_item')
    f_type = request.args.get('filter_type')
    export = request.args.get('export')

    # Подготовка данных из счета для быстрого создания расхода
    invoice_id = request.args.get('invoice_id')
    prefill_invoice = None
    if invoice_id:
        prefill_invoice = PaymentInvoice.query.get(invoice_id)

    # Если открыли вкладку "Счета" — подготавливаем список счетов
    invoices = []
    invoice_summary = None
    if tab == 'invoices':
        invoices = PaymentInvoice.query.filter(PaymentInvoice.status != 'paid').order_by(
            PaymentInvoice.due_date.asc(),
            PaymentInvoice.priority.desc()
        ).all()
        invoice_summary = {
            'high': 0, 'normal': 0, 'low': 0, 'total': 0,
            'count_high': 0, 'count_normal': 0, 'count_low': 0, 'count_total': len(invoices)
        }
        # Pre-aggregate paid sums for all invoices in one query
        inv_paid_agg = db.session.query(
            Expense.invoice_id, func.sum(Expense.amount)
        ).filter(Expense.invoice_id.isnot(None)).group_by(Expense.invoice_id).all()
        inv_paid_map = {r[0]: r[1] or Decimal(0) for r in inv_paid_agg}

        for inv in invoices:
            if isinstance(inv.due_date, str):
                try:
                    inv.due_date = datetime.strptime(inv.due_date, '%Y-%m-%d').date()
                except Exception:
                    pass
            elif isinstance(inv.due_date, datetime):
                inv.due_date = inv.due_date.date()

            paid_sum = inv_paid_map.get(inv.id, Decimal(0))
            inv.paid_sum = float(paid_sum)
            inv.remaining = float(inv.amount - paid_sum)

            # Сводная статистика: считаем остаток к оплачиваению (не всю сумму счета)
            val = inv.remaining
            p = inv.priority
            invoice_summary[p] += val
            invoice_summary[f'count_{p}'] += 1
            invoice_summary['total'] += val

    # Если открыта вкладка "Расходы" — подготавливаем журнал расходов
    expenses_list = []
    if tab == 'expenses':
        query = Expense.query
        if start_date:
            query = query.filter(func.date(Expense.date) >= start_date)
        if end_date:
            query = query.filter(func.date(Expense.date) <= end_date)
        if f_item:
            query = query.filter(Expense.budget_item_id == int(f_item))
        if f_type:
            query = query.filter(Expense.payment_type == f_type)
        
        from sqlalchemy.orm import joinedload
        expenses_list = (
            query.options(
                joinedload(Expense.item),
                joinedload(Expense.employee),
                joinedload(Expense.project_link),
            )
            .order_by(Expense.date.desc())
            .all()
        )
    
    if export == 'pdf' and tab == 'expenses':
        font_path = os.path.join(current_app.root_path, 'static', 'DejaVuSans.ttf').replace('\\', '/')
        rendered = render_template('expenses_pdf.html', expenses=expenses_list, font_path=font_path)
        return create_pdf_response(rendered, "expenses.pdf")
        
    if request.method == 'POST':
        if current_user.role == 'executive':
            flash('У вас права только на просмотр')
            return redirect(url_for('finance.expenses', tab='expenses'))
        
        # 1. Удаление обычного расхода
        if 'delete_expense' in request.form:
            try:
                exp_id = int(request.form.get('expense_id') or 0)
            except (TypeError, ValueError):
                exp_id = 0
            exp = Expense.query.get(exp_id) if exp_id else None
            if exp is None:
                flash('Расход не найден')
            else:
                try:
                    # Если расход был импортирован из ТГ-чата, на него ссылается
                    # ChatExpenseMessage.expense_id (FK). Отвязываем такие карточки
                    # и возвращаем их в очередь админа (status='pending'), чтобы
                    # их можно было повторно проимпортировать или отклонить.
                    linked = ChatExpenseMessage.query.filter_by(expense_id=exp.id).all()
                    for row in linked:
                        row.expense_id = None
                        if row.status == 'imported':
                            row.status = 'pending'
                    db.session.delete(exp)
                    db.session.commit()
                    flash('Расход удален')
                    log_action(f"Удалил расход #{exp.id}")
                except Exception as e:
                    db.session.rollback()
                    current_app.logger.exception('expenses: delete failed')
                    flash(f'Не удалось удалить расход: {e}')
            return redirect(url_for('finance.expenses', tab='expenses'))
            
        # 2. НОВЫЙ БЛОК: Удаление или пометка счета как "Оплачен"
        elif request.form.get('action') in ['delete', 'mark_paid']:
            inv_id = request.form.get('id')
            inv = PaymentInvoice.query.get(inv_id)
            if inv:
                action = request.form.get('action')
                if action == 'delete':
                    # Удаляем файл с диска
                    inv_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'invoices')
                    file_path = os.path.join(inv_dir, inv.filename)
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    db.session.delete(inv)
                    flash('Счет полностью удален')
                    log_action(f"Удалил счет: {inv.original_name}")
                else:
                    inv.status = 'paid'
                    flash('Счет помечен как оплаченный')
                    log_action(f"Пометил счет как оплаченный: {inv.original_name}")
                    try:
                        from app.vium_inbox import maybe_enqueue as _vium_enqueue
                        _vium_enqueue(inv)
                    except Exception:
                        current_app.logger.exception('vium_inbox.maybe_enqueue (mark_paid) failed')
                db.session.commit()
            return redirect(url_for('finance.expenses', tab='invoices'))

        elif 'due_date' in request.form or 'add_invoice' in request.form:
            # === ЭТО ЛОГИКА ЗАГРУЗКИ СЧЕТА НА ОПЛАТУ ===
            try:
                file = request.files.get('file')
                if not file or not file.filename:
                    flash('Ошибка: Файл счета не выбран')
                    return redirect(url_for('finance.expenses', tab='invoices'))

                filename = secure_filename(file.filename)
                save_name = f"inv_{int(msk_now().timestamp())}_{filename}"
                
                inv_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'invoices')
                if not os.path.exists(inv_dir):
                    os.makedirs(inv_dir)
                    
                file.save(os.path.join(inv_dir, save_name))
                
                due_date_str = request.form.get('due_date')
                d_val = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else None
                
                b_id = request.form.get('budget_item_id') or request.form.get('budget_item')
                
                vim = (request.form.get('vium_intake_mode') or 'auto').strip().lower()
                if vim not in ('auto', 'force', 'skip'):
                    vim = 'auto'
                inv = PaymentInvoice(
                    filename=save_name,
                    original_name=file.filename,
                    budget_item_id=int(b_id) if b_id else None,
                    amount=float(request.form.get('amount') or 0),
                    due_date=d_val,
                    priority=request.form.get('priority', 'normal'),
                    comment=request.form.get('comment', ''),
                    vium_intake_mode=(vim if vim != 'auto' else None),
                )
                db.session.add(inv)
                db.session.commit()
                flash('Счет успешно загружен')
                log_action('Загрузил новый счет на оплату')
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка при загрузке счета: {e}')
            return redirect(url_for('finance.expenses', tab='invoices'))
            
        elif 'edit_expense' in request.form:
            # === РЕДАКТИРОВАНИЕ СУЩЕСТВУЮЩЕГО РАСХОДА (только admin) ===
            if current_user.role != 'admin':
                flash('Редактирование расходов доступно только администратору')
                return redirect(url_for('finance.expenses', tab='expenses'))

            try:
                exp_id = int(request.form.get('expense_id') or 0)
            except (TypeError, ValueError):
                exp_id = 0
            exp = Expense.query.get(exp_id) if exp_id else None
            if exp is None:
                flash('Расход не найден')
                return redirect(url_for('finance.expenses', tab='expenses'))

            try:
                date_str = (request.form.get('date') or '').strip()
                if not date_str:
                    flash('Ошибка: Укажите дату расхода!')
                    return redirect(url_for('finance.expenses', tab='expenses'))
                exp.date = datetime.strptime(date_str, '%Y-%m-%d')

                bi = request.form.get('budget_item')
                if bi:
                    exp.budget_item_id = int(bi)
                exp.description = request.form.get('description') or exp.description
                amount_raw = request.form.get('amount')
                if amount_raw not in (None, ''):
                    exp.amount = float(amount_raw)
                ptype = request.form.get('payment_type')
                if ptype:
                    exp.payment_type = ptype

                def _opt_int(name):
                    v = request.form.get(name)
                    if v in (None, ''):
                        return None
                    try:
                        return int(v)
                    except (TypeError, ValueError):
                        return None

                exp.employee_id = _opt_int('employee_id')
                exp.order_id = _opt_int('order_id')
                exp.project_id = _opt_int('project_id')
                exp.project_budget_id = _opt_int('project_budget_id')
                exp.barter_order_id = _opt_int('barter_order_id')
                exp.target_month = _opt_int('target_month')
                exp.target_year = _opt_int('target_year')

                inv_id_raw = request.form.get('invoice_id')
                old_invoice_id = exp.invoice_id
                if inv_id_raw in (None, ''):
                    exp.invoice_id = None
                else:
                    try:
                        exp.invoice_id = int(inv_id_raw)
                    except (TypeError, ValueError):
                        exp.invoice_id = None

                db.session.commit()

                # Пересчёт статуса связанных счетов (старый и новый, если менялся)
                touched_invoice_ids = {iid for iid in (old_invoice_id, exp.invoice_id) if iid}
                paid_invoices = []
                for iid in touched_invoice_ids:
                    inv = PaymentInvoice.query.get(iid)
                    if not inv:
                        continue
                    paid_sum = db.session.query(func.sum(Expense.amount)).filter(
                        Expense.invoice_id == inv.id
                    ).scalar() or Decimal(0)
                    if paid_sum >= inv.amount:
                        inv.status = 'paid'
                        paid_invoices.append(inv)
                    elif paid_sum > 0:
                        inv.status = 'partial'
                        paid_invoices.append(inv)
                    else:
                        inv.status = 'new'
                if touched_invoice_ids:
                    try:
                        from app.vium_inbox import maybe_enqueue as _vium_enqueue
                        for inv in paid_invoices:
                            _vium_enqueue(inv, exp)
                    except Exception:
                        current_app.logger.exception('vium_inbox.maybe_enqueue (edit) failed')
                    db.session.commit()

                flash('Расход обновлен')
                log_action(f"Отредактировал расход #{exp.id}")
            except Exception as e:
                db.session.rollback()
                current_app.logger.exception('expenses: edit failed')
                flash(f'Ошибка при сохранении расхода: {e}')
            return redirect(url_for('finance.expenses', tab='expenses'))

        else:
            # === ЭТО ЛОГИКА СОЗДАНИЯ ОБЫЧНОГО РАСХОДА ===
            try:
                date_str = request.form.get('date')
                if not date_str:
                    flash('Ошибка: Укажите дату расхода!')
                    return redirect(url_for('finance.expenses', tab='expenses'))
                    
                date_val = datetime.strptime(date_str, '%Y-%m-%d')
                
                # Получаем ID (если пустая строка, будет None)
                oid = request.form.get('order_id')
                pid = request.form.get('project_id')
                pbid = request.form.get('project_budget_id') # Плановая статья проекта
                inv_id = request.form.get('invoice_id')
                barter_oid = request.form.get('barter_order_id')
                
                t_month = request.form.get('target_month')
                t_year = request.form.get('target_year')

                exp = Expense(
                    date=date_val, 
                    budget_item_id=int(request.form.get('budget_item')), 
                    description=request.form.get('description'), 
                    amount=float(request.form.get('amount')), 
                    payment_type=request.form.get('payment_type'), 
                    employee_id=(int(request.form.get('employee_id')) if request.form.get('employee_id') else None),
                    order_id=(int(oid) if oid else None),
                    project_id=(int(pid) if pid else None),
                    project_budget_id=(int(pbid) if pbid else None),
                    barter_order_id=(int(barter_oid) if barter_oid else None),
                    target_month=(int(t_month) if t_month else None),
                    target_year=(int(t_year) if t_year else None)
                )

                if inv_id:
                    inv = PaymentInvoice.query.get(int(inv_id))
                    if inv:
                        exp.invoice_id = inv.id

                db.session.add(exp)
                db.session.commit()

                # Обновляем статус счета (paid/partial/new) в зависимости от суммы созданных расходов
                if inv_id:
                    inv = PaymentInvoice.query.get(int(inv_id))
                    if inv:
                        paid_sum = db.session.query(func.sum(Expense.amount)).filter(Expense.invoice_id == inv.id).scalar() or Decimal(0)
                        if paid_sum >= inv.amount:
                            inv.status = 'paid'
                        elif paid_sum > 0:
                            inv.status = 'partial'
                        else:
                            inv.status = 'new'
                        try:
                            if inv.status in ('paid', 'partial'):
                                from app.vium_inbox import maybe_enqueue as _vium_enqueue
                                _vium_enqueue(inv, exp)
                        except Exception:
                            current_app.logger.exception('vium_inbox.maybe_enqueue (create) failed')
                        db.session.commit()

                flash('Добавлен')
                log_action("Добавил расход")
            except Exception as e: 
                flash(f'Ошибка: {e}')
        return redirect(url_for('finance.expenses', tab='expenses'))
        
    active_orders = Order.query.filter(Order.status != 'canceled', Order.is_deleted == False).order_by(Order.id.desc()).limit(50).all()
    barter_orders = Order.query.filter(Order.is_barter == True, Order.status != 'canceled', Order.is_deleted == False).order_by(Order.id.desc()).all()
    # Загружаем проекты и их бюджетные статьи (для JS)
    active_projects = Project.query.filter_by(status='active').order_by(Project.name).all()

    return render_template('finance/expenses.html', 
                           active_tab=tab,
                           expenses=expenses_list, 
                           invoices=invoices,
                           invoice_summary=invoice_summary,
                           budget_items=sorted(BudgetItem.query.all(), key=natural_key),
                           employees=sorted(Employee.query.filter_by(is_active=True).all(), key=lambda x: x.name),
                           active_orders=active_orders,
                           barter_orders=barter_orders,
                           active_projects=active_projects, 
                           today=msk_today(), 
                           prefill_invoice=prefill_invoice,
                           filters={'start': start_date, 'end': end_date, 'item': f_item, 'type': f_type})

@bp.route('/expenses/download_template')
@login_required
def expenses_download_template():
    wb = Workbook()
    ws = wb.active
    ws.append(["Дата", "Код", "Описание", "Сумма", "Тип"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='expenses_template.xlsx', as_attachment=True)

@bp.route('/expenses/import', methods=['POST'])
@login_required
def expenses_import():
    if current_user.role != 'admin':
        return redirect(url_for('finance.expenses'))
    
    file = request.files.get('file')
    if not file:
        return redirect(url_for('finance.expenses'))
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        budget_map = {str(b.code).strip(): b.id for b in BudgetItem.query.all()}
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_date = row[0]
            code_val = str(row[1]).strip() if row[1] else None
            desc = row[2]
            amount = row[3]
            type_val = row[4]
            
            if not raw_date or not amount or not code_val:
                continue
            
            final_date = msk_today()
            if isinstance(raw_date, datetime):
                final_date = raw_date.date()
            elif isinstance(raw_date, str):
                try:
                    final_date = datetime.strptime(raw_date, '%d.%m.%Y').date()
                except:
                    pass
            
            b_item_id = budget_map.get(code_val)
            if not b_item_id and code_val.isdigit():
                 check_id = int(code_val)
                 if BudgetItem.query.get(check_id):
                     b_item_id = check_id
            
            if not b_item_id:
                continue
            
            p_type = 'cashless'
            tv_str = str(type_val).strip()
            if tv_str == '1' or tv_str.lower() == 'нал':
                p_type = 'cash'
            elif tv_str == '2' or tv_str.lower() == 'безнал':
                p_type = 'cashless'
            
            exp = Expense(date=final_date, budget_item_id=b_item_id, description=desc, amount=float(amount), payment_type=p_type, employee_id=None)
            db.session.add(exp)
            count += 1
            
        db.session.commit()
        log_action(f'Импорт {count} расходов')
        flash(f'Успешно загружено {count} расходов')
    except Exception as e: 
        db.session.rollback()
        flash(f'Ошибка: {e}')
    return redirect(url_for('finance.expenses'))

@bp.route('/expenses/export')
@login_required
def expenses_export():
    if current_user.role not in ['admin', 'executive']:
        return redirect(url_for('main.index'))

    # Получаем те же фильтры, что и при просмотре
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    f_item = request.args.get('filter_item')
    f_type = request.args.get('filter_type')

    query = Expense.query
    if start_date: query = query.filter(func.date(Expense.date) >= start_date)
    if end_date: query = query.filter(func.date(Expense.date) <= end_date)
    if f_item: query = query.filter(Expense.budget_item_id == int(f_item))
    if f_type: query = query.filter(Expense.payment_type == f_type)
    
    expenses = query.order_by(Expense.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Расходы"

    # Шапка
    headers = ["Дата", "Статья бюджета", "Код", "Сумма", "Тип", "Сотрудник", "Проект", "Заказ", "Описание"]
    ws.append(headers)
    
    # Стили шапки
    header_fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

    for e in expenses:
        emp_name = e.employee.name if e.employee else ""
        proj_name = e.project_link.name if e.project_id else ""
        ord_name = str(e.order_id) if e.order_id else ""
        
        ws.append([
            e.date.strftime('%d.%m.%Y'),
            e.item.name,
            e.item.code,
            float(e.amount),
            'Нал' if e.payment_type == 'cash' else 'Безнал',
            emp_name,
            proj_name,
            ord_name,
            e.description
        ])

    # Ширина колонок
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['I'].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf, 
        download_name=f'expenses_{msk_now().strftime("%Y%m%d")}.xlsx', 
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def _budget_period_to_months(tm):
    """Преобразует параметр месяца/периода в список месяцев и подпись."""
    if tm == 'all':
        return list(range(1, 13)), 'За год'
    if tm == 'q1':
        return [1, 2, 3], '1 кв.'
    if tm == 'q2':
        return [4, 5, 6], '2 кв.'
    if tm == 'q3':
        return [7, 8, 9], '3 кв.'
    if tm == 'q4':
        return [10, 11, 12], '4 кв.'
    if tm == 'spring':
        return [1, 2, 3, 4, 5, 6], 'Весна'
    if tm == 'autumn':
        return [7, 8, 9, 10, 11, 12], 'Осень'
    try:
        m = int(tm)
        if 1 <= m <= 12:
            return [m], MONTH_NAMES.get(m, f"Месяц {m}")
    except Exception:
        pass
    return list(range(1, 13)), 'За год'


def _resolve_selected_months(args, fallback_target_month='all'):
    """Извлекает выбранные месяцы из request.args.

    Поддерживает два механизма:
      • multi-select `?months=2&months=3&months=4` — приоритетный, нужен
        для нового фильтра «несколько месяцев».
      • старый `?month=q1|spring|all|<число>` — fallback, чтобы сохранить
        совместимость со старыми ссылками и кнопками-периодами.

    Возвращает кортеж (selected_months: List[int], period_label: str).
    Список всегда отсортирован по возрастанию, без дублей.
    """
    raw_list = args.getlist('months') or []
    cleaned = []
    seen = set()
    for v in raw_list:
        v = (v or '').strip()
        if not v.isdigit():
            continue
        m = int(v)
        if 1 <= m <= 12 and m not in seen:
            cleaned.append(m)
            seen.add(m)

    if cleaned:
        cleaned.sort()
        if cleaned == list(range(1, 13)):
            label = 'За год'
        elif len(cleaned) == 1:
            label = MONTH_NAMES.get(cleaned[0], f"Месяц {cleaned[0]}")
        elif len(cleaned) <= 3:
            # «Фев · Март · Апр» — короткая подпись, влезает в шапку
            label = ' · '.join(
                MONTH_NAMES.get(m, str(m))[:3] for m in cleaned
            )
        else:
            label = f"{len(cleaned)} мес. ({MONTH_NAMES.get(cleaned[0], '')[:3]}–{MONTH_NAMES.get(cleaned[-1], '')[:3]})"
        return cleaned, label

    target_month = args.get('month', fallback_target_month)
    return _budget_period_to_months(target_month)


@bp.route('/budget', methods=['GET', 'POST'])
@login_required
def budget():
    if current_user.role not in ['admin', 'executive']: 
        return redirect(url_for('orders.orders_list'))
    
    # Параметры фильтрации
    year = int(request.args.get('year', msk_now().year))
    # По умолчанию берем "all" (весь год) или конкретный месяц/период
    default_month = 'all'  # Или str(msk_now().month) для фокуса на текущем
    target_month = request.args.get('month', default_month)
    tab = request.args.get('tab', 'budget')

    export = request.args.get('export')

    # Поддерживаем два механизма выбора периода:
    #   • multi-select «месяцы» (?months=2&months=3) — приоритетный,
    #   • старые быстрые кнопки (Год / Q1-Q4 / Весна / Осень) через ?month=…
    selected_months, period_label = _resolve_selected_months(
        request.args, fallback_target_month=target_month
    )
    period_end_month = max(selected_months)
    # Чтобы шаблон мог отметить выбранные option-ы в multiselect.
    selected_months_set = set(selected_months)
    is_custom_months = bool(request.args.getlist('months'))
    
    # Загрузка данных
    items = sorted(BudgetItem.query.all(), key=natural_key)
    plans = BudgetPlan.query.filter_by(year=year).all()
    
    # Умное определение периода: берем target_month, если он есть, иначе берем месяц из даты
    actual_month = func.coalesce(Expense.target_month, func.extract('month', Expense.date))
    actual_year = func.coalesce(Expense.target_year, func.extract('year', Expense.date))
    
    expenses_q = db.session.query(
        Expense.budget_item_id, 
        actual_month.label('month'), 
        func.sum(Expense.amount)
    ).filter(actual_year == year).group_by(Expense.budget_item_id, actual_month).all()

    # Структура данных
    # data = { item_id: { item: Obj, months: {1:..., 2:...}, total: {...} } }
    data = {i.id: {'item': i, 'months': {m: {'plan': Decimal(0), 'fact': Decimal(0), 'diff': Decimal(0)} for m in range(1, 13)}, 'total': {'plan':Decimal(0), 'fact':Decimal(0), 'diff':Decimal(0), 'percent': 0}} for i in items}
    
    for p in plans: 
        if p.budget_item_id in data:
            data[p.budget_item_id]['months'][p.month]['plan'] = p.amount
    for e in expenses_q: 
        if e[0] in data:
            data[e[0]]['months'][e[1]]['fact'] = e[2]
        
    grand_totals = {m: {'plan':Decimal(0), 'fact':Decimal(0), 'diff':Decimal(0)} for m in range(1, 14)} # 13 - это год
    
    # Расчет итогов
    for row in data.values():
        tp, tf = Decimal(0), Decimal(0)
        for m in range(1, 13):
            d = row['months'][m]
            d['diff'] = d['plan'] - d['fact']
            tp += d['plan']
            tf += d['fact']
            
            # Глобальные итоги по месяцам
            grand_totals[m]['plan'] += d['plan']
            grand_totals[m]['fact'] += d['fact']
            grand_totals[m]['diff'] += d['diff']
            
        row['total'] = {'plan': tp, 'fact': tf, 'diff': tp - tf}

        # Итог за выбранный период (месяц/кв/сезон/multi-select).
        period_plan = Decimal(0)
        period_fact = Decimal(0)
        for m in selected_months:
            period_plan += row['months'][m]['plan']
            period_fact += row['months'][m]['fact']
        if period_plan > 0:
            period_pct = (period_fact / period_plan) * 100
        else:
            period_pct = Decimal(0)
        row['period'] = {
            'plan': period_plan,
            'fact': period_fact,
            'diff': period_plan - period_fact,
            'percent': min(period_pct, Decimal(100)),
            'percent_raw': period_pct,
        }
        
        # Процент выполнения (для прогресс-бара)
        if tp > 0:
            pct = (tf / tp) * 100
            row['total']['percent'] = min(pct, 100) # Ограничиваем для визуализации
            row['total']['percent_raw'] = pct
        else:
            row['total']['percent'] = 0
            row['total']['percent_raw'] = 0

        grand_totals[13]['plan'] += tp
        grand_totals[13]['fact'] += tf
        grand_totals[13]['diff'] += (tp - tf)

    # Итоги для выбранного периода (месяц/кв./сезон/год)
    period_totals = {'plan': Decimal(0), 'fact': Decimal(0), 'diff': Decimal(0), 'pct': 0}
    for m in selected_months:
        period_totals['plan'] += grand_totals[m]['plan']
        period_totals['fact'] += grand_totals[m]['fact']
        period_totals['diff'] += grand_totals[m]['diff']
    if period_totals['plan'] > 0:
        period_totals['pct'] = float((period_totals['fact'] / period_totals['plan']) * 100)

    # CASHFLOW: планируемые поступления и фактические поступления
    cashflow_data = {m: {'plan': Decimal(0), 'fact': Decimal(0)} for m in range(1, 13)}
    cashflow_records = CashflowPlan.query.filter_by(year=year).all()
    for r in cashflow_records:
        if r.month in cashflow_data:
            cashflow_data[r.month]['plan'] = r.amount

    payments_q = db.session.query(func.extract('month', Payment.date).label('month'), func.sum(Payment.amount)).filter(func.extract('year', Payment.date) == year).group_by('month').all()
    for m, total in payments_q:
        if int(m) in cashflow_data:
            cashflow_data[int(m)]['fact'] = total

    cashflow_total_plan = sum(cashflow_data[m]['plan'] for m in range(1, 13))
    cashflow_total_fact = sum(cashflow_data[m]['fact'] for m in range(1, 13))

    # Итоги cashflow для выбранного периода
    cashflow_period_plan = sum(cashflow_data[m]['plan'] for m in selected_months)
    cashflow_period_fact = sum(cashflow_data[m]['fact'] for m in selected_months)

    # Накопительные (кумулятивные) показатели
    cumulative = {m: {'plan_balance': Decimal(0), 'fact_balance': Decimal(0)} for m in range(1, 13)}
    cum_in_plan = Decimal(0)
    cum_in_fact = Decimal(0)
    cum_out_plan = Decimal(0)
    cum_out_fact = Decimal(0)

    for m in range(1, 13):
        cum_in_plan += cashflow_data[m]['plan']
        cum_in_fact += cashflow_data[m]['fact']
        cum_out_plan += grand_totals[m]['plan']
        cum_out_fact += grand_totals[m]['fact']

        cumulative[m]['plan_balance'] = cum_in_fact - cum_out_plan  # фактический вход - плановый расход
        cumulative[m]['fact_balance'] = cum_in_fact - cum_out_fact  # фактический вход - фактический расход

    # Экспорт PDF — учитывает текущий период (multi-select или target_month)
    # и активный таб (бюджет / кешфлоу).
    if export == 'pdf':
        font_path = os.path.join(current_app.root_path, 'static', 'DejaVuSans.ttf').replace('\\', '/')
        rendered = render_template(
            'finance/budget_pdf.html',
            year=year,
            tab=tab,
            data=data,
            grand_totals=grand_totals,
            items=items,
            month_names=MONTH_NAMES,
            selected_months=selected_months,
            period_label=period_label,
            period_totals=period_totals,
            cashflow_data=cashflow_data,
            cashflow_period_plan=cashflow_period_plan,
            cashflow_period_fact=cashflow_period_fact,
            font_path=font_path,
            generated_at=msk_now(),
        )
        period_safe = (period_label or 'period').replace(' ', '_').replace('/', '-').replace('·', '')
        suffix = 'Cashflow' if tab == 'cashflow' else 'Budget'
        return create_pdf_response(rendered, f"{suffix}_{year}_{period_safe}.pdf")
        
    # Обработка сохранения (POST)
    if request.method == 'POST':
        if current_user.role == 'executive':
            flash('Только просмотр')
            return redirect(url_for('finance.budget', year=year, month=target_month, tab=tab))

        act = request.form.get('action')
        
        if act == 'add_item':
            db.session.add(BudgetItem(
                code=request.form.get('code'),
                name=request.form.get('name'),
                is_amortization=bool(request.form.get('is_amortization')),
                is_vium_source=bool(request.form.get('is_vium_source')),
            ))
            db.session.commit()
            flash('Статья добавлена')
        elif act == 'edit_item':
            item = BudgetItem.query.get(request.form.get('id'))
            if item:
                item.code = request.form.get('code')
                item.name = request.form.get('name')
                item.is_amortization = bool(request.form.get('is_amortization'))
                item.is_vium_source = bool(request.form.get('is_vium_source'))
                db.session.commit()
                flash('Статья обновлена')
        elif act == 'delete_item':
            item_id = request.form.get('id')
            if not Expense.query.filter_by(budget_item_id=item_id).first():
                BudgetItem.query.filter_by(id=item_id).delete()
                db.session.commit()
                flash('Статья удалена')
            else: flash('Нельзя удалить статью с расходами')
        elif 'save_plan' in request.form:
            for k, v in request.form.items():
                if k.startswith('plan['):
                    match = re.findall(r'\[(\d+)\]\[(\d+)\]', k) # Ищем plan[ID][MONTH]
                    if match:
                        iid, m = int(match[0][0]), int(match[0][1])
                        p = BudgetPlan.query.filter_by(year=year, budget_item_id=iid, month=m).first()
                        if not p: 
                            p = BudgetPlan(year=year, budget_item_id=iid, month=m)
                            db.session.add(p)
                        try: p.amount = float(v.replace(' ', '').replace(',', '.') or 0)
                        except: pass
            db.session.commit()
            flash('План сохранен')
        elif 'save_cashflow' in request.form:
            for k, v in request.form.items():
                if k.startswith('cashflow['):
                    match = re.findall(r'\[(\d+)\]', k) # Ищем cashflow[MONTH]
                    if match:
                        m = int(match[0])
                        p = CashflowPlan.query.filter_by(year=year, month=m).first()
                        if not p:
                            p = CashflowPlan(year=year, month=m)
                            db.session.add(p)
                        try:
                            # Извлекаем только цифры, запятую и точку перед конвертацией
                            val_str = re.sub(r'[^\d.,-]', '', str(v)).replace(',', '.')
                            p.amount = float(val_str or 0)
                        except Exception as e:
                            print(f"Error parsing cashflow value: {v} -> {e}")
            db.session.commit()
            flash('План поступлений сохранен')
            
        return redirect(url_for('finance.budget', year=year, month=target_month, tab=tab))
        
    return render_template('finance/budget.html',
                           year=year,
                           data=data,
                           grand_totals=grand_totals,
                           items=items,
                           month_names=MONTH_NAMES,
                           target_month=target_month,
                           period_label=period_label,
                           selected_months=selected_months,
                           selected_months_set=selected_months_set,
                           is_custom_months=is_custom_months,
                           period_end_month=period_end_month,
                           period_totals=period_totals,
                           active_tab=tab,
                           cashflow_data=cashflow_data,
                           cashflow_total_plan=cashflow_total_plan,
                           cashflow_total_fact=cashflow_total_fact,
                           cashflow_period_plan=cashflow_period_plan,
                           cashflow_period_fact=cashflow_period_fact,
                           cumulative=cumulative)

@bp.route('/budget/export')
@login_required
def budget_export():
    """Excel-выгрузка таблицы бюджета или кешфлоу.

    Учитывает выбор пользователя на странице:
      • активный таб `tab` (`budget` / `cashflow`),
      • период через `?months=2&months=3` (multi-select) или старое
        `?month=all|q1|spring|<число>`.

    Дизайн новый — компактный, читаемый, с цветной подсветкой отклонений
    и итогами за период. На пустом всё ещё корректно отдаёт пустой шаблон.
    """
    if current_user.role not in ['admin', 'executive']:
        return redirect(url_for('orders.orders_list'))

    year = int(request.args.get('year', msk_now().year))
    tab = request.args.get('tab', 'budget')
    target_month = request.args.get('month', 'all')
    selected_months, period_label = _resolve_selected_months(
        request.args, fallback_target_month=target_month
    )

    # --- Загрузка данных бюджета ------------------------------------------
    items = sorted(BudgetItem.query.all(), key=natural_key)
    plans = BudgetPlan.query.filter_by(year=year).all()

    actual_month = func.coalesce(Expense.target_month, func.extract('month', Expense.date))
    actual_year = func.coalesce(Expense.target_year, func.extract('year', Expense.date))
    expenses_q = db.session.query(
        Expense.budget_item_id,
        actual_month.label('month'),
        func.sum(Expense.amount)
    ).filter(actual_year == year).group_by(Expense.budget_item_id, actual_month).all()

    data = {
        i.id: {
            'item': i,
            'months': {m: {'plan': 0.0, 'fact': 0.0} for m in range(1, 13)},
        }
        for i in items
    }
    for p in plans:
        if p.budget_item_id in data:
            data[p.budget_item_id]['months'][p.month]['plan'] = float(p.amount or 0)
    for e in expenses_q:
        if e[0] in data:
            data[e[0]]['months'][int(e[1])]['fact'] = float(e[2] or 0)

    # --- Загрузка кешфлоу --------------------------------------------------
    cashflow_data = {m: {'plan': 0.0, 'fact': 0.0} for m in range(1, 13)}
    for r in CashflowPlan.query.filter_by(year=year).all():
        if r.month in cashflow_data:
            cashflow_data[r.month]['plan'] = float(r.amount or 0)
    payments_q = (
        db.session.query(
            func.extract('month', Payment.date).label('month'),
            func.sum(Payment.amount),
        )
        .filter(func.extract('year', Payment.date) == year)
        .group_by('month')
        .all()
    )
    for m, total in payments_q:
        if int(m) in cashflow_data:
            cashflow_data[int(m)]['fact'] = float(total or 0)

    # --- Стили (общие) ----------------------------------------------------
    style_title = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    style_subtitle = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    style_table_header = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    style_total = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    fill_red = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    fill_green = PatternFill(start_color="DCEDC8", end_color="DCEDC8", fill_type="solid")
    fill_zebra = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")

    # Цветовая схема План / Факт.
    # План — холодный синий («лимит»), Факт — тёплый янтарный («израсходовано»).
    # У каждой палитры есть оттенки для обычной / зебра-строк / шапки / итога.
    fill_plan = PatternFill(start_color="EAF4FB", end_color="EAF4FB", fill_type="solid")
    fill_plan_alt = PatternFill(start_color="D6E9F6", end_color="D6E9F6", fill_type="solid")
    fill_plan_head = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    fill_plan_total = PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid")
    fill_fact = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
    fill_fact_alt = PatternFill(start_color="FFE9C2", end_color="FFE9C2", fill_type="solid")
    fill_fact_head = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    fill_fact_total = PatternFill(start_color="FFCC80", end_color="FFCC80", fill_type="solid")

    font_title = Font(bold=True, color="FFFFFF", size=14)
    font_subtitle = Font(bold=True, color="1B5E20", size=11)
    font_th = Font(bold=True, color="000000")
    font_total = Font(bold=True, size=11)
    font_plan = Font(color="0D47A1")
    font_plan_head = Font(bold=True, color="0D47A1")
    font_plan_total = Font(bold=True, color="0D47A1", size=11)
    font_fact = Font(color="E65100")
    font_fact_head = Font(bold=True, color="E65100")
    font_fact_total = Font(bold=True, color="E65100", size=11)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    money_fmt = '#,##0 "₽"'

    def paint_plan_cell(cell, zebra=False):
        cell.fill = fill_plan_alt if zebra else fill_plan
        cell.font = font_plan

    def paint_fact_cell(cell, zebra=False):
        cell.fill = fill_fact_alt if zebra else fill_fact
        cell.font = font_fact

    wb = Workbook()
    ws = wb.active

    def write_title(row, span, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = style_title
        c.font = font_title
        c.alignment = align_center
        ws.row_dimensions[row].height = 26

    def write_subtitle(row, span, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = style_subtitle
        c.font = font_subtitle
        c.alignment = align_center
        ws.row_dimensions[row].height = 22

    if tab == 'cashflow':
        # =================================================================
        # CASHFLOW: «Месяц · План · Факт · Δ · % исп.»
        # =================================================================
        ws.title = f"Cashflow {year}"

        # Колонки:
        ws.column_dimensions['A'].width = 4   # №
        ws.column_dimensions['B'].width = 22  # Месяц
        ws.column_dimensions['C'].width = 18  # План
        ws.column_dimensions['D'].width = 18  # Факт
        ws.column_dimensions['E'].width = 18  # Δ
        ws.column_dimensions['F'].width = 12  # % исп.

        write_title(1, 6, f"Cashflow {year} · {period_label}")
        write_subtitle(2, 6, f"Дата выгрузки: {msk_now().strftime('%d.%m.%Y %H:%M')}")
        row_idx = 4

        headers = ["№", "Месяц", "План, ₽", "Факт, ₽", "Δ (Факт − План)", "% исп."]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=i, value=h)
            c.fill = style_table_header
            c.font = font_th
            c.border = thin
            c.alignment = align_center
        # Подкрашиваем шапку «План»/«Факт» в свои цвета — для быстрой читаемости.
        plan_h = ws.cell(row=row_idx, column=3)
        plan_h.fill = fill_plan_head
        plan_h.font = font_plan_head
        fact_h = ws.cell(row=row_idx, column=4)
        fact_h.fill = fill_fact_head
        fact_h.font = font_fact_head
        row_idx += 1

        period_plan = 0.0
        period_fact = 0.0
        for n, m in enumerate(selected_months, 1):
            plan_v = float(cashflow_data[m]['plan'] or 0)
            fact_v = float(cashflow_data[m]['fact'] or 0)
            delta = fact_v - plan_v
            pct = (fact_v / plan_v * 100) if plan_v else 0
            period_plan += plan_v
            period_fact += fact_v
            zebra = fill_zebra if n % 2 == 0 else None

            cells = [
                (1, n, align_center),
                (2, MONTH_NAMES.get(m, str(m)), Alignment(horizontal="left", vertical="center")),
                (3, plan_v, align_right),
                (4, fact_v, align_right),
                (5, delta, align_right),
                (6, round(pct, 1) if plan_v else None, align_center),
            ]
            for col, val, alg in cells:
                c = ws.cell(row=row_idx, column=col, value=val)
                c.border = thin
                c.alignment = alg
                if zebra:
                    c.fill = zebra
                if col in (3, 4, 5):
                    c.number_format = money_fmt
                if col == 6 and val is not None:
                    c.number_format = '0.0"%"'
            # Цветовое разделение колонок План/Факт.
            paint_plan_cell(ws.cell(row=row_idx, column=3), zebra=bool(zebra))
            paint_fact_cell(ws.cell(row=row_idx, column=4), zebra=bool(zebra))
            # Подсветка дельты: положительная — зелёная, отрицательная — красная.
            delta_cell = ws.cell(row=row_idx, column=5)
            if delta < 0:
                delta_cell.fill = fill_red
                delta_cell.font = Font(bold=True, color="B71C1C")
            elif delta > 0:
                delta_cell.fill = fill_green
                delta_cell.font = Font(bold=True, color="1B5E20")
            row_idx += 1

        # ИТОГО за период
        delta_total = period_fact - period_plan
        pct_total = (period_fact / period_plan * 100) if period_plan else 0
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        c = ws.cell(row=row_idx, column=1, value=f"ИТОГО за период ({period_label})")
        c.fill = style_total
        c.font = font_total
        c.alignment = align_right
        c.border = thin
        for col, val in [(3, period_plan), (4, period_fact), (5, delta_total)]:
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = style_total
            c.font = font_total
            c.alignment = align_right
            c.border = thin
            c.number_format = money_fmt
        # Итоговые «План»/«Факт» — насыщенные синий/янтарный для контраста.
        plan_total_cell = ws.cell(row=row_idx, column=3)
        plan_total_cell.fill = fill_plan_total
        plan_total_cell.font = font_plan_total
        fact_total_cell = ws.cell(row=row_idx, column=4)
        fact_total_cell.fill = fill_fact_total
        fact_total_cell.font = font_fact_total
        c = ws.cell(row=row_idx, column=6, value=round(pct_total, 1) if period_plan else None)
        c.fill = style_total
        c.font = font_total
        c.alignment = align_center
        c.border = thin
        c.number_format = '0.0"%"'

    else:
        # =================================================================
        # БЮДЖЕТ РАСХОДОВ
        # Если выбран 1 месяц — компактная таблица «Статья · План · Факт · Откл · %».
        # Если несколько месяцев — добавляем колонки разбивки по месяцам.
        # =================================================================
        ws.title = f"Бюджет {year}"

        n_months = len(selected_months)
        # Колонки:
        # 1: Код, 2: Статья, 3: План, 4: Факт, 5: Откл, 6: %
        # 7+: для каждого месяца — П/Ф (2 колонки)
        cols_total = 6 + (n_months * 2 if n_months > 1 else 0)

        ws.column_dimensions['A'].width = 8   # Код
        ws.column_dimensions['B'].width = 36  # Статья
        for i in range(3, 7):
            ws.column_dimensions[get_column_letter(i)].width = 14
        for i in range(7, cols_total + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12

        write_title(1, cols_total, f"Бюджет расходов {year} · {period_label}")
        write_subtitle(2, cols_total, f"Дата выгрузки: {msk_now().strftime('%d.%m.%Y %H:%M')}")
        row_idx = 4

        # Шапка: верхняя строка с группами + нижняя с подзаголовками.
        # Делаем merge для группы «Итого за период» и для каждого месяца.
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + 1, end_column=1)
        c = ws.cell(row=row_idx, column=1, value="Код")
        c.fill = style_table_header
        c.font = font_th
        c.border = thin
        c.alignment = align_center

        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx + 1, end_column=2)
        c = ws.cell(row=row_idx, column=2, value="Статья")
        c.fill = style_table_header
        c.font = font_th
        c.border = thin
        c.alignment = align_center

        ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)
        c = ws.cell(row=row_idx, column=3, value=f"ИТОГО ЗА ПЕРИОД ({period_label})")
        c.fill = style_total
        c.font = font_th
        c.border = thin
        c.alignment = align_center

        if n_months > 1:
            for idx, m in enumerate(selected_months):
                col = 7 + idx * 2
                ws.merge_cells(start_row=row_idx, start_column=col, end_row=row_idx, end_column=col + 1)
                c = ws.cell(row=row_idx, column=col, value=MONTH_NAMES.get(m, str(m)))
                c.fill = style_table_header
                c.font = font_th
                c.border = thin
                c.alignment = align_center

        row_idx += 1
        sub_headers = [
            ("План", fill_plan_head, font_plan_head),
            ("Факт", fill_fact_head, font_fact_head),
            ("Откл", style_total, font_th),
            ("% исп.", style_total, font_th),
        ]
        for i, (h, fill, fnt) in enumerate(sub_headers, 1):
            c = ws.cell(row=row_idx, column=2 + i, value=h)
            c.fill = fill
            c.font = fnt
            c.border = thin
            c.alignment = align_center
        if n_months > 1:
            for idx, m in enumerate(selected_months):
                col = 7 + idx * 2
                # «План» — синий, «Факт» — янтарный, чтобы взгляд сразу делил пары.
                month_subs = [
                    ("План", fill_plan_head, font_plan_head),
                    ("Факт", fill_fact_head, font_fact_head),
                ]
                for j, (h, fill, fnt) in enumerate(month_subs):
                    c = ws.cell(row=row_idx, column=col + j, value=h)
                    c.fill = fill
                    c.font = fnt
                    c.border = thin
                    c.alignment = align_center

        row_idx += 1

        # --- Строки статей ----
        period_plan_total = 0.0
        period_fact_total = 0.0
        n_row = 0
        for item in items:
            row_data = data.get(item.id)
            if not row_data:
                continue

            row_plan = sum(row_data['months'][m]['plan'] for m in selected_months)
            row_fact = sum(row_data['months'][m]['fact'] for m in selected_months)
            # Пустые строки скрываем — не нужны в выгрузке.
            if row_plan == 0 and row_fact == 0:
                continue
            n_row += 1
            row_diff = row_plan - row_fact
            row_pct = (row_fact / row_plan * 100) if row_plan else 0
            period_plan_total += row_plan
            period_fact_total += row_fact

            zebra = fill_zebra if n_row % 2 == 0 else None

            cells = [
                (1, item.code, align_center),
                (2, item.name, Alignment(horizontal="left", vertical="center", wrap_text=True)),
                (3, row_plan, align_right),
                (4, row_fact, align_right),
                (5, row_diff, align_right),
                (6, round(row_pct, 1) if row_plan else None, align_center),
            ]
            for col, val, alg in cells:
                c = ws.cell(row=row_idx, column=col, value=val)
                c.border = thin
                c.alignment = alg
                if zebra:
                    c.fill = zebra
                if col in (3, 4, 5):
                    c.number_format = money_fmt
                if col == 6 and val is not None:
                    c.number_format = '0.0"%"'

            # Цветовое разделение колонок «План» и «Факт».
            paint_plan_cell(ws.cell(row=row_idx, column=3), zebra=bool(zebra))
            paint_fact_cell(ws.cell(row=row_idx, column=4), zebra=bool(zebra))

            # Подсветка отклонения: отрицательное всегда красное,
            # в том числе при План=0 и Факт>0.
            if row_diff < 0:
                c = ws.cell(row=row_idx, column=5)
                c.fill = fill_red
                c.font = Font(bold=True, color="B71C1C")
            elif row_diff > 0:
                c = ws.cell(row=row_idx, column=5)
                c.fill = fill_green

            # Месячная разбивка
            if n_months > 1:
                for idx, m in enumerate(selected_months):
                    col = 7 + idx * 2
                    plan_v = float(row_data['months'][m]['plan'] or 0)
                    fact_v = float(row_data['months'][m]['fact'] or 0)
                    for j, val in enumerate([plan_v, fact_v]):
                        c = ws.cell(row=row_idx, column=col + j, value=val)
                        c.border = thin
                        c.alignment = align_right
                        c.number_format = money_fmt
                        if zebra:
                            c.fill = zebra
                        # Внутри пары: левая — план, правая — факт.
                        if j == 0:
                            paint_plan_cell(c, zebra=bool(zebra))
                        else:
                            paint_fact_cell(c, zebra=bool(zebra))

            row_idx += 1

        # --- ИТОГО за период
        period_diff = period_plan_total - period_fact_total
        period_pct = (period_fact_total / period_plan_total * 100) if period_plan_total else 0

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        c = ws.cell(row=row_idx, column=1, value=f"ИТОГО за период ({period_label})")
        c.fill = style_total
        c.font = font_total
        c.alignment = align_right
        c.border = thin
        for col, val in [(3, period_plan_total), (4, period_fact_total), (5, period_diff)]:
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = style_total
            c.font = font_total
            c.alignment = align_right
            c.border = thin
            c.number_format = money_fmt
        # Итоговые «План»/«Факт» — насыщенные синий/янтарный.
        plan_total_cell = ws.cell(row=row_idx, column=3)
        plan_total_cell.fill = fill_plan_total
        plan_total_cell.font = font_plan_total
        fact_total_cell = ws.cell(row=row_idx, column=4)
        fact_total_cell.fill = fill_fact_total
        fact_total_cell.font = font_fact_total
        c = ws.cell(row=row_idx, column=6, value=round(period_pct, 1) if period_plan_total else None)
        c.fill = style_total
        c.font = font_total
        c.alignment = align_center
        c.border = thin
        c.number_format = '0.0"%"'

        if n_months > 1:
            for idx, m in enumerate(selected_months):
                col = 7 + idx * 2
                m_plan = sum(row_data['months'][m]['plan']
                             for row_data in data.values()
                             if any((row_data['months'][mm]['plan'] or row_data['months'][mm]['fact'])
                                    for mm in selected_months))
                m_fact = sum(row_data['months'][m]['fact']
                             for row_data in data.values()
                             if any((row_data['months'][mm]['plan'] or row_data['months'][mm]['fact'])
                                    for mm in selected_months))
                for j, val in enumerate([m_plan, m_fact]):
                    c = ws.cell(row=row_idx, column=col + j, value=val)
                    c.alignment = align_right
                    c.border = thin
                    c.number_format = money_fmt
                    # «План» — синий total, «Факт» — янтарный total.
                    if j == 0:
                        c.fill = fill_plan_total
                        c.font = font_plan_total
                    else:
                        c.fill = fill_fact_total
                        c.font = font_fact_total

        # Замораживаем шапку и колонку «Статья»
        ws.freeze_panes = 'C5'

    # --- Отдача файла ------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    period_safe = period_label.replace(' ', '_').replace('/', '-').replace('·', '')
    name_prefix = 'Cashflow' if tab == 'cashflow' else 'Budget'
    filename = f'{name_prefix}_{year}_{period_safe}.xlsx'
    return send_file(
        buf,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

_COST_PERIOD_LABELS = {
    '': 'Весь год',
    'q1': '1 квартал (янв–мар)',
    'q2': '2 квартал (апр–июн)',
    'q3': '3 квартал (июл–сен)',
    'q4': '4 квартал (окт–дек)',
    'spring': 'Весна (янв–июн)',
    'autumn': 'Осень (июл–дек)',
}


def _build_cost_report_xlsx(
    *,
    visible_years,
    op_rows,
    amort_rows,
    cd,
    impact_rows,
    impact_total_qty,
    impact_table_total,
    selected_year,
    selected_period,
):
    """Формирует xlsx: свод себестоимости + детализация партий (без грядок)."""
    period_label = _COST_PERIOD_LABELS.get(selected_period or '', selected_period or 'Весь год')
    exported_at = msk_now().strftime('%d.%m.%Y %H:%M')

    fill_title = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    fill_subtitle = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    fill_header = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    fill_section = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')
    fill_total = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
    fill_zebra = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
    fill_impact_hdr = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    fill_impact_total = PatternFill(start_color='FFECB3', end_color='FFECB3', fill_type='solid')

    font_title = Font(bold=True, color='FFFFFF', size=14)
    font_subtitle = Font(bold=True, color='1B5E20', size=11)
    font_th = Font(bold=True, color='000000', size=10)
    font_section = Font(bold=True, color='1B5E20', size=10)
    font_total = Font(bold=True, size=11)
    thin = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    money_fmt = '#,##0.00 "₽"'
    money_int_fmt = '#,##0 "₽"'
    qty_fmt = '#,##0'

    def paint_row(ws, row_idx, ncol, *, fill=None, font=None, number_cols=None, qty_cols=None):
        for col in range(1, ncol + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = thin
            if fill:
                c.fill = fill
            if font:
                c.font = font
            if number_cols and col in number_cols:
                c.alignment = align_right
                c.number_format = money_fmt
            elif qty_cols and col in qty_cols:
                c.alignment = align_right
                c.number_format = qty_fmt
            elif col == 1:
                c.alignment = align_left
            else:
                c.alignment = align_right

    def autosize(ws, max_col, min_w=10, max_w=42):
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            best = min_w
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        best = max(best, min(len(str(cell.value)) + 2, max_w))
            ws.column_dimensions[letter].width = best

    wb = Workbook()
    cols_total = 1 + len(visible_years) * 2

    # --- Лист 1: Свод ---
    ws = wb.active
    ws.title = 'Свод себестоимости'

    def write_banner(ws, row_title, row_sub, title_text, subtitle_text):
        ws.merge_cells(start_row=row_title, start_column=1, end_row=row_title, end_column=cols_total)
        c = ws.cell(row=row_title, column=1, value=title_text)
        c.fill = fill_title
        c.font = font_title
        c.alignment = align_center
        ws.row_dimensions[row_title].height = 28
        ws.merge_cells(start_row=row_sub, start_column=1, end_row=row_sub, end_column=cols_total)
        c2 = ws.cell(row=row_sub, column=1, value=subtitle_text)
        c2.fill = fill_subtitle
        c2.font = font_subtitle
        c2.alignment = align_center
        ws.row_dimensions[row_sub].height = 22

    write_banner(
        ws, 1, 2,
        'Себестоимость — свод по годам',
        f'Год отчёта: {selected_year} · Период расходов: {period_label} · Выгрузка: {exported_at}',
    )

    header_row = 4
    ws.cell(row=header_row, column=1, value='Статья / показатель')
    col = 2
    for y in visible_years:
        qty_y = cd['qty_by_year'].get(y) or 0
        ws.cell(row=header_row, column=col, value=f'{y} — сумма')
        ws.cell(row=header_row, column=col + 1, value=f'{y} — на ед. ({qty_y} шт)')
        col += 2
    paint_row(ws, header_row, cols_total, fill=fill_header, font=font_th)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    row_idx = header_row + 1

    def append_section_title(text):
        nonlocal row_idx
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=cols_total)
        c = ws.cell(row=row_idx, column=1, value=text)
        c.fill = fill_section
        c.font = font_section
        c.alignment = align_left
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    def append_data_row(name, data_getter, *, is_total=False):
        nonlocal row_idx
        line = [name]
        for y in visible_years:
            d = data_getter(y)
            line.extend([float(d.get('sum', 0)), float(d.get('per_unit', 0))])
        for col_idx, val in enumerate(line, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        paint_row(
            ws, row_idx, cols_total,
            fill=fill_total if is_total else (fill_zebra if row_idx % 2 == 0 else None),
            font=font_total if is_total else None,
            number_cols=set(range(2, cols_total + 1)),
        )
        row_idx += 1

    append_section_title('Операционные расходы')
    for row in op_rows:
        append_data_row(row['name'], lambda y, r=row: r['data'].get(y, {'sum': 0, 'per_unit': 0}))
    append_data_row(
        'ИТОГО операционные',
        lambda y: {'sum': float(cd['summary_totals'][y]['op']), 'per_unit': float(cd['summary_totals'][y]['op_unit'])},
        is_total=True,
    )
    row_idx += 1

    append_section_title('Амортизация')
    for row in amort_rows:
        append_data_row(row['name'], lambda y, r=row: r['data'].get(y, {'sum': 0, 'per_unit': 0}))
    append_data_row(
        'ИТОГО амортизация',
        lambda y: {'sum': float(cd['summary_totals'][y]['amort']), 'per_unit': float(cd['summary_totals'][y]['amort_unit'])},
        is_total=True,
    )
    row_idx += 1

    append_section_title('Итоговая себестоимость')
    grand_line = ['ВСЕГО себестоимость (на 1 шт.)']
    for y in visible_years:
        grand_line.extend(['', float(cd['summary_totals'][y]['total_unit'])])
    for col_idx, val in enumerate(grand_line, 1):
        ws.cell(row=row_idx, column=col_idx, value=val)
    paint_row(ws, row_idx, cols_total, fill=fill_total, font=font_total, number_cols=set(range(2, cols_total + 1)))
    autosize(ws, cols_total)

    # --- Лист 2: Детализация партий ---
    ws2 = wb.create_sheet('Влияние на цену')
    impact_cols = 11
    write_banner(
        ws2, 1, 2,
        f'Влияние на цену — партии на {selected_year} год',
        f'Период расходов: {period_label} · Поля «Грядки» исключены · Выгрузка: {exported_at}',
    )
    ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=impact_cols)
    note = ws2.cell(
        row=3, column=1,
        value=(
            f'Строк: {len(impact_rows)} · Кол-во (без грядок): {impact_total_qty} шт · '
            f'Сумма себестоимости партий: {float(impact_table_total):,.2f} ₽'
        ),
    )
    note.fill = fill_subtitle
    note.font = Font(size=10, color='1B5E20')
    note.alignment = align_left

    hdr_row = 5
    impact_headers = [
        'Растение', 'Поле', 'Год партии', 'Кол-во, шт',
        'Цена закупки', 'Накоп. опер.', 'Накоп. аморт.',
        'ВиУМ ₽/шт', 'ФОТ ₽/шт', 'Итого за ед.', 'Итого сумма',
    ]
    for i, h in enumerate(impact_headers, 1):
        ws2.cell(row=hdr_row, column=i, value=h)
    paint_row(ws2, hdr_row, impact_cols, fill=fill_impact_hdr, font=font_th)
    ws2.freeze_panes = ws2.cell(row=hdr_row + 1, column=1)

    money_cols = {5, 6, 7, 8, 9, 10, 11}
    row_idx = hdr_row + 1
    for n, r in enumerate(impact_rows):
        line = [
            r['name'], r['field'], r['year'], r['quantity'],
            float(r['purchase_price']), float(r['accum_opex']), float(r['accum_amort']),
            float(r.get('vium_unit') or 0), float(r.get('fot_unit') or 0),
            float(r['total_unit_cost']), float(r['total_cost']),
        ]
        for col_idx, val in enumerate(line, 1):
            ws2.cell(row=row_idx, column=col_idx, value=val)
        paint_row(
            ws2, row_idx, impact_cols,
            fill=fill_zebra if n % 2 else None,
            number_cols=money_cols,
            qty_cols={4},
        )
        row_idx += 1

    for col_idx, val in enumerate(
        ['ИТОГО', '', '', impact_total_qty, '', '', '', '', '', '', float(impact_table_total)],
        1,
    ):
        ws2.cell(row=row_idx, column=col_idx, value=val if val != '' else None)
    paint_row(
        ws2, row_idx, impact_cols,
        fill=fill_impact_total, font=font_total,
        number_cols={11}, qty_cols={4},
    )
    autosize(ws2, impact_cols)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.route('/cost', methods=['GET', 'POST'])
@login_required
def cost_report():
    if current_user.role not in ['admin', 'executive']: 
        return redirect(url_for('orders.orders_list'))
    
    current_year = msk_now().year
    years = list(range(2017, current_year + 1))
    
    if request.method == 'POST':
        if current_user.role == 'executive':
            flash('Только просмотр')
            return redirect(url_for('finance.cost_report'))

    if request.method == 'POST' and (
        'update_purchase_price' in request.form
        or (request.is_json and (request.get_json(silent=True) or {}).get('action') == 'bulk_update_purchase_price')
    ):
        # Партия: растение + поле + год. На вкладке «Влияние» — все размеры.
        # На «Контейнерная площадка» можно уточнить size_id / lot_id.
        from app.models import StockPurchaseLot

        def _apply_one_purchase_price(item: dict) -> int:
            p_id = int(item.get('plant_id') or 0) or None
            f_id = int(item.get('field_id') or 0) or None
            batch_year = int(item.get('year') or 0) or None
            size_id = int(item['size_id']) if item.get('size_id') not in (None, '', 0, '0') else None
            lot_id = int(item['lot_id']) if item.get('lot_id') not in (None, '', 0, '0') else None
            raw_price = str(item.get('purchase_price') or '').replace(',', '.').replace(' ', '').replace('\xa0', '').strip()
            val = float(raw_price) if raw_price else 0.0
            updated = 0

            if lot_id:
                lot = StockPurchaseLot.query.get(lot_id)
                if lot:
                    lot.purchase_price = val
                    updated += 1
                    sb = StockBalance.query.filter_by(
                        plant_id=lot.plant_id,
                        size_id=lot.size_id,
                        field_id=lot.field_id,
                        year=lot.year,
                    ).first()
                    if sb:
                        sb.purchase_price = val
                        updated += 1
                return updated

            sb_q = StockBalance.query.filter_by(
                plant_id=p_id, field_id=f_id, year=batch_year
            )
            if size_id:
                sb_q = sb_q.filter_by(size_id=size_id)
            sbs = sb_q.all()
            for sb in sbs:
                sb.purchase_price = val
                updated += 1

            lot_q = StockPurchaseLot.query.filter_by(
                plant_id=p_id, field_id=f_id, year=batch_year
            )
            if size_id:
                lot_q = lot_q.filter_by(size_id=size_id)
            for lot in lot_q.all():
                lot.purchase_price = val
                updated += 1

            if updated == 0 and size_id and p_id and f_id and batch_year:
                qty = int(item.get('quantity') or 0)
                sb = StockBalance.query.filter_by(
                    plant_id=p_id, size_id=size_id, field_id=f_id, year=batch_year
                ).first()
                if sb:
                    sb.purchase_price = val
                    qty = max(qty, int(sb.quantity or 0))
                    updated += 1
                lot = StockPurchaseLot(
                    plant_id=p_id,
                    size_id=size_id,
                    field_id=f_id,
                    year=batch_year,
                    purchase_price=val,
                    quantity=max(qty, 1),
                )
                db.session.add(lot)
                updated += 1
            return updated

        payload = request.get_json(silent=True) if request.is_json else None
        if payload and payload.get('action') == 'bulk_update_purchase_price':
            ok_n = 0
            err_n = 0
            for it in payload.get('items') or []:
                try:
                    if _apply_one_purchase_price(it) > 0:
                        ok_n += 1
                    else:
                        err_n += 1
                except (TypeError, ValueError):
                    err_n += 1
            if ok_n:
                db.session.commit()
            else:
                db.session.rollback()
            return jsonify({
                'status': 'ok' if ok_n else 'error',
                'updated': ok_n,
                'errors': err_n,
            })

        p_id = request.form.get('plant_id', type=int)
        f_id = request.form.get('field_id', type=int)
        batch_year = request.form.get('year', type=int)
        size_id = request.form.get('size_id', type=int)
        lot_id = request.form.get('lot_id', type=int)
        raw_price = (request.form.get('purchase_price') or '').replace(',', '.').replace(' ', '').replace('\xa0', '').strip()
        try:
            updated = _apply_one_purchase_price({
                'plant_id': p_id,
                'field_id': f_id,
                'year': batch_year,
                'size_id': size_id,
                'lot_id': lot_id,
                'purchase_price': raw_price,
                'quantity': request.form.get('quantity', type=int) or 0,
            })
            if updated:
                db.session.commit()
                flash(f'Цена закупки обновлена', 'success')
            else:
                flash('Не нашли позиции этой партии в остатках — цена не записана.', 'warning')
        except (TypeError, ValueError):
            flash('Не удалось разобрать цену закупки.', 'danger')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'status': 'ok'})
        return redirect(request.referrer or url_for('finance.cost_report', tab='impact'))
    
    selected_year = int(request.args.get('filter_year', current_year))
    active_tab = request.args.get('tab', 'summary')
    selected_period = request.args.get('period', '')

    cd = calculate_cost_data(selected_year, period=selected_period)
    
    years_with_data = [y for y in years if cd['summary_totals'][y]['total'] > 0]
    visible_years = sorted(list(set(years_with_data) | {current_year, selected_year}))
    if 'hide_year' in request.args: 
        visible_years = [y for y in visible_years if y not in [int(x) for x in request.args.getlist('hide_year')]]
    
    active_amort_years = [y for y in years if cd['raw_amort_source'].get(y, Decimal(0)) > 0]
    
    f_plants = [int(x) for x in request.args.getlist('filter_plant')]
    f_fields = [int(x) for x in request.args.getlist('filter_field')]
    need_impact_data = (active_tab == 'impact') or (request.args.get('export') == 'excel')
    need_container_data = (active_tab == 'container')
    bed_field_ids = get_bed_field_ids()
    container_field_ids = get_container_field_ids()
    excluded_impact_field_ids = bed_field_ids | container_field_ids

    impact_rows = []
    impact_total_qty = 0
    impact_table_total = Decimal(0)
    if need_impact_data:
        raw_rows = get_detailed_stock_at_year_end(selected_year)
        # «Влияние на цену»: грядки и контейнерная площадка не участвуют.
        impact_rows_source = [r for r in raw_rows if r['field_id'] not in excluded_impact_field_ids]
        if f_plants:
            impact_rows_source = [r for r in impact_rows_source if r['plant_id'] in f_plants]
        if f_fields:
            impact_rows_source = [r for r in impact_rows_source if r['field_id'] in f_fields]

        from app.models import StockPurchaseLot, Supplier
        from decimal import Decimal as _D
        suppliers_map = {s.id: s.name for s in Supplier.query.all()}
        lot_q = StockPurchaseLot.query.filter(StockPurchaseLot.quantity > 0)
        if f_plants:
            lot_q = lot_q.filter(StockPurchaseLot.plant_id.in_(f_plants))
        if f_fields:
            lot_q = lot_q.filter(StockPurchaseLot.field_id.in_(f_fields))
        else:
            lot_q = lot_q.filter(~StockPurchaseLot.field_id.in_(list(excluded_impact_field_ids) or [-1]))
        lots = lot_q.all()

        # Ключ партии себестоимости: растение + поле + год + поставщик + цена
        grouped_map = {}
        covered_sb_keys = set()  # (plant, field, year) закрытые лотами

        for lot in lots:
            if lot.field_id in excluded_impact_field_ids:
                continue
            pp = _D(str(lot.purchase_price or 0))
            key = (lot.plant_id, lot.field_id, lot.year, lot.supplier_id or 0, f'{pp:.2f}')
            if key not in grouped_map:
                plant_name = next((r['name'] for r in impact_rows_source if r['plant_id'] == lot.plant_id), None)
                field_name = next((r['field'] for r in impact_rows_source if r['field_id'] == lot.field_id), None)
                if plant_name is None:
                    from app.models import Plant as _Plant, Field as _Field
                    pl = _Plant.query.get(lot.plant_id)
                    plant_name = pl.name if pl else f'#{lot.plant_id}'
                if field_name is None:
                    from app.models import Field as _Field
                    fld = _Field.query.get(lot.field_id)
                    field_name = fld.name if fld else f'#{lot.field_id}'
                grouped_map[key] = {
                    'plant_id': lot.plant_id, 'field_id': lot.field_id, 'year': lot.year,
                    'name': plant_name, 'field': field_name,
                    'supplier_id': lot.supplier_id,
                    'supplier_name': suppliers_map.get(lot.supplier_id, '') if lot.supplier_id else '',
                    'total_qty': 0, 'total_purchase_val': _D(0),
                    'sample_size_id': lot.size_id,
                    'lot_price': pp,
                }
            grouped_map[key]['total_qty'] += int(lot.quantity or 0)
            grouped_map[key]['total_purchase_val'] += pp * _D(int(lot.quantity or 0))
            covered_sb_keys.add((lot.plant_id, lot.field_id, lot.year))

        # Остатки без лотов — как раньше (средняя по SB)
        for r in impact_rows_source:
            sb_key = (r['plant_id'], r['field_id'], r['year'])
            if sb_key in covered_sb_keys:
                continue
            key = (r['plant_id'], r['field_id'], r['year'], 0, 'sb')
            if key not in grouped_map:
                grouped_map[key] = {
                    'plant_id': r['plant_id'], 'field_id': r['field_id'], 'year': r['year'],
                    'name': r['name'], 'field': r['field'],
                    'supplier_id': None, 'supplier_name': '',
                    'total_qty': 0, 'total_purchase_val': Decimal(0),
                    'sample_size_id': r['size_id'],
                    'lot_price': None,
                }
            grouped_map[key]['total_qty'] += r['quantity']
            grouped_map[key]['total_purchase_val'] += r['purchase_price'] * Decimal(r['quantity'])

        for v in grouped_map.values():
            qty = v['total_qty']
            if qty == 0:
                continue
            avg_purchase_price = v['total_purchase_val'] / Decimal(qty)
            batch_year = v['year']
            ac_opex = cd['accum_opex_map'].get(batch_year, Decimal(0))
            ac_amort = cd['accum_amort_map'].get(batch_year, Decimal(0))
            vium_unit = cd.get('vium_unit_by_year', {}).get(batch_year, Decimal(0)) or Decimal(0)
            fot_unit = cd.get('fot_unit_by_year', {}).get(batch_year, Decimal(0)) or Decimal(0)
            total_unit_cost = avg_purchase_price + ac_opex + ac_amort + vium_unit + fot_unit
            total_cost_batch = total_unit_cost * Decimal(qty)

            label_name = v['name']
            if v.get('supplier_name'):
                label_name = f"{v['name']} · {v['supplier_name']}"

            impact_rows.append({
                'plant_id': v['plant_id'],
                'field_id': v['field_id'],
                'year': batch_year,
                'name': label_name,
                'field': v['field'],
                'quantity': qty,
                'purchase_price': avg_purchase_price,
                'accum_opex': ac_opex,
                'accum_amort': ac_amort,
                'vium_unit': vium_unit,
                'fot_unit': fot_unit,
                'total_unit_cost': total_unit_cost,
                'total_cost': total_cost_batch,
                'supplier_name': v.get('supplier_name') or '',
                'sample_size_id': v.get('sample_size_id'),
            })
            impact_total_qty += qty
            impact_table_total += total_cost_batch

        impact_rows.sort(key=lambda x: (x['name'], x['field'], x['year']))
    impact_summary_card = {'total_qty': impact_total_qty, 'year_cost': cd['summary_totals'][selected_year]['total'], 'unit_cost': cd['unit_cost_by_year'].get(selected_year, Decimal(0))}

    container_project_ids = get_cost_container_project_ids()
    container_cd = {
        'budget_items': {},
        'all_expenses': [],
        'qty_by_year': {},
        'summary_totals': {selected_year: {'total': Decimal(0), 'qty': 0, 'total_unit': Decimal(0)}},
        'unit_cost_by_year': {},
        'cumulative_cost': Decimal(0),
        'container_field_id': None,
        'container_field_label': None,
        'stock_as_of': None,
    }
    container_rows, container_total_qty, container_table_total = [], 0, Decimal(0)
    container_op_rows = []
    container_projects = []
    container_project_map = {}
    container_summary_card = {
        'total_qty': 0,
        'year_cost': Decimal(0),
        'unit_cost': Decimal(0),
    }
    if need_container_data:
        container_cd = calculate_container_cost_data(
            project_ids=container_project_ids,
            selected_year=selected_year,
            period=selected_period,
        )
        container_rows, container_total_qty, container_table_total = build_container_cost_table_rows(
            container_cd, plant_ids=f_plants or None,
        )

        for item in sorted(container_cd['budget_items'].values(), key=lambda x: natural_key(x.name)):
            row = {'name': item.name, 'data': {}}
            for y in visible_years:
                fact = Decimal(0)
                for yr, iid, amt in container_cd['all_expenses']:
                    if int(yr) == y and iid == item.id:
                        fact += amt
                qty = container_cd['qty_by_year'].get(y) or 0
                row['data'][y] = {
                    'sum': fact,
                    'per_unit': fact / Decimal(qty) if qty else Decimal(0),
                }
            if any(row['data'][y]['sum'] > 0 for y in visible_years):
                container_op_rows.append(row)

        container_projects = Project.query.order_by(Project.name.asc()).all()
        container_project_map = {p.id: p.name for p in container_projects}
        container_summary_card = {
            'total_qty': container_total_qty,
            'year_cost': container_cd['summary_totals'][selected_year]['total'],
            'unit_cost': container_cd['unit_cost_by_year'].get(selected_year, Decimal(0)),
        }
    
    all_plants = sorted(Plant.query.all(), key=lambda x: x.name) if need_impact_data else []
    all_items = sorted(cd['budget_items'].values(), key=lambda x: natural_key(x.name))
    
    op_rows, amort_rows = [], []
    for item in all_items:
        row = {'name': item.name, 'data': {}}
        for y in visible_years:
            fact = Decimal(0)
            for yr, iid, amt in cd['all_expenses']: 
                if int(yr) == y and iid == item.id: fact = amt
            qty = cd['qty_by_year'].get(y) or 1
            row['data'][y] = {'sum': fact, 'per_unit': fact / Decimal(qty) if cd['qty_by_year'].get(y) else Decimal(0)}
        (amort_rows if item.is_amortization else op_rows).append(row)
        
    if request.args.get('export') == 'excel':
        buf = _build_cost_report_xlsx(
            visible_years=visible_years,
            op_rows=op_rows,
            amort_rows=amort_rows,
            cd=cd,
            impact_rows=impact_rows,
            impact_total_qty=impact_total_qty,
            impact_table_total=impact_table_total,
            selected_year=selected_year,
            selected_period=selected_period,
        )
        period_suffix = f"_{selected_period}" if selected_period else ""
        return send_file(
            buf,
            download_name=f'cost_report_{selected_year}{period_suffix}.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    # ФОТ-плашка: даты с выкопкой без часов (только за выбранный год).
    fot_missing_dates = []
    try:
        from app import vium_fot
        from datetime import date as _d
        if selected_year:
            yr_start = _d(int(selected_year), 1, 1)
            today = msk_now().date()
            yr_end = _d(int(selected_year), 12, 31)
            if yr_end > today:
                yr_end = today
            if yr_end >= yr_start:
                fot_missing_dates = vium_fot.missing_hours_dates(yr_start, yr_end)
    except Exception:
        current_app.logger.exception('vium_fot.missing_hours_dates failed')

    return render_template('finance/cost.html', 
                           years=years, visible_years=visible_years, hidden_years=[], hidden_items=[], 
                           all_items=all_items, op_rows=op_rows, amort_rows=amort_rows,
                           summary_totals=cd['summary_totals'], waterfall=cd['amort_waterfall'], 
                           raw_source=cd['raw_amort_source'], active_amort_years=active_amort_years, 
                           impact_rows=impact_rows, cumulative_cost=cd['cumulative_cost'], 
                           selected_year=selected_year, selected_period=selected_period, current_year=current_year,
                           impact_summary_card=impact_summary_card, impact_table_total=impact_table_total, active_tab=active_tab, 
                           all_plants=all_plants, selected_plants=f_plants,
                           all_fields=Field.query.all() if need_impact_data else [], selected_fields=f_fields,
                           fot_missing_dates=fot_missing_dates,
                           vium_unit_by_year=cd.get('vium_unit_by_year', {}),
                           fot_unit_by_year=cd.get('fot_unit_by_year', {}),
                           container_rows=container_rows,
                           container_total_qty=container_total_qty,
                           container_table_total=container_table_total,
                           container_summary_card=container_summary_card,
                           container_cd=container_cd,
                           container_op_rows=container_op_rows,
                           container_project_ids=container_project_ids,
                           container_projects=container_projects,
                           container_project_map=container_project_map,
                           container_field_ids=container_field_ids,
                           container_field_label=container_cd.get('container_field_label'))

@bp.route('/cost/save_container_projects', methods=['POST'])
@login_required
def save_cost_container_projects():
    if current_user.role != 'admin':
        flash('Только администратор может менять проекты для контейнерной себестоимости')
        return redirect(url_for('finance.cost_report', tab='container'))

    raw_ids = request.form.getlist('container_project_id')
    try:
        ids = save_cost_container_project_ids(raw_ids)
        flash(f'Проекты для контейнерной себестоимости сохранены ({len(ids)} шт.)', 'success')
        log_action(f'Обновил проекты контейнерной себестоимости: {ids}')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}', 'danger')
    return redirect(url_for(
        'finance.cost_report',
        tab='container',
        filter_year=request.form.get('return_year', msk_now().year),
        period=request.form.get('return_period', ''),
    ))


@bp.route('/cost/save_override', methods=['POST'])
@login_required
def save_cost_override():
    if current_user.role != 'admin': return jsonify({'status': 'error'})
    try:
        year = int(request.form.get('year'))
        val_total = float(request.form.get('amount')) if request.form.get('amount') else None
        val_amort = float(request.form.get('amortization')) if request.form.get('amortization') else None

        override = UnitCostOverride.query.filter_by(year=year).first()
        if not override:
            override = UnitCostOverride(year=year)
            db.session.add(override)
        
        override.amount = val_total
        override.amortization = val_amort

        if override.amount is None and override.amortization is None:
            db.session.delete(override)
        
        db.session.commit()
        log_action(f"Обновил ручную себестоимость для {year} года")
        flash(f'Данные за {year} год обновлены')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка: {e}')
    
    return redirect(url_for('finance.cost_report', tab='summary', filter_year=request.form.get('return_year', msk_now().year)))

# --- ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ДЛЯ СВЕРКИ ---
def get_reconciliation_data(c_id, f_start, f_end, mode, use_fixed_balance=False):
    rows = []
    opening_balance = Decimal(0)
    current_bal = Decimal(0)
    
    client = Client.query.get(c_id)
    fixed_balance = Decimal(str(client.fixed_balance)) if client and client.fixed_balance is not None else None
    fixed_balance_date = client.fixed_balance_date if client else None
    
    # 1. РАСЧЕТ ВХОДЯЩЕГО САЛЬДО
    if f_start:
        f_start_date = datetime.strptime(f_start, '%Y-%m-%d').date()
        
        if use_fixed_balance and fixed_balance is not None and fixed_balance_date is not None and fixed_balance_date < f_start_date:
            base_balance = fixed_balance
            base_date_filter = func.date(Document.date) > fixed_balance_date
            base_date_filter_order = func.date(Order.date) > fixed_balance_date
            base_date_filter_payment = func.date(Payment.date) > fixed_balance_date
        else:
            base_balance = Decimal(0)
            base_date_filter = True
            base_date_filter_order = True
            base_date_filter_payment = True

        past_shipments_real = db.session.query(func.sum(OrderItem.price * DocumentRow.quantity))\
            .select_from(Document)\
            .join(Order, Document.order_id == Order.id)\
            .join(DocumentRow, DocumentRow.document_id == Document.id)\
            .join(OrderItem, and_(
                OrderItem.order_id == Document.order_id,
                OrderItem.plant_id == DocumentRow.plant_id,
                OrderItem.size_id == DocumentRow.size_id,
                OrderItem.field_id == DocumentRow.field_from_id
            ))\
            .filter(Order.client_id == c_id, Document.doc_type == 'shipment', func.date(Document.date) < f_start, base_date_filter).scalar() or Decimal(0)

        past_shipments_ghost = db.session.query(func.sum(OrderItem.price * OrderItem.quantity))\
            .join(Order)\
            .filter(Order.client_id == c_id, Order.status == 'ghost', func.date(Order.date) < f_start, base_date_filter_order).scalar() or Decimal(0)

        past_payments = db.session.query(func.sum(Payment.amount))\
            .join(Order)\
            .filter(Order.client_id == c_id, func.date(Payment.date) < f_start, base_date_filter_payment).scalar() or Decimal(0)
        
        opening_balance = base_balance + (Decimal(str(past_shipments_real)) + Decimal(str(past_shipments_ghost))) - Decimal(str(past_payments))
        
        rows.append({
            'date': datetime.strptime(f_start, '%Y-%m-%d'), 
            'sort_date': datetime.strptime(f_start, '%Y-%m-%d'),
            'desc': 'Входящее сальдо', 'debit': Decimal(0), 'credit': Decimal(0), 
            'is_balance': True, 'doc_items': [], 'invoice_info': '', 'balance': opening_balance,
            'has_payments': False
        })
    elif use_fixed_balance and fixed_balance is not None and fixed_balance_date is not None:
        opening_balance = fixed_balance
        dt = datetime.combine(fixed_balance_date, datetime.min.time())
        rows.append({
            'date': dt, 
            'sort_date': dt,
            'desc': 'Фиксированное сальдо (Корректировка)', 'debit': Decimal(0), 'credit': Decimal(0), 
            'is_balance': True, 'doc_items': [], 'invoice_info': '', 'balance': opening_balance,
            'has_payments': False
        })

    # 2. СБОР ДАННЫХ
    raw_rows = []
    
    # Предварительно загружаем ID заказов, у которых есть оплаты (для кнопки гармошки)
    # Это нужно, чтобы знать, рисовать ли кнопку "показать оплаты" у отгрузки
    paid_orders_query = db.session.query(Payment.order_id).join(Order).filter(Order.client_id == c_id).distinct()
    paid_orders_ids = set(r[0] for r in paid_orders_query.all())

    # А. Документы (Отгрузки) — pre-load order items for price lookup
    ship_query = db.session.query(Document).join(Order).filter(Order.client_id == c_id, Document.doc_type == 'shipment')
    if f_start: ship_query = ship_query.filter(func.date(Document.date) >= f_start)
    if f_end: ship_query = ship_query.filter(func.date(Document.date) <= f_end)
    if use_fixed_balance and fixed_balance_date is not None:
        ship_query = ship_query.filter(func.date(Document.date) > fixed_balance_date)
    
    ship_docs = ship_query.all()
    ship_order_ids = list(set(d.order_id for d in ship_docs if d.order_id))
    oi_price_map = {}
    if ship_order_ids:
        oi_rows = OrderItem.query.filter(OrderItem.order_id.in_(ship_order_ids)).all()
        for oi in oi_rows:
            oi_price_map[(oi.order_id, oi.plant_id, oi.size_id, oi.field_id)] = oi.price

    for doc in ship_docs:
        doc_sum = Decimal(0); doc_items_list = []
        for r in doc.rows:
            raw_price = oi_price_map.get((doc.order_id, r.plant_id, r.size_id, r.field_from_id))
            price = Decimal(str(raw_price)) if raw_price is not None else Decimal(0)
            row_sum = price * Decimal(r.quantity)
            doc_sum += row_sum
            doc_items_list.append({'date': doc.date, 'name': r.plant.name if r.plant else '-', 'size': r.size.name if r.size else '-', 'field': r.field_from.name if r.field_from else '-', 'qty': r.quantity, 'price': price, 'sum': row_sum})
        
        inv_info = f"Счет № {doc.order.invoice_number} от {doc.order.invoice_date.strftime('%d.%m.%Y')}" if (doc.order and doc.order.invoice_number and doc.order.invoice_date) else (f"Счет № {doc.order.invoice_number}" if (doc.order and doc.order.invoice_number) else "")
        
        raw_rows.append({
            'date': doc.date, 
            'sort_date': doc.date, 
            'desc': f"Отгрузка (Заказ #{doc.order_id})", 
            'invoice_info': inv_info, 
            'debit': doc_sum, 
            'credit': Decimal(0), 
            'order_id': doc.order_id, 
            'doc_items': doc_items_list, 
            'is_grouped': False,
            'has_payments': (doc.order_id in paid_orders_ids) # Флаг наличия оплат
        })

    # Б. Ghost (Архивные отгрузки)
    ghost_query = Order.query.filter(Order.client_id == c_id, Order.status == 'ghost')
    if f_start: ghost_query = ghost_query.filter(func.date(Order.date) >= f_start)
    if f_end: ghost_query = ghost_query.filter(func.date(Order.date) <= f_end)
    if use_fixed_balance and fixed_balance_date is not None:
        ghost_query = ghost_query.filter(func.date(Order.date) > fixed_balance_date)
    
    for o in ghost_query.all():
        g_sum = Decimal(0); ghost_items_list = []
        for i in o.items:
            price = Decimal(str(i.price)) if i.price else Decimal(0)
            row_sum = price * Decimal(i.quantity)
            g_sum += row_sum
            ghost_items_list.append({'date': o.date, 'name': i.plant.name, 'size': i.size.name, 'field': i.field.name, 'qty': i.quantity, 'price': price, 'sum': row_sum})
        
        inv_info = f"Счет № {o.invoice_number} от {o.invoice_date.strftime('%d.%m.%Y')}" if (o.invoice_number and o.invoice_date) else (f"Счет № {o.invoice_number}" if o.invoice_number else "")
        
        raw_rows.append({
            'date': o.date, 
            'sort_date': o.date, 
            'desc': f"Архив. отгрузка #{o.id}", 
            'invoice_info': inv_info, 
            'debit': g_sum, 
            'credit': Decimal(0), 
            'order_id': o.id, 
            'doc_items': ghost_items_list, 
            'is_grouped': False,
            'has_payments': (o.id in paid_orders_ids) # Флаг наличия оплат
        })

    # В. Оплаты
    # Изменили запрос: берем Payment и Order целиком, чтобы достать invoice_number
    pay_query = db.session.query(Payment, Order).join(Order).filter(Order.client_id == c_id)
    if f_start: pay_query = pay_query.filter(func.date(Payment.date) >= f_start)
    if f_end: pay_query = pay_query.filter(func.date(Payment.date) <= f_end)
    if use_fixed_balance and fixed_balance_date is not None:
        pay_query = pay_query.filter(func.date(Payment.date) > fixed_balance_date)
    
    for p, o in pay_query.all():
        display_date = datetime.combine(p.date, datetime.min.time())
        sort_date_val = (o.date + timedelta(seconds=1)) if o.date else display_date
        
        # Формируем информацию о счете так же, как в отгрузках, для группировки
        inv_info = f"Счет № {o.invoice_number} от {o.invoice_date.strftime('%d.%m.%Y')}" if (o.invoice_number and o.invoice_date) else (f"Счет № {o.invoice_number}" if o.invoice_number else "")
        
        # Создаем элемент детализации для оплаты, чтобы показать его внутри группы
        payment_item_detail = {
            'date': display_date,
            'name': f"Оплата от {p.date.strftime('%d.%m.%Y')}",
            'size': '-',
            'field': '-',
            'qty': 1,
            'price': Decimal(str(p.amount)),
            'sum': Decimal(str(p.amount)),
            'type': 'payment' # Метка типа строки
        }

        raw_rows.append({
            'date': display_date, 
            'sort_date': sort_date_val, 
            'desc': f"Оплата (по заказу #{p.order_id})", 
            'debit': Decimal(0), 
            'credit': Decimal(str(p.amount)), 
            'order_id': p.order_id, 
            'doc_items': [payment_item_detail], # Кладем детализацию сюда
            'invoice_info': inv_info, # Теперь здесь есть номер счета!
            'is_grouped': False,
            'has_payments': False,
            'is_payment_row': True 
        })

    # 3. ГРУППИРОВКА И СОРТИРОВКА
    if mode == 'grouped':
        grouped_map = {}
        final_rows = []
        
        for r in raw_rows:
            # Группируем всё (и отгрузки, и оплаты), если есть номер счета
            if r.get('invoice_info'):
                key = r['invoice_info']
                
                if key not in grouped_map:
                    # Создаем новую группу
                    new_group = r.copy()
                    new_group['is_grouped'] = True
                    # Если первая запись была оплатой, меняем описание, иначе оставляем "Реализация..." или меняем на общее
                    new_group['desc'] = "Операции по счету" 
                    new_group['doc_items'] = list(r['doc_items'])
                    # Инициализируем суммы, так как r.copy() взяло только текущие значения
                    new_group['debit'] = r['debit']
                    new_group['credit'] = r['credit']
                    
                    grouped_map[key] = new_group
                else:
                    # Добавляем к существующей группе
                    grouped_map[key]['debit'] += r['debit']
                    grouped_map[key]['credit'] += r['credit'] # Суммируем оплаты
                    grouped_map[key]['doc_items'].extend(r['doc_items'])
                    
                    # Обновляем даты по самой ранней операции
                    if r['sort_date'] < grouped_map[key]['sort_date']: grouped_map[key]['sort_date'] = r['sort_date']
                    if r['date'] < grouped_map[key]['date']: grouped_map[key]['date'] = r['date']
                    
                    # Если это была запись отгрузки (есть order_id с флагом), сохраняем флаг
                    if r.get('has_payments'): grouped_map[key]['has_payments'] = True
            else:
                # Если счета нет - оставляем как есть (например, входящее сальдо или оплата без привязки)
                final_rows.append(r)
        
        # Сортируем детали внутри групп по дате
        for grp in grouped_map.values(): 
            grp['doc_items'].sort(key=lambda x: x['date'])
            
        final_rows.extend(grouped_map.values())
        raw_rows = final_rows

    # Сортировка по дате привязки
    raw_rows.sort(key=lambda x: x['sort_date'])
    
    # 4. РАСЧЕТ ИТОГОВОГО САЛЬДО И ОБОРОТОВ
    current_bal = opening_balance
    
    # Переменные для итогов по колонкам
    total_debit = Decimal(0)
    total_credit = Decimal(0)

    for r in raw_rows:
        # Суммируем обороты за период (не считая входящее сальдо, если оно реализовано как строка)
        # В твоем коде входящее сальдо имеет флаг 'is_balance', его в обороты включать не надо
        if not r.get('is_balance'):
            total_debit += r['debit']
            total_credit += r['credit']

        current_bal += r['debit'] - r['credit']
        r['balance'] = current_bal
        rows.append(r)
        
    return rows, current_bal, total_debit, total_credit

@bp.route('/reports/reconciliation')
@login_required
def reports_reconciliation():
    clients = Client.query.all()
    cid = request.args.get('client_id')
    f_start = request.args.get('start_date')
    f_end = request.args.get('end_date')
    mode = request.args.get('mode', 'grouped')
    
    rows = []
    total_bal = Decimal(0)
    t_debit = Decimal(0)
    t_credit = Decimal(0)
    client_name = ''
    barter_orders = []
    
    current_url = request.url

    if cid:
        try:
            client_obj = Client.query.get(int(cid))
            client_name = client_obj.name if client_obj else "Неизвестный"
            rows, total_bal, t_debit, t_credit = get_reconciliation_data(int(cid), f_start, f_end, mode)
            
            # Получаем бартерные заказы клиента
            b_orders = Order.query.filter(Order.client_id == int(cid), Order.is_barter == True, Order.status != 'canceled', Order.is_deleted == False).all()
            for bo in b_orders:
                bo_total = bo.total_sum
                bo_paid = sum(e.amount for e in bo.barter_expenses)
                barter_orders.append({
                    'order': bo,
                    'total': bo_total,
                    'paid': bo_paid,
                    'remaining': bo_total - bo_paid
                })
        except Exception as e:
            flash(f"Ошибка: {str(e)}")
            print(f"RECONCILIATION ERROR: {e}")

    # Передаем debit и credit в totals
    return render_template('finance/reports.html', active_tab='reconciliation', clients=clients, rows=rows, 
                           totals={'balance': total_bal, 'debit': t_debit, 'credit': t_credit}, 
                           filters={'client_id': cid, 'start': f_start, 'end': f_end},
                           client_name=client_name, mode=mode,
                           current_url=current_url,
                           barter_orders=barter_orders)

@bp.route('/reports/reconciliation/export')
@login_required
def reports_reconciliation_export():
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    cid = request.args.get('client_id')
    f_start = request.args.get('start_date')
    f_end = request.args.get('end_date')
    mode = request.args.get('mode', 'grouped')
    
    if not cid: return redirect(url_for('finance.reports_reconciliation'))
    
    client_obj = Client.query.get(int(cid))
    client_name = client_obj.name if client_obj else "Неизвестный"
    
    # Получаем данные (ИСПРАВЛЕНО: принимаем 4 значения)
    rows, final_balance, t_debit, t_credit = get_reconciliation_data(int(cid), f_start, f_end, mode)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Акт сверки"
    
    # Стили
    bold_font = Font(bold=True)
    italic_font = Font(italic=True, size=10, color="555555") # Для деталей
    header_fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
    detail_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Фон для деталей
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # Заголовок
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Акт сверки с {client_name}"
    ws['A1'].font = Font(size=14, bold=True, color="2E7D32")
    ws['A1'].alignment = align_center
    
    ws.merge_cells('A2:F2')
    period_str = f"Период: {datetime.strptime(f_start, '%Y-%m-%d').strftime('%d.%m.%Y') if f_start else '...'} - {datetime.strptime(f_end, '%Y-%m-%d').strftime('%d.%m.%Y') if f_end else '...'}"
    ws['A2'] = period_str
    ws['A2'].alignment = align_center
    
    # Шапка таблицы
    headers = ["Дата", "Документ / Операция", "Основание (Счет)", "Дебет (Отгрузка)", "Кредит (Оплата)", "Сальдо"]
    ws.append([]) 
    ws.append(headers) 
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = align_center

    # Данные
    for r in rows:
        # ОСНОВНАЯ СТРОКА
        row_cells = [
            r['date'].strftime('%d.%m.%Y'),
            r['desc'],
            r['invoice_info'] or '-',
            float(r['debit']) if r['debit'] else 0,
            float(r['credit']) if r['credit'] else 0,
            float(r['balance'])
        ]
        ws.append(row_cells)
        curr_row = ws.max_row
        
        # Стилизация основной строки
        for i in range(1, 7):
            cell = ws.cell(row=curr_row, column=i)
            cell.border = border_style
            if i in [4, 5, 6]: 
                cell.number_format = '#,##0.00'
                cell.alignment = align_right
            
            if r.get('is_grouped') or r.get('is_balance'):
                cell.font = bold_font
            
            if r.get('is_balance'): cell.fill = PatternFill(start_color="FFF3E0", fill_type="solid")
            elif r.get('is_grouped'): cell.fill = PatternFill(start_color="FFF8E1", fill_type="solid") # Желтоватый для группы

        # ДЕТАЛИЗАЦИЯ (ГАРМОШКА) В EXCEL
        if r.get('is_grouped') and r.get('doc_items'):
            for item in r['doc_items']:
                # Формируем строку детализации
                # Смещаем данные: 
                # A: Дата отгрузки
                # B: Название + Размер + Поле
                # C: Пусто
                # D: Сумма (в дебет)
                # E-F: Пусто
                
                item_desc = f"   ↳ {item['name']} ({item['size']}) / {item['field']} / {item['qty']} шт x {float(item['price'])}"
                
                detail_cells = [
                    item['date'].strftime('%d.%m.%Y'), # Дата конкретной отгрузки
                    item_desc,
                    "",
                    float(item['sum']),
                    "",
                    ""
                ]
                ws.append(detail_cells)
                det_row = ws.max_row
                
                # Стилизация детали
                for i in range(1, 7):
                    cell = ws.cell(row=det_row, column=i)
                    cell.font = italic_font
                    cell.fill = detail_fill
                    cell.border = border_style # Можно убрать border, если хочется "воздуха"
                    
                    if i == 4: # Сумма
                        cell.number_format = '#,##0.00'
                        cell.alignment = align_right

    # Итоги
    ws.append([])
    # ИСПРАВЛЕНО: Добавляем t_debit и t_credit в строку итогов
    ws.append(["", "", "ИТОГО ОБОРОТЫ И САЛЬДО:", float(t_debit), float(t_credit), float(final_balance)])
    last_row = ws.max_row
    
    ws.cell(row=last_row, column=3).alignment = align_right
    ws.cell(row=last_row, column=3).font = bold_font
    
    # Стилизуем итоги оборотов (Дебет и Кредит)
    for col_idx in [4, 5]:
        cell = ws.cell(row=last_row, column=col_idx)
        cell.font = bold_font
        cell.number_format = '#,##0.00'
        cell.border = border_style
        cell.alignment = align_right

    # Стилизуем Сальдо
    bal_cell = ws.cell(row=last_row, column=6)
    bal_cell.font = Font(bold=True, color="C62828" if final_balance > 0 else "2E7D32")
    bal_cell.number_format = '#,##0.00'
    bal_cell.border = border_style

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50 # Шире для деталей
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"Act_{client_name}_{f_start or ''}_{f_end or ''}.xlsx"
    return send_file(buf, download_name=filename, as_attachment=True)

@bp.route('/reports/turnover')
@login_required
def reports_turnover():
    if current_user.role == 'user2':
        return redirect(url_for('main.index'))
    
    f_plant = request.args.get('plant_id', type=int)
    f_field = request.args.get('field_id', type=int)
    f_year = request.args.get('year', type=int)
    f_start = request.args.get('start_date')
    f_end = request.args.get('end_date')
    
    rows =[]
    opening_balance = 0
    closing_balance = 0

    if f_plant and f_field and f_year:
        doc_rows = db.session.query(DocumentRow, Document).join(Document).filter(
            DocumentRow.plant_id == f_plant, DocumentRow.year == f_year, 
            or_(DocumentRow.field_from_id == f_field, DocumentRow.field_to_id == f_field)
        ).all()
        
        ghosts = db.session.query(OrderItem, Order).join(Order).filter(
            OrderItem.plant_id == f_plant, OrderItem.field_id == f_field, OrderItem.year == f_year, Order.status == 'ghost'
        ).all()
        
        all_events =[]
        for r, d in doc_rows:
            qty = r.quantity
            is_income = False
            is_expense = False
            doc_name = ""
            
            if d.doc_type in ('income', 'correction'):
                if r.field_to_id == f_field:
                    is_income = True
                    # НОВАЯ ЛОГИКА: если год <= 2025, то это "Первоначальный импорт остатков"
                    if d.date.year <= 2025:
                        doc_name = "Первоначальный импорт остатков"
                    else:
                        doc_name = "Поступление" if d.doc_type == 'income' else "Импорт"
            elif d.doc_type in ('writeoff', 'shipment'):
                if r.field_from_id == f_field:
                    is_expense = True
                    doc_name = "Списание" if d.doc_type == 'writeoff' else f"Отгрузка (Заказ #{d.order_id})"
            elif d.doc_type == 'move':
                if r.field_to_id == f_field:
                    is_income = True
                    doc_name = "Перемещение (ВХ)"
                elif r.field_from_id == f_field:
                    is_expense = True
                    doc_name = "Перемещение (ИСХ)"
            
            if is_income or is_expense:
                all_events.append({
                    'date': d.date, 
                    'type': 'real', 
                    'doc_name': doc_name, 
                    'income': qty if is_income else 0, 
                    'expense': qty if is_expense else 0, 
                    'author': d.user.username if d.user else '', 
                    'comment': d.comment, 
                    'size_name': r.size.name if r.size else ''
                })

        for item, order in ghosts:
            all_events.append({
                'date': order.date, 
                'type': 'ghost', 
                'doc_name': f"Ghost #{order.id}", 
                'income': 0, 
                'expense': item.quantity, 
                'author': 'Ghost', 
                'comment': order.client.name, 
                'size_name': item.size.name if item.size else ''
            })

        all_events.sort(key=lambda x: x.get('date'))
        
        current_balance = 0
        start_dt = datetime.strptime(f_start, '%Y-%m-%d') if f_start else datetime.min
        end_dt = datetime.strptime(f_end, '%Y-%m-%d').replace(hour=23, minute=59) if f_end else datetime.max

        for ev in all_events:
            if ev.get('type') == 'real':
                current_balance += ev.get('income', 0) - ev.get('expense', 0)
            
            ev['balance'] = current_balance
            
            if ev.get('date') < start_dt and ev.get('type') == 'real':
                opening_balance = current_balance
            
            if start_dt <= ev.get('date') <= end_dt:
                rows.append(ev)
        
        closing_balance = current_balance

    return render_template('finance/reports.html', active_tab='turnover', rows=rows, opening_balance=opening_balance, closing_balance=closing_balance, plants=Plant.query.all(), sizes=Size.query.all(), fields=Field.query.all(), filters=request.args)

@bp.route('/reports/financial')
@login_required
def reports_financial():
    if current_user.role not in ['admin', 'executive']: return redirect(url_for('main.index'))
    year = int(request.args.get('year', msk_now().year))
    
    # Исправление: Добавлена обработка NULL (None) при суммировании
    expenses = db.session.query(Expense.payment_type, func.sum(Expense.amount)).filter(func.extract('year', Expense.date) == year).group_by(Expense.payment_type).all()
    
    # Безопасное получение сумм: если расхода нет, ставим 0
    ec = next((x[1] for x in expenses if x[0] == 'cash'), Decimal(0))
    if ec is None: ec = Decimal(0)
    
    ecl = next((x[1] for x in expenses if x[0] == 'cashless'), Decimal(0))
    if ecl is None: ecl = Decimal(0)
    
    cd_prev = calculate_cost_data(year - 1)
    debt, breakdown = calculate_investor_debt(year, cd_prev['accumulated_costs_map'])
    
    total_exp = ec + ecl + (ec * Decimal('0.15')) + debt
    
    orders = Order.query.filter(func.extract('year', Order.date) == year, Order.status != 'canceled', Order.is_deleted == False).all()
    
    val_shipped = Decimal(0); val_reserved = Decimal(0)
    for o in orders:
        for item in o.items:
            # Защита от None в цене
            price = item.price if item.price is not None else Decimal(0)
            val_shipped += Decimal(item.shipped_quantity) * price
            reserved_qty = item.quantity - item.shipped_quantity
            if reserved_qty > 0: val_reserved += Decimal(reserved_qty) * price
            
    data = {
        'exp_cash': ec, 
        'exp_cashless': ecl, 
        'tax_cash_15': ec * Decimal('0.15'), 
        'investor_debt': debt, 
        'investor_breakdown': breakdown, 
        'total_expenses': total_exp, 
        'realized_revenue': val_shipped, 
        'net_profit': val_shipped - total_exp, 
        'val_reserved': val_reserved, 
        'val_shipped': val_shipped, 
        'val_partial_shipped': 0, 
        'val_partial_unshipped': 0, 
        'total_sales_potential': val_shipped + val_reserved
    }
    return render_template('finance/reports.html', active_tab='financial', years=range(2017, 2051), selected_year=year, data=data)

@bp.route('/reports/margin')
@login_required
def reports_margin():
    if current_user.role not in ['admin', 'executive']: return redirect(url_for('main.index'))
    year = int(request.args.get('year', msk_now().year))
    view = (request.args.get('view') or 'snapshot').lower()
    if view not in ('snapshot', 'dynamics'):
        view = 'snapshot'

    # Получаем списки ID для фильтрации
    f_plants = [int(x) for x in request.args.getlist('filter_plant')]
    f_sizes = [int(x) for x in request.args.getlist('filter_size')]
    f_fields = [int(x) for x in request.args.getlist('filter_field')]

    rows = []
    dynamics = None

    if view == 'snapshot':
        cost_data = calculate_cost_data(year)
        accum = cost_data['accumulated_costs_map']

        for r in get_detailed_stock_at_year_end(year):
            if f_plants and r['plant_id'] not in f_plants: continue
            if f_sizes and r['size_id'] not in f_sizes: continue
            if f_fields and r['field_id'] not in f_fields: continue

            tc = r['purchase_price'] + accum.get(r['year'], Decimal(0))
            margin = r['selling_price'] - tc
            margin_percent = (margin / r['selling_price'] * 100) if r['selling_price'] > 0 else 0

            rows.append({
                'name': r['name'], 'size': r['size'], 'field': r['field'],
                'price': r['selling_price'], 'cost': tc,
                'margin': margin, 'margin_percent': margin_percent,
                'size_name': r['size'],
            })
    else:
        dynamics = _build_margin_dynamics(
            year_from=int(request.args.get('year_from') or (year - 4)),
            year_to=int(request.args.get('year_to') or year),
            f_plants=f_plants, f_sizes=f_sizes, f_fields=f_fields,
        )

    return render_template('finance/reports.html', active_tab='margin', years=range(2017, 2051),
                           selected_year=year, rows=rows,
                           view=view, dynamics=dynamics,
                           all_plants=Plant.query.order_by(Plant.name).all(),
                           all_sizes=Size.query.all(),
                           all_fields=Field.query.all(),
                           filters={'plants': f_plants, 'sizes': f_sizes, 'fields': f_fields})


def _build_margin_dynamics(year_from, year_to, f_plants=None, f_sizes=None, f_fields=None):
    """Динамика фактической маржинальности в разрезе растения × года отгрузки.

    Маржа считается по фактически отгруженным позициям (shipped_quantity > 0):
    - выручка = price * shipped_quantity (из OrderItem);
    - себестоимость = (purchase_price партии + накопленная себестоимость до year_of_sale - 1) * qty;
    - год отгрузки определяем по Order.date.year (исключаем canceled/ghost/is_deleted).

    Возвращает структуру, удобную для рендера:
    {
        'years': [2020, 2021, ...],
        'plants': [ {plant_id, name, cells: {year: {qty, revenue, cost, margin, pct}},
                     total_qty, total_revenue, total_cost, total_margin, total_pct,
                     trend_delta, spark_points: [(idx, pct), ...],
                     sizes: [ { size_name, cells: {...}, total_* } ]
                   }, ...],
        'year_totals': {year: {qty, revenue, cost, margin, pct}},
        'grand_total': {qty, revenue, cost, margin, pct},
    }
    """
    from sqlalchemy.orm import joinedload

    if year_from > year_to:
        year_from, year_to = year_to, year_from
    years = list(range(year_from, year_to + 1))
    f_plants = f_plants or []
    f_sizes = f_sizes or []
    f_fields = f_fields or []

    # Кеш цен закупки по партиям
    stock_prices = {
        (sb.plant_id, sb.size_id, sb.field_id, sb.year): (sb.purchase_price or Decimal(0))
        for sb in StockBalance.query.all()
    }
    # Кеш накопленной себестоимости: {year_basis: {batch_year: Decimal}}
    costs_cache = {}

    q = db.session.query(OrderItem, Order).join(Order).options(
        joinedload(OrderItem.plant), joinedload(OrderItem.size)
    ).filter(
        Order.status != 'canceled',
        Order.status != 'ghost',
        Order.is_deleted == False,  # noqa: E712
        OrderItem.shipped_quantity > 0,
        func.extract('year', Order.date) >= year_from,
        func.extract('year', Order.date) <= year_to,
    )
    if f_plants:
        q = q.filter(OrderItem.plant_id.in_(f_plants))
    if f_sizes:
        q = q.filter(OrderItem.size_id.in_(f_sizes))
    if f_fields:
        q = q.filter(OrderItem.field_id.in_(f_fields))

    def _empty_cell():
        return {'qty': Decimal(0), 'revenue': Decimal(0), 'cost': Decimal(0),
                'margin': Decimal(0), 'pct': 0.0}

    plants_map = {}  # plant_id -> {name, cells{year:_cell}, sizes{size_name:_cell}, ...}

    for item, order in q.all():
        ysale = order.date.year
        if ysale not in years:
            continue
        qty = Decimal(item.shipped_quantity or 0)
        if qty <= 0:
            continue
        price = item.price if item.price is not None else Decimal(0)
        revenue = price * qty

        basis = ysale - 1
        if basis not in costs_cache:
            costs_cache[basis] = calculate_cost_data(basis).get('accumulated_costs_map', {})
        accum_unit = costs_cache[basis].get(item.year, Decimal(0))
        purch_unit = stock_prices.get((item.plant_id, item.size_id, item.field_id, item.year), Decimal(0))
        cost_unit = (purch_unit or Decimal(0)) + (accum_unit or Decimal(0))
        cost = cost_unit * qty

        pid = item.plant_id
        pdata = plants_map.setdefault(pid, {
            'plant_id': pid,
            'name': item.plant.name if item.plant else f'Растение #{pid}',
            'cells': {y: _empty_cell() for y in years},
            'sizes_map': {},
            'total_qty': Decimal(0), 'total_revenue': Decimal(0),
            'total_cost': Decimal(0), 'total_margin': Decimal(0),
        })

        c = pdata['cells'][ysale]
        c['qty'] += qty
        c['revenue'] += revenue
        c['cost'] += cost
        c['margin'] = c['revenue'] - c['cost']

        pdata['total_qty'] += qty
        pdata['total_revenue'] += revenue
        pdata['total_cost'] += cost

        # Детализация по размеру
        size_name = item.size.name if item.size else '—'
        sdata = pdata['sizes_map'].setdefault(size_name, {
            'size_name': size_name,
            'cells': {y: _empty_cell() for y in years},
            'total_qty': Decimal(0), 'total_revenue': Decimal(0),
            'total_cost': Decimal(0), 'total_margin': Decimal(0),
        })
        sc = sdata['cells'][ysale]
        sc['qty'] += qty
        sc['revenue'] += revenue
        sc['cost'] += cost
        sc['margin'] = sc['revenue'] - sc['cost']
        sdata['total_qty'] += qty
        sdata['total_revenue'] += revenue
        sdata['total_cost'] += cost

    # Постобработка: проценты, тренд, sparkline
    def _finalize(d):
        for y, c in d['cells'].items():
            c['pct'] = float(c['margin'] / c['revenue'] * 100) if c['revenue'] > 0 else 0.0
        d['total_margin'] = d['total_revenue'] - d['total_cost']
        d['total_pct'] = float(d['total_margin'] / d['total_revenue'] * 100) if d['total_revenue'] > 0 else 0.0
        non_empty = [(y, d['cells'][y]['pct']) for y in years if d['cells'][y]['revenue'] > 0]
        if len(non_empty) >= 2:
            d['trend_delta'] = non_empty[-1][1] - non_empty[0][1]
            d['first_year'] = non_empty[0][0]
            d['last_year'] = non_empty[-1][0]
        else:
            d['trend_delta'] = 0.0
            d['first_year'] = d['last_year'] = None
        d['spark_points'] = [(i, d['cells'][y]['pct'] if d['cells'][y]['revenue'] > 0 else None)
                             for i, y in enumerate(years)]

    year_totals = {y: _empty_cell() for y in years}
    plants = []
    for pid, pdata in plants_map.items():
        _finalize(pdata)
        # Сортируем размеры по выручке
        sizes = sorted(pdata['sizes_map'].values(), key=lambda s: s['total_revenue'], reverse=True)
        for s in sizes:
            _finalize(s)
        pdata['sizes'] = sizes
        pdata.pop('sizes_map', None)
        plants.append(pdata)
        # Копим общие итоги по годам
        for y in years:
            c = pdata['cells'][y]
            yt = year_totals[y]
            yt['qty'] += c['qty']
            yt['revenue'] += c['revenue']
            yt['cost'] += c['cost']
            yt['margin'] = yt['revenue'] - yt['cost']

    for y, yt in year_totals.items():
        yt['pct'] = float(yt['margin'] / yt['revenue'] * 100) if yt['revenue'] > 0 else 0.0

    grand = {'qty': Decimal(0), 'revenue': Decimal(0), 'cost': Decimal(0), 'margin': Decimal(0), 'pct': 0.0}
    for yt in year_totals.values():
        grand['qty'] += yt['qty']
        grand['revenue'] += yt['revenue']
        grand['cost'] += yt['cost']
    grand['margin'] = grand['revenue'] - grand['cost']
    grand['pct'] = float(grand['margin'] / grand['revenue'] * 100) if grand['revenue'] > 0 else 0.0

    # Сортируем растения по выручке (от большего к меньшему)
    plants.sort(key=lambda p: p['total_revenue'], reverse=True)

    return {
        'years': years,
        'year_from': year_from,
        'year_to': year_to,
        'plants': plants,
        'year_totals': year_totals,
        'grand_total': grand,
    }

@bp.route('/reports/investor')
@login_required
def reports_investor():
    if current_user.role not in ['admin', 'executive']: return redirect(url_for('main.index'))
    f_start = request.args.get('start_date')
    f_end = request.args.get('end_date')
    real_picture = request.args.get('real_picture', '0') == '1'
    use_fixed_balance = not real_picture
    
    # Списки ID для мульти-фильтра
    f_fields = [int(x) for x in request.args.getlist('filter_field')]
    f_plants = [int(x) for x in request.args.getlist('filter_plant')]
    
    # Условия выборки
    conds = [
        Order.status != 'canceled', 
        Order.is_deleted == False, 
        OrderItem.shipped_quantity > 0, 
        Field.investor_id.isnot(None),
        or_(
            func.extract('year', Order.date) < 2025, 
            Order.client_id != Field.investor_id
        )
    ]
    
    if f_start: conds.append(func.date(Order.date) >= f_start)
    if f_end: conds.append(func.date(Order.date) <= f_end)
    if f_fields: conds.append(OrderItem.field_id.in_(f_fields))
    if f_plants: conds.append(OrderItem.plant_id.in_(f_plants))
    
    # Запрос детализации (Главный запрос)
    details_query = db.session.query(OrderItem, Order, Field, Client)\
        .select_from(OrderItem)\
        .join(Order).join(Field).join(Client, Order.client_id == Client.id)\
        .filter(and_(*conds))\
        .order_by(Order.date.desc())
    
    raw_details = details_query.all()
    
    # Кэши
    stock_prices = { (sb.plant_id, sb.size_id, sb.field_id, sb.year): sb.purchase_price for sb in StockBalance.query.all() }
    costs_cache = {} # { year: { batch_year: amount } }
    
    investor_details = []
    investors_summary = {} # { inv_id: {name, share, balance_act, net_result} }
    
    total_period_share = Decimal(0) # Общая сумма доли за период по всем строкам

    for item, order, field, client in raw_details:
        # 1. Основные данные
        qty_det = Decimal(item.shipped_quantity)
        price_det = item.price if item.price is not None else Decimal(0)
        is_return = (field.investor_id == order.client_id)
        
        # 2. Цена закупки (из партии)
        purch_price = stock_prices.get((item.plant_id, item.size_id, item.field_id, item.year), Decimal(0))
        
        # 3. Накопленная себестоимость (База = Год продажи - 1)
        # Мы НЕ смотрим в базу (StockBalance.current_total_cost), мы считаем динамически через сервис
        ship_year = order.date.year
        calc_basis_year = ship_year - 1
        
        if calc_basis_year not in costs_cache:
            costs_cache[calc_basis_year] = calculate_cost_data(calc_basis_year)['accumulated_costs_map']
            
        accum_cost = costs_cache[calc_basis_year].get(item.year, Decimal(0))
        
        # 4. Финальная математика
        full_cost_unit = purch_price + accum_cost
        share_unit = (price_det / 2) - full_cost_unit
        total_share_row = share_unit * qty_det
        
        # 5. Сбор данных в список
        investor_details.append({
            'date': order.date, 
            'order_id': order.id, 
            'plant': item.plant.name, 
            'size': item.size.name,
            'field': field.name, 
            'batch_year': item.year, # Год партии
            'investor': field.investor.name if field.investor else '-',
            'buyer': client.name, 
            'type_code': 2 if is_return else 1,
            'qty': int(qty_det), 
            'price': price_det,
            
            # Новые поля
            'full_cost_unit': full_cost_unit, # Полная себестоимость
            'share_unit': share_unit,         # Доля с 1 шт
            'total_share': total_share_row    # Итого доля строки
        })
        
        # 6. Агрегация для сводки
        total_period_share += total_share_row
        
        if field.investor_id:
            inv_id = field.investor_id
            if inv_id not in investors_summary:
                investors_summary[inv_id] = {'name': field.investor.name, 'share': Decimal(0), 'balance_act': Decimal(0), 'net_result': Decimal(0)}
            investors_summary[inv_id]['share'] += total_share_row

    # 7. Догружаем баланс из Акта сверки (Сальдо)
    for inv_id, data in investors_summary.items():
        # Берем сальдо из акта сверки именно за этот период
        _, balance, _, _ = get_reconciliation_data(inv_id, f_start, f_end, 'grouped', use_fixed_balance=use_fixed_balance)
        data['balance_act'] = balance
        # Итого = Сальдо (Долг клиента) - Доля (Наш долг)
        data['net_result'] = balance - data['share']

    return render_template('finance/reports.html', active_tab='investor', 
                           investor_summary=investors_summary,
                           investor_details=investor_details,
                           total_period_share=total_period_share,
                           filters={'start': f_start, 'end': f_end, 'fields': f_fields, 'plants': f_plants}, 
                           all_fields=Field.query.all(), all_plants=Plant.query.all(),
                           real_picture=real_picture)

@bp.route('/reports/investor/export')
@login_required
def reports_investor_export():
    if current_user.role not in ['admin', 'executive']: return redirect(url_for('main.index'))
    
    # Получаем параметры (так же как в view)
    f_start = request.args.get('start_date')
    f_end = request.args.get('end_date')
    real_picture = request.args.get('real_picture', '0') == '1'
    use_fixed_balance = not real_picture
    f_fields = [int(x) for x in request.args.getlist('filter_field')]
    f_plants = [int(x) for x in request.args.getlist('filter_plant')]
    
    # Базовый запрос
    conds = [
        Order.status != 'canceled', 
        Order.is_deleted == False, 
        OrderItem.shipped_quantity > 0, 
        Field.investor_id.isnot(None),
        or_(func.extract('year', Order.date) < 2025, Order.client_id != Field.investor_id)
    ]
    
    if f_start: conds.append(func.date(Order.date) >= f_start)
    if f_end: conds.append(func.date(Order.date) <= f_end)
    if f_fields: conds.append(OrderItem.field_id.in_(f_fields))
    if f_plants: conds.append(OrderItem.plant_id.in_(f_plants))
    
    raw_details = db.session.query(OrderItem, Order, Field, Client)\
        .select_from(OrderItem)\
        .join(Order).join(Field).join(Client, Order.client_id == Client.id)\
        .filter(and_(*conds))\
        .order_by(Order.date.desc()).all()
        
    stock_prices = { (sb.plant_id, sb.size_id, sb.field_id, sb.year): sb.purchase_price for sb in StockBalance.query.all() }
    costs_cache = {} 

    wb = Workbook()
    ws = wb.active
    ws.title = "Партнер"
    
    # Шапка
    headers = ["Дата", "Партнер", "Покупатель", "Растение", "Размер", "Поле", "Партия", "Цена (Прайс)", "Себест (Полная)", "Доля (Прибыль/шт)", "Кол-во", "Сумма Доли"]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True)

    for item, order, field, client in raw_details:
        purch_price = stock_prices.get((item.plant_id, item.size_id, item.field_id, item.year), Decimal(0))
        
        calc_basis_year = order.date.year - 1
        if calc_basis_year not in costs_cache:
            costs_cache[calc_basis_year] = calculate_cost_data(calc_basis_year)['accumulated_costs_map']
        
        accum_cost = costs_cache[calc_basis_year].get(item.year, Decimal(0))
        full_cost_unit = purch_price + accum_cost
        
        price_det = item.price if item.price is not None else Decimal(0)
        share_unit = (price_det / 2) - full_cost_unit
        total_share = share_unit * Decimal(item.shipped_quantity)
        
        ws.append([
            order.date.strftime('%d.%m.%Y'),
            field.investor.name if field.investor else '-',
            client.name,
            item.plant.name,
            item.size.name,
            field.name,
            item.year,
            float(price_det),
            float(full_cost_unit),
            float(share_unit),
            item.shipped_quantity,
            float(total_share)
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return send_file(
        buf, 
        download_name=f'investor_report.xlsx', 
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/reports/calculator', methods=['GET', 'POST'])
@login_required
def reports_calculator():
    if current_user.role not in ['admin', 'executive']: return redirect(url_for('main.index'))
    
    calc_results = []
    current_y = msk_now().year
    
    if request.method == 'POST':
        # Получаем списки данных из формы
        p_ids = request.form.getlist('plant_id[]')
        s_ids = request.form.getlist('size_id[]')
        f_ids = request.form.getlist('field_id[]')
        prices = request.form.getlist('price[]')
        years_input = request.form.getlist('year[]')
        
        # Загружаем данные для расчетов
        # Берем прошлый год для себестоимости (т.к. текущий еще не закрыт), либо текущий если надо
        calc_cost_year = current_y - 1 
        cost_data = calculate_cost_data(calc_cost_year)
        accum_costs_map = cost_data['accumulated_costs_map'] # Здесь сидят и Опер. расходы, и Амортизация
        
        # Кэшируем справочники для быстрого доступа
        plants_d = {p.id: p.name for p in Plant.query.all()}
        sizes_d = {s.id: s.name for s in Size.query.all()}
        fields_d = {f.id: f for f in Field.query.all()} 
        
        for i in range(len(p_ids)):
            try:
                # Парсим данные строки
                pid = int(p_ids[i])
                sid = int(s_ids[i])
                fid = int(f_ids[i])
                in_price = Decimal(prices[i]) if prices[i] else Decimal(0)
                
                # Год партии: если не выбран, берем 2017 (начало)
                y_val = years_input[i]
                year = int(y_val) if y_val else 2017
                
                f_obj = fields_d.get(fid)
                is_investor = (f_obj.investor_id is not None)
                
                # 1. Цена закупки (из остатков конкретной партии)
                sb = StockBalance.query.filter_by(plant_id=pid, size_id=sid, field_id=fid, year=year).first()
                purchase_price = sb.purchase_price if sb else Decimal(0)
                
                # 2. Накопленная себестоимость (Опер + Аморт) до года партии
                # Используем accum_costs_map[year], который содержит сумму unit_cost за все годы жизни растения
                accum_cost = accum_costs_map.get(year, Decimal(0))
                
                # Полная себестоимость
                total_cost = purchase_price + accum_cost
                
                # 3. Расчет результата
                if is_investor:
                    # Формула: (Цена продажи / 2) - Полная себестоимость
                    share = (in_price / 2) - total_cost
                    status_text = f"Партнер: {f_obj.investor.name}"
                    formula_text = f"({in_price} / 2) - {total_cost}"
                else:
                    # Собственное: Прибыль = Цена продажи (ни с кем не делим)
                    share = in_price
                    status_text = "Собственность (Без партнера)"
                    formula_text = "100% от цены"

                calc_results.append({
                    'plant': plants_d.get(pid), 
                    'size': sizes_d.get(sid), 
                    'field': f_obj.name, 
                    'year': year, 
                    'status': status_text,
                    'input_price': in_price, 
                    'total_cost': total_cost, 
                    'purchase_price': purchase_price,
                    'accum_cost': accum_cost,
                    'formula': formula_text,
                    'result_price': share,
                    'is_investor': is_investor
                })
            except Exception as e: 
                print(f"Calc error row {i}: {e}")
            
    return render_template('finance/reports.html', active_tab='calculator', calc_results=calc_results, 
                           plants=Plant.query.all(), sizes=Size.query.all(), fields=Field.query.all(), 
                           years_range=range(2017, msk_now().year+1))

@bp.route('/cost/save_to_db', methods=['POST'])
@login_required
def save_cost_to_db():
    if current_user.role != 'admin': 
        return redirect(url_for('main.index'))
    
    # 1. Считаем за прошлый год (чтобы взять полные закрытые данные)
    # Или текущий - 1
    calc_year = msk_now().year - 1 
    
    # 2. Получаем сложную математику из utils.py
    data = calculate_cost_data(calc_year)
    accum_costs_map = data['accumulated_costs_map']

    container_project_ids = get_cost_container_project_ids()
    container_data = calculate_container_cost_data(
        project_ids=container_project_ids,
        selected_year=calc_year,
    )
    container_accum_map = container_data['accumulated_costs_map']
    container_field_ids = get_container_field_ids()
    
    stocks = StockBalance.query.filter(StockBalance.quantity > 0).all()
    count = 0
    container_count = 0
    
    for sb in stocks:
        if sb.field_id in container_field_ids:
            ac = container_accum_map.get(sb.year, Decimal(0))
            container_count += 1
        else:
            ac = accum_costs_map.get(sb.year, Decimal(0))
        sb.current_total_cost = sb.purchase_price + ac
        count += 1
        
    db.session.commit()
    flash(
        f'Записана себестоимость для {count} позиций '
        f'(база: {calc_year} г., контейнерная площадка: {container_count}).'
    )
    log_action("Пересчитал и сохранил себестоимость в БД")
    
    # Возвращаем пользователя на вкладку "Влияние на цену"
    return redirect(url_for('finance.cost_report', tab='impact'))

@bp.route('/reports/projects')
@login_required
def reports_projects():
    if current_user.role not in ['admin', 'executive']: 
        return redirect(url_for('main.index'))
    
    # Фильтр по году создания
    f_year = request.args.get('year', msk_now().year, type=int)
    
    # Берем проекты
    projects_db = Project.query.filter(
        or_(
            func.extract('year', Project.created_at) == f_year,
            Project.status == 'active'
        )
    ).order_by(Project.created_at.desc()).all()
    
    report_data = []
    
    for p in projects_db:
        # ВСЯ МАГИЯ ТЕПЕРЬ ТУТ:
        eco = p.get_economics()
        
        report_data.append({
            'id': p.id,
            'name': p.name,
            'status': p.status,
            'revenue': eco['revenue'],
            'plants_cost': eco['plants_cost'],
            'direct_expenses': eco['direct_expenses'],
            'plan_expenses': sum(b.amount for b in p.budget_items),
            'profit': eco['profit'],
            'margin': eco['margin'],
            'orders_count': len(p.orders)
        })
        
    return render_template('finance/reports.html', active_tab='projects', projects=report_data, selected_year=f_year, years=range(2020, 2030))

def _potting_item_label(item):
    plant = item.plant.name if item.plant else '—'
    size = item.size.name if item.size else '—'
    return f'{plant} · {size}'


def _potting_effective_workers(d_str, meta_saved, meta_counts, timesheet_counts):
    """Число работников за день: сохранённое в meta, иначе из табеля."""
    if d_str in meta_saved:
        return meta_counts.get(d_str, 0)
    return timesheet_counts.get(d_str, 0)


def _project_potting_day_workers_context(project, view_date):
    """Значение поля «Работников» и подсказка из табеля для ежедневного ввода."""
    day_meta = ProjectPottingDayMeta.query.filter_by(
        project_id=project.id, log_date=view_date,
    ).first()
    ts_count = timesheet_workers_count_on_date(view_date)
    if day_meta is not None:
        display = day_meta.workers_count
    else:
        display = ts_count if ts_count else ''
    return {
        'potting_day_workers': display,
        'potting_timesheet_workers': ts_count,
    }


def _potting_day_qty_for_form(project, view_date, skip_load=False):
    """Поля «За этот день» — только приращение; не подставляем уже сохранённое за дату."""
    return {}


def _parse_potting_qty_raw(q_raw):
    q_raw = (q_raw or '').strip()
    if not q_raw:
        return 0
    try:
        return max(0, int(q_raw))
    except (TypeError, ValueError):
        return 0


def apply_daily_potting_save(project, log_date, item_ids, qtys, user_id):
    """Добавляет введённое количество к записи за день (пустое = 0, строка пропускается)."""
    saved = 0
    cleared = 0
    for i in range(len(item_ids)):
        try:
            iid = int(item_ids[i])
        except (TypeError, ValueError):
            continue
        item = ProjectItem.query.filter_by(id=iid, project_id=project.id).first()
        if not item:
            continue
        entered = _parse_potting_qty_raw(qtys[i] if i < len(qtys) else '')
        if entered == 0:
            continue
        existing = ProjectPottingLog.query.filter_by(
            project_item_id=iid,
            log_date=log_date,
        ).first()
        existing_qty = int(existing.quantity or 0) if existing else 0
        new_qty = existing_qty + entered
        if new_qty <= 0:
            if existing:
                db.session.delete(existing)
                cleared += 1
            continue
        if existing:
            existing.quantity = new_qty
        else:
            db.session.add(ProjectPottingLog(
                project_id=project.id,
                project_item_id=iid,
                log_date=log_date,
                quantity=new_qty,
                created_by_id=user_id,
            ))
        saved += 1
    return saved, cleared


def _potting_log_label(row):
    item = row.project_item
    plant = item.plant.name if item and item.plant else '—'
    size = item.size.name if item and item.size else '—'
    d = row.log_date.strftime('%d.%m.%Y') if row.log_date else '—'
    return f'{d}, {plant} · {size}'


def apply_potting_log_edit(project, log_id, qty_raw):
    """Редактирование или удаление (qty=0) одной строки журнала посадки."""
    row = ProjectPottingLog.query.filter_by(id=log_id, project_id=project.id).first()
    if not row:
        return False, 'Запись посадки не найдена', 'warning'
    new_qty = _parse_potting_qty_raw(qty_raw)
    label = _potting_log_label(row)
    old_qty = int(row.quantity or 0)
    if new_qty == 0:
        db.session.delete(row)
        db.session.commit()
        log_action(
            f'Удалена запись посадки в горшки, проект {project.id}: '
            f'{label} ({old_qty} шт)'
        )
        return True, 'Запись посадки удалена', 'success'
    row.quantity = new_qty
    db.session.commit()
    log_action(
        f'Изменена посадка в горшки, проект {project.id}: '
        f'{label}, {old_qty} → {new_qty} шт'
    )
    return True, 'Запись посадки обновлена', 'success'


def delete_potting_log_row(project, log_id):
    """Удаление одной строки журнала посадки (admin/executive)."""
    row = ProjectPottingLog.query.filter_by(id=log_id, project_id=project.id).first()
    if not row:
        return False, 'Запись посадки не найдена', 'warning'
    label = _potting_log_label(row)
    old_qty = int(row.quantity or 0)
    db.session.delete(row)
    db.session.commit()
    log_action(
        f'Удалена запись посадки в горшки, проект {project.id}: '
        f'{label} ({old_qty} шт)'
    )
    return True, 'Запись посадки удалена', 'success'


def _project_potting_analytics(project):
    """Данные графика посадки в горшки (по образцу аналитики выкопки)."""
    empty = {
        'has_data': False,
        'chart_data': None,
        'summary': None,
        'plant_options': [],
        'chart_start': None,
        'chart_end': None,
        'chart_start_fmt': None,
        'chart_end_fmt': None,
    }
    bounds = db.session.query(
        func.min(ProjectPottingLog.log_date),
        func.max(ProjectPottingLog.log_date),
    ).filter(ProjectPottingLog.project_id == project.id).first()
    if not bounds or not bounds[0]:
        return empty

    start_date, end_date = bounds[0], bounds[1]
    items = {i.id: i for i in project.items}
    if not items:
        log_items = db.session.query(ProjectItem).join(
            ProjectPottingLog, ProjectPottingLog.project_item_id == ProjectItem.id,
        ).filter(ProjectPottingLog.project_id == project.id).distinct().all()
        items = {i.id: i for i in log_items}
    if not items:
        return empty

    daily_totals = db.session.query(
        ProjectPottingLog.log_date,
        func.sum(ProjectPottingLog.quantity).label('qty'),
    ).filter(
        ProjectPottingLog.project_id == project.id,
        ProjectPottingLog.log_date >= start_date,
        ProjectPottingLog.log_date <= end_date,
    ).group_by(ProjectPottingLog.log_date).all()
    total_map = {
        (r.log_date.strftime('%Y-%m-%d') if hasattr(r.log_date, 'strftime') else str(r.log_date)): int(r.qty or 0)
        for r in daily_totals
    }

    item_daily = db.session.query(
        ProjectPottingLog.log_date,
        ProjectPottingLog.project_item_id,
        func.sum(ProjectPottingLog.quantity).label('qty'),
    ).filter(
        ProjectPottingLog.project_id == project.id,
        ProjectPottingLog.log_date >= start_date,
        ProjectPottingLog.log_date <= end_date,
    ).group_by(ProjectPottingLog.log_date, ProjectPottingLog.project_item_id).all()
    item_data_map = {}
    for r in item_daily:
        d_str = r.log_date.strftime('%Y-%m-%d') if hasattr(r.log_date, 'strftime') else str(r.log_date)
        item_data_map.setdefault(d_str, {})[r.project_item_id] = int(r.qty or 0)

    meta_rows = ProjectPottingDayMeta.query.filter(
        ProjectPottingDayMeta.project_id == project.id,
        ProjectPottingDayMeta.log_date >= start_date,
        ProjectPottingDayMeta.log_date <= end_date,
    ).all()
    meta_saved = set()
    meta_counts = {}
    for m in meta_rows:
        d_str = timelog_date_str(m.log_date)
        meta_saved.add(d_str)
        meta_counts[d_str] = int(m.workers_count or 0)
    timesheet_counts = timesheet_workers_count_by_date(start_date, end_date)

    days = []
    cur = start_date
    while cur <= end_date:
        days.append(cur)
        cur += timedelta(days=1)

    labels = [d.strftime('%d.%m') for d in days]
    potted_per_day = []
    workers_per_day = []
    efficiency_per_day = []
    for d in days:
        d_str = d.strftime('%Y-%m-%d')
        qty = total_map.get(d_str, 0)
        w = _potting_effective_workers(d_str, meta_saved, meta_counts, timesheet_counts)
        potted_per_day.append(qty)
        workers_per_day.append(w)
        if w and w > 0 and qty > 0:
            efficiency_per_day.append(round(qty / w, 1))
        else:
            efficiency_per_day.append(None)

    palette = [
        '#3366CC', '#DC3912', '#FF9900', '#109618', '#990099',
        '#0099C6', '#DD4477', '#66AA00', '#B82E2E', '#316395',
    ]
    plant_series = []
    plant_options = []
    for idx, (iid, item) in enumerate(sorted(items.items(), key=lambda x: _potting_item_label(x[1]))):
        label = _potting_item_label(item)
        data = []
        total = 0
        for d in days:
            d_str = d.strftime('%Y-%m-%d')
            q = item_data_map.get(d_str, {}).get(iid, 0)
            data.append(q)
            total += q
        plant_options.append({'id': iid, 'label': label, 'total': total})
        plant_series.append({
            'itemId': iid,
            'label': label,
            'data': data,
            'type': 'line',
            'borderColor': palette[idx % len(palette)],
            'backgroundColor': palette[idx % len(palette)],
            'fill': False,
            'hidden': True,
            'yAxisID': 'y1',
            'tension': 0.2,
            'pointRadius': 2,
        })

    days_with_potting = sum(1 for q in potted_per_day if q > 0)
    total_potted = sum(potted_per_day)
    total_workers_days = sum(workers_per_day)
    eff_valid = [e for e in efficiency_per_day if e is not None and e > 0]
    avg_efficiency_by_day = round(sum(eff_valid) / len(eff_valid), 1) if eff_valid else 0
    peak_efficiency = max(eff_valid) if eff_valid else 0

    summary = {
        'days': len(days),
        'days_with_potting': days_with_potting,
        'total_workers': total_workers_days,
        'total_potted': total_potted,
        'avg_potted_per_day': int(round(total_potted / days_with_potting)) if days_with_potting > 0 else 0,
        'avg_potted_per_worker': round(total_potted / total_workers_days, 1) if total_workers_days > 0 else 0,
        'avg_efficiency_by_day': avg_efficiency_by_day,
        'peak_efficiency': peak_efficiency,
    }

    chart_data = {
        'labels': labels,
        'workers': workers_per_day,
        'potted': potted_per_day,
        'efficiency': efficiency_per_day,
        'plantSeries': plant_series,
    }

    return {
        'has_data': True,
        'chart_data': chart_data,
        'summary': summary,
        'plant_options': plant_options,
        'chart_start': start_date.isoformat(),
        'chart_end': end_date.isoformat(),
        'chart_start_fmt': start_date.strftime('%d.%m.%Y'),
        'chart_end_fmt': end_date.strftime('%d.%m.%Y'),
    }


def _save_potting_day_workers(project, log_date, workers_raw):
    """Сохраняет число работников за день (admin). Пустое поле — не трогаем запись."""
    if workers_raw is None:
        return
    raw = str(workers_raw).strip()
    if raw == '':
        return
    try:
        workers = max(0, int(raw))
    except (TypeError, ValueError):
        return
    meta = ProjectPottingDayMeta.query.filter_by(
        project_id=project.id,
        log_date=log_date,
    ).first()
    if workers == 0 and not meta:
        return
    if meta:
        meta.workers_count = workers
    else:
        db.session.add(ProjectPottingDayMeta(
            project_id=project.id,
            log_date=log_date,
            workers_count=workers,
        ))


def _project_potting_context(project):
    """Сводка поставки vs посадки в горшки для карточки проекта."""
    items = list(project.items)
    item_ids = [i.id for i in items]
    potted_map = {}
    if item_ids:
        rows = db.session.query(
            ProjectPottingLog.project_item_id,
            func.sum(ProjectPottingLog.quantity),
        ).filter(
            ProjectPottingLog.project_item_id.in_(item_ids)
        ).group_by(ProjectPottingLog.project_item_id).all()
        potted_map = {pid: int(s or 0) for pid, s in rows}

    supply_rows = []
    field_id = _project_potting_stock_field_id(project)
    stock_year = msk_today().year
    for item in items:
        potted = potted_map.get(item.id, 0)
        stock_qty = None
        if field_id and item.size_id:
            sb = StockBalance.query.filter_by(
                plant_id=item.plant_id,
                size_id=item.size_id,
                field_id=field_id,
                year=stock_year,
            ).first()
            stock_qty = int(sb.quantity or 0) if sb else 0
        supply_rows.append({
            'item': item,
            'potted': potted,
            'deviation': potted - (item.quantity or 0),
            'stock_qty': stock_qty,
        })

    potting_daily = []
    logs = (
        ProjectPottingLog.query.filter_by(project_id=project.id)
        .order_by(ProjectPottingLog.log_date.desc(), ProjectPottingLog.id.desc())
        .all()
    )
    by_date = {}
    for log in logs:
        by_date.setdefault(log.log_date, []).append(log)
    meta_saved = set()
    meta_counts = {}
    timesheet_counts = {}
    if by_date:
        dates_list = list(by_date.keys())
        metas = ProjectPottingDayMeta.query.filter(
            ProjectPottingDayMeta.project_id == project.id,
            ProjectPottingDayMeta.log_date.in_(dates_list),
        ).all()
        for m in metas:
            d_str = timelog_date_str(m.log_date)
            meta_saved.add(d_str)
            meta_counts[d_str] = int(m.workers_count or 0)
        timesheet_counts = timesheet_workers_count_by_date(min(dates_list), max(dates_list))
    for d in sorted(by_date.keys(), reverse=True):
        entries = by_date[d]
        d_str = timelog_date_str(d)
        potting_daily.append({
            'date': d,
            'entries': entries,
            'total': sum(e.quantity for e in entries),
            'workers': _potting_effective_workers(d_str, meta_saved, meta_counts, timesheet_counts),
        })

    supply_total = sum(i.quantity or 0 for i in items)
    potted_total = sum(potted_map.values())
    _, potting_field_label = _project_potting_stock_field_label(project)
    return {
        'supply_rows': supply_rows,
        'potting_daily': potting_daily,
        'potting_totals': {
            'supply': supply_total,
            'potted': potted_total,
            'deviation': potted_total - supply_total,
        },
        'potting_default_date': msk_today().isoformat(),
        'potting_stock_field_label': potting_field_label,
    }


POTTING_RECOUNT_DOC_TYPE = 'potting_recount'
PROJECT_NAME_MAX_LEN = 200
PROJECT_DESCRIPTION_MAX_LEN = 500


def apply_project_settings_from_form(project, form):
    """Обновляет название, описание, статус и поле пересчёта посадки из формы update_project."""
    raw_name = (form.get('name') or '').strip()
    if not raw_name:
        return False, 'Укажите название проекта', 'warning'
    if len(raw_name) > PROJECT_NAME_MAX_LEN:
        return False, f'Название не длиннее {PROJECT_NAME_MAX_LEN} символов', 'warning'

    desc = (form.get('description') or '').strip()
    if len(desc) > PROJECT_DESCRIPTION_MAX_LEN:
        return False, f'Описание не длиннее {PROJECT_DESCRIPTION_MAX_LEN} символов', 'warning'

    status = (form.get('status') or '').strip()
    if status not in ('active', 'closed'):
        return False, 'Некорректный статус проекта', 'warning'

    pf = (form.get('potting_stock_field_id') or '').strip()
    potting_field_id = None
    if pf:
        try:
            potting_field_id = int(pf)
        except (TypeError, ValueError):
            return False, 'Некорректное поле склада для пересчёта', 'warning'

    old_name = project.name
    project.name = raw_name
    project.description = desc or None
    project.status = status
    project.potting_stock_field_id = potting_field_id
    db.session.commit()

    if old_name != raw_name:
        log_action(f'Переименовал проект {project.id}: «{old_name}» → «{raw_name}»')
    return True, 'Настройки проекта обновлены', 'success'


def _default_potting_stock_field_id():
    """Глобальное поле для посадки: AppSetting → контейнерные площадки → «Грядки»."""
    setting = AppSetting.query.get('potting_stock_field_id')
    if setting and setting.value:
        try:
            return int(setting.value)
        except (TypeError, ValueError):
            pass
    try:
        from app.services import get_container_field_id, list_container_yard_fields
        yards = list_container_yard_fields()
        if yards:
            return get_container_field_id() or yards[0].id
    except Exception:
        pass
    for pattern in ('контейнерная площадка', 'контейнер'):
        field = Field.query.filter(func.lower(Field.name).like(f'%{pattern}%')).first()
        if field:
            return field.id
    field = Field.query.filter(func.lower(Field.name).like('%грядк%')).first()
    if field:
        return field.id
    return None


def _project_potting_stock_field_id(project):
    """Поле склада для посадки в горшки: проект → глобальная настройка / автопоиск."""
    if project.potting_stock_field_id:
        return project.potting_stock_field_id
    return _default_potting_stock_field_id()


def _project_potting_stock_field_label(project):
    field_id = _project_potting_stock_field_id(project)
    if not field_id:
        return None, 'не задано'
    field = Field.query.get(field_id)
    return field_id, (field.name if field else f'#{field_id}')


def resolve_project_id_for_yard_fields(field_ids):
    """Проект, у которого контейнерная площадка совпадает с одним из полей позиций."""
    ids = []
    for fid in field_ids or []:
        try:
            n = int(fid)
        except (TypeError, ValueError):
            continue
        if n:
            ids.append(n)
    if not ids:
        return None
    row = (
        Project.query
        .filter(Project.potting_stock_field_id.in_(ids))
        .order_by(Project.id.asc())
        .first()
    )
    return row.id if row else None


def _get_project_potting_recount_doc(project):
    return Document.query.filter_by(
        doc_type=POTTING_RECOUNT_DOC_TYPE,
        project_id=project.id,
    ).order_by(Document.id.asc()).first()


def _ensure_potting_recount_doc(project, user_id, comment=None):
    doc = _get_project_potting_recount_doc(project)
    created = False
    if not doc:
        doc = Document(
            doc_type=POTTING_RECOUNT_DOC_TYPE,
            project_id=project.id,
            user_id=user_id,
            date=msk_now(),
            comment=comment or f'Пересчёт по посадке в горшки: {project.name}',
        )
        db.session.add(doc)
        db.session.flush()
        created = True
    return doc, created


def _apply_potting_item_stock_dolly(project, item, field_id, year, doc, applied_sums, rmap):
    """Тележка: при нулевом остатке записать на склад «Посажено» и учесть отклонение."""
    qty = int(item.potted_total())
    if qty <= 0:
        return {'ok': True, 'action': 'skip', 'reason': 'no_planted'}

    if not item.size_id:
        return {'ok': False, 'error': f'Позиция #{item.id}: не указан размер'}

    stock = get_or_create_stock(item.plant_id, item.size_id, field_id, year)
    old_qty = int(stock.quantity or 0)
    if old_qty != 0:
        return {'ok': True, 'action': 'defer'}

    already = int(applied_sums.get(item.id, 0) or 0)
    pending = int(item.potting_deviation) - already

    reserved = int(rmap.get((item.plant_id, item.size_id, field_id, year), 0) or 0)
    if qty < reserved:
        return {
            'ok': False,
            'error': (
                f'{_potting_item_label(item)}: нельзя установить {qty} шт — '
                f'в резерве {reserved} шт'
            ),
        }

    stock.quantity = qty
    delta = qty - old_qty
    row = None
    if delta != 0:
        row = DocumentRow(
            document_id=doc.id,
            plant_id=item.plant_id,
            size_id=item.size_id,
            field_to_id=field_id,
            year=year,
            quantity=delta,
        )
        db.session.add(row)
        db.session.flush()

    if pending != 0 and not ProjectPottingRecountLine.query.filter_by(
        project_id=project.id,
        project_item_id=item.id,
        quantity_delta=pending,
    ).first():
        db.session.add(ProjectPottingRecountLine(
            project_id=project.id,
            project_item_id=item.id,
            quantity_delta=pending,
            document_row_id=row.id if row is not None else None,
        ))
        applied_sums[item.id] = already + pending

    return {'ok': True, 'action': 'dolly', 'qty': qty, 'pending': pending}


def apply_project_potting_recount(project, user_id):
    """Создаёт/дополняет карточку пересчёта по посадке в горшки.

    При нулевом остатке на позиции — как тележка (остаток = «Посажено»).
    Иначе — дельта отклонения (посажено − поставка), уже учтённая ранее.
    """
    from app.stock_helpers import get_reserved_map

    field_id = _project_potting_stock_field_id(project)
    if not field_id:
        return {
            'ok': False,
            'message': (
                'Укажите поле склада для посадки в настройках проекта '
                'или создайте поле «Контейнерная площадка» в справочнике полей.'
            ),
        }

    field_obj = Field.query.get(field_id)
    if not field_obj:
        return {'ok': False, 'message': 'Поле склада для пересчёта не найдено.'}

    year = msk_today().year
    doc, created_doc = _ensure_potting_recount_doc(project, user_id)

    applied_sums = dict(
        db.session.query(
            ProjectPottingRecountLine.project_item_id,
            func.sum(ProjectPottingRecountLine.quantity_delta),
        )
        .filter_by(project_id=project.id)
        .group_by(ProjectPottingRecountLine.project_item_id)
        .all()
    )

    rmap = get_reserved_map()
    added = dolly_count = skipped = zero = 0
    errors = []

    for item in project.items:
        if not item.size_id:
            if int(item.potted_total()) > 0 or item.potting_deviation != 0:
                errors.append(f'Позиция #{item.id}: не указан размер')
            continue

        dolly_result = _apply_potting_item_stock_dolly(
            project, item, field_id, year, doc, applied_sums, rmap,
        )
        if not dolly_result.get('ok'):
            errors.append(dolly_result.get('error') or 'Ошибка прихода')
            continue
        if dolly_result.get('action') == 'dolly':
            dolly_count += 1
            added += 1
            continue
        if dolly_result.get('action') == 'skip' and item.potting_deviation == 0:
            zero += 1
            continue

        current_dev = item.potting_deviation
        if current_dev == 0:
            zero += 1
            continue

        already = int(applied_sums.get(item.id, 0) or 0)
        pending = int(current_dev) - already
        if pending == 0:
            skipped += 1
            continue

        if ProjectPottingRecountLine.query.filter_by(
            project_id=project.id,
            project_item_id=item.id,
            quantity_delta=pending,
        ).first():
            skipped += 1
            continue

        if pending < 0:
            stock = get_or_create_stock(item.plant_id, item.size_id, field_id, year)
            available = int(stock.quantity or 0)
            reserved = int(rmap.get((item.plant_id, item.size_id, field_id, year), 0) or 0)
            free = available - reserved
            if free + pending < 0:
                label = _potting_item_label(item)
                errors.append(
                    f'{label}: списание {abs(pending)} шт превышает свободный остаток ({free} шт)'
                )
                continue

        stock = get_or_create_stock(item.plant_id, item.size_id, field_id, year)
        stock.quantity = int(stock.quantity or 0) + pending

        row = DocumentRow(
            document_id=doc.id,
            plant_id=item.plant_id,
            size_id=item.size_id,
            field_to_id=field_id,
            year=year,
            quantity=pending,
        )
        db.session.add(row)
        db.session.flush()

        db.session.add(ProjectPottingRecountLine(
            project_id=project.id,
            project_item_id=item.id,
            quantity_delta=pending,
            document_row_id=row.id,
        ))
        applied_sums[item.id] = already + pending
        added += 1

    if errors:
        db.session.rollback()
        return {'ok': False, 'message': '; '.join(errors)}

    if created_doc and added == 0:
        doc_id = None
        if doc.rows:
            doc_id = doc.id
        elif doc.id:
            db.session.delete(doc)
    else:
        doc_id = doc.id

    db.session.commit()

    if added == 0 and not doc_id:
        msg = f'Новых строк нет (пропущено: {skipped}, без отклонения: {zero}).'
    elif created_doc and added > 0:
        msg = (
            f'Создана карточка пересчёта #{doc_id}: строк — {added}, '
            f'пропущено — {skipped}.'
        )
    elif added > 0:
        msg = (
            f'Карточка пересчёта #{doc_id} обновлена: строк — {added}, '
            f'пропущено — {skipped}.'
        )
    else:
        msg = f'Карточка пересчёта #{doc_id} уже актуальна (пропущено: {skipped}).'

    if dolly_count:
        msg += f' Приход «посажено» (как тележка): {dolly_count} поз.'

    if doc_id and added > 0:
        log_action(
            f'Пересчёт посадки в горшки, проект {project.id}, док #{doc_id}: '
            f'+{added} (тележка {dolly_count}) / skip {skipped}'
        )
        try:
            from app.anomaly_engine import sync_recount_anomaly_for_doc
            sync_recount_anomaly_for_doc(doc_id)
        except Exception:
            pass

    return {
        'ok': True,
        'message': msg,
        'doc_id': doc_id,
        'added': added,
        'dolly_count': dolly_count,
        'skipped': skipped,
        'created_doc': bool(created_doc and added > 0),
    }


def apply_project_item_potting_receipt(project, item_id, user_id):
    """Приход одной позиции на поле посадки: остаток = «Посажено» (перезапись, не +)."""
    from app.stock_helpers import get_reserved_map

    item = ProjectItem.query.filter_by(id=item_id, project_id=project.id).first()
    if not item:
        return {'ok': False, 'message': 'Позиция не найдена'}
    if not item.size_id:
        return {'ok': False, 'message': 'У позиции не указан размер'}

    qty = int(item.potted_total())
    if qty <= 0:
        return {'ok': False, 'message': 'Посажено 0 шт — нечего приходовать на склад'}

    field_id = _project_potting_stock_field_id(project)
    if not field_id:
        return {
            'ok': False,
            'message': (
                'Не задано поле склада. Укажите в настройках проекта или создайте '
                'поле «Контейнерная площадка» в справочнике.'
            ),
        }

    field_obj = Field.query.get(field_id)
    if not field_obj:
        return {'ok': False, 'message': 'Поле склада для прихода не найдено'}

    year = msk_today().year
    stock = get_or_create_stock(item.plant_id, item.size_id, field_id, year)
    old_qty = int(stock.quantity or 0)

    if qty < old_qty:
        rmap = get_reserved_map()
        reserved = int(rmap.get((item.plant_id, item.size_id, field_id, year), 0) or 0)
        if qty < reserved:
            return {
                'ok': False,
                'message': (
                    f'{_potting_item_label(item)}: нельзя установить {qty} шт — '
                    f'в резерве {reserved} шт на поле «{field_obj.name}»'
                ),
            }

    doc, _ = _ensure_potting_recount_doc(
        project, user_id, comment=f'Приход посадки в горшки: {project.name}',
    )
    applied_sums = {
        item.id: int(
            db.session.query(func.coalesce(func.sum(ProjectPottingRecountLine.quantity_delta), 0))
            .filter_by(project_id=project.id, project_item_id=item.id)
            .scalar() or 0
        )
    }
    rmap = get_reserved_map()

    if old_qty == 0:
        result = _apply_potting_item_stock_dolly(
            project, item, field_id, year, doc, applied_sums, rmap,
        )
        if not result.get('ok'):
            return {'ok': False, 'message': result.get('error') or 'Ошибка прихода'}
    else:
        stock.quantity = qty
        delta = qty - old_qty
        row = None
        if delta != 0:
            row = DocumentRow(
                document_id=doc.id,
                plant_id=item.plant_id,
                size_id=item.size_id,
                field_to_id=field_id,
                year=year,
                quantity=delta,
            )
            db.session.add(row)
            db.session.flush()

        already = applied_sums.get(item.id, 0)
        pending = int(item.potting_deviation) - already
        if pending != 0 and not ProjectPottingRecountLine.query.filter_by(
            project_id=project.id,
            project_item_id=item.id,
            quantity_delta=pending,
        ).first():
            db.session.add(ProjectPottingRecountLine(
                project_id=project.id,
                project_item_id=item.id,
                quantity_delta=pending,
                document_row_id=row.id if row is not None else None,
            ))

    db.session.commit()

    label = _potting_item_label(item)
    msg = (
        f'«{label}»: на «{field_obj.name}» ({year}) записано {qty} шт '
        f'(было {old_qty}, перезапись).'
    )
    log_action(f'Приход посадки, проект {project.id}: {label} → {qty} шт на {field_obj.name}')

    doc_id = doc.id if doc else None
    if doc_id:
        try:
            from app.anomaly_engine import sync_recount_anomaly_for_doc
            sync_recount_anomaly_for_doc(doc_id)
        except Exception:
            pass

    return {
        'ok': True,
        'message': msg,
        'doc_id': doc_id,
        'qty': qty,
        'old_qty': old_qty,
        'field_name': field_obj.name,
    }


@bp.route('/project/<int:project_id>/potting-recount', methods=['POST'])
@login_required
def project_potting_recount(project_id):
    if current_user.role not in ['admin', 'executive']:
        flash('Недостаточно прав', 'warning')
        return redirect(url_for('main.index'))

    project = Project.query.get_or_404(project_id)
    result = apply_project_potting_recount(project, current_user.id)
    flash(result['message'], 'success' if result['ok'] else 'warning')
    return redirect(url_for('finance.project_detail', project_id=project.id))


@bp.route('/project/<int:project_id>', methods=['GET', 'POST'])
@login_required
def project_detail(project_id):
    if current_user.role not in ['admin', 'executive']:
        return redirect(url_for('main.index'))
        
    project = Project.query.get_or_404(project_id)
    redirect_potting_date = None

    if request.method == 'POST':
        # 1. Редактирование проекта
        if 'seedling_action' in request.form:
            from app.seedlings import handle_seedling_project_form
            from app.utils import log_action as _log_action
            ok, msg, category = handle_seedling_project_form(project, request.form, current_user.id)
            flash(msg, category)
            if ok:
                _log_action(f'Саженцы/проект #{project.id}: {msg}')
            return redirect(url_for('finance.project_detail', project_id=project.id))

        if 'update_project' in request.form:
            ok, msg, category = apply_project_settings_from_form(project, request.form)
            flash(msg, category)

        # 2. Добавление Растений
        elif 'add_item' in request.form:
            plant_id = request.form.get('plant_id')
            size_id = request.form.get('size_id')
            qty = int(request.form.get('quantity') or 0)
            db.session.add(ProjectItem(project_id=project.id, plant_id=plant_id, size_id=size_id, quantity=qty))
            db.session.commit()

        # 2b. Массовое добавление растений (несколько строк за один POST)
        elif 'bulk_add_items' in request.form:
            plant_ids = request.form.getlist('bulk_plant_id[]')
            size_ids = request.form.getlist('bulk_size_id[]')
            qtys = request.form.getlist('bulk_quantity[]')
            added = 0
            skipped = 0
            for i in range(len(plant_ids)):
                pid = (plant_ids[i] or '').strip()
                sid = (size_ids[i] or '').strip()
                q_raw = (qtys[i] if i < len(qtys) else '').strip()
                if not pid or not sid or not q_raw:
                    skipped += 1
                    continue
                try:
                    pid_i = int(pid); sid_i = int(sid); q_i = int(q_raw)
                except (TypeError, ValueError):
                    skipped += 1
                    continue
                if q_i <= 0:
                    skipped += 1
                    continue
                db.session.add(ProjectItem(
                    project_id=project.id,
                    plant_id=pid_i, size_id=sid_i, quantity=q_i,
                ))
                added += 1
            if added:
                db.session.commit()
                flash(f'Добавлено позиций: {added}' + (f' (пропущено {skipped})' if skipped else ''))
                log_action(f'Массовое добавление в проект {project.id}: {added} строк')
            else:
                flash('Не добавлено ни одной строки. Проверьте заполнение.', 'warning')

        elif 'delete_item' in request.form:
            ProjectItem.query.filter_by(id=request.form.get('item_id')).delete()
            db.session.commit()

        # 2c. Посадка в горшки (ежедневный ввод)
        elif 'save_daily_potting' in request.form:
            d_raw = (request.form.get('potting_date') or '').strip()
            try:
                log_date = datetime.strptime(d_raw, '%Y-%m-%d').date()
            except (TypeError, ValueError):
                log_date = msk_today()

            saved, cleared = apply_daily_potting_save(
                project,
                log_date,
                request.form.getlist('potting_item_id[]'),
                request.form.getlist('potting_qty[]'),
                current_user.id,
            )
            _save_potting_day_workers(
                project, log_date, request.form.get('potting_workers'),
            )
            db.session.commit()
            if saved or cleared:
                flash(
                    f'Посадка в горшки за {log_date.strftime("%d.%m.%Y")}: '
                    f'добавлено позиций — {saved}'
                    + (f', очищено — {cleared}' if cleared else '')
                )
                log_action(f'Посадка в горшки, проект {project.id}, дата {log_date}')
                redirect_potting_date = log_date.isoformat()
            elif request.form.get('potting_workers', '').strip() != '':
                flash(
                    f'Число работников за {log_date.strftime("%d.%m.%Y")} сохранено',
                    'success',
                )
                redirect_potting_date = log_date.isoformat()
            else:
                flash('Нет данных для сохранения', 'warning')

        elif 'edit_potting_log' in request.form:
            ok, msg, category = apply_potting_log_edit(
                project,
                request.form.get('potting_log_id'),
                request.form.get('potting_log_qty'),
            )
            flash(msg, category)

        elif 'delete_potting_log' in request.form:
            ok, msg, category = delete_potting_log_row(
                project, request.form.get('potting_log_id'),
            )
            flash(msg, category)

        elif 'receipt_item_stock' in request.form:
            result = apply_project_item_potting_receipt(
                project, request.form.get('item_id'), current_user.id,
            )
            flash(result['message'], 'success' if result['ok'] else 'warning')

        # 3. БЮДЖЕТ (ПЛАНОВЫЕ РАСХОДЫ)
        elif 'add_budget' in request.form:
            name = request.form.get('name')
            amount = float(request.form.get('amount') or 0)
            db.session.add(ProjectBudget(project_id=project.id, name=name, amount=amount))
            db.session.commit()
            flash('Статья бюджета добавлена')

        elif 'delete_budget' in request.form:
            ProjectBudget.query.filter_by(id=request.form.get('budget_id')).delete()
            db.session.commit()

        # 4. ПРИВЯЗКА ЗАКАЗА К ПРОЕКТУ
        elif 'link_order' in request.form:
            oid = request.form.get('order_id')
            order = Order.query.get(oid)
            if order:
                order.project_id = project.id
                db.session.commit()
                flash(f'Заказ #{oid} привязан к проекту')
        
        elif 'unlink_order' in request.form:
            oid = request.form.get('order_id')
            order = Order.query.get(oid)
            if order and order.project_id == project.id:
                order.project_id = None
                db.session.commit()
                flash('Заказ отвязан')

        if redirect_potting_date:
            return redirect(url_for(
                'finance.project_detail',
                project_id=project.id,
                potting_date=redirect_potting_date,
                potting_day_cleared=1,
            ))
        return redirect(url_for('finance.project_detail', project_id=project.id))

    # Сбор статистики через новый единый метод модели
    eco = project.get_economics()
    
    # Подготовка данных для таблицы бюджета (План/Факт по статьям)
    budget_comparison = []
    total_plan = Decimal(0)
    
    for item in project.budget_items:
        # Считаем сумму расходов, привязанных к ЭТОЙ статье бюджета
        item_fact = db.session.query(func.sum(Expense.amount))\
            .filter(Expense.project_budget_id == item.id).scalar() or Decimal(0)
        
        total_plan += item.amount
        
        budget_comparison.append({
            'id': item.id,
            'name': item.name,
            'plan': item.amount,
            'fact': item_fact,
            'diff': item.amount - item_fact # Положительное - экономия, Отрицательное - перерасход
        })
    
    # Формируем объект статистики из данных get_economics
    stats = {
        'total_plan': total_plan,
        'total_fact': eco['direct_expenses'],
        # Алиасы для шаблона: исторически шаблон называл план/факт расходов
        # `plan_expenses`/`fact_expenses`. Чтобы не ломать совместимость и
        # не плодить дубли, отдаём оба имени.
        'plan_expenses': total_plan,
        'fact_expenses': eco['direct_expenses'],
        'unallocated_fact': eco['unallocated_fact'],
        'revenue': eco['revenue'],
        'plants_cost': eco['plants_cost'],
        'profit': eco['profit'],
        'margin': eco['margin'],
        'items_count': sum(i.quantity for i in project.items),
        'cost_per_item': 0
    }
    
    if stats['items_count'] > 0:
        stats['cost_per_item'] = stats['total_fact'] / Decimal(stats['items_count'])
        
    # ИЗМЕНЕНИЕ: Загружаем ВСЕ активные заказы (кроме отмененных и удаленных), 
    # чтобы можно было перепривязать заказ из другого проекта сюда.
    all_orders = Order.query.filter(
        Order.status != 'canceled', 
        Order.is_deleted == False
    ).order_by(Order.date.desc()).all()

    potting_ctx = _project_potting_context(project)
    potting_analytics = _project_potting_analytics(project)
    from app.seedlings import seedling_context_for_project
    seedling_ctx = seedling_context_for_project(project)
    view_date_raw = request.args.get('potting_date', '').strip()
    try:
        view_date = datetime.strptime(view_date_raw, '%Y-%m-%d').date() if view_date_raw else msk_today()
    except ValueError:
        view_date = msk_today()
    potting_day_qty = _potting_day_qty_for_form(
        project, view_date,
        skip_load=request.args.get('potting_day_cleared') == '1',
    )

    potting_recount_doc = _get_project_potting_recount_doc(project)

    return render_template(
        'finance/project_detail.html',
        project=project,
        stats=stats,
        budget_comparison=budget_comparison,
        all_orders=all_orders,
        plants=sorted(Plant.query.all(), key=lambda x: x.name),
        sizes=Size.query.all(),
        fields=sorted(Field.query.all(), key=natural_key),
        potting_recount_doc=potting_recount_doc,
        potting_view_date=view_date.isoformat(),
        potting_day_qty=potting_day_qty,
        **_project_potting_day_workers_context(project, view_date),
        **potting_ctx,
        **potting_analytics,
        **seedling_ctx,
    )


@bp.route('/project/<int:project_id>/seedling_dieback.xlsx')
@login_required
def seedling_dieback_export(project_id):
    if current_user.role not in ['admin', 'executive', 'user']:
        return redirect(url_for('main.index'))
    project = Project.query.get_or_404(project_id)
    from app.seedlings import export_dieback_workbook
    wb = export_dieback_workbook(project.id)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"vypad_project_{project.id}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/project/<int:project_id>/seedling_dieback_summary.xlsx')
@login_required
def seedling_dieback_summary_export(project_id):
    if current_user.role not in ['admin', 'executive', 'user']:
        return redirect(url_for('main.index'))
    project = Project.query.get_or_404(project_id)
    from app.seedlings import export_dieback_summary_workbook
    from app.utils import msk_now
    year = int(request.args.get('year') or msk_now().year)
    period_type = (request.args.get('period') or 'month').strip()
    try:
        period_n = int(request.args.get('n') or msk_now().month)
    except (TypeError, ValueError):
        period_n = msk_now().month
    try:
        wb, period_label = export_dieback_summary_workbook(project, year, period_type, period_n)
    except ValueError as exc:
        flash(str(exc))
        return redirect(url_for('finance.project_detail', project_id=project.id, seed_tab='seedTabDieback'))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = period_label.replace(' ', '_')
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"vypad_svod_{project.id}_{safe}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/project/<int:project_id>/yard_stock.xlsx')
@login_required
def seedling_yard_stock_export(project_id):
    if current_user.role not in ['admin', 'executive', 'user']:
        return redirect(url_for('main.index'))
    project = Project.query.get_or_404(project_id)
    from app.seedlings import export_yard_stock_workbook
    try:
        wb = export_yard_stock_workbook(project)
    except ValueError as exc:
        flash(str(exc))
        return redirect(url_for('finance.project_detail', project_id=project.id))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'yard_stock_p{project.id}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/finance/invoices', methods=['GET', 'POST'])
@login_required
def invoices_list():
    # Перенаправляем на объединенную вкладку "Счета и расходы"
    return redirect(url_for('finance.expenses', tab='invoices'))

@bp.route('/finance/invoices/download/<int:inv_id>')
@login_required
def invoice_download(inv_id):
    if current_user.role not in ['admin', 'executive']:
        return redirect(url_for('main.index'))
    
    inv = PaymentInvoice.query.get_or_404(inv_id)
    inv_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'invoices')
    try:
        return send_from_directory(inv_dir, inv.filename, as_attachment=True, download_name=inv.original_name)
    except Exception:
        flash("Файл не найден на сервере")
        return redirect(url_for('finance.expenses', tab='invoices'))

@bp.route('/finance/summary', methods=['GET'])
@login_required
def invoices_summary():
    # Доступ разрешен всем авторизованным (включая менеджеров 'user')
    # Фильтры
    f_client = request.args.get('client_id')
    f_start = request.args.get('start_date')
    f_end = request.args.get('end_date')
    
    # Базовый запрос: Все активные заказы
    query = Order.query.filter(Order.is_deleted == False)
    
    # Применяем фильтры
    if f_client:
        query = query.filter(Order.client_id == int(f_client))
    if f_start:
        query = query.filter(func.date(Order.date) >= f_start)
    if f_end:
        query = query.filter(func.date(Order.date) <= f_end)
        
    orders = query.order_by(Order.date.desc()).all()
    
    # Группировка по (client_id, invoice_number)
    # Используем словарь для объединения
    invoices_map = {}
    
    from decimal import Decimal
    
    for o in orders:
        # Если номера счета нет, группируем под ключом "no_invoice" или пропускаем
        # Чтобы отчет был красивым, лучше группировать пустые тоже, но отдельно
        inv_num = o.invoice_number if o.invoice_number else "Без счета"
        
        # Ключ уникальности: Клиент + Номер счета
        # (Разные клиенты могут иметь счета с одинаковым номером "1", "2" и т.д.)
        key = (o.client_id, inv_num)
        
        if key not in invoices_map:
            invoices_map[key] = {
                'client_id': o.client_id,
                'client_name': o.client.name,
                'invoice_number': inv_num,
                'has_invoice': bool(o.invoice_number), # Флаг для ссылки
                'date': o.invoice_date if o.invoice_date else o.date.date(), # Если даты счета нет, берем дату заказа
                'orders_count': 0,
                'total_sum': Decimal(0),
                'paid_sum': Decimal(0),
                'shipped_sum': Decimal(0),
                'debt': Decimal(0),
                'status': 'paid' # По умолчанию paid, если найдем долг - сменим
            }
            
        row = invoices_map[key]
        row['orders_count'] += 1
        row['total_sum'] += o.total_sum
        row['paid_sum'] += o.paid_sum
        
        # Считаем отгруженное (сумма)
        shipped_val = Decimal(0)
        for item in o.items:
            price = item.price if item.price is not None else Decimal(0)
            shipped_val += price * item.shipped_quantity
        row['shipped_sum'] += shipped_val
        
        # Если есть хотя бы одна дата счета в группе, берем её (приоритет у заполненной)
        if o.invoice_date and row['invoice_number'] != "Без счета":
            row['date'] = o.invoice_date

    # Превращаем словарь в список
    report_rows = list(invoices_map.values())
    
    # Досчитываем долги и статусы
    for r in report_rows:
        r['debt'] = r['total_sum'] - r['paid_sum']
        
        if r['debt'] <= 0:
            r['status'] = 'paid' # Оплачен
        elif r['paid_sum'] > 0:
            r['status'] = 'partial' # Частично
        else:
            r['status'] = 'unpaid' # Не оплачен

    # Сортировка (по умолчанию по дате убывания)
    report_rows.sort(key=lambda x: x['date'], reverse=True)
    
    return render_template(
        'finance/invoices_summary.html',
        invoices=report_rows,
        clients=Client.query.order_by(Client.name).all(),
        filters={'client_id': f_client, 'start': f_start, 'end': f_end}
    )

# --- УПРАВЛЕНИЕ ПРОЕКТАМИ ---

@bp.route('/projects', methods=['GET', 'POST'])
@login_required
def projects_list():
    if current_user.role not in ['admin', 'executive']:
        return redirect(url_for('main.index'))

    tab = request.args.get('tab', 'list')

    if request.method == 'POST':
        name = request.form.get('name')
        desc = request.form.get('description')
        if name:
            new_proj = Project(name=name, description=desc, status='active')
            db.session.add(new_proj)
            db.session.commit()
            flash(f'Проект "{name}" создан')
            log_action(f"Создал проект {name}")
        return redirect(url_for('finance.projects_list', tab='list'))

    # Фильтр статуса
    show_closed = request.args.get('show_closed') == '1'

    # Таб "Список"
    projects_data = []
    if tab == 'list':
        query = Project.query
        if not show_closed:
            query = query.filter_by(status='active')
        projects_db = query.order_by(Project.created_at.desc()).all()

        for p in projects_db:
            eco = p.get_economics()
            projects_data.append({
                'id': p.id,
                'name': p.name,
                'description': p.description,
                'status': p.status,
                'total_revenue': eco['revenue'],
                'total_expenses': eco['direct_expenses'],
                'profit': eco['profit']
            })

    # Таб "Отчеты"
    report_data = []
    selected_year = request.args.get('year', msk_now().year, type=int)
    if tab == 'reports':
        projects_db = Project.query.filter(
            or_(
                func.extract('year', Project.created_at) == selected_year,
                Project.status == 'active'
            )
        ).order_by(Project.created_at.desc()).all()

        for p in projects_db:
            eco = p.get_economics()
            report_data.append({
                'id': p.id,
                'name': p.name,
                'status': p.status,
                'revenue': eco['revenue'],
                'plants_cost': eco['plants_cost'],
                'direct_expenses': eco['direct_expenses'],
                'plan_expenses': sum(b.amount for b in p.budget_items),
                'profit': eco['profit'],
                'margin': eco['margin'],
                'orders_count': len(p.orders)
            })

    return render_template('finance/projects_list.html', 
                           active_tab=tab,
                           projects=projects_data,
                           report_data=report_data,
                           show_closed=show_closed,
                           selected_year=selected_year,
                           current_year=msk_now().year)

from app.models import (
    db, Expense, BudgetItem, BudgetPlan, CashflowPlan, Employee, UnitCostOverride, 
    Plant, StockBalance, Order, OrderItem, Field, Payment, Document, DocumentRow, Size,
    Client, PaymentInvoice, Project, ProjectItem, ProjectPottingLog, ProjectPottingDayMeta, ProjectBudget
)