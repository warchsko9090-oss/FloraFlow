import io
import os
import re
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path
from urllib.parse import unquote, urlparse

from flask import make_response, has_app_context, current_app
from flask_login import current_user
from sqlalchemy import func
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models import (
    db, StockBalance, Order, OrderItem, BudgetItem, ActionLog, PriceHistory,
    Plant, Size, Field, Client, Supplier
)

# --- КОНСТАНТЫ ---
MONTH_NAMES = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь', 
               7: 'Июль', 8: 'Август', 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}
RATE_TYPES_LABELS = {'norm': 'Обычный', 'norm_over': 'Обычный (Перераб)', 'spec': 'Копка/Стрижка', 'spec_over': 'Копка/Стрижка (Перераб)'}
EMPLOYEE_ROLES = {'worker': 'Рабочий', 'brigadier': 'Бригадир', 'assistant': 'Помощник', 'manager': 'Менеджер', 'mechanic': 'Механик'}

# --- ФУНКЦИЯ ВРЕМЕНИ (MSK UTC+3) ---
def msk_now():
    return datetime.utcnow() + timedelta(hours=3)

def msk_today():
    return msk_now().date()

def timelog_date_str(d):
    """YYYY-MM-DD для даты из TimeLog (защита от разных типов SQLite)."""
    return d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else str(d)

def timesheet_worker_sets_by_date(start_date, end_date):
    """Уникальные employee_id по дням, где сумма часов (все ставки) > 1 — как в аналитике выкопки."""
    from app.models import TimeLog
    tl = db.session.query(
        TimeLog.date,
        TimeLog.employee_id,
        (TimeLog.hours_norm + TimeLog.hours_norm_over + TimeLog.hours_spec + TimeLog.hours_spec_over).label('hours'),
    ).filter(TimeLog.date >= start_date, TimeLog.date <= end_date).all()
    worked = {}
    for row in tl:
        if row.hours and row.hours > 1:
            worked.setdefault(timelog_date_str(row.date), set()).add(row.employee_id)
    return worked

def timesheet_workers_count_by_date(start_date, end_date):
    sets = timesheet_worker_sets_by_date(start_date, end_date)
    return {d: len(s) for d, s in sets.items()}

def timesheet_workers_count_on_date(d):
    return timesheet_workers_count_by_date(d, d).get(timelog_date_str(d), 0)

# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ---
def log_action(text):
    """Записывает действие текущего пользователя в лог."""
    try:
        if current_user.is_authenticated:
            log = ActionLog(user_id=current_user.id, action=text, date=msk_now())
            db.session.add(log)
            db.session.commit()
    except Exception as e:
        print(f"Log error: {e}")

def get_or_create_stock(plant_id, size_id, field_id, year):
    """Ищет запись остатка, если нет - создает нулевую."""
    stock = StockBalance.query.filter_by(plant_id=plant_id, size_id=size_id, field_id=field_id, year=year).first()
    if not stock:
        stock = StockBalance(plant_id=plant_id, size_id=size_id, field_id=field_id, year=year, quantity=0, price=0, purchase_price=0)
        db.session.add(stock)
        db.session.flush()
    return stock

def natural_key(obj):
    """Для сортировки (C1, C2, C10...)"""
    text = obj.name if hasattr(obj, 'name') else str(obj)
    if isinstance(obj, BudgetItem): text = obj.code
    return [int(s) if s.isdigit() else s.lower() for s in re.split('([0-9]+)', text)]

def check_stock_availability(plant_id, size_id, field_id, year, quantity_needed, exclude_item_id=None):
    """Проверяет, хватает ли свободного остатка (с учётом резервов).

    Для размеров-саженцев дефицит разрешён всегда (рост быстрый, минус чинят
    в заказе через предложение других промеров).
    """
    from app.stock_helpers import compute_free
    from app.models import Size
    from app.seedlings import allows_order_deficit

    _fact, _reserved, free_qty = compute_free(
        plant_id, size_id, field_id, year, exclude_item_id=exclude_item_id
    )
    size = Size.query.get(size_id)
    if size and allows_order_deficit(size.name):
        return True, free_qty
    if free_qty < quantity_needed:
        return False, free_qty
    return True, free_qty

def get_actual_price(plant_id, size_id, field_id, check_year=None):
    """Получает цену товара из истории или текущих остатков."""
    curr_year = check_year if check_year else msk_now().year
    hist = PriceHistory.query.filter_by(
        plant_id=plant_id, size_id=size_id, field_id=field_id, year=curr_year,
    ).first()
    if hist:
        return hist.price
    hist = PriceHistory.query.filter_by(
        plant_id=plant_id, size_id=size_id, field_id=field_id,
    ).order_by(PriceHistory.year.desc()).first()
    if hist:
        return hist.price
    hist = PriceHistory.query.filter_by(
        plant_id=plant_id, size_id=size_id, year=curr_year,
    ).first()
    if hist:
        return hist.price
    sb = StockBalance.query.filter_by(
        plant_id=plant_id, size_id=size_id, field_id=field_id, year=curr_year,
    ).first()
    if sb and sb.price:
        return sb.price
    sb = StockBalance.query.filter_by(
        plant_id=plant_id, size_id=size_id, field_id=field_id,
    ).order_by(StockBalance.year.desc()).first()
    return sb.price if sb else Decimal('0.00')

PDF_FONT_FILES = (
    ('DejaVuSans', 'DejaVuSans.ttf'),
    ('DejaVuSans-Bold', 'DejaVuSans-Bold.ttf'),
    ('DejaVuSans-Oblique', 'DejaVuSans-Oblique.ttf'),
)


def _pdf_static_dir():
    return Path(current_app.root_path) / 'static'


def _iter_dejavu_font_paths(filename: str):
    """Пути к TTF: app/static и системные каталоги Linux (Amvera)."""
    yield _pdf_static_dir() / filename
    for base in (
        Path('/usr/share/fonts/truetype/dejavu'),
        Path('/usr/share/fonts/dejavu'),
        Path('/usr/local/share/fonts'),
    ):
        yield base / filename


def ensure_dejavu_fonts():
    """Регистрирует TTF в ReportLab — запасной путь для xhtml2pdf."""
    from reportlab.lib.fonts import addMapping
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    registered = set(pdfmetrics.getRegisteredFontNames())

    for name, filename in PDF_FONT_FILES:
        if name in registered:
            continue
        for path in _iter_dejavu_font_paths(filename):
            if path.is_file():
                pdfmetrics.registerFont(TTFont(name, str(path.resolve())))
                registered.add(name)
                break

    # Если Bold/Oblique не найдены — используем regular, чтобы PDF не падал.
    regular_path = next((p for p in _iter_dejavu_font_paths('DejaVuSans.ttf') if p.is_file()), None)
    if regular_path:
        regular = str(regular_path.resolve())
        for alias in ('DejaVuSans-Bold', 'DejaVuSans-Oblique'):
            if alias not in registered:
                pdfmetrics.registerFont(TTFont(alias, regular))
                registered.add(alias)

    if 'DejaVuSans' in registered:
        addMapping('DejaVuSans', 0, 0, 'DejaVuSans')
        addMapping('DejaVuSans', 1, 0, 'DejaVuSans-Bold')
        addMapping('DejaVuSans', 0, 1, 'DejaVuSans-Oblique')
        addMapping('DejaVuSans', 1, 1, 'DejaVuSans-Bold')


def pdf_resource_callback(uri, rel):
    """Разрешает url() в CSS (@font-face) в локальные файлы шрифтов."""
    if not has_app_context():
        return uri

    static = _pdf_static_dir()
    bare = uri.split('?')[0].split('/')[-1]
    if bare in {f for _, f in PDF_FONT_FILES}:
        path = static / bare
        if path.is_file():
            return str(path.resolve())

    rel_uri = uri.replace('\\', '/').lstrip('/')
    if rel_uri.startswith('data-sidebar/'):
        candidates = [_pdf_static_dir() / rel_uri]
        if has_app_context():
            candidates.append(Path(current_app.config['UPLOAD_FOLDER']) / rel_uri)
        for path in candidates:
            if path.is_file():
                return str(path.resolve())

    sidebar = static / 'data-sidebar' / bare
    if sidebar.is_file():
        return str(sidebar.resolve())

    if uri.startswith('file:'):
        path = unquote(urlparse(uri).path)
        if os.name == 'nt' and path.startswith('/') and len(path) > 2 and path[2] == ':':
            path = path[1:]
        if os.path.isfile(path):
            return path

    return uri


def create_pdf_response(html_content, filename, *, page_bg=None, page_margin='1cm'):
    """PDF через xhtml2pdf; DejaVu регистрируется напрямую (без @font-face — Windows)."""
    from reportlab.lib.colors import HexColor
    from reportlab.platypus.frames import Frame
    from xhtml2pdf.context import pisaContext
    from xhtml2pdf.document import pisaStory
    from xhtml2pdf.files import cleanFiles
    from xhtml2pdf.util import getBox
    from xhtml2pdf.xhtml2pdf_reportlab import PmlBaseDoc, PmlPageTemplate

    link_cb = pdf_resource_callback if has_app_context() else None
    ctx = pisaContext()
    if has_app_context():
        ensure_dejavu_fonts()
        for font_name, _ in PDF_FONT_FILES:
            ctx.registerFont(font_name)
    if link_cb:
        ctx.pathCallback = link_cb

    ctx = pisaStory(
        io.BytesIO(html_content.encode('UTF-8')),
        link_callback=link_cb,
        encoding='UTF-8',
        context=ctx,
    )
    if ctx.err:
        cleanFiles()
        return 'Error generating PDF'

    out = io.BytesIO()
    doc = PmlBaseDoc(
        out,
        pagesize=ctx.pageSize,
        author=ctx.meta.get('author', '').strip(),
        subject=ctx.meta.get('subject', '').strip(),
        keywords=[x.strip() for x in ctx.meta.get('keywords', '').strip().split(',') if x],
        title=ctx.meta.get('title', '').strip(),
        showBoundary=0,
        allowSplitting=1,
    )

    margin_box = page_margin if page_margin != '0' else '0 0 0 0'

    if 'body' in ctx.templateList:
        body = ctx.templateList['body']
        del ctx.templateList['body']
    else:
        x, y, w, h = getBox(margin_box, ctx.pageSize)
        body = PmlPageTemplate(
            id='body',
            frames=[
                Frame(
                    x, y, w, h, id='body',
                    leftPadding=0, rightPadding=0, bottomPadding=0, topPadding=0,
                )
            ],
            pagesize=ctx.pageSize,
        )

    if page_bg:
        fill = HexColor(page_bg)
        prev_before = getattr(body, 'beforeDrawPage', None)

        def _paint_page_bg(canvas, doc):
            canvas.saveState()
            canvas.setFillColor(fill)
            pw, ph = doc.pagesize
            canvas.rect(0, 0, pw, ph, fill=1, stroke=0)
            canvas.restoreState()
            if prev_before:
                prev_before(canvas, doc)

        body.beforeDrawPage = _paint_page_bg

    doc.addPageTemplates([body, *list(ctx.templateList.values())])
    if ctx.multiBuild:
        doc.multiBuild(ctx.story)
    else:
        doc.build(ctx.story)

    cleanFiles()
    response = make_response(out.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

def apply_excel_styles(ws):
    header_fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
    header_font = Font(bold=True, color="004D40")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border
    for col in ws.columns:
        max_length = 0; column = col[0].column_letter
        for cell in col:
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except Exception: pass
        ws.column_dimensions[column].width = max_length + 2

# --- ФУНКЦИИ ФОРМАТИРОВАНИЯ ---
def format_money(value):
    if value is None: return ""
    try: return "{:,.2f} ₽".format(float(value)).replace(',', ' ')
    except (ValueError, TypeError): return str(value)

def format_money_int(value):
    if value is None: return ""
    try: return "{:,.0f} ₽".format(float(value)).replace(',', ' ')
    except (ValueError, TypeError): return str(value)

def dateru(value):
    if value is None: return ""
    if isinstance(value, str): return value 
    return value.strftime('%d.%m.%Y')

# --- EXCEL IMPORT MAP HELPERS ---
def get_plant_map():
    return {p.name.strip().lower(): p.id for p in Plant.query.all()}

def get_size_map():
    return {s.name.strip().lower(): s.id for s in Size.query.all()}

def get_field_map():
    return {f.name.strip().lower(): f.id for f in Field.query.all()}

def get_client_map():
    return {c.name.strip().lower(): c.id for c in Client.query.all()}

def get_supplier_map():
    return {s.name.strip().lower(): s.id for s in Supplier.query.all()}