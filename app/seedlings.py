"""Контур саженцев: контейнер → товарность → промер.

Остаток хранится в StockBalance через «умные» Size.name:

  Нетов · C5
  Саженцы · C5
  Саженцы · C5 · 20-30

Движения оформляются документами regrading (совместимо с текущим леджером).
ВиУМ / техкарты в этом апдейте не трогаем.
"""

from __future__ import annotations

import re
from typing import Optional

from sqlalchemy.orm import joinedload

from app.models import (
    Document,
    DocumentRow,
    SeedlingContainer,
    Size,
    StockBalance,
    db,
)
from app.utils import get_or_create_stock, msk_now

SEEDLING_SEP = ' · '
# Канон в Size.name (после переименования в БД). Старое «Саженцы» тоже распознаём.
STAGE_NETOV = 'Нетов'
STAGE_SAZHENCY = 'Товарный'
STAGE_SAZHENCY_LEGACY = 'Саженцы'
# PDF остатков / Excel проектов и площадки: стадия всегда как «Саженцы».
STAGE_SAZHENCY_EXPORT = 'Саженцы'
# Подпись в UI менеджера (если в БД ещё лежит legacy «Саженцы»).
STAGE_SAZHENCY_UI = 'Товарный'

DEFAULT_CONTAINERS = (
    'C2', 'C3', 'C5', 'C7.5', 'C10', 'C15', 'C20', 'C25', 'C35', 'C50',
)

_GRADE_RE = re.compile(r'^[\w\dА-Яа-яЁё.\-+/\s]{1,40}$', re.UNICODE)


def normalize_container_name(name: str) -> str:
    return (name or '').strip()


def normalize_grade(name: str) -> str:
    return re.sub(r'\s+', ' ', (name or '').strip())


def is_bare_netov_name(name: str) -> bool:
    low = re.sub(r'[\s.]+', '', (name or '').lower())
    return low == 'нетов'


def is_bare_sazhency_name(name: str) -> bool:
    """Голый размер товарной стадии: «Товарный» или legacy «Саженцы»."""
    low = (name or '').strip().lower()
    return low in ('товарный', 'саженцы')


def _is_sazhency_stage_token(token: str) -> bool:
    low = (token or '').lower().replace('.', '')
    return 'товарн' in low or 'саженц' in low


def stage_ui_label(stage: str) -> str:
    """Подпись стадии в UI: всегда «Товарный» (и legacy «Саженцы»)."""
    raw = (stage or '').strip()
    if not raw:
        return '—'
    if raw == STAGE_SAZHENCY or raw == STAGE_SAZHENCY_LEGACY:
        return STAGE_SAZHENCY_UI
    if raw.startswith(STAGE_SAZHENCY_LEGACY + ' '):
        return STAGE_SAZHENCY_UI + raw[len(STAGE_SAZHENCY_LEGACY):]
    if raw.startswith(STAGE_SAZHENCY_LEGACY + '('):
        return STAGE_SAZHENCY_UI + ' ' + raw[len(STAGE_SAZHENCY_LEGACY):]
    if raw.startswith(STAGE_SAZHENCY + ' '):
        return STAGE_SAZHENCY_UI + raw[len(STAGE_SAZHENCY):]
    if raw.startswith(STAGE_SAZHENCY + '('):
        return STAGE_SAZHENCY_UI + ' ' + raw[len(STAGE_SAZHENCY):]
    return raw


def size_name_ui_label(name: str) -> str:
    """UI: legacy «Саженцы · …» → «Товарный · …»; уже «Товарный» оставляем."""
    raw = (name or '').strip()
    if not raw:
        return raw
    if raw == STAGE_SAZHENCY or raw == STAGE_SAZHENCY_LEGACY:
        return STAGE_SAZHENCY_UI
    if raw.startswith(STAGE_SAZHENCY_LEGACY + SEEDLING_SEP):
        return STAGE_SAZHENCY_UI + raw[len(STAGE_SAZHENCY_LEGACY):]
    if raw.startswith(STAGE_SAZHENCY_LEGACY + ' '):
        return STAGE_SAZHENCY_UI + raw[len(STAGE_SAZHENCY_LEGACY):]
    return raw


def size_name_export_label(name: str) -> str:
    """PDF/Excel: «Товарный · …» → «Саженцы · …»."""
    raw = (name or '').strip()
    if not raw:
        return raw
    if raw == STAGE_SAZHENCY:
        return STAGE_SAZHENCY_EXPORT
    if raw.startswith(STAGE_SAZHENCY + SEEDLING_SEP):
        return STAGE_SAZHENCY_EXPORT + raw[len(STAGE_SAZHENCY):]
    if raw.startswith(STAGE_SAZHENCY + ' '):
        return STAGE_SAZHENCY_EXPORT + raw[len(STAGE_SAZHENCY):]
    return raw


def stage_export_label(stage: str) -> str:
    """Колонка «Стадия» в Excel площадки / PDF: Товарный → Саженцы."""
    raw = (stage or '').strip()
    if not raw:
        return '—'
    if raw in (STAGE_SAZHENCY, STAGE_SAZHENCY_LEGACY, STAGE_SAZHENCY_UI):
        return STAGE_SAZHENCY_EXPORT
    if raw.startswith(STAGE_SAZHENCY + ' ') or raw.startswith(STAGE_SAZHENCY + '('):
        return STAGE_SAZHENCY_EXPORT + raw[len(STAGE_SAZHENCY):]
    if raw.startswith(STAGE_SAZHENCY_LEGACY + ' ') or raw.startswith(STAGE_SAZHENCY_LEGACY + '('):
        return STAGE_SAZHENCY_EXPORT + raw[len(STAGE_SAZHENCY_LEGACY):]
    return raw


def parse_seedling_size(name: str) -> Optional[dict]:
    """Разбирает Size.name саженцевой линейки.

    Returns:
      None — не саженцевая линейка контейнеров.
      dict: stage ('Нетов'|'Товарный'), container, grade (или None),
            measured (bool), bare (bool для голых «Товарный»/«Нетов» / legacy «Саженцы»).
    """
    raw = (name or '').strip()
    if not raw:
        return None

    if is_bare_sazhency_name(raw):
        return {
            'stage': STAGE_SAZHENCY,
            'container': None,
            'grade': None,
            'measured': False,
            'bare': True,
        }
    if is_bare_netov_name(raw):
        return {
            'stage': STAGE_NETOV,
            'container': None,
            'grade': None,
            'measured': False,
            'bare': True,
        }

    parts = [p.strip() for p in raw.split(SEEDLING_SEP) if p.strip()]
    if len(parts) < 2:
        return None

    stage_raw = parts[0]
    stage_low = stage_raw.lower().replace('.', '')
    if stage_low.startswith('нетов'):
        stage = STAGE_NETOV
    elif _is_sazhency_stage_token(stage_raw):
        stage = STAGE_SAZHENCY
    else:
        return None

    container = parts[1]
    grade = parts[2] if len(parts) >= 3 else None
    measured = bool(stage == STAGE_SAZHENCY and container and grade)
    return {
        'stage': stage,
        'container': container,
        'grade': grade,
        'measured': measured,
        'bare': False,
    }


def format_seedling_size(stage: str, container: str, grade: str | None = None) -> str:
    container = normalize_container_name(container)
    if not container:
        raise ValueError('Не указан контейнер')
    if stage == STAGE_NETOV or (stage or '').lower().startswith('нетов'):
        return f'{STAGE_NETOV}{SEEDLING_SEP}{container}'
    # Нормализуем legacy «Саженцы» / любой товарный токен к канону «Товарный».
    if stage in (STAGE_SAZHENCY, STAGE_SAZHENCY_LEGACY, STAGE_SAZHENCY_UI) or _is_sazhency_stage_token(stage or ''):
        stage = STAGE_SAZHENCY
    else:
        stage = STAGE_SAZHENCY
    if grade:
        return f'{STAGE_SAZHENCY}{SEEDLING_SEP}{container}{SEEDLING_SEP}{normalize_grade(grade)}'
    return f'{STAGE_SAZHENCY}{SEEDLING_SEP}{container}'


def is_seedling_size_name(name: str) -> bool:
    """Любой размер саженцевого контура (вкл. голые «Саженцы»/«Нетов»)."""
    return parse_seedling_size(name) is not None


def is_measured_seedling_size_name(name: str) -> bool:
    parsed = parse_seedling_size(name)
    return bool(parsed and parsed.get('measured'))


def is_product_seedling_size_name(name: str) -> bool:
    """Товарный саженец: стадия «Саженцы» + контейнер (промер/сорт необязателен).

    Примеры товарных: «Саженцы · C5», «Саженцы · C5 · 20-30».
    Нетоварные: «Нетов · C5», голый «Саженцы» / «Нетов» без контейнера.
    """
    parsed = parse_seedling_size(name)
    if not parsed or parsed.get('bare'):
        return False
    return bool(
        parsed.get('stage') == STAGE_SAZHENCY
        and parsed.get('container')
    )


def is_excluded_from_product_stock(name: str) -> bool:
    """Что скрываем из «товарных» остатков/обычного прайса и витрины.

    Товарные саженцы (Саженцы · контейнер [· промер]) — НЕ исключаем.
    Нетов (в т.ч. «Нетов · C5», «Нетов (без контейнера)») — всегда исключаем.
    """
    raw = (name or '').strip()
    if not raw:
        return False
    if is_bare_netov_name(raw) or is_bare_sazhency_name(raw):
        return True
    # Legacy/UI: «Нетов (без контейнера)» не парсятся через « · », но это нетоварный контур.
    low_compact = re.sub(r'[\s.]+', '', raw.lower())
    if low_compact.startswith('нетов') and parse_seedling_size(raw) is None:
        return True
    parsed = parse_seedling_size(raw)
    if not parsed:
        return False
    if is_product_seedling_size_name(raw):
        return False
    return True

def allows_order_deficit(size_name: str) -> bool:
    """Минус в заказе разрешён только для размеров-саженцев."""
    parsed = parse_seedling_size(size_name)
    if not parsed:
        return False
    # Голый «Нетов» без контейнера — не продаём в минус; саженцы — да
    if parsed.get('stage') == STAGE_NETOV and not parsed.get('container'):
        return False
    if parsed.get('stage') == STAGE_NETOV:
        return False
    return parsed.get('stage') == STAGE_SAZHENCY


def get_or_create_size_by_name(name: str) -> Size:
    name = (name or '').strip()
    if not name:
        raise ValueError('Пустое имя размера')
    size = Size.query.filter_by(name=name).first()
    if size:
        return size
    # Unique — ловим гонку
    size = Size(name=name)
    db.session.add(size)
    try:
        db.session.flush()
    except Exception:
        db.session.rollback()
        size = Size.query.filter_by(name=name).first()
        if not size:
            raise
    return size


def get_or_create_seedling_size(stage: str, container: str, grade: str | None = None) -> Size:
    return get_or_create_size_by_name(format_seedling_size(stage, container, grade))


def ensure_seedling_containers() -> list[SeedlingContainer]:
    """Создаёт таблицу/дефолтные контейнеры при первом обращении."""
    try:
        rows = SeedlingContainer.query.order_by(
            SeedlingContainer.sort_order.asc(),
            SeedlingContainer.name.asc(),
        ).all()
    except Exception:
        db.session.rollback()
        db.create_all()
        rows = SeedlingContainer.query.order_by(
            SeedlingContainer.sort_order.asc(),
            SeedlingContainer.name.asc(),
        ).all()

    if rows:
        return rows

    for i, name in enumerate(DEFAULT_CONTAINERS):
        db.session.add(SeedlingContainer(name=name, sort_order=i * 10, is_active=True))
    db.session.commit()
    return SeedlingContainer.query.order_by(
        SeedlingContainer.sort_order.asc(),
        SeedlingContainer.name.asc(),
    ).all()


def list_active_containers() -> list[SeedlingContainer]:
    ensure_seedling_containers()
    return (
        SeedlingContainer.query
        .filter_by(is_active=True)
        .order_by(SeedlingContainer.sort_order.asc(), SeedlingContainer.name.asc())
        .all()
    )


def _fact_qty(plant_id: int, size_id: int, field_id: int, year: int) -> int:
    sb = StockBalance.query.filter_by(
        plant_id=plant_id, size_id=size_id, field_id=field_id, year=year,
    ).first()
    return int(sb.quantity or 0) if sb else 0


def _add_regrading_row(
    doc: Document,
    plant_id: int,
    size_from_id: int,
    size_to_id: int,
    field_from_id: int,
    year: int,
    qty: int,
    field_to_id: int | None = None,
) -> None:
    """Пересортировка/пересадка. Нельзя увести источник ниже 0."""
    if qty <= 0:
        return
    if size_from_id == size_to_id and (field_to_id is None or field_to_id == field_from_id):
        return
    dst_field = field_to_id if field_to_id is not None else field_from_id

    src = get_or_create_stock(plant_id, size_from_id, field_from_id, year)
    available = int(src.quantity or 0)
    if qty > available:
        from app.models import Plant, Size, Field
        pl = Plant.query.get(plant_id)
        sz = Size.query.get(size_from_id)
        fld = Field.query.get(field_from_id)
        raise ValueError(
            f'Нельзя списать {qty} шт: доступно {available} '
            f'(«{pl.name if pl else plant_id}» / «{sz.name if sz else size_from_id}» / '
            f'«{fld.name if fld else field_from_id}» / {year})'
        )

    db.session.add(DocumentRow(
        document_id=doc.id,
        plant_id=plant_id,
        size_id=size_from_id,
        size_to_id=size_to_id,
        field_from_id=field_from_id,
        field_to_id=dst_field if dst_field != field_from_id else None,
        year=year,
        quantity=qty,
    ))
    src.quantity = available - qty
    if src.quantity < 0:
        raise ValueError('Внутренняя ошибка: остаток ушёл в минус')
    dst = get_or_create_stock(plant_id, size_to_id, dst_field, year)
    dst.quantity = int(dst.quantity or 0) + qty


def log_seedling_event(
    *,
    action: str,
    message: str,
    user_id: int | None,
    project_id: int | None = None,
    document_id: int | None = None,
) -> None:
    from app.models import SeedlingEventLog
    db.session.add(SeedlingEventLog(
        action=action,
        message=message or '',
        user_id=user_id,
        project_id=project_id,
        document_id=document_id,
        created_at=msk_now(),
    ))


def list_project_seedling_events(project_id: int, limit: int = 200) -> list:
    from app.models import SeedlingEventLog
    ensure_seedling_schema()
    return (
        SeedlingEventLog.query
        .filter_by(project_id=project_id)
        .order_by(SeedlingEventLog.created_at.desc(), SeedlingEventLog.id.desc())
        .limit(limit)
        .all()
    )


def _parse_doc_date(raw):
    from datetime import datetime as _dt
    text = (raw or '').strip()
    if not text:
        return msk_now()
    for fmt in ('%Y-%m-%d', '%d.%m.%Y'):
        try:
            d = _dt.strptime(text, fmt)
            now = msk_now()
            return d.replace(hour=now.hour, minute=now.minute, second=now.second, microsecond=0)
        except ValueError:
            continue
    return msk_now()


def _new_regrading_doc(
    user_id: int,
    comment: str,
    project_id: int | None = None,
    doc_date=None,
) -> Document:
    doc = Document(
        doc_type='regrading',
        date=doc_date if doc_date is not None else msk_now(),
        user_id=user_id,
        comment=comment,
        project_id=project_id,
    )
    db.session.add(doc)
    db.session.flush()
    return doc


def transplant_to_container(
    *,
    plant_id: int,
    source_size_id: int,
    container: str,
    field_from_id: int,
    year: int,
    quantity: int,
    user_id: int,
    field_to_id: int | None = None,
    project_id: int | None = None,
    doc_date=None,
    doc: Document | None = None,
    commit: bool = True,
) -> Document:
    """Пересадка: с исходного размера/поля → «Нетов · контейнер» на поле назначения."""
    qty = int(quantity or 0)
    if qty <= 0:
        raise ValueError('Количество должно быть больше 0')
    container = normalize_container_name(container)
    if not container:
        raise ValueError('Укажите контейнер')

    src_size = Size.query.get(source_size_id)
    if not src_size:
        raise ValueError('Исходный размер не найден')

    dst_field = int(field_to_id or field_from_id)
    dst_size = get_or_create_seedling_size(STAGE_NETOV, container)
    if doc is None:
        doc = _new_regrading_doc(
            user_id,
            'Саженцы: пересадка',
            project_id=project_id,
            doc_date=doc_date,
        )
    _add_regrading_row(
        doc, plant_id, source_size_id, dst_size.id,
        int(field_from_id), year, qty, field_to_id=dst_field,
    )
    if commit:
        db.session.commit()
    return doc


def transplant_batch(
    *,
    lines: list[dict],
    user_id: int,
    project_id: int | None = None,
    doc_date=None,
) -> Document:
    if not lines:
        raise ValueError('Добавьте хотя бы одну позицию')
    doc = _new_regrading_doc(
        user_id,
        f'Саженцы: пересадка ({len(lines)} поз.)',
        project_id=project_id,
        doc_date=doc_date,
    )
    for line in lines:
        transplant_to_container(
            plant_id=int(line['plant_id']),
            source_size_id=int(line['source_size_id']),
            container=line['container'],
            field_from_id=int(line['field_from_id']),
            field_to_id=int(line.get('field_to_id') or line['field_from_id']),
            year=int(line['year']),
            quantity=int(line['quantity']),
            user_id=user_id,
            project_id=project_id,
            doc=doc,
            commit=False,
        )
    log_seedling_event(
        action='transplant',
        message=f'Пересадка: {len(lines)} поз., док. #{doc.id}',
        user_id=user_id,
        project_id=project_id,
        document_id=doc.id,
    )
    db.session.commit()
    return doc


def split_commercial(
    *,
    plant_id: int,
    container: str,
    field_id: int,
    year: int,
    quantity: int,
    to_commercial: bool,
    user_id: int,
    project_id: int | None = None,
    doc_date=None,
    doc: Document | None = None,
    commit: bool = True,
) -> Document:
    """Товарность: Нетов·C ↔ Саженцы·C."""
    qty = int(quantity or 0)
    if qty <= 0:
        raise ValueError('Количество должно быть больше 0')
    container = normalize_container_name(container)
    netov = get_or_create_seedling_size(STAGE_NETOV, container)
    sazh = get_or_create_seedling_size(STAGE_SAZHENCY, container)
    if to_commercial:
        src, dst = netov, sazh
    else:
        src, dst = sazh, netov

    if doc is None:
        doc = _new_regrading_doc(
            user_id,
            f'Саженцы: товарность ({container})',
            project_id=project_id,
            doc_date=doc_date,
        )
    _add_regrading_row(doc, plant_id, src.id, dst.id, field_id, year, qty)
    if commit:
        db.session.commit()
    return doc


def split_commercial_batch(
    *,
    lines: list[dict],
    user_id: int,
    project_id: int | None = None,
    doc_date=None,
) -> Document:
    if not lines:
        raise ValueError('Добавьте хотя бы одну позицию')
    doc = _new_regrading_doc(
        user_id,
        f'Саженцы: товарность ({len(lines)} поз.)',
        project_id=project_id,
        doc_date=doc_date,
    )
    for line in lines:
        split_commercial(
            plant_id=int(line['plant_id']),
            container=line['container'],
            field_id=int(line['field_id']),
            year=int(line['year']),
            quantity=int(line['quantity']),
            to_commercial=bool(line.get('to_commercial')),
            user_id=user_id,
            project_id=project_id,
            doc=doc,
            commit=False,
        )
    log_seedling_event(
        action='split_commercial',
        message=f'Товарность: {len(lines)} поз., док. #{doc.id}',
        user_id=user_id,
        project_id=project_id,
        document_id=doc.id,
    )
    db.session.commit()
    return doc


def snapshot_for_container(
    plant_id: int, container: str, field_id: int, year: int,
) -> dict:
    """Текущий срез: нетов + саженцы без промера + промеры."""
    container = normalize_container_name(container)
    netov_name = format_seedling_size(STAGE_NETOV, container)
    pool_name = format_seedling_size(STAGE_SAZHENCY, container)
    netov_size = Size.query.filter_by(name=netov_name).first()
    pool_size = Size.query.filter_by(name=pool_name).first()
    grades = []
    for sz in Size.query.all():
        parsed = parse_seedling_size(sz.name)
        if not parsed or parsed.get('bare'):
            continue
        if parsed.get('container') != container:
            continue
        if parsed.get('stage') != STAGE_SAZHENCY or not parsed.get('grade'):
            continue
        qty = _fact_qty(plant_id, sz.id, field_id, year)
        grades.append({
            'size_id': sz.id,
            'grade': parsed['grade'],
            'size_name': sz.name,
            'quantity': qty,
        })
    grades.sort(key=lambda g: g['grade'])
    return {
        'container': container,
        'netov': {
            'size_id': netov_size.id if netov_size else None,
            'size_name': netov_name,
            'quantity': _fact_qty(plant_id, netov_size.id, field_id, year) if netov_size else 0,
        },
        'pool': {
            'size_id': pool_size.id if pool_size else None,
            'size_name': pool_name,
            'quantity': _fact_qty(plant_id, pool_size.id, field_id, year) if pool_size else 0,
        },
        'grades': grades,
    }


def commercial_snapshot(plant_id: int, container: str, field_id: int, year: int) -> dict:
    """Срез только товарных саженцев: пул без промера + промеры (Нетов не трогаем)."""
    snap = snapshot_for_container(plant_id, container, field_id, year)
    commercial_total = int(snap['pool']['quantity']) + sum(int(g['quantity']) for g in snap['grades'])
    return {
        **snap,
        'commercial_total': commercial_total,
    }


def apply_measure_line(
    *,
    doc: Document,
    plant_id: int,
    container: str,
    field_id: int,
    year: int,
    grade_qtys: dict[str, int],
) -> int:
    """Промер одной позиции: только Саженцы·C и Саженцы·C·размер.

    Возвращает итоговый коммерческий объём, который разнесли.
    """
    from app.models import Plant

    container = normalize_container_name(container)
    cleaned_grades: dict[str, int] = {}
    for g, q in (grade_qtys or {}).items():
        g_norm = normalize_grade(g)
        if not g_norm:
            continue
        if not _GRADE_RE.match(g_norm):
            raise ValueError(f'Некорректный размер промера: {g_norm}')
        q_int = max(0, int(q or 0))
        if q_int > 0:
            cleaned_grades[g_norm] = cleaned_grades.get(g_norm, 0) + q_int

    snap = commercial_snapshot(plant_id, container, field_id, year)
    current_total = int(snap['commercial_total'])
    target_grades = sum(cleaned_grades.values())
    pl = Plant.query.get(plant_id)
    label = f'«{pl.name if pl else plant_id}» · {container}'

    if current_total <= 0:
        raise ValueError(f'{label}: нет товарных саженцев для промера')
    if target_grades <= 0:
        raise ValueError(f'{label}: укажите хотя бы один размер промера')
    if target_grades > current_total:
        raise ValueError(
            f'{label}: сумма промеров {target_grades} больше доступных '
            f'товарных {current_total}'
        )
    leftover_pool = current_total - target_grades

    pool = get_or_create_seedling_size(STAGE_SAZHENCY, container)

    # 1) Все промеренные → в пул
    for g in snap['grades']:
        q = int(g['quantity'] or 0)
        if q > 0 and g['size_id'] != pool.id:
            _add_regrading_row(doc, plant_id, g['size_id'], pool.id, field_id, year, q)

    db.session.flush()
    pool_qty = _fact_qty(plant_id, pool.id, field_id, year)
    if pool_qty < target_grades:
        raise ValueError(f'{label}: внутренний сбой промера (пул {pool_qty} < {target_grades})')

    # 2) Из пула → размеры
    for grade, q in cleaned_grades.items():
        dst = get_or_create_seedling_size(STAGE_SAZHENCY, container, grade)
        _add_regrading_row(doc, plant_id, pool.id, dst.id, field_id, year, q)

    # leftover остаётся в «Саженцы · C»
    if leftover_pool < 0:
        raise ValueError(f'{label}: остаток пула ушёл в минус')
    return current_total


def apply_measure_batch(
    *,
    lines: list[dict],
    user_id: int,
    project_id: int | None = None,
    doc_date=None,
) -> Document:
    if not lines:
        raise ValueError('Добавьте хотя бы одну позицию промера')
    doc = _new_regrading_doc(
        user_id,
        f'Саженцы: промер ({len(lines)} поз.)',
        project_id=project_id,
        doc_date=doc_date,
    )
    total_moved = 0
    for line in lines:
        total_moved += apply_measure_line(
            doc=doc,
            plant_id=int(line['plant_id']),
            container=line['container'],
            field_id=int(line['field_id']),
            year=int(line['year']),
            grade_qtys=line.get('grade_qtys') or {},
        )
    log_seedling_event(
        action='measure',
        message=f'Промер: {len(lines)} поз., {total_moved} шт, док. #{doc.id}',
        user_id=user_id,
        project_id=project_id,
        document_id=doc.id,
    )
    db.session.commit()
    return doc


# Совместимость со старым API (один промер с netov — больше не используется в UI)
def apply_measure_snapshot(
    *,
    plant_id: int,
    container: str,
    field_id: int,
    year: int,
    netov_qty: int = 0,
    grade_qtys: dict[str, int] | None = None,
    user_id: int,
    project_id: int | None = None,
    doc_date=None,
) -> Document:
    return apply_measure_batch(
        lines=[{
            'plant_id': plant_id,
            'container': container,
            'field_id': field_id,
            'year': year,
            'grade_qtys': grade_qtys or {},
        }],
        user_id=user_id,
        project_id=project_id,
        doc_date=doc_date,
    )

def suggest_deficit_sources(
    plant_id: int,
    size_id: int,
    field_id: int | None,
    year: int | None,
    need_qty: int,
) -> list[dict]:
    """Откуда взять минус: другие промеры той же культуры+контейнера."""
    size = Size.query.get(size_id)
    if not size:
        return []
    parsed = parse_seedling_size(size.name)
    if not parsed or not parsed.get('container') or parsed.get('stage') != STAGE_SAZHENCY:
        return []

    container = parsed['container']
    from app.stock_helpers import compute_free

    suggestions = []
    sizes = Size.query.all()
    for sz in sizes:
        if sz.id == size_id:
            continue
        p = parse_seedling_size(sz.name)
        if not p or p.get('container') != container:
            continue
        if p.get('stage') != STAGE_SAZHENCY or not p.get('grade'):
            continue

        stocks = (
            StockBalance.query
            .options(joinedload(StockBalance.field), joinedload(StockBalance.size))
            .filter(
                StockBalance.plant_id == plant_id,
                StockBalance.size_id == sz.id,
                StockBalance.quantity > 0,
            )
            .all()
        )
        for st in stocks:
            fact, reserved, free = compute_free(
                plant_id, sz.id, st.field_id, st.year,
            )
            if free <= 0:
                continue
            suggestions.append({
                'plant_id': plant_id,
                'size_id': sz.id,
                'size_name': sz.name,
                'grade': p.get('grade'),
                'field_id': st.field_id,
                'field_name': st.field.name if st.field else '',
                'year': st.year,
                'free': free,
                'fact': fact,
                'reserved': reserved,
                'same_field': field_id is not None and st.field_id == int(field_id),
                'same_year': year is not None and st.year == int(year),
            })

    suggestions.sort(key=lambda x: (
        0 if x['same_field'] else 1,
        0 if x['same_year'] else 1,
        -x['free'],
        x['size_name'],
    ))
    return suggestions


def list_source_batches(plant_id: int, size_id: int, preferred_field_id: int | None = None) -> list[dict]:
    """Партии с остатком для источника пересадки."""
    from app.models import Field
    stocks = (
        StockBalance.query
        .filter(
            StockBalance.plant_id == plant_id,
            StockBalance.size_id == size_id,
            StockBalance.quantity > 0,
        )
        .all()
    )
    out = []
    for st in stocks:
        field = Field.query.get(st.field_id)
        out.append({
            'field_id': st.field_id,
            'field_name': field.name if field else f'#{st.field_id}',
            'year': st.year,
            'qty': int(st.quantity or 0),
            'preferred': preferred_field_id is not None and st.field_id == preferred_field_id,
        })
    out.sort(key=lambda x: (0 if x['preferred'] else 1, x['field_name'], -x['year']))
    return out


def list_measure_years(plant_id: int, container: str, field_id: int) -> list[dict]:
    """Годы с ненулевым товарным срезом (Саженцы · C / промеры)."""
    container = normalize_container_name(container)
    years: dict[int, int] = {}
    for sz in Size.query.all():
        parsed = parse_seedling_size(sz.name)
        if not parsed or parsed.get('bare'):
            continue
        if parsed.get('container') != container:
            continue
        if parsed.get('stage') != STAGE_SAZHENCY:
            continue
        stocks = StockBalance.query.filter_by(
            plant_id=plant_id, size_id=sz.id, field_id=field_id,
        ).filter(StockBalance.quantity != 0).all()
        for st in stocks:
            years[st.year] = years.get(st.year, 0) + int(st.quantity or 0)
    return [
        {'year': y, 'qty': years[y]}
        for y in sorted(years.keys(), reverse=True)
    ]


def parse_grades_blob(raw: str) -> dict[str, int]:
    """'20-30:10;30-40:5' → dict."""
    out: dict[str, int] = {}
    for part in (raw or '').replace(',', ';').split(';'):
        part = part.strip()
        if not part:
            continue
        if ':' in part:
            name, qty_s = part.rsplit(':', 1)
        elif '=' in part:
            name, qty_s = part.rsplit('=', 1)
        else:
            continue
        try:
            q = int(qty_s.strip())
        except (TypeError, ValueError):
            continue
        g = normalize_grade(name)
        if g and q > 0:
            out[g] = out.get(g, 0) + q
    return out


def build_container_yard_stock(project) -> dict:
    """Все остатки на контейнерной площадке, привязанной к проекту."""
    from app.finance import _project_potting_stock_field_label
    from app.models import Plant
    from app.stock_helpers import get_reserved_map
    from app.utils import natural_key

    field_id, field_label = _project_potting_stock_field_label(project)
    if not field_id:
        return {
            'field_id': None,
            'field_label': field_label,
            'groups': [],
            'totals': {'qty': 0, 'reserved': 0, 'free': 0, 'sum': 0.0},
            'filter_stages': [],
            'filter_containers': [],
            'filter_years': [],
        }

    reserved_map = get_reserved_map(field_id=field_id)
    sizes = {s.id: s for s in Size.query.all()}
    stocks = (
        StockBalance.query
        .filter(
            StockBalance.field_id == field_id,
            StockBalance.quantity != 0,
        )
        .all()
    )
    plants = {p.id: p for p in Plant.query.all()}
    grouped = {}
    totals = {'qty': 0, 'reserved': 0, 'free': 0, 'sum': 0.0}
    for st in stocks:
        qty = int(st.quantity or 0)
        if qty == 0:
            continue
        reserved = int(reserved_map.get((st.plant_id, st.size_id, st.field_id, st.year), 0) or 0)
        free = qty - reserved
        plant = plants.get(st.plant_id)
        size = sizes.get(st.size_id)
        size_name = size.name if size else ''
        parsed = parse_seedling_size(size_name)
        key = st.plant_id
        if key not in grouped:
            grouped[key] = {
                'plant_id': st.plant_id,
                'plant_name': plant.name if plant else f'#{st.plant_id}',
                'rows': [],
                'totals': {'qty': 0, 'reserved': 0, 'free': 0, 'sum': 0.0},
            }
        if parsed:
            stage = parsed.get('stage') or ''
            container = parsed.get('container') or '—'
            grade = parsed.get('grade') or ('—' if not parsed.get('bare') else 'без контейнера')
            if parsed.get('bare'):
                label = f"{size_name} (без контейнера)"
                container = '—'
            elif parsed.get('measured'):
                label = f"{container} · {grade}"
            elif stage == STAGE_NETOV:
                label = f"Нетов · {container}"
            else:
                label = f"{STAGE_SAZHENCY_UI} · {container} (без промера)"
        else:
            stage, container, grade = '', '—', ''
            label = size_name_ui_label(size_name) or '—'

        price = float(st.price or 0)
        row_sum = price * qty
        grouped[key]['rows'].append({
            'size_id': st.size_id,
            'size_name': size_name,
            'label': label,
            'stage': stage or '—',
            'stage_ui': stage_ui_label(stage) if stage else '—',
            'container': container,
            'grade': grade or '',
            'year': st.year,
            'qty': qty,
            'reserved': reserved,
            'free': free,
            'price': price,
            'sum': row_sum,
        })
        grouped[key]['totals']['qty'] += qty
        grouped[key]['totals']['reserved'] += reserved
        grouped[key]['totals']['free'] += free
        grouped[key]['totals']['sum'] += row_sum
        totals['qty'] += qty
        totals['reserved'] += reserved
        totals['free'] += free
        totals['sum'] = totals.get('sum', 0) + row_sum

    stages = set()
    containers = set()
    years = set()
    groups = []
    for g in grouped.values():
        g['rows'].sort(key=lambda r: (
            0 if r['stage'] == STAGE_SAZHENCY else (1 if r['stage'] == STAGE_NETOV else 2),
            natural_key(r['container']),
            natural_key(r['grade'] or ''),
            -int(r['year'] or 0),
        ))
        for r in g['rows']:
            if r.get('stage'):
                stages.add(r['stage'])
            if r.get('container'):
                containers.add(r['container'])
            if r.get('year') is not None:
                years.add(int(r['year']))
        groups.append(g)
    groups.sort(key=lambda g: natural_key(g['plant_name']))
    return {
        'field_id': field_id,
        'field_label': field_label,
        'groups': groups,
        'totals': totals,
        'filter_stages': sorted(stages, key=lambda s: s.lower()),
        'filter_containers': sorted(containers, key=natural_key),
        'filter_years': sorted(years, reverse=True),
    }


def list_grade_sizes_from_handbook() -> list[Size]:
    """Обычные размеры из справочника (без саженцевой линейки) для промера."""
    from app.utils import natural_key
    return sorted(
        [s for s in Size.query.all() if not is_seedling_size_name(s.name)],
        key=lambda s: natural_key(s.name),
    )


def list_plant_seedling_locations(plant_id: int) -> list[dict]:
    """Все ячейки саженцевой линейки по растению: где лежат (поле/размер/год)."""
    from app.models import Field
    from app.stock_helpers import compute_free
    from app.utils import natural_key

    sizes = {s.id: s for s in Size.query.all() if parse_seedling_size(s.name) is not None}
    if not sizes:
        return []
    stocks = (
        StockBalance.query
        .filter(
            StockBalance.plant_id == plant_id,
            StockBalance.size_id.in_(list(sizes.keys())),
            StockBalance.quantity != 0,
        )
        .all()
    )
    fields = {f.id: f for f in Field.query.all()}
    out = []
    for st in stocks:
        qty = int(st.quantity or 0)
        if qty == 0:
            continue
        sz = sizes.get(st.size_id)
        parsed = parse_seedling_size(sz.name if sz else '') or {}
        fact, reserved, free = compute_free(plant_id, st.size_id, st.field_id, st.year)
        fld = fields.get(st.field_id)
        out.append({
            'plant_id': plant_id,
            'size_id': st.size_id,
            'size_name': sz.name if sz else '',
            'size_name_ui': size_name_ui_label(sz.name if sz else ''),
            'field_id': st.field_id,
            'field_name': fld.name if fld else f'#{st.field_id}',
            'year': st.year,
            'qty': qty,
            'reserved': reserved,
            'free': free,
            'stage': parsed.get('stage') or '',
            'container': parsed.get('container') or '',
            'grade': parsed.get('grade') or '',
            'bare': bool(parsed.get('bare')),
            'measured': bool(parsed.get('measured')),
        })
    out.sort(key=lambda r: (
        natural_key(r['field_name']),
        natural_key(r['size_name']),
        -int(r['year'] or 0),
    ))
    return out


def build_plant_pipeline_view(plant_id: int, preferred_field_id: int | None = None) -> dict:
    """Сводка по цепочке Пересадка → Товарность → Промер для выбранного растения."""
    from app.models import Plant

    plant = Plant.query.get(plant_id)
    items = list_plant_seedling_locations(plant_id)
    if preferred_field_id:
        preferred = [i for i in items if int(i.get('field_id') or 0) == int(preferred_field_id)]
        # Показываем все, но сначала площадку проекта.
        items = preferred + [i for i in items if i not in preferred]

    stages = [
        {
            'key': 'transplant',
            'step': 1,
            'title': 'Пересадка',
            'hint': 'Нетов · контейнер',
            'action': 'Посадка в контейнер с поля / из «Нетов без контейнера»',
            'qty': 0,
            'free': 0,
            'lines': [],
        },
        {
            'key': 'commercial',
            'step': 2,
            'title': 'Товарность',
            'hint': 'Саженцы · контейнер',
            'action': 'Нетов · C → Саженцы · C (товарные без промера)',
            'qty': 0,
            'free': 0,
            'lines': [],
        },
        {
            'key': 'measure',
            'step': 3,
            'title': 'Промер',
            'hint': 'Саженцы · C · размер',
            'action': 'Промер товарных саженцев по размерному ряду',
            'qty': 0,
            'free': 0,
            'lines': [],
        },
    ]
    by_key = {s['key']: s for s in stages}
    other_lines = []

    for it in items:
        qty = int(it.get('qty') or 0)
        free = int(it.get('free') if it.get('free') is not None else qty)
        line = {
            'label': it.get('size_name') or '',
            'field_name': it.get('field_name') or '',
            'field_id': it.get('field_id'),
            'year': it.get('year'),
            'container': it.get('container') or '',
            'grade': it.get('grade') or '',
            'qty': qty,
            'free': free,
            'on_project_field': bool(
                preferred_field_id and int(it.get('field_id') or 0) == int(preferred_field_id)
            ),
        }
        if it.get('measured'):
            key = 'measure'
        elif it.get('stage') == STAGE_SAZHENCY and it.get('container') and not it.get('grade'):
            key = 'commercial'
        elif it.get('stage') == STAGE_NETOV and (it.get('container') or it.get('bare')):
            key = 'transplant'
        elif it.get('bare') and it.get('stage') == STAGE_SAZHENCY:
            # Голые «Саженцы» — ещё до контейнерной схемы; показываем у пересадки.
            key = 'transplant'
            line['label'] = f"{line['label']} (без контейнера)"
        else:
            other_lines.append(line)
            continue

        st = by_key[key]
        st['qty'] += qty
        st['free'] += free
        st['lines'].append(line)

    # Текущая стадия: последняя с остатком (куда «дошли»), иначе первая пустая.
    current_key = 'transplant'
    for s in stages:
        if s['qty'] > 0:
            current_key = s['key']
    if stages[0]['qty'] == 0 and stages[1]['qty'] == 0 and stages[2]['qty'] == 0:
        current_key = 'transplant'

    for s in stages:
        s['active'] = s['key'] == current_key
        s['has_stock'] = s['qty'] > 0

    return {
        'plant_id': plant_id,
        'plant_name': plant.name if plant else f'#{plant_id}',
        'current_key': current_key,
        'stages': stages,
        'other_lines': other_lines,
        'total_qty': sum(s['qty'] for s in stages),
    }


DIEBACK_COMMENT_PREFIX = 'Саженцы: выпад'


def dieback_batch(
    *,
    lines: list[dict],
    user_id: int,
    project_id: int | None = None,
    doc_date=None,
) -> Document:
    """Выпад: списание факта без ухода в минус. doc_type=writeoff."""
    if not lines:
        raise ValueError('Добавьте хотя бы одну позицию выпада')

    doc = Document(
        doc_type='writeoff',
        date=doc_date if doc_date is not None else msk_now(),
        user_id=user_id,
        comment=f'{DIEBACK_COMMENT_PREFIX} ({len(lines)} поз.)',
        project_id=project_id,
    )
    db.session.add(doc)
    db.session.flush()

    from app.models import Plant, Field
    total = 0
    for line in lines:
        plant_id = int(line['plant_id'])
        size_id = int(line['size_id'])
        field_id = int(line['field_id'])
        year = int(line['year'])
        qty = int(line['quantity'])
        if qty <= 0:
            continue
        available = _fact_qty(plant_id, size_id, field_id, year)
        if qty > available:
            pl = Plant.query.get(plant_id)
            sz = Size.query.get(size_id)
            fld = Field.query.get(field_id)
            raise ValueError(
                f'Выпад: нельзя списать {qty} шт — доступно {available} '
                f'(«{pl.name if pl else plant_id}» / «{sz.name if sz else size_id}» / '
                f'«{fld.name if fld else field_id}» / {year})'
            )
        stock = get_or_create_stock(plant_id, size_id, field_id, year)
        stock.quantity = available - qty
        if stock.quantity < 0:
            raise ValueError('Внутренняя ошибка выпада: остаток ушёл в минус')
        db.session.add(DocumentRow(
            document_id=doc.id,
            plant_id=plant_id,
            size_id=size_id,
            field_from_id=field_id,
            year=year,
            quantity=qty,
        ))
        total += qty

    if total <= 0:
        db.session.rollback()
        raise ValueError('Нет валидных строк выпада')

    log_seedling_event(
        action='dieback',
        message=f'Выпад: {len(lines)} поз., {total} шт, док. #{doc.id}',
        user_id=user_id,
        project_id=project_id,
        document_id=doc.id,
    )
    db.session.commit()
    return doc


def list_project_dieback_documents(project_id: int, limit: int = 100) -> list[Document]:
    return (
        Document.query
        .options(joinedload(Document.user), joinedload(Document.rows))
        .filter(
            Document.project_id == project_id,
            Document.doc_type == 'writeoff',
            Document.comment.ilike(f'{DIEBACK_COMMENT_PREFIX}%'),
        )
        .order_by(Document.date.desc(), Document.id.desc())
        .limit(limit)
        .all()
    )


def build_dieback_export_rows(project_id: int) -> list[dict]:
    """Строки для Excel-отчёта по выпаду проекта."""
    from app.models import Plant, Field, User
    docs = list_project_dieback_documents(project_id, limit=500)
    plants = {p.id: p for p in Plant.query.all()}
    sizes = {s.id: s for s in Size.query.all()}
    fields = {f.id: f for f in Field.query.all()}
    users = {u.id: u for u in User.query.all()}
    rows = []
    for doc in docs:
        user = users.get(doc.user_id)
        for r in doc.rows:
            pl = plants.get(r.plant_id)
            sz = sizes.get(r.size_id)
            fld = fields.get(r.field_from_id)
            rows.append({
                'doc_id': doc.id,
                'date': doc.date,
                'user': user.username if user else '',
                'plant': pl.name if pl else f'#{r.plant_id}',
                'size': size_name_export_label(sz.name if sz else '') or (sz.name if sz else f'#{r.size_id}'),
                'field': fld.name if fld else '',
                'year': r.year,
                'quantity': int(r.quantity or 0),
                'comment': doc.comment or '',
            })
    return rows


def export_dieback_workbook(project_id: int):
    """Workbook Excel по выпаду проекта."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Выпад'
    headers = ['Дата', 'Кто', 'Растение', 'Размер', 'Поле', 'Год', 'Кол-во', '№ док', 'Комментарий']
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')
    for idx, row in enumerate(build_dieback_export_rows(project_id), start=2):
        vals = [
            row['date'].strftime('%d.%m.%Y %H:%M') if row.get('date') else '',
            row.get('user') or '',
            row.get('plant') or '',
            row.get('size') or '',
            row.get('field') or '',
            row.get('year') or '',
            row.get('quantity') or 0,
            row.get('doc_id') or '',
            row.get('comment') or '',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=idx, column=col, value=val)
            c.border = thin
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = 16
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['I'].width = 32
    return wb


def size_export_parts(size_name: str) -> dict:
    """Разбивка Size.name на колонки Excel: Стадия / Контейнер / промер.

    Стадия в файле всегда «Саженцы» (даже если в БД «Товарный»).
    """
    parsed = parse_seedling_size(size_name)
    if parsed:
        stage = parsed.get('stage') or ''
        export_stage = stage_export_label(stage)
        if parsed.get('bare'):
            return {
                'stage': f'{export_stage} (без контейнера)',
                'container': '—',
                'grade': '—',
                'bare': True,
            }
        return {
            'stage': export_stage,
            'container': parsed.get('container') or '—',
            'grade': parsed.get('grade') or '—',
            'bare': False,
        }
    return {
        'stage': size_name_export_label(size_name) or '—',
        'container': '—',
        'grade': '—',
        'bare': False,
    }


def resolve_period_bounds(year: int, period_type: str, period_n: int):
    """Границы периода: month(1-12) / quarter(1-4) / half(1-2)."""
    from calendar import monthrange
    from datetime import datetime, time

    year = int(year)
    period_type = (period_type or 'month').strip().lower()
    period_n = int(period_n or 1)

    if period_type == 'quarter':
        if period_n < 1 or period_n > 4:
            raise ValueError('Квартал: 1–4')
        start_m = (period_n - 1) * 3 + 1
        end_m = start_m + 2
        label = f'{period_n} кв. {year}'
    elif period_type in ('half', 'halfyear', 'semester'):
        if period_n < 1 or period_n > 2:
            raise ValueError('Полугодие: 1–2')
        start_m = 1 if period_n == 1 else 7
        end_m = 6 if period_n == 1 else 12
        label = f'{period_n} полугодие {year}'
    else:
        if period_n < 1 or period_n > 12:
            raise ValueError('Месяц: 1–12')
        start_m = end_m = period_n
        months = [
            '', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
            'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь',
        ]
        label = f'{months[period_n]} {year}'

    start = datetime(year, start_m, 1, 0, 0, 0)
    end_day = monthrange(year, end_m)[1]
    end = datetime(year, end_m, end_day, 23, 59, 59)
    return start, end, label


def _unit_cost_parts(stock: StockBalance | None, year: int, costs_cache: dict) -> tuple:
    """(purchase, additive, final) — устаревший путь (полевая себестоимость)."""
    from decimal import Decimal
    from app.services import calculate_cost_data

    purchase = Decimal(str(stock.purchase_price or 0)) if stock else Decimal(0)
    total = Decimal(str(stock.current_total_cost or 0)) if stock else Decimal(0)
    if total > 0:
        additive = max(Decimal(0), total - purchase)
        return purchase, additive, total

    batch_year = int(stock.year) if stock else year
    basis = max(2017, batch_year - 1) if batch_year else year - 1
    if basis not in costs_cache:
        costs_cache[basis] = calculate_cost_data(basis).get('accumulated_costs_map', {})
    additive = Decimal(str(costs_cache[basis].get(batch_year, 0) or 0))
    return purchase, additive, purchase + additive


def _journal_income_keys(field_id: int) -> set[tuple]:
    """Позиции (plant, size, field, year) с обычным поступлением через журнал (income)."""
    rows = (
        db.session.query(
            DocumentRow.plant_id,
            DocumentRow.size_id,
            DocumentRow.field_to_id,
            DocumentRow.year,
        )
        .join(Document, DocumentRow.document_id == Document.id)
        .filter(
            Document.doc_type == 'income',
            DocumentRow.field_to_id == field_id,
            DocumentRow.quantity > 0,
        )
        .all()
    )
    return {(int(r[0]), int(r[1]), int(r[2]), int(r[3])) for r in rows if r[2] is not None}


def _journal_income_purchase_map(field_id: int) -> dict:
    """Средняя цена закупа из документов поступления (income) по ключу позиции."""
    from decimal import Decimal

    rows = (
        db.session.query(
            DocumentRow.plant_id,
            DocumentRow.size_id,
            DocumentRow.field_to_id,
            DocumentRow.year,
            DocumentRow.quantity,
            DocumentRow.purchase_price,
        )
        .join(Document, DocumentRow.document_id == Document.id)
        .filter(
            Document.doc_type == 'income',
            DocumentRow.field_to_id == field_id,
            DocumentRow.quantity > 0,
        )
        .all()
    )
    agg: dict[tuple, list] = {}
    for plant_id, size_id, field_to_id, year, qty, pp in rows:
        if field_to_id is None:
            continue
        key = (int(plant_id), int(size_id), int(field_to_id), int(year))
        agg.setdefault(key, []).append((int(qty or 0), Decimal(str(pp or 0))))
    result = {}
    for key, parts in agg.items():
        total_qty = sum(q for q, _ in parts)
        if total_qty <= 0:
            continue
        # Если в строках поступления цена не заполнена — берём 0 (не подменяем отчётом)
        priced = [(q, p) for q, p in parts if p > 0]
        if not priced:
            result[key] = Decimal(0)
            continue
        pq = sum(q for q, _ in priced)
        result[key] = sum(p * Decimal(q) for q, p in priced) / Decimal(pq)
    return result


def _container_yard_cost_maps(project, field_id: int, selected_year: int | None = None) -> dict:
    """Карты цен для Excel площадки: как вкладка «Себестоимость → Контейнерная площадка»."""
    from decimal import Decimal
    from app.models import StockPurchaseLot
    from app.services import calculate_container_cost_data, get_cost_container_project_ids
    from app.utils import msk_now

    year = int(selected_year or msk_now().year)
    project_ids = list(get_cost_container_project_ids() or [])
    if project and getattr(project, 'id', None) and project.id not in project_ids:
        project_ids.append(int(project.id))
    if not project_ids and project and getattr(project, 'id', None):
        project_ids = [int(project.id)]

    container_cd = calculate_container_cost_data(
        project_ids=project_ids,
        selected_year=year,
    )
    accum_opex_map = container_cd.get('accum_opex_map') or {}

    # Цена закупа как в отчёте себестоимости (лоты → иначе StockBalance)
    purchase_by_key: dict[tuple, Decimal] = {}
    for sb in StockBalance.query.filter(StockBalance.field_id == field_id).all():
        key = (sb.plant_id, sb.size_id, sb.field_id, sb.year)
        purchase_by_key[key] = Decimal(str(sb.purchase_price or 0))

    lot_purchase: dict[tuple, list] = {}
    for lot in StockPurchaseLot.query.filter(
        StockPurchaseLot.field_id == field_id,
        StockPurchaseLot.quantity > 0,
    ).all():
        key = (lot.plant_id, lot.size_id, lot.field_id, lot.year)
        lot_purchase.setdefault(key, []).append(lot)

    for key, lots in lot_purchase.items():
        total_qty = sum(int(l.quantity or 0) for l in lots)
        if total_qty <= 0:
            continue
        total_val = sum(
            Decimal(str(l.purchase_price or 0)) * Decimal(int(l.quantity or 0))
            for l in lots
        )
        purchase_by_key[key] = total_val / Decimal(total_qty)

    return {
        'accum_opex_map': accum_opex_map,
        'purchase_by_key': purchase_by_key,
        'income_keys': _journal_income_keys(field_id),
        'income_purchase_by_key': _journal_income_purchase_map(field_id),
    }


def _yard_cost_parts(plant_id, size_id, field_id, year, cost_maps: dict) -> tuple:
    """(purchase, additive, final) для остатков/выпада площадки.

    Цена закупа из отчёта «Себестоимость → Контейнерная площадка» —
    только если не было обычного поступления (income). Если поступление было —
    в колонку цены идёт цена из журнала поступления.
    Доб. себестоимость — накопленная opex из отчёта контейнерной себестоимости.
    """
    from decimal import Decimal

    key = (int(plant_id), int(size_id), int(field_id), int(year))
    additive = Decimal(str((cost_maps.get('accum_opex_map') or {}).get(int(year), 0) or 0))
    purchase_map = cost_maps.get('purchase_by_key') or {}
    income_map = cost_maps.get('income_purchase_by_key') or {}
    has_income = key in (cost_maps.get('income_keys') or set())
    if has_income:
        purchase = Decimal(str(income_map.get(key, 0) or 0))
    else:
        purchase = Decimal(str(purchase_map.get(key, 0) or 0))
    return purchase, additive, purchase + additive


def build_dieback_summary_rows(project_id: int, date_from, date_to, cost_maps: dict | None = None) -> list[dict]:
    """Сводные строки выпада за период (как в примере Excel)."""
    from decimal import Decimal
    from app.models import Plant, Field

    docs = (
        Document.query
        .options(joinedload(Document.rows))
        .filter(
            Document.project_id == project_id,
            Document.doc_type == 'writeoff',
            Document.comment.ilike(f'{DIEBACK_COMMENT_PREFIX}%'),
            Document.date >= date_from,
            Document.date <= date_to,
        )
        .order_by(Document.date.asc(), Document.id.asc())
        .all()
    )
    plants = {p.id: p for p in Plant.query.all()}
    sizes = {s.id: s for s in Size.query.all()}
    fields = {f.id: f for f in Field.query.all()}
    costs_cache = {}
    stock_map = {
        (sb.plant_id, sb.size_id, sb.field_id, sb.year): sb
        for sb in StockBalance.query.all()
    }

    agg = {}
    for doc in docs:
        for r in doc.rows:
            qty = int(r.quantity or 0)
            if qty <= 0:
                continue
            pl = plants.get(r.plant_id)
            sz = sizes.get(r.size_id)
            fld = fields.get(r.field_from_id)
            parts = size_export_parts(sz.name if sz else '')
            key = (
                r.plant_id,
                r.field_from_id,
                parts['stage'],
                parts['container'] or '',
                parts['grade'] or '',
            )
            if key not in agg:
                stock = stock_map.get((r.plant_id, r.size_id, r.field_from_id, r.year))
                if not stock:
                    for (pid, sid, fid, _), sb in stock_map.items():
                        if pid == r.plant_id and sid == r.size_id and fid == r.field_from_id:
                            stock = sb
                            break
                if cost_maps is not None and r.field_from_id:
                    _, _, final = _yard_cost_parts(
                        r.plant_id, r.size_id, r.field_from_id,
                        int(r.year or date_from.year), cost_maps,
                    )
                else:
                    _, _, final = _unit_cost_parts(stock, int(r.year or date_from.year), costs_cache)
                agg[key] = {
                    'plant': pl.name if pl else f'#{r.plant_id}',
                    'field': fld.name if fld else '',
                    'stage': parts['stage'],
                    'container': parts['container'],
                    'grade': parts['grade'],
                    'qty': 0,
                    'unit_cost': final,
                }
            agg[key]['qty'] += qty

    rows = list(agg.values())
    rows.sort(key=lambda x: ((x['plant'] or '').lower(), x['stage'] or '', x['container'] or '', x['grade'] or ''))
    return rows


def export_dieback_summary_workbook(project, year: int, period_type: str, period_n: int):
    """Сводный Excel выпада за месяц/квартал/полугодие."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.finance import _project_potting_stock_field_label

    date_from, date_to, period_label = resolve_period_bounds(year, period_type, period_n)
    field_id, field_label = _project_potting_stock_field_label(project)
    cost_maps = _container_yard_cost_maps(project, field_id, selected_year=year) if field_id else None
    rows = build_dieback_summary_rows(project.id, date_from, date_to, cost_maps=cost_maps)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Выпад'
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    title_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    title_font = Font(bold=True, color='FFFFFF', size=14)
    header_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    zebra_fill = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')
    bare_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    total_fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    title = f'Выпад на {date_to.strftime("%d.%m.%Y")} ({field_label or "Контейнерная площадка"}) · {period_label}'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 26

    headers = [
        'Растение', 'Поле', 'Стадия', 'Контейнер', 'Товарность',
        'Списание, шт', 'Конеч. цена ед., руб', 'Итого, руб',
    ]
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = align_center
    ws.row_dimensions[2].height = 32

    prev_plant = None
    plant_block = 0
    for idx, row in enumerate(rows, start=3):
        plant = row.get('plant') or ''
        show_plant = plant if plant != prev_plant else ''
        if plant != prev_plant:
            prev_plant = plant
            plant_block += 1
        stage = row.get('stage') or '—'
        container = row.get('container') or '—'
        grade = row.get('grade') or '—'
        is_bare = 'без контейнера' in str(stage).lower() or container in ('', '—', None)

        vals = [
            show_plant,
            row.get('field') or '',
            stage,
            container if container not in (None, '') else '—',
            grade if grade not in (None, '') else '—',
            int(row.get('qty') or 0),
            round(float(row.get('unit_cost') or 0), 2),
        ]
        row_fill = bare_fill if is_bare else (zebra_fill if plant_block % 2 == 0 else None)
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=idx, column=col, value=val)
            c.border = thin
            c.alignment = align_left if col == 1 else (align_right if col >= 6 else align_center)
            if row_fill:
                c.fill = row_fill
            if col == 7:
                c.number_format = '0.00'
            if show_plant and col == 1:
                c.font = Font(bold=True)
        total_cell = ws.cell(row=idx, column=8, value=f'=G{idx}*F{idx}')
        total_cell.border = thin
        total_cell.number_format = '0.00'
        total_cell.alignment = align_right
        if row_fill:
            total_cell.fill = row_fill

    # Итого
    if rows:
        total_row = 3 + len(rows)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
        label = ws.cell(row=total_row, column=1, value='ИТОГО')
        label.font = Font(bold=True)
        label.fill = total_fill
        label.border = thin
        for col in range(2, 6):
            c = ws.cell(row=total_row, column=col)
            c.fill = total_fill
            c.border = thin
        qty_total = ws.cell(row=total_row, column=6, value=f'=SUM(F3:F{total_row - 1})')
        qty_total.font = Font(bold=True)
        qty_total.fill = total_fill
        qty_total.border = thin
        qty_total.alignment = align_right
        empty = ws.cell(row=total_row, column=7)
        empty.fill = total_fill
        empty.border = thin
        sum_total = ws.cell(row=total_row, column=8, value=f'=SUM(H3:H{total_row - 1})')
        sum_total.font = Font(bold=True)
        sum_total.fill = total_fill
        sum_total.border = thin
        sum_total.number_format = '0.00'
        sum_total.alignment = align_right

    widths = [36, 24, 22, 12, 12, 12, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return wb, period_label


def export_yard_stock_workbook(project):
    """Excel остатков площадки проекта (включая нетоварные размеры)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.finance import _project_potting_stock_field_label
    from app.models import Plant, Field
    from app.utils import msk_now, natural_key

    field_id, field_label = _project_potting_stock_field_label(project)
    if not field_id:
        raise ValueError('Контейнерная площадка проекта не задана')

    plants = {p.id: p for p in Plant.query.all()}
    sizes = {s.id: s for s in Size.query.all()}
    field = Field.query.get(field_id)
    field_name = field.name if field else (field_label or '')
    cost_maps = _container_yard_cost_maps(project, field_id, selected_year=msk_now().year)
    stocks = (
        StockBalance.query
        .filter(StockBalance.field_id == field_id, StockBalance.quantity != 0)
        .all()
    )

    rows = []
    for st in stocks:
        qty = int(st.quantity or 0)
        if qty == 0:
            continue
        pl = plants.get(st.plant_id)
        sz = sizes.get(st.size_id)
        parts = size_export_parts(sz.name if sz else '')
        batch_year = int(st.year or msk_now().year)
        purch, additive, final = _yard_cost_parts(
            st.plant_id, st.size_id, field_id, batch_year, cost_maps,
        )
        rows.append({
            'plant': pl.name if pl else f'#{st.plant_id}',
            'field': field_name,
            'stage': parts['stage'],
            'container': parts['container'],
            'grade': parts['grade'],
            'bare': bool(parts.get('bare')),
            'purchase': round(float(purch or 0), 2),
            'additive': round(float(additive or 0), 2),
            'qty': qty,
            'sort': (
                natural_key(pl.name if pl else ''),
                parts['stage'] or '',
                parts['container'] or '',
                parts['grade'] or '',
                -batch_year,
            ),
        })
    rows.sort(key=lambda r: r['sort'])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Остатки'
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    title_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    title_font = Font(bold=True, color='FFFFFF', size=14)
    header_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    zebra_fill = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')
    bare_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    total_fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    today = msk_now()
    title_ru = f'Остатки на {today.strftime("%d.%m.%Y")} ({field_name or "Контейнерная площадка"})'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1, value=title_ru)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 26

    # Подзаголовок-легенда
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    note = ws.cell(
        row=2, column=1,
        value=(
            'Конечная цена за ед. = Цена (закуп) + Доб. себес. (формула Excel =F+G). '
            'Сумма = конечная × факт (=H*I). '
            'Стадия «Нетов» = нетоварные (ещё без промера). '
            '«Нетов (без контейнера)» — устаревший размер: назначьте контейнер на вкладке «Пересадка».'
        ),
    )
    note.font = Font(italic=True, size=9, color='555555')
    note.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 28

    headers = [
        'Растение', 'Поле', 'Стадия', 'Контейнер', 'Товарность',
        'Цена (закуп)', 'Доб. себес., руб', 'Конеч. цена за ед., руб',
        'Факт, шт', 'Сумма, руб',
    ]
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = align_center
    ws.row_dimensions[3].height = 32

    prev_plant = None
    plant_block = 0
    data_start = 4
    for offset, row in enumerate(rows):
        idx = data_start + offset
        plant = row['plant']
        show_plant = plant if plant != prev_plant else ''
        if plant != prev_plant:
            prev_plant = plant
            plant_block += 1

        row_fill = bare_fill if row.get('bare') else (zebra_fill if plant_block % 2 == 0 else None)
        purch = float(row['purchase'] or 0)
        additive = float(row['additive'] or 0)
        qty = int(row['qty'] or 0)
        vals = [
            show_plant,
            row['field'],
            row['stage'],
            row['container'],
            row['grade'],
            purch,
            additive,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=idx, column=col, value=val)
            c.border = thin
            c.alignment = align_left if col == 1 else (align_right if col >= 6 else align_center)
            if row_fill:
                c.fill = row_fill
            if col in (6, 7):
                c.number_format = '0.00'
            if show_plant and col == 1:
                c.font = Font(bold=True)

        # Конечная цена = F(закуп) + G(доб. себес.); сумма = H × I.
        final_cell = ws.cell(row=idx, column=8, value=f'=F{idx}+G{idx}')
        final_cell.border = thin
        final_cell.number_format = '0.00'
        final_cell.alignment = align_right
        if row_fill:
            final_cell.fill = row_fill

        qty_cell = ws.cell(row=idx, column=9, value=qty)
        qty_cell.border = thin
        qty_cell.alignment = align_right
        if row_fill:
            qty_cell.fill = row_fill

        sum_cell = ws.cell(row=idx, column=10, value=f'=H{idx}*I{idx}')
        sum_cell.border = thin
        sum_cell.number_format = '0.00'
        sum_cell.alignment = align_right
        if row_fill:
            sum_cell.fill = row_fill

    if rows:
        total_row = data_start + len(rows)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
        label = ws.cell(row=total_row, column=1, value='ИТОГО')
        label.font = Font(bold=True)
        label.fill = total_fill
        label.border = thin
        label.alignment = align_left
        # MergedCell (col 2–8): только стиль, value читать/писать нельзя
        for col in range(2, 9):
            c = ws.cell(row=total_row, column=col)
            c.fill = total_fill
            c.border = thin
        last_data = total_row - 1
        qty_total = ws.cell(row=total_row, column=9, value=f'=SUM(I{data_start}:I{last_data})')
        qty_total.font = Font(bold=True)
        qty_total.fill = total_fill
        qty_total.border = thin
        qty_total.alignment = align_right
        sum_total = ws.cell(row=total_row, column=10, value=f'=SUM(J{data_start}:J{last_data})')
        sum_total.font = Font(bold=True)
        sum_total.fill = total_fill
        sum_total.border = thin
        sum_total.number_format = '0.00'
        sum_total.alignment = align_right

    widths = [36, 24, 22, 12, 12, 12, 14, 16, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Чтобы Excel/LibreOffice сразу посчитали =F+G и =H*I при открытии
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    return wb


def seedling_context_for_project(project) -> dict:
    """Данные для блока UI посадки."""
    ensure_seedling_schema()
    try:
        from app.services import ensure_numbered_container_yards
        ensure_numbered_container_yards()
    except Exception:
        pass
    containers = list_active_containers()
    from app.models import Plant, Field
    from app.finance import _project_potting_stock_field_label
    from app.utils import natural_key
    plants = sorted(Plant.query.all(), key=lambda p: (p.name or '').lower())
    fields = sorted(Field.query.all(), key=lambda f: (f.name or '').lower())
    source_sizes = []
    for sz in Size.query.order_by(Size.name).all():
        parsed = parse_seedling_size(sz.name)
        if not parsed:
            continue
        if parsed.get('bare') or (
            parsed.get('stage') in (STAGE_NETOV, STAGE_SAZHENCY)
            and parsed.get('container')
            and not parsed.get('grade')
        ):
            label = size_name_ui_label(sz.name)
            if parsed.get('bare'):
                label = f'{size_name_ui_label(sz.name)} (без контейнера)'
            source_sizes.append({'id': sz.id, 'name': sz.name, 'label': label})
    grade_sizes = list_grade_sizes_from_handbook()
    # Для выпада: все размеры (обычные + саженцевые), чтобы выбрать ячейку
    all_sizes = sorted(Size.query.all(), key=lambda s: natural_key(s.name))
    default_field_id, default_field_label = _project_potting_stock_field_label(project)
    return {
        'seedling_containers': containers,
        'seedling_plants': plants,
        'seedling_fields': fields,
        'seedling_source_sizes': source_sizes,
        'seedling_grade_sizes': grade_sizes,
        'seedling_all_sizes': all_sizes,
        'seedling_year_default': msk_now().year,
        'seedling_month_default': msk_now().month,
        'seedling_default_field_id': default_field_id,
        'seedling_default_field_label': default_field_label,
        'seedling_doc_date_default': msk_now().strftime('%Y-%m-%d'),
        'seedling_yard_stock': build_container_yard_stock(project),
        'seedling_events': list_project_seedling_events(project.id) if project and project.id else [],
        'seedling_dieback_docs': list_project_dieback_documents(project.id) if project and project.id else [],
    }


def ensure_seedling_schema() -> None:
    """create_all + удлинение size.name под составные имена саженцев."""
    from sqlalchemy import inspect, text

    db.create_all()
    try:
        insp = inspect(db.engine)
        if not insp.has_table('size'):
            return
        cols = {c['name']: c for c in insp.get_columns('size')}
        col = cols.get('name')
        if not col:
            return
        col_type = str(col.get('type') or '').upper()
        if '120' in col_type or '255' in col_type or 'TEXT' in col_type:
            return
        dialect = db.engine.dialect.name
        if dialect == 'postgresql':
            db.session.execute(text('ALTER TABLE size ALTER COLUMN name TYPE VARCHAR(120)'))
            db.session.commit()
        elif dialect == 'sqlite':
            pass
    except Exception:
        db.session.rollback()


def _collect_parallel_rows(form, keys: list[str]) -> list[dict]:
    """Собирает строки из form arrays plant_id[] / …; одиночные поля тоже поддерживает."""
    lists = {}
    max_len = 0
    for key in keys:
        vals = form.getlist(f'{key}[]')
        if not vals and form.get(key) not in (None, ''):
            vals = [form.get(key)]
        lists[key] = vals
        max_len = max(max_len, len(vals))
    rows = []
    for i in range(max_len):
        row = {}
        empty = True
        for key in keys:
            vals = lists[key]
            row[key] = vals[i] if i < len(vals) else ''
            if str(row[key]).strip():
                empty = False
        if not empty:
            rows.append(row)
    return rows


def handle_seedling_project_form(project, form, user_id: int) -> tuple[bool, str, str]:
    """Обработка POST из блока саженцев. Возвращает (ok, message, category)."""
    from app.models import SeedlingContainer

    ensure_seedling_schema()
    action = (form.get('seedling_action') or '').strip()
    if not action:
        return False, 'Не указано действие', 'warning'
    doc_date = _parse_doc_date(form.get('doc_date'))

    try:
        if action == 'add_container':
            name = normalize_container_name(form.get('container_name'))
            if not name:
                return False, 'Укажите имя контейнера', 'warning'
            existing = SeedlingContainer.query.filter_by(name=name).first()
            if existing:
                existing.is_active = True
                existing.sort_order = int(form.get('sort_order') or existing.sort_order or 0)
                db.session.commit()
                return True, f'Контейнер «{name}» активирован', 'success'
            db.session.add(SeedlingContainer(
                name=name,
                sort_order=int(form.get('sort_order') or 100),
                is_active=True,
            ))
            log_seedling_event(
                action='container_add',
                message=f'Добавлен контейнер «{name}»',
                user_id=user_id,
                project_id=project.id,
            )
            db.session.commit()
            return True, f'Контейнер «{name}» добавлен', 'success'

        if action == 'deactivate_container':
            row = SeedlingContainer.query.get(int(form.get('container_id') or 0))
            if not row:
                return False, 'Контейнер не найден', 'warning'
            row.is_active = False
            log_seedling_event(
                action='container_off',
                message=f'Отключён контейнер «{row.name}»',
                user_id=user_id,
                project_id=project.id,
            )
            db.session.commit()
            return True, f'Контейнер «{row.name}» отключён', 'success'

        if action == 'transplant':
            raw_rows = _collect_parallel_rows(form, [
                'plant_id', 'source_size_id', 'container',
                'field_from_id', 'field_to_id', 'field_id', 'year', 'quantity',
            ])
            lines = []
            for r in raw_rows:
                q = int(r.get('quantity') or 0)
                if q <= 0:
                    continue
                field_from = int(r.get('field_from_id') or r.get('field_id') or 0)
                field_to = int(r.get('field_to_id') or field_from)
                lines.append({
                    'plant_id': int(r['plant_id']),
                    'source_size_id': int(r['source_size_id']),
                    'container': r['container'],
                    'field_from_id': field_from,
                    'field_to_id': field_to,
                    'year': int(r['year']),
                    'quantity': q,
                })
            doc = transplant_batch(
                lines=lines,
                user_id=user_id,
                project_id=project.id,
                doc_date=doc_date,
            )
            return True, f'Пересадка проведена: {len(lines)} поз. (док. #{doc.id})', 'success'

        if action == 'split_commercial':
            raw_rows = _collect_parallel_rows(form, [
                'plant_id', 'container', 'field_id', 'year', 'quantity', 'to_commercial',
            ])
            lines = []
            for r in raw_rows:
                q = int(r.get('quantity') or 0)
                if q <= 0:
                    continue
                lines.append({
                    'plant_id': int(r['plant_id']),
                    'container': r['container'],
                    'field_id': int(r['field_id']),
                    'year': int(r['year']),
                    'quantity': q,
                    'to_commercial': str(r.get('to_commercial') or '1') == '1',
                })
            doc = split_commercial_batch(
                lines=lines,
                user_id=user_id,
                project_id=project.id,
                doc_date=doc_date,
            )
            return True, f'Товарность обновлена: {len(lines)} поз. (док. #{doc.id})', 'success'

        if action == 'measure':
            raw_rows = _collect_parallel_rows(form, [
                'plant_id', 'container', 'field_id', 'year', 'measure_grades',
            ])
            lines = []
            for r in raw_rows:
                if not r.get('plant_id'):
                    continue
                grades = parse_grades_blob(r.get('measure_grades') or '')
                if not grades:
                    continue
                lines.append({
                    'plant_id': int(r['plant_id']),
                    'container': r['container'],
                    'field_id': int(r['field_id']),
                    'year': int(r['year']),
                    'grade_qtys': grades,
                })
            if not lines:
                names = form.getlist('grade_name[]')
                qtys = form.getlist('grade_qty[]')
                grade_qtys = {}
                for i, name in enumerate(names):
                    try:
                        q = int((qtys[i] if i < len(qtys) else '0') or 0)
                    except (TypeError, ValueError):
                        q = 0
                    g = normalize_grade(name)
                    if g and q > 0:
                        grade_qtys[g] = grade_qtys.get(g, 0) + q
                if form.get('plant_id') and grade_qtys:
                    lines = [{
                        'plant_id': int(form.get('plant_id')),
                        'container': form.get('container'),
                        'field_id': int(form.get('field_id')),
                        'year': int(form.get('year')),
                        'grade_qtys': grade_qtys,
                    }]
            doc = apply_measure_batch(
                lines=lines,
                user_id=user_id,
                project_id=project.id,
                doc_date=doc_date,
            )
            return True, f'Промер проведён: {len(lines)} поз. (док. #{doc.id})', 'success'

        if action == 'dieback':
            raw_rows = _collect_parallel_rows(form, [
                'plant_id', 'size_id', 'field_id', 'year', 'quantity',
            ])
            lines = []
            for r in raw_rows:
                try:
                    q = int(r.get('quantity') or 0)
                    pid = int(r.get('plant_id') or 0)
                    sid = int(r.get('size_id') or 0)
                    fid = int(r.get('field_id') or 0)
                    yr = int(r.get('year') or 0)
                except (TypeError, ValueError):
                    continue
                if q <= 0 or not pid or not sid or not fid or not yr:
                    continue
                lines.append({
                    'plant_id': pid,
                    'size_id': sid,
                    'field_id': fid,
                    'year': yr,
                    'quantity': q,
                })
            doc = dieback_batch(
                lines=lines,
                user_id=user_id,
                project_id=project.id,
                doc_date=doc_date,
            )
            return True, f'Выпад проведён: {len(lines)} поз. (док. #{doc.id})', 'success'

        return False, f'Неизвестное действие: {action}', 'warning'
    except ValueError as exc:
        db.session.rollback()
        return False, str(exc), 'warning'
    except Exception as exc:
        db.session.rollback()
        return False, f'Ошибка: {exc}', 'danger'
