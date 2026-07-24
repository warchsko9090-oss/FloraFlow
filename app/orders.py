import json
import io
import os
import requests
import time
from decimal import Decimal
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app
from werkzeug.utils import secure_filename
from flask_login import login_required, current_user
from sqlalchemy import func, or_, and_
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from app.models import db, Order, OrderItem, Client, Plant, Size, Field, Payment, StockBalance, Document, DocumentRow, OrderItemHistory, DiggingLog, ActionLog, User, TgTask
from app.utils import msk_now, check_stock_availability, get_actual_price, log_action, natural_key, dateru
from app.shop_prices import order_default_price, resolve_shop_price, get_shop_price_map
from app.models import Project

bp = Blueprint('orders', __name__)

from app.telegram import send_message as _tg_send_msg, send_photo as _tg_send_photo, send_photo_album as _tg_send_photo_album

def send_tg_message_orders(text):
    _tg_send_msg(text, chat_type="orders")

def send_tg_photo_orders(photo_path, caption=""):
    _tg_send_photo(photo_path, caption, chat_type="orders")

def send_tg_photo_album_orders(photo_paths):
    _tg_send_photo_album(photo_paths, chat_type="orders")
# ---------------------------------


def _build_order_item_snapshot(item):
    return {
        'plant_id': item.plant_id,
        'size_id': item.size_id,
        'field_id': item.field_id,
        'year': item.year,
        'price': float(item.price or 0),
        'plant_name': item.plant.name if item.plant else '',
        'size_name': item.size.name if item.size else '',
        'field_name': item.field.name if item.field else '',
    }


def _record_order_item_history(order_id, item, action_type, before_qty, after_qty):
    before_val = int(before_qty) if before_qty is not None else None
    after_val = int(after_qty) if after_qty is not None else None
    delta = 0
    if before_val is not None and after_val is not None:
        delta = after_val - before_val
    history = OrderItemHistory(
        order_id=order_id,
        order_item_id=item.id if item else None,
        action_type=action_type,
        before_quantity=before_val,
        after_quantity=after_val,
        delta_quantity=delta,
        snapshot_payload=json.dumps(_build_order_item_snapshot(item), ensure_ascii=False) if item else None,
        changed_by_user_id=current_user.id if current_user.is_authenticated else None,
        created_at=msk_now()
    )
    db.session.add(history)


def _order_paid_total(order_id):
    """Сумма всех оплат по заказу (Decimal). 0, если оплат нет."""
    total = (db.session.query(func.coalesce(func.sum(Payment.amount), 0))
             .filter(Payment.order_id == order_id)
             .scalar())
    try:
        return Decimal(total or 0)
    except Exception:
        return Decimal(0)


def _is_order_locked_for_manager(order, user):
    """True, если для роли 'user' (менеджер по продажам) заказ нужно
    защитить от удаления и уменьшения позиций (есть хотя бы одна оплата).

    Админ и executive — без ограничений.
    """
    if order is None or user is None:
        return False
    if not getattr(user, 'is_authenticated', False):
        return False
    role = getattr(user, 'role', None)
    if role not in ('user', 'shop_manager'):
        return False
    return _order_paid_total(order.id) > 0


_AUDIT_FIELD_LABELS = {
    'qty': 'кол-во',
    'price': 'цена',
    'field_id': 'поле',
    'year': 'год партии',
    'size_id': 'размер',
}


def _audit_paid_order_change(order, action, summary, item_id=None):
    """Создаёт/обновляет TgTask с аудитом изменений менеджером в оплаченном заказе.

    action: 'save_items' | 'add_item' | 'split_item'
    summary: dict — payload с описанием изменений (см. вызовы).
    item_id: id затронутой позиции (опционально).

    Карточки группируются по дню (dedup_key per (order, date)) — внутри одной
    карточки накапливается список действий менеджера, чтобы не плодить
    дубликаты на каждое сохранение.
    """
    try:
        if order is None or current_user is None or not getattr(current_user, 'is_authenticated', False):
            return
        if getattr(current_user, 'role', None) != 'user':
            return
        if not _is_order_locked_for_manager(order, current_user):
            return

        today = msk_now().date()
        dedup_key = f'paid_order_change:{order.id}:{today.isoformat()}'

        existing = (TgTask.query
                    .filter_by(dedup_key=dedup_key)
                    .order_by(TgTask.id.desc())
                    .first())

        entry_text = _format_audit_entry(action, summary)
        if not entry_text:
            return

        username = getattr(current_user, 'username', None) or 'manager'
        ts = msk_now().strftime('%H:%M')
        new_line = f'<div class="small mb-1"><span class="text-muted">[{ts}]</span> <strong>{username}</strong>: {entry_text}</div>'

        if existing and existing.status == 'new':
            existing.details = (existing.details or '') + new_line
            existing.last_seen_at = msk_now()
            try:
                payload = json.loads(existing.action_payload or '{}')
            except Exception:
                payload = {}
            payload.setdefault('changes', []).append({
                'action': action,
                'item_id': item_id,
                'ts': msk_now().isoformat(),
                'by': username,
                'summary': summary,
            })
            existing.action_payload = json.dumps(payload, ensure_ascii=False, default=str)
            return

        title = f'Менеджер изменил оплаченный заказ #{order.id}'
        client_name = ''
        try:
            client_name = order.client.name if order.client else ''
        except Exception:
            client_name = ''
        details_header = (
            f'<div class="mb-2"><strong>Заказ #{order.id}</strong>'
            + (f' — {client_name}' if client_name else '')
            + '</div>'
        )
        details = details_header + new_line

        payload = {
            'order_id': order.id,
            'changes': [{
                'action': action,
                'item_id': item_id,
                'ts': msk_now().isoformat(),
                'by': username,
                'summary': summary,
            }],
        }

        task = TgTask(
            raw_text=f'[paid_order_change:{order.id}]',
            title=title,
            details=details,
            action_type='paid_order_change',
            action_payload=json.dumps(payload, ensure_ascii=False, default=str),
            status='new',
            assignee_role='admin,executive',
            sender_name='system.audit',
            source='audit',
            severity='warning',
            dedup_key=dedup_key,
            first_seen_at=msk_now(),
            last_seen_at=msk_now(),
        )
        db.session.add(task)
    except Exception as exc:
        # Аудит не должен ломать сохранение заказа.
        try:
            current_app.logger.warning('audit_paid_order_change failed: %s', exc)
        except Exception:
            pass


def _format_audit_entry(action, summary):
    """Превращает summary в читаемую html-строку. summary зависит от action."""
    if not summary:
        return ''

    def _name(model, obj_id):
        if not obj_id:
            return ''
        try:
            obj = model.query.get(int(obj_id))
            return obj.name if obj else f'#{obj_id}'
        except Exception:
            return f'#{obj_id}'

    if action == 'save_items':
        # summary = {'item_id', 'plant_name', 'size_name', 'changes': [{'field':'qty','old','new'}, ...]}
        plant = summary.get('plant_name') or ''
        size = summary.get('size_name') or ''
        item_id = summary.get('item_id')
        head = f'позиция #{item_id} — {plant} {size}'.strip()
        parts = []
        for ch in (summary.get('changes') or []):
            field = ch.get('field')
            label = _AUDIT_FIELD_LABELS.get(field, field)
            old = ch.get('old')
            new = ch.get('new')
            if field == 'field_id':
                old = _name(Field, old) or '—'
                new = _name(Field, new) or '—'
            elif field == 'size_id':
                old = _name(Size, old) or '—'
                new = _name(Size, new) or '—'
            parts.append(f'{label}: <code>{old}</code> → <code>{new}</code>')
        if not parts:
            return ''
        return f'изменил {head}: ' + '; '.join(parts)

    if action == 'add_item':
        plant = summary.get('plant_name') or ''
        size = summary.get('size_name') or ''
        field = summary.get('field_name') or ''
        year = summary.get('year') or ''
        qty = summary.get('quantity') or 0
        price = summary.get('price') or 0
        return (f'добавил позицию: {plant} {size}, поле {field}, год {year}, '
                f'<code>{qty}</code> шт по <code>{price}</code> ₽')

    if action == 'split_item':
        src = summary.get('source') or {}
        parts = summary.get('parts') or []
        head = (f'разделил позицию #{summary.get("item_id")} '
                f'({src.get("plant_name","")} {src.get("size_name","")}, '
                f'поле {src.get("field_name","")}, год {src.get("year","")}, '
                f'было <code>{src.get("quantity",0)}</code> шт)')
        bullets = []
        for p in parts:
            bullets.append(
                f'• {p.get("size_name","")}, поле {p.get("field_name","")}, год {p.get("year","")}: '
                f'<code>{p.get("quantity",0)}</code> шт × <code>{p.get("price",0)}</code> ₽'
            )
        return head + ('<br>' + '<br>'.join(bullets) if bullets else '')

    return ''


def _parse_split_parts_from_form(form):
    """Достаёт массив 'частей' из формы split_item.
    Поля: split_size_id[], split_field_id[], split_year[], split_quantity[], split_price[].

    Возвращает [{'size_id', 'field_id', 'year', 'quantity', 'price'}, ...]
    Невалидные строки (нет qty, qty<=0, не парсится число) — отбрасывает.
    Дубли по (size_id, field_id, year) — суммируются по quantity, цена берётся максимальная.
    """
    size_ids = form.getlist('split_size_id[]')
    field_ids = form.getlist('split_field_id[]')
    years = form.getlist('split_year[]')
    qtys = form.getlist('split_quantity[]')
    prices = form.getlist('split_price[]')

    n = max(len(size_ids), len(field_ids), len(years), len(qtys), len(prices))
    result = {}
    for i in range(n):
        try:
            qty = int((qtys[i] if i < len(qtys) else '0') or 0)
            if qty <= 0:
                continue
            sid = int(size_ids[i])
            fid = int(field_ids[i])
            yr = int(years[i])
            try:
                price = Decimal(str(prices[i] or '0'))
            except Exception:
                price = Decimal('0')
        except (ValueError, TypeError, IndexError):
            continue
        key = (sid, fid, yr)
        if key in result:
            result[key]['quantity'] += qty
            if price > result[key]['price']:
                result[key]['price'] = price
        else:
            result[key] = {
                'size_id': sid,
                'field_id': fid,
                'year': yr,
                'quantity': qty,
                'price': price,
            }
    return list(result.values())


def _available_batches_for_product(plant_id, size_id):
    stocks = StockBalance.query.filter_by(plant_id=plant_id, size_id=size_id).all()
    if not stocks:
        return []
    # Single aggregated query for all reserved quantities for this plant+size
    reserved_rows = (
        db.session.query(
            OrderItem.field_id, OrderItem.year,
            func.sum(OrderItem.quantity - OrderItem.shipped_quantity)
        )
        .join(Order)
        .filter(
            OrderItem.plant_id == plant_id,
            OrderItem.size_id == size_id,
            Order.status != 'canceled',
            Order.status != 'ghost',
            Order.is_deleted == False
        )
        .group_by(OrderItem.field_id, OrderItem.year)
        .all()
    )
    reserved_map = {(r[0], r[1]): int(r[2] or 0) for r in reserved_rows}
    
    result = []
    from app.shop_prices import resolve_shop_price, get_shop_price_map
    price_overrides = get_shop_price_map()
    for st in stocks:
        reserved_qty = reserved_map.get((st.field_id, st.year), 0)
        free_qty = int(st.quantity or 0) - reserved_qty
        if free_qty > 0:
            wholesale = float(st.price or 0)
            retail = resolve_shop_price(st.plant_id, st.size_id, wholesale, price_overrides)
            result.append({
                'field_id': st.field_id,
                'field_name': st.field.name if st.field else '-',
                'year': st.year,
                'free': free_qty,
                'price': retail,
            })
    result.sort(key=lambda x: x['free'], reverse=True)
    return result


def _get_client_catalog_rows():
    """Каталог остатков для клиента (как на /shop)."""
    reserved_rows = (
        db.session.query(
            OrderItem.plant_id,
            OrderItem.size_id,
            OrderItem.field_id,
            OrderItem.year,
            func.sum(OrderItem.quantity - OrderItem.shipped_quantity),
        )
        .join(Order)
        .filter(
            Order.status != "canceled",
            Order.status != "ghost",
            Order.is_deleted == False,
        )
        .group_by(OrderItem.plant_id, OrderItem.size_id, OrderItem.field_id, OrderItem.year)
        .all()
    )
    reserved_map = {(r[0], r[1], r[2], r[3]): int(r[4] or 0) for r in reserved_rows}

    grouped = {}
    from app.seedlings import is_excluded_from_product_stock
    for st in StockBalance.query.all():
        size_name = st.size.name if st.size else ""
        if is_excluded_from_product_stock(size_name):
            continue

        free = int(st.quantity or 0) - reserved_map.get((st.plant_id, st.size_id, st.field_id, st.year), 0)
        if free <= 0:
            continue

        key = (st.plant_id, st.size_id)
        if key not in grouped:
            grouped[key] = {
                "plant_name": st.plant.name if st.plant else "-",
                "size_name": st.size.name if st.size else "-",
                "free_qty": 0,
                "price": None,
            }

        grouped[key]["free_qty"] += free
        p = float(st.price or 0)
        grouped[key]["price"] = p if grouped[key]["price"] is None else min(grouped[key]["price"], p)

    items = list(grouped.values())
    items.sort(key=lambda x: (natural_key(x["plant_name"]), natural_key(x["size_name"])))
    return items


def _parse_client_draft_comment(comment):
    from app.client_draft_approval import parse_client_draft_meta
    return parse_client_draft_meta(comment)


def _site_request_badges():
    """Счётчики заявок с сайта для бейджей в разделах заказов."""
    from app.models import ShopOnRequest

    cutoff = msk_now() - timedelta(hours=24)

    on_request_new = ShopOnRequest.query.filter(
        ShopOnRequest.status == ShopOnRequest.STATUS_NEW
    ).count()
    on_request_overdue = ShopOnRequest.query.filter(
        ShopOnRequest.status == ShopOnRequest.STATUS_NEW,
        ShopOnRequest.created_at <= cutoff,
    ).count()

    drafts_new = Document.query.filter(
        Document.doc_type == 'client_draft'
    ).count()
    drafts_overdue = Document.query.filter(
        Document.doc_type == 'client_draft',
        Document.date <= cutoff,
    ).count()

    return {
        'on_request': {
            'new': on_request_new,
            'overdue': on_request_overdue,
        },
        'drafts': {
            'new': drafts_new,
            'overdue': drafts_overdue,
        },
    }


@bp.route('/orders', methods=['GET', 'POST'])
@login_required
def orders_list():
    from app.digging import ensure_payment_file_column
    ensure_payment_file_column()
    mode = request.args.get('mode', 'active')
    f_client = request.args.get('filter_client')
    f_status = request.args.get('filter_status')
    f_date_start = request.args.get('start_date')
    f_date_end = request.args.get('end_date')
    # Multi-select по номерам заказов: пользователь может отметить один или
    # несколько номеров в выпадающем фильтре, и в списке останутся только
    # они. Кнопка «Excel» учитывает этот фильтр и выгружает ровно то,
    # что сейчас на странице.
    f_ids_raw = request.args.getlist('filter_ids')
    f_ids = []
    for v in f_ids_raw:
        v = (v or '').strip()
        if v.isdigit():
            f_ids.append(int(v))
    sort = request.args.get('sort', 'name')  # По умолчанию сортировка по названию (имени клиента)
    
    if request.method == 'POST' and current_user.role == 'admin':
        oid = request.form.get('order_id')
        act = request.form.get('action')
        o = Order.query.get(oid)
        if o:
            if act == 'delete':
                o.is_deleted = True
                flash('Заказ скрыт')
            elif act == 'restore':
                o.is_deleted = False
                flash('Заказ восстановлен')
            elif act == 'purge':
                # Полное удаление из БД (только для админа)
                db.session.delete(o)
                flash('Заказ удален навсегда')
            db.session.commit()
            log_action(f"Спрятал/показал/удалил заказ {o.id}")
        return redirect(url_for('orders.orders_list', mode=mode))
        
    # Автоматически скрываем старые отмененные заказы (через ~4 месяца)
    cutoff = msk_now() - timedelta(days=120)
    Order.query.filter(
        Order.status == 'canceled',
        Order.is_deleted == False,
        or_(
            and_(Order.canceled_at != None, Order.canceled_at < cutoff),
            and_(Order.canceled_at == None, Order.date < cutoff)
        )
    ).update({Order.is_deleted: True}, synchronize_session=False)
    db.session.commit()

    from sqlalchemy.orm import joinedload

    # флаг скрытых (old 'trash')
    show_hidden = mode in ('hidden', 'trash')
    q = Order.query.options(
        joinedload(Order.client),
        joinedload(Order.items).joinedload(OrderItem.plant),
        joinedload(Order.items).joinedload(OrderItem.size),
    ).filter_by(is_deleted=show_hidden)
        
    if f_client: q = q.filter(Order.client_id == int(f_client))
    if f_status: q = q.filter(Order.status == f_status)
    if not f_status: q = q.filter(Order.status != 'ghost')
    if f_date_start: q = q.filter(func.date(Order.date) >= f_date_start)
    if f_date_end: q = q.filter(func.date(Order.date) <= f_date_end)
    if f_ids:
        q = q.filter(Order.id.in_(f_ids))

    # Применяем сортировку
    if sort == 'date':
        orders = q.order_by((Order.status == 'canceled'), Order.date.desc()).all()
    else:  # sort == 'name' (по умолчанию)
        orders = q.order_by((Order.status == 'canceled')).all()
        # Сортируем по имени клиента в памяти, чтобы использовать natural_key
        orders.sort(key=lambda x: natural_key(x.client.name))
    
    total_stats = {'sum': 0, 'shipped_fact': 0}
    orders_view = []
    for o in orders:
        if not o.items:
            orders_view.append({'order': o, 'rowspan': 1, 'order_items': [], 'is_done': False})
            continue
        if o.status != 'canceled' and not o.is_deleted:
            total_stats['sum'] += o.total_sum
            total_stats['shipped_fact'] += sum(i.shipped_quantity for i in o.items)

        o.items.sort(key=lambda x: natural_key(x.size.name))

        # Полностью завершён = оплачен полностью + всё выкопано + всё отгружено.
        # Отменённые и призрачные не считаем "завершёнными".
        is_done = False
        if o.status not in ('canceled', 'ghost') and not o.is_deleted and o.items:
            fully_paid = (o.payment_status == 'paid')
            fully_shipped = all((i.shipped_quantity or 0) >= (i.quantity or 0) for i in o.items)
            fully_dug = all((i.dug_total or 0) >= (i.quantity or 0) for i in o.items)
            is_done = fully_paid and fully_shipped and fully_dug

        orders_view.append({'order': o, 'rowspan': len(o.items), 'order_items': o.items, 'is_done': is_done})

    # Финальный порядок: сначала активные незавершённые (в выбранном порядке),
    # затем полностью завершённые зелёные, затем отменённые/скрытые.
    # Внутри каждой группы исходный порядок сохраняется (stable partition).
    def _bucket(v):
        o = v['order']
        if o.status == 'canceled' or o.is_deleted:
            return 2
        if v['is_done']:
            return 1
        return 0
    orders_view.sort(key=_bucket)

    # Список всех заказов для multiselect-фильтра «По номеру заказа».
    # Берём из той же видимой группы (активные/скрытые), исключаем 'ghost',
    # сортируем по убыванию id, чтобы свежие номера были сверху списка.
    # Имя клиента подтягиваем заранее, чтобы шаблон не делал N+1 запросов.
    filter_orders_q = (Order.query
                       .filter(Order.is_deleted == show_hidden,
                               Order.status != 'ghost')
                       .order_by(Order.id.desc()))
    all_orders_for_filter = [
        {
            'id': o.id,
            'client_name': (o.client.name if o.client else '—'),
        }
        for o in filter_orders_q.all()
    ]

    return render_template('orders/orders.html',
                           orders=orders_view,
                           stats=total_stats,
                           site_badges=_site_request_badges(),
                           show_hidden=show_hidden,
                           all_clients=Client.query.all(),
                           all_orders_for_filter=all_orders_for_filter,
                           filters={
                               'clients': [int(f_client)] if f_client else [],
                               'statuses': [f_status] if f_status else [],
                               'start': f_date_start,
                               'end': f_date_end,
                               'ids': f_ids,
                           },
                           sort=sort,
                           mode=mode)


@bp.route('/orders/client_drafts')
@login_required
def client_drafts_list():
    if current_user.role == 'user2':
        flash('Доступ запрещен')
        return redirect(url_for('orders.orders_list'))

    mode = request.args.get('mode', 'new')
    q = Document.query.filter(Document.doc_type.in_(['client_draft', 'client_draft_approved', 'client_draft_rejected']))
    if mode == 'new':
        q = q.filter(Document.doc_type == 'client_draft')
    elif mode == 'approved':
        q = q.filter(Document.doc_type == 'client_draft_approved')
    elif mode == 'rejected':
        q = q.filter(Document.doc_type == 'client_draft_rejected')

    docs = q.order_by(Document.date.desc(), Document.id.desc()).all()
    from app.client_draft_approval import approval_status
    view = []
    for d in docs:
        meta = _parse_client_draft_comment(d.comment)
        appr = approval_status(meta)
        view.append({
            'doc': d,
            'meta': meta,
            'approval': appr,
            'items_count': len(d.rows or []),
            'items_qty': sum((r.quantity or 0) for r in (d.rows or []))
        })
    return render_template(
        'orders/client_drafts.html',
        drafts=view,
        mode=mode,
        site_badges=_site_request_badges(),
    )


@bp.route('/orders/client_draft/<int:doc_id>', methods=['GET', 'POST'])
@login_required
def client_draft_detail(doc_id):
    if current_user.role == 'user2':
        flash('Доступ запрещен')
        return redirect(url_for('orders.orders_list'))

    from app.client_draft_approval import (
        approval_status,
        build_allocation_rows_view,
        can_approve_as_shop_manager,
        can_approve_as_user,
        can_reject_draft,
        deserialize_allocations,
        finalize_client_draft,
        load_draft_payload,
        parse_allocations_from_form,
        resolve_or_create_client,
        save_draft_payload,
        stamp_shop_manager_approval,
        stamp_user_approval,
        validate_allocations,
    )

    doc = Document.query.get_or_404(doc_id)
    if doc.doc_type not in ('client_draft', 'client_draft_approved', 'client_draft_rejected'):
        flash('Это не клиентский черновик')
        return redirect(url_for('orders.client_drafts_list'))

    meta = _parse_client_draft_comment(doc.comment)
    appr = approval_status(meta)
    role = current_user.role

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'reject':
            if doc.doc_type != 'client_draft':
                flash('Черновик уже обработан')
                return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))
            if not can_reject_draft(role, doc.doc_type):
                flash('Отклонить заявку могут менеджер питомника, менеджер сайта или администратор')
                return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))
            doc.doc_type = 'client_draft_rejected'
            db.session.commit()
            flash('Заявка отклонена')
            return redirect(url_for('orders.client_drafts_list', mode='rejected'))

        if action == 'approve_user':
            if doc.doc_type != 'client_draft':
                flash('Заявка уже обработана')
                return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))
            if not can_approve_as_user(role, doc.doc_type):
                flash('Распределение и согласование менеджера питомника доступны только менеджеру питомника или администратору')
                return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))

            client = resolve_or_create_client(
                request.form.get('client_id', type=int),
                request.form.get('new_client_name'),
                meta,
            )
            if not client:
                flash('Выберите или укажите клиента')
                return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))

            allocations = parse_allocations_from_form(request.form)
            ok, err = validate_allocations(doc, allocations)
            if not ok:
                flash(err)
                return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))

            payload = load_draft_payload(doc.comment)
            stamp_user_approval(payload, current_user, client, allocations)
            save_draft_payload(doc, payload)

            try:
                if approval_status(payload)['complete']:
                    order = finalize_client_draft(doc, payload, current_user.id)
                    db.session.commit()
                    flash(f'Заявка полностью согласована. Создан заказ #{order.id} (резерв)')
                    return redirect(url_for('orders.order_detail', order_id=order.id))
                db.session.commit()
                flash('Согласование менеджера питомника сохранено. Ожидается согласование менеджера сайта.')
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка: {e}')
            return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))

        if action == 'approve_shop_manager':
            if doc.doc_type != 'client_draft':
                flash('Заявка уже обработана')
                return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))
            if not can_approve_as_shop_manager(role, doc.doc_type, appr['slots']['shop_manager']['done']):
                flash('Согласование менеджера сайта доступно менеджеру сайта или администратору')
                return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))

            payload = load_draft_payload(doc.comment)
            stamp_shop_manager_approval(payload, current_user)
            save_draft_payload(doc, payload)

            try:
                if approval_status(payload)['complete']:
                    order = finalize_client_draft(doc, payload, current_user.id)
                    db.session.commit()
                    flash(f'Заявка полностью согласована. Создан заказ #{order.id} (резерв)')
                    return redirect(url_for('orders.order_detail', order_id=order.id))
                db.session.commit()
                flash('Согласование менеджера сайта сохранено. Ожидается согласование менеджера питомника.')
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка: {e}')
            return redirect(url_for('orders.client_draft_detail', doc_id=doc.id))

    rows_view = []
    line_prices = {}
    for line in meta.get('lines') or []:
        try:
            line_prices[(int(line['plant_id']), int(line['size_id']))] = float(line['price'])
        except (KeyError, TypeError, ValueError):
            continue

    stored_allocations = deserialize_allocations(
        (meta.get('approvals') or {}).get('user', {}).get('allocations')
    )
    stored_client_id = (meta.get('approvals') or {}).get('user', {}).get('client_id')

    for r in doc.rows:
        row_price = line_prices.get((r.plant_id, r.size_id))
        row_stored = stored_allocations.get(r.id) or []
        rows_view.append({
            'row': r,
            'price': row_price,
            'line_sum': (row_price * r.quantity) if row_price is not None else None,
            'batches': _available_batches_for_product(r.plant_id, r.size_id),
            'stored_allocations': row_stored,
        })

    can_approve_user = can_approve_as_user(role, doc.doc_type)
    can_approve_shop = can_approve_as_shop_manager(
        role, doc.doc_type, appr['slots']['shop_manager']['done']
    )
    can_reject = can_reject_draft(role, doc.doc_type)
    allocation_rows = build_allocation_rows_view(doc, meta)

    return render_template(
        'orders/client_draft_detail.html',
        doc=doc,
        meta=meta,
        approval=appr,
        rows_view=rows_view,
        allocation_rows=allocation_rows,
        clients=sorted(Client.query.all(), key=lambda x: natural_key(x.name)),
        can_approve_user=can_approve_user,
        can_approve_shop=can_approve_shop,
        can_reject=can_reject,
        stored_client_id=stored_client_id,
    )


@bp.route('/orders/shop_on_requests')
@login_required
def shop_on_requests_list():
    if current_user.role == 'user2':
        flash('Доступ запрещен')
        return redirect(url_for('orders.orders_list'))

    from app.models import ShopOnRequest
    from sqlalchemy.orm import joinedload

    mode = request.args.get('mode', 'new')
    q = ShopOnRequest.query.options(
        joinedload(ShopOnRequest.plant),
        joinedload(ShopOnRequest.size),
        joinedload(ShopOnRequest.reviewed_by),
    )
    if mode == 'new':
        q = q.filter(ShopOnRequest.status == ShopOnRequest.STATUS_NEW)
    elif mode == 'approved':
        q = q.filter(ShopOnRequest.status == ShopOnRequest.STATUS_APPROVED)
    elif mode == 'rejected':
        q = q.filter(ShopOnRequest.status == ShopOnRequest.STATUS_REJECTED)

    rows = q.order_by(ShopOnRequest.created_at.desc(), ShopOnRequest.id.desc()).all()
    return render_template(
        'orders/shop_on_requests.html',
        requests=rows,
        mode=mode,
        site_badges=_site_request_badges(),
    )


@bp.route('/orders/shop_on_request/<int:req_id>', methods=['GET', 'POST'])
@login_required
def shop_on_request_detail(req_id):
    if current_user.role == 'user2':
        flash('Доступ запрещен')
        return redirect(url_for('orders.orders_list'))

    from app.models import ShopOnRequest, ShopOnRequestLog
    from app.shop_on_request import review_on_request
    from sqlalchemy.orm import joinedload

    row = ShopOnRequest.query.options(
        joinedload(ShopOnRequest.plant),
        joinedload(ShopOnRequest.size),
        joinedload(ShopOnRequest.reviewed_by),
        joinedload(ShopOnRequest.logs).joinedload(ShopOnRequestLog.user),
    ).get_or_404(req_id)

    if request.method == 'POST':
        action = request.form.get('action')
        comment = request.form.get('manager_comment', '')
        try:
            review_on_request(row.id, action, comment)
            db.session.commit()
            log_action(
                f'Заявка «По запросу» #{row.id}: {action} '
                f'({row.plant.name if row.plant else "?"}, {row.size.name if row.size else "?"})'
            )
            if action == 'approve':
                flash('Заявка согласована')
                return redirect(url_for('orders.shop_on_requests_list', mode='approved'))
            flash('Заявка отклонена')
            return redirect(url_for('orders.shop_on_requests_list', mode='rejected'))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc))
        except Exception as exc:
            db.session.rollback()
            flash(f'Ошибка: {exc}')

    return render_template('orders/shop_on_request_detail.html', req=row)


@bp.route('/order/create', methods=['GET', 'POST'])
@login_required
def order_create():
    if request.method == 'POST':
        c_id = request.form.get('client')
        is_barter = request.form.get('is_barter') == 'on'
        o = Order(
            client_id=c_id,
            date=msk_now(),
            is_barter=is_barter,
            created_by_user_id=current_user.id,
        )
        db.session.add(o)
        # Пока не коммитим, чтобы если будет ошибка товара, можно было откатить
        
        p_ids = request.form.getlist('plant[]')
        s_ids = request.form.getlist('size[]')
        f_ids = request.form.getlist('field[]')
        y_ids = request.form.getlist('year[]')
        q_ids = request.form.getlist('quantity[]')
        
        created_items = []
        deficit_notes = []
        for i in range(len(p_ids)):
            if int(q_ids[i]) > 0:
                ok, free = check_stock_availability(p_ids[i], s_ids[i], f_ids[i], int(y_ids[i]), int(q_ids[i]))
                if not ok: 
                    # ИСПРАВЛЕНО: Откат транзакции перед редиректом
                    db.session.rollback()
                    flash(f"Ошибка! Недостаточно товара (доступно: {free}). Заказ не создан.")
                    return redirect(url_for('orders.order_create'))

                need = int(q_ids[i])
                if free < need:
                    from app.seedlings import suggest_deficit_sources, allows_order_deficit
                    from app.models import Size as _Size, Plant as _Plant
                    sz = _Size.query.get(int(s_ids[i]))
                    pl = _Plant.query.get(int(p_ids[i]))
                    if sz and allows_order_deficit(sz.name):
                        short = need - max(free, 0)
                        srcs = suggest_deficit_sources(
                            int(p_ids[i]), int(s_ids[i]), int(f_ids[i]), int(y_ids[i]), short,
                        )
                        hint = '; '.join(
                            f"{s['size_name']} / {s['field_name']} / {s['year']}: {s['free']} шт"
                            for s in srcs[:5]
                        ) or 'других промеров с остатком нет'
                        deficit_notes.append(
                            f"{(pl.name if pl else '?')} · {(sz.name if sz else '?')}: "
                            f"в минус на {short} шт. Откуда взять: {hint}"
                        )
                
                wholesale = get_actual_price(int(p_ids[i]), int(s_ids[i]), int(f_ids[i]))
                p = order_default_price(int(p_ids[i]), int(s_ids[i]), wholesale)
                new_item = OrderItem(
                    order_id=o.id,
                    plant_id=int(p_ids[i]),
                    size_id=int(s_ids[i]),
                    field_id=int(f_ids[i]),
                    year=int(y_ids[i]),
                    quantity=int(q_ids[i]),
                    price=p
                )
                db.session.add(new_item)
                created_items.append(new_item)

        db.session.flush()
        for created_item in created_items:
            _record_order_item_history(
                order_id=o.id,
                item=created_item,
                action_type='initial_item',
                before_qty=0,
                after_qty=created_item.quantity
            )

        # Если позиции с контейнерной площадки проекта — привязываем заказ к выручке проекта
        if not o.project_id and created_items:
            from app.finance import resolve_project_id_for_yard_fields
            linked = resolve_project_id_for_yard_fields([it.field_id for it in created_items])
            if linked:
                o.project_id = linked

        # Защита от гонки двух менеджеров. Заказ только что создан, поэтому
        # baseline пустой — любое overcommit на партии = новое нарушение.
        try:
            from app.stock_helpers import assert_stock_for_order
            assert_stock_for_order(o.id, baseline={})
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc))
            return redirect(url_for('orders.order_create'))

        db.session.commit()
        log_action(f"Создал заказ #{o.id}")
        for note in deficit_notes:
            flash(note, 'warning')
        if deficit_notes:
            flash(
                'Заказ создан с минусом по саженцам. Скорректируйте позиции '
                'или проведите перемер — см. предупреждения выше.',
                'warning',
            )
        # Синхронная проверка цен: если есть позиции без цены, сразу уведомим
        # админа/менеджера карточкой в дашборде. Не ломает основной поток при ошибке.
        try:
            from app.anomaly_engine import sync_price_anomaly_for_order
            sync_price_anomaly_for_order(o.id)
        except Exception:
            pass
        # Год партии vs фактический остаток — раннее предупреждение.
        try:
            from app.anomaly_engine import sync_year_mismatch_for_order
            sync_year_mismatch_for_order(o.id)
        except Exception:
            pass
        return redirect(url_for('orders.orders_list'))

    return render_template('orders/order_form.html', clients=Client.query.all(), plants=Plant.query.all(), sizes=Size.query.all(), fields=Field.query.all())

@bp.route('/api/stock_info')
@login_required
def api_stock_info():
    p, s, f = request.args.get('plant'), request.args.get('size'), request.args.get('field')
    # Добавил проверку на наличие параметров
    if not p or not s or not f:
        return jsonify({'batches': []})

    # exclude_item_id — исключить указанный OrderItem из учёта резервов
    # (нужно для split: исходная позиция сама держит резерв на партии).
    # include_item_id — гарантированно вернуть год партии этой позиции
    # даже если free_qty == 0 (граничный случай).
    try:
        exclude_item_id = int(request.args.get('exclude_item_id') or 0) or None
    except (TypeError, ValueError):
        exclude_item_id = None
    try:
        include_item_id = int(request.args.get('include_item_id') or 0) or None
    except (TypeError, ValueError):
        include_item_id = None

    stocks = StockBalance.query.filter_by(plant_id=p, size_id=s, field_id=f).order_by(StockBalance.year.asc()).all()
    # Используем единые правила «активного заказа» (см. app.stock_helpers).
    # Раньше тут не отфильтровывались ghost-заказы и расчёт расходился с
    # check_stock_availability — могло мерцать «свободно» между API.
    from app.stock_helpers import get_reserved_map as _get_reserved_map
    reserved_by_year_full = _get_reserved_map(
        plant_id=p, size_id=s, field_id=f, exclude_item_id=exclude_item_id
    )
    reserved_map = {
        key[3]: qty for key, qty in reserved_by_year_full.items()
    }

    forced_year = None
    forced_price = None
    if include_item_id:
        forced_item = OrderItem.query.get(include_item_id)
        if forced_item:
            forced_year = forced_item.year
            try:
                forced_price = float(forced_item.price or 0)
            except Exception:
                forced_price = 0.0

    res = []
    seen_years = set()
    price_overrides = get_shop_price_map()
    for st in stocks:
        free_qty = st.quantity - reserved_map.get(st.year, 0)
        if free_qty > 0 or st.year == forced_year:
            wholesale = float(st.price or 0)
            retail = resolve_shop_price(st.plant_id, st.size_id, wholesale, price_overrides)
            res.append({
                'year': st.year,
                'free': max(free_qty, 0),
                'price': retail,
                'wholesale_price': wholesale,
                'retail_price': retail,
            })
            seen_years.add(st.year)

    if forced_year is not None and forced_year not in seen_years:
        # партии нет в StockBalance — добавим виртуальную, free=0,
        # цену возьмём из позиции
        res.append({
            'year': forced_year,
            'free': 0,
            'price': forced_price or 0.0,
            'wholesale_price': forced_price or 0.0,
            'retail_price': forced_price or 0.0,
        })
        res.sort(key=lambda x: x['year'])

    return jsonify({'batches': res})


@bp.route('/api/seedling_snapshot')
@login_required
def api_seedling_snapshot():
    from app.seedlings import commercial_snapshot, snapshot_for_container
    try:
        plant_id = int(request.args.get('plant_id') or 0)
        field_id = int(request.args.get('field_id') or 0)
        year = int(request.args.get('year') or 0)
        container = (request.args.get('container') or '').strip()
        mode = (request.args.get('mode') or 'commercial').strip()
    except (TypeError, ValueError):
        return jsonify({'error': 'Некорректные параметры'}), 400
    if not plant_id or not field_id or not year or not container:
        return jsonify({'error': 'Заполните все поля'}), 400
    try:
        if mode == 'full':
            return jsonify(snapshot_for_container(plant_id, container, field_id, year))
        return jsonify(commercial_snapshot(plant_id, container, field_id, year))
    except Exception as exc:
        return jsonify({'error': str(exc)}), 400


@bp.route('/api/seedling_source_batches')
@login_required
def api_seedling_source_batches():
    from app.seedlings import list_source_batches
    try:
        plant_id = int(request.args.get('plant_id') or 0)
        size_id = int(request.args.get('size_id') or 0)
        preferred = int(request.args.get('preferred_field_id') or 0) or None
    except (TypeError, ValueError):
        return jsonify({'items': []})
    if not plant_id or not size_id:
        return jsonify({'items': []})
    return jsonify({'items': list_source_batches(plant_id, size_id, preferred)})


@bp.route('/api/seedling_measure_years')
@login_required
def api_seedling_measure_years():
    from app.seedlings import list_measure_years
    try:
        plant_id = int(request.args.get('plant_id') or 0)
        field_id = int(request.args.get('field_id') or 0)
        container = (request.args.get('container') or '').strip()
    except (TypeError, ValueError):
        return jsonify({'items': []})
    if not plant_id or not field_id or not container:
        return jsonify({'items': []})
    return jsonify({'items': list_measure_years(plant_id, container, field_id)})


@bp.route('/api/seedling_deficit_sources')
@login_required
def api_seedling_deficit_sources():
    from app.seedlings import suggest_deficit_sources
    try:
        plant_id = int(request.args.get('plant_id') or 0)
        size_id = int(request.args.get('size_id') or 0)
        field_id = int(request.args.get('field_id') or 0) or None
        year = int(request.args.get('year') or 0) or None
        need = int(request.args.get('need') or 0)
    except (TypeError, ValueError):
        return jsonify({'items': []})
    if not plant_id or not size_id:
        return jsonify({'items': []})
    return jsonify({
        'items': suggest_deficit_sources(plant_id, size_id, field_id, year, need),
    })


@bp.route('/api/seedling_plant_locations')
@login_required
def api_seedling_plant_locations():
    from app.seedlings import list_plant_seedling_locations
    try:
        plant_id = int(request.args.get('plant_id') or 0)
    except (TypeError, ValueError):
        return jsonify({'items': []})
    if not plant_id:
        return jsonify({'items': []})
    return jsonify({'items': list_plant_seedling_locations(plant_id)})


@bp.route('/api/seedling_pipeline')
@login_required
def api_seedling_pipeline():
    """Цепочка Пересадка → Товарность → Промер по растению (для инфо-блока в проекте)."""
    from app.seedlings import build_plant_pipeline_view
    try:
        plant_id = int(request.args.get('plant_id') or 0)
        preferred = int(request.args.get('preferred_field_id') or 0) or None
    except (TypeError, ValueError):
        return jsonify({'error': 'bad params'}), 400
    if not plant_id:
        return jsonify({'error': 'plant_id required'}), 400
    return jsonify(build_plant_pipeline_view(plant_id, preferred_field_id=preferred))


@bp.route('/api/stock_availability')
@login_required
def api_stock_availability():
    """
    Возвращает доступные остатки по всем полям для конкретного Растения и Размера.
    """
    p_id = request.args.get('plant_id')
    s_id = request.args.get('size_id')
    
    if not p_id or not s_id:
        return jsonify({'items': []})
        
    stocks = StockBalance.query.filter_by(plant_id=p_id, size_id=s_id).all()
    if not stocks:
        return jsonify({'items': []})
    
    reserved_rows = (
        db.session.query(OrderItem.field_id, OrderItem.year, func.sum(OrderItem.quantity - OrderItem.shipped_quantity))
        .join(Order)
        .filter(
            OrderItem.plant_id == p_id, OrderItem.size_id == s_id,
            Order.status != 'canceled', Order.status != 'ghost', Order.is_deleted == False
        )
        .group_by(OrderItem.field_id, OrderItem.year).all()
    )
    reserved_map = {(r[0], r[1]): int(r[2] or 0) for r in reserved_rows}
    
    result = []
    price_overrides = get_shop_price_map()
    for st in stocks:
        free_qty = st.quantity - reserved_map.get((st.field_id, st.year), 0)
        if free_qty > 0:
            wholesale = float(st.price or 0)
            retail = resolve_shop_price(st.plant_id, st.size_id, wholesale, price_overrides)
            result.append({
                'field_id': st.field_id,
                'field_name': st.field.name if st.field else '-',
                'year': st.year,
                'free': free_qty,
                'price': retail,
                'wholesale_price': wholesale,
                'retail_price': retail,
            })
            
    result.sort(key=lambda x: x['free'], reverse=True)
    return jsonify({'items': result})

@bp.route('/api/turnover_years')
@login_required
def api_turnover_years():
    p = request.args.get('plant')
    f = request.args.get('field')
    if not p or not f: 
        return jsonify(list())
    
    unique_years = set()
    stocks = db.session.query(StockBalance).filter_by(plant_id=p, field_id=f).all()
    for st in stocks: 
        if st.year: unique_years.add(st.year)
    ghosts = db.session.query(OrderItem).filter_by(plant_id=p, field_id=f).all()
    for g in ghosts: 
        if g.year: unique_years.add(g.year)
    docs_from = db.session.query(DocumentRow).filter_by(plant_id=p, field_from_id=f).all()
    for d in docs_from: 
        if d.year: unique_years.add(d.year)
    docs_to = db.session.query(DocumentRow).filter_by(plant_id=p, field_to_id=f).all()
    for d in docs_to: 
        if d.year: unique_years.add(d.year)
            
    return jsonify(sorted(list(unique_years)))

@bp.route('/order/<int:order_id>', methods=['GET', 'POST'])
@login_required
def order_detail(order_id):
    from app.digging import ensure_payment_file_column
    ensure_payment_file_column()
    o = Order.query.get_or_404(order_id)
    
    return_to = request.form.get('return_to') or request.args.get('return_to')
    if return_to == 'None': return_to = None
    if current_user.role == 'user2' and request.method == 'POST': 
        flash('Доступ запрещен')
        if return_to:
            return redirect(return_to)
        return redirect(url_for('orders.order_detail', order_id=order_id))
    
    if request.method == 'POST':
        if 'link_project' in request.form:
            pid = request.form.get('project_id')
            if pid:
                o.project_id = int(pid)
            else:
                o.project_id = None
            db.session.commit()
            flash('Привязка к проекту обновлена')

        if 'update_info' in request.form:
            new_client_id = request.form.get('client_id')
            if new_client_id: o.client_id = int(new_client_id)
            o.invoice_number = request.form.get('invoice_number')
            d = request.form.get('invoice_date')
            o.invoice_date = datetime.strptime(d, '%Y-%m-%d') if d else None
            
            if current_user.role in ['admin', 'executive']:
                o.is_barter = request.form.get('is_barter') == 'on'
                
            db.session.commit()
            flash('Обновлено')
            # При выставлении счёта может закрыться карточка «копают без счёта».
            try:
                from app.anomaly_engine import check_digging_started_without_invoice
                check_digging_started_without_invoice(o.id)
            except Exception:
                pass

        elif 'delete_payment_file' in request.form and current_user.role in ['admin', 'executive', 'user', 'shop_manager']:
            p = Payment.query.get(request.form.get('payment_id'))
            if p and p.file_path:
                try:
                    if os.path.exists(p.file_path):
                        os.remove(p.file_path)
                    p.file_path = None
                    db.session.commit()
                    flash('Файл удален')
                    log_action(f"Удалил файл оплаты в заказе #{o.id}")
                except Exception as e:
                    flash(f'Ошибка удаления файла: {e}')
                    
        elif 'upload_payment_file' in request.form and current_user.role in ['admin', 'executive', 'user', 'shop_manager']:
            p = Payment.query.get(request.form.get('payment_id'))
            file = request.files.get('payment_file')
            if p and file and file.filename:
                # Удаляем старый файл, если был
                if p.file_path and os.path.exists(p.file_path):
                    try: os.remove(p.file_path)
                    except OSError: pass
                
                date_str = msk_now().strftime('%Y-%m-%d')
                upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'payment', date_str)
                os.makedirs(upload_dir, exist_ok=True)
                fname = secure_filename(file.filename)
                base, ext = os.path.splitext(fname)
                unique_fname = f"pay_{p.id}_{int(time.time())}{ext}"
                fpath = os.path.join(upload_dir, unique_fname)
                file.save(fpath)
                p.file_path = fpath
                db.session.commit()
                flash('Файл загружен')
                log_action(f"Загрузил файл оплаты в заказ #{o.id}")
                
        elif 'delete_payment' in request.form and current_user.role in ['admin', 'executive', 'user', 'shop_manager']:
            p = Payment.query.get(request.form.get('payment_id'))
            if p:
                db.session.delete(p)
                db.session.commit()
                flash('Оплата удалена')
                log_action(f"Удалил оплату в заказе #{o.id}")

        elif 'edit_payment' in request.form and current_user.role in ['admin', 'executive', 'user', 'shop_manager']:
            p = Payment.query.get(request.form.get('payment_id'))
            if p:
                new_type = request.form.get('payment_type') or 'cashless'
                new_comment = (request.form.get('comment') or '').strip()
                if new_type == 'writeoff' or (p.payment_type == 'writeoff'):
                    if current_user.role != 'admin':
                        flash('Списание долга доступно только администратору', 'danger')
                        return redirect(url_for('orders.order_detail', order_id=o.id))
                    if new_type == 'writeoff' and not new_comment:
                        flash('Комментарий обязателен для списания долга', 'danger')
                        return redirect(url_for('orders.order_detail', order_id=o.id))
                old_amount = p.amount
                old_date = p.date
                old_comment = p.comment

                p.amount = float(request.form.get('amount'))
                p.date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
                p.payment_type = new_type
                p.comment = new_comment
                db.session.commit()
                flash('Оплата изменена')
                log_action(f"Изменил оплату в заказе #{o.id}: {old_amount}→{p.amount} ({old_date}→{p.date})")

        elif 'add_payment' in request.form:
            p_type = request.form.get('payment_type') or 'cashless'
            comment = (request.form.get('comment') or '').strip()
            if p_type == 'writeoff':
                if current_user.role != 'admin':
                    flash('Списание долга доступно только администратору', 'danger')
                    return redirect(url_for('orders.order_detail', order_id=o.id))
                if not comment:
                    flash('Комментарий обязателен для списания долга', 'danger')
                    return redirect(url_for('orders.order_detail', order_id=o.id))
            p = Payment(order_id=o.id, amount=float(request.form.get('amount')), date=datetime.strptime(request.form.get('date'), '%Y-%m-%d'), payment_type=p_type, comment=comment)
            db.session.add(p)
            db.session.commit() # Сохраняем, чтобы получить ID

            file = request.files.get('payment_file')
            if file and file.filename:
                date_str = msk_now().strftime('%Y-%m-%d')
                upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'payment', date_str)
                os.makedirs(upload_dir, exist_ok=True)
                fname = secure_filename(file.filename)
                base, ext = os.path.splitext(fname)
                unique_fname = f"pay_{p.id}_{int(time.time())}{ext}"
                fpath = os.path.join(upload_dir, unique_fname)
                file.save(fpath)
                p.file_path = fpath
                db.session.commit()

            flash('Оплата добавлена')
            log_action(f"Добавил оплату в заказ #{o.id}")

        elif 'cancel_order' in request.form: 
            o.status = 'canceled'
            o.canceled_at = msk_now()
            db.session.commit()
            flash('Отменен')
            log_action(f"Отменил заказ #{o.id}")
            if return_to:
                return redirect(return_to)
            if return_to:
                return redirect(return_to)
            return redirect(url_for('orders.orders_list'))

        elif 'restore_order' in request.form: 
            o.status = 'reserved'
            o.canceled_at = None
            db.session.commit()
            flash('Восстановлен')
            log_action(f"Восстановил заказ #{o.id}")

        elif 'save_items' in request.form:
            ids = request.form.getlist('item_id[]')
            prices = request.form.getlist('price[]')
            qtys = request.form.getlist('quantity[]')
            flds = request.form.getlist('field_id[]')
            yrs = request.form.getlist('year[]')
            prepared_updates = []

            # Snapshot для защиты от гонки: смотрим, какие нарушения уже есть
            # ДО изменений менеджера. Финальный assert ниже сравнит с этим.
            from app.stock_helpers import snapshot_overcommit_for_order
            stock_baseline = snapshot_overcommit_for_order(o.id)

            order_locked = _is_order_locked_for_manager(o, current_user)

            for i, iid in enumerate(ids):
                it = OrderItem.query.get(int(iid))
                if not it:
                    continue

                new_qty = int(qtys[i])
                new_field_id = int(flds[i])
                new_year = int(yrs[i])

                # Защита оплаченных заказов от менеджера: нельзя уменьшать кол-во,
                # менять поле/год/цену. Только увеличение и добавление новых позиций.
                if order_locked:
                    old_qty = int(it.quantity or 0)
                    old_field_id = int(it.field_id or 0)
                    old_year = int(it.year or 0)
                    try:
                        old_price = Decimal(it.price or 0)
                    except Exception:
                        old_price = Decimal(0)
                    try:
                        new_price_dec = Decimal(str(prices[i] or '0'))
                    except Exception:
                        new_price_dec = old_price

                    violation = None
                    if new_qty < old_qty:
                        violation = (f'Нельзя уменьшить кол-во позиции "{it.plant.name} {it.size.name}" '
                                     f'({old_qty} → {new_qty}): заказ уже частично оплачен.')
                    elif new_field_id != old_field_id:
                        violation = (f'Нельзя сменить поле для позиции "{it.plant.name} {it.size.name}": '
                                     'заказ оплачен. Используйте кнопку "Разделить позицию".')
                    elif new_year != old_year:
                        violation = (f'Нельзя сменить год для позиции "{it.plant.name} {it.size.name}": '
                                     'заказ оплачен. Используйте кнопку "Разделить позицию".')
                    elif new_price_dec != old_price:
                        violation = (f'Нельзя сменить цену для позиции "{it.plant.name} {it.size.name}": '
                                     'заказ оплачен.')

                    if violation:
                        flash(violation)
                        if return_to:
                            return redirect(return_to)
                        return redirect(url_for('orders.order_detail', order_id=order_id))

                # Нельзя уменьшать план ниже уже отгруженного количества.
                if new_qty < int(it.shipped_quantity or 0):
                    flash(
                        f'Нельзя установить кол-во меньше отгруженного для позиции '
                        f'"{it.plant.name} {it.size.name}". '
                        f'Отгружено: {it.shipped_quantity}, попытка: {new_qty}.'
                    )
                    if return_to:
                        return redirect(return_to)
                    return redirect(url_for('orders.order_detail', order_id=order_id))

                # Проверка остатка с учетом новой партии и исключением текущей строки.
                ok, free = check_stock_availability(
                    it.plant_id,
                    it.size_id,
                    new_field_id,
                    new_year,
                    new_qty,
                    exclude_item_id=it.id
                )
                if not ok:
                    field_obj = Field.query.get(new_field_id)
                    field_name = field_obj.name if field_obj else new_field_id
                    flash(
                        f'Недостаточно остатка для позиции "{it.plant.name} {it.size.name}" '
                        f'(поле: {field_name}, год: {new_year}). Доступно: {free}, запрошено: {new_qty}.'
                    )
                    if return_to:
                        return redirect(return_to)
                    return redirect(url_for('orders.order_detail', order_id=order_id))

                prepared_updates.append({
                    'item': it,
                    'new_qty': new_qty,
                    'new_price': float(prices[i]),
                    'new_field_id': new_field_id,
                    'new_year': new_year
                })

            for upd in prepared_updates:
                it = upd['item']
                old_qty = int(it.quantity or 0)
                old_field_id = int(it.field_id or 0)
                old_year = int(it.year or 0)
                try:
                    old_price_dec = Decimal(it.price or 0)
                except Exception:
                    old_price_dec = Decimal(0)

                it.quantity = upd['new_qty']
                it.price = upd['new_price']
                it.field_id = upd['new_field_id']
                it.year = upd['new_year']
                if old_qty != upd['new_qty']:
                    _record_order_item_history(
                        order_id=o.id,
                        item=it,
                        action_type='qty_change',
                        before_qty=old_qty,
                        after_qty=upd['new_qty']
                    )

                # Аудит для админа: фиксируем что именно менял менеджер в оплаченном заказе.
                try:
                    new_price_dec = Decimal(str(upd['new_price']))
                except Exception:
                    new_price_dec = old_price_dec
                changes = []
                if old_qty != upd['new_qty']:
                    changes.append({'field': 'qty', 'old': old_qty, 'new': upd['new_qty']})
                if old_field_id != upd['new_field_id']:
                    changes.append({'field': 'field_id', 'old': old_field_id, 'new': upd['new_field_id']})
                if old_year != upd['new_year']:
                    changes.append({'field': 'year', 'old': old_year, 'new': upd['new_year']})
                if old_price_dec != new_price_dec:
                    changes.append({'field': 'price', 'old': str(old_price_dec), 'new': str(new_price_dec)})
                if changes:
                    _audit_paid_order_change(o, 'save_items', {
                        'item_id': it.id,
                        'plant_name': it.plant.name if it.plant else '',
                        'size_name': it.size.name if it.size else '',
                        'changes': changes,
                    }, item_id=it.id)

            # Защита от гонки: финальный re-check относительно baseline'а.
            try:
                db.session.flush()
                from app.stock_helpers import assert_stock_for_order
                assert_stock_for_order(o.id, baseline=stock_baseline)
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc))
                if return_to:
                    return redirect(return_to)
                return redirect(url_for('orders.order_detail', order_id=order_id))

            db.session.commit()
            flash('Сохранено')
            log_action(f"Обновил товары в заказе #{o.id}")

        elif 'delete_item' in request.form:
            # Защита оплаченных заказов: менеджеру нельзя удалять позиции
            if _is_order_locked_for_manager(o, current_user):
                flash('Заказ частично оплачен — менеджер не может удалять позиции. '
                      'Обратитесь к администратору.')
                if return_to:
                    return redirect(return_to)
                return redirect(url_for('orders.order_detail', order_id=order_id))

            it = OrderItem.query.get(request.form.get('item_id'))
            if it and it.shipped_quantity == 0:
                try:
                    _record_order_item_history(
                        order_id=o.id,
                        item=it,
                        action_type='delete_item',
                        before_qty=int(it.quantity or 0),
                        after_qty=0
                    )

                    # FK-safe: история и логи выкопки могут ссылаться на удаляемую позицию.
                    # Перед удалением отвязываем ссылки (order_item_id -> NULL), сохраняя сами записи.
                    OrderItemHistory.query.filter_by(order_item_id=it.id).update(
                        {OrderItemHistory.order_item_id: None},
                        synchronize_session=False
                    )

                    related_logs = DiggingLog.query.filter_by(order_item_id=it.id).all()
                    for log_entry in related_logs:
                        if log_entry.plant_id is None:
                            log_entry.plant_id = it.plant_id
                        if log_entry.size_id is None:
                            log_entry.size_id = it.size_id
                        if log_entry.field_id is None:
                            log_entry.field_id = it.field_id
                        if log_entry.year is None:
                            log_entry.year = it.year
                        log_entry.order_item_id = None

                    db.session.delete(it)
                    db.session.commit()
                    flash('Позиция удалена')
                except Exception as exc:
                    db.session.rollback()
                    flash(f'Ошибка удаления позиции: {exc}')

        elif 'add_new_item' in request.form:
            from app.stock_helpers import snapshot_overcommit_for_order
            stock_baseline = snapshot_overcommit_for_order(o.id)
            p, s, f, y, q = request.form.get('plant'), request.form.get('size'), request.form.get('field'), request.form.get('year'), int(request.form.get('quantity'))
            ok, free = check_stock_availability(p, s, f, int(y), q)
            if ok: 
                wholesale = get_actual_price(p, s, f)
                new_item = OrderItem(
                    order_id=o.id,
                    plant_id=p,
                    size_id=s,
                    field_id=f,
                    year=int(y),
                    quantity=q,
                    price=order_default_price(p, s, wholesale)
                )
                db.session.add(new_item)
                db.session.flush()
                if not o.project_id:
                    from app.finance import resolve_project_id_for_yard_fields
                    linked = resolve_project_id_for_yard_fields([f])
                    if linked:
                        o.project_id = linked
                _record_order_item_history(
                    order_id=o.id,
                    item=new_item,
                    action_type='add_item',
                    before_qty=0,
                    after_qty=q
                )
                # Аудит: менеджер добавил позицию в оплаченный заказ
                try:
                    plant_obj = Plant.query.get(p)
                    size_obj = Size.query.get(s)
                    field_obj = Field.query.get(f)
                    _audit_paid_order_change(o, 'add_item', {
                        'item_id': new_item.id,
                        'plant_name': plant_obj.name if plant_obj else '',
                        'size_name': size_obj.name if size_obj else '',
                        'field_name': field_obj.name if field_obj else '',
                        'year': int(y),
                        'quantity': q,
                        'price': str(new_item.price),
                    }, item_id=new_item.id)
                except Exception:
                    pass

                # Защита от гонки: финальный re-check относительно baseline'а.
                try:
                    from app.stock_helpers import assert_stock_for_order
                    assert_stock_for_order(o.id, baseline=stock_baseline)
                except ValueError as exc:
                    db.session.rollback()
                    flash(str(exc))
                    if return_to:
                        return redirect(return_to)
                    return redirect(url_for('orders.order_detail', order_id=order_id))

                db.session.commit()
            else: 
                flash(f'Недостаточно товара. Доступно: {free}')

        elif 'split_item' in request.form:
            # === Разделение позиции ===
            # Менеджер уточняет поля/размеры по факту бирковки. Защита денег клиента:
            # суммарное кол-во новых частей >= исходного, и суммарная стоимость
            # (qty * price) тоже >= исходной. Если по позиции уже что-то выкопано
            # или отгружено — split запрещён.
            from app.stock_helpers import snapshot_overcommit_for_order
            stock_baseline = snapshot_overcommit_for_order(o.id)
            try:
                source_id = int(request.form.get('item_id'))
            except (TypeError, ValueError):
                flash('Не указана позиция для разделения')
                if return_to:
                    return redirect(return_to)
                return redirect(url_for('orders.order_detail', order_id=order_id))

            it = OrderItem.query.get(source_id)
            if not it or it.order_id != o.id:
                flash('Позиция не найдена')
                if return_to:
                    return redirect(return_to)
                return redirect(url_for('orders.order_detail', order_id=order_id))

            if int(it.shipped_quantity or 0) > 0:
                flash('Нельзя разделить позицию: по ней уже была отгрузка. Обратитесь к администратору.')
                if return_to:
                    return redirect(return_to)
                return redirect(url_for('orders.order_detail', order_id=order_id))

            try:
                dug_total = int(it.dug_total or 0)
            except Exception:
                dug_total = 0
            if dug_total > 0:
                flash('Нельзя разделить позицию: по ней уже есть выкопка. Обратитесь к администратору.')
                if return_to:
                    return redirect(return_to)
                return redirect(url_for('orders.order_detail', order_id=order_id))

            parts = _parse_split_parts_from_form(request.form)
            if not parts:
                flash('Не указаны новые части позиции (нужно хотя бы одну с кол-вом > 0).')
                if return_to:
                    return redirect(return_to)
                return redirect(url_for('orders.order_detail', order_id=order_id))

            old_qty = int(it.quantity or 0)
            try:
                old_price = Decimal(it.price or 0)
            except Exception:
                old_price = Decimal(0)
            old_value = Decimal(old_qty) * old_price

            new_qty_total = sum(p['quantity'] for p in parts)
            new_value_total = sum(
                (Decimal(p['quantity']) * p['price'] for p in parts),
                Decimal(0),
            )

            if new_qty_total < old_qty:
                flash(f'Сумма штук в частях ({new_qty_total}) меньше исходного количества ({old_qty}).')
                if return_to:
                    return redirect(return_to)
                return redirect(url_for('orders.order_detail', order_id=order_id))

            # допустимая погрешность 1 копейка
            if new_value_total + Decimal('0.01') < old_value:
                flash(f'Стоимость частей ({new_value_total} ₽) меньше исходной ({old_value} ₽). '
                      'Заказ оплачен — стоимость уменьшать нельзя.')
                if return_to:
                    return redirect(return_to)
                return redirect(url_for('orders.order_detail', order_id=order_id))

            orig_key = (int(it.size_id), int(it.field_id), int(it.year))
            orig_part = None
            for p in parts:
                if (p['size_id'], p['field_id'], p['year']) == orig_key:
                    orig_part = p
                    break

            # Проверка наличия остатков на каждую часть.
            # Исходная позиция исключается из расчёта резервов всегда:
            # её резерв либо обновится (если совпадает ключ), либо снимется
            # (исходная удалится). Это даёт менеджеру возможность перебиркивать
            # позицию даже когда свободный остаток сейчас 0 из-за её же резерва.
            for p in parts:
                ok, free = check_stock_availability(
                    it.plant_id, p['size_id'], p['field_id'], p['year'],
                    p['quantity'], exclude_item_id=it.id,
                )
                if not ok:
                    field_obj = Field.query.get(p['field_id'])
                    field_name = field_obj.name if field_obj else p['field_id']
                    size_obj = Size.query.get(p['size_id'])
                    size_name = size_obj.name if size_obj else p['size_id']
                    flash(
                        f'Недостаточно остатка для части "{it.plant.name} {size_name}" '
                        f'(поле: {field_name}, год: {p["year"]}). '
                        f'Доступно: {free}, запрошено: {p["quantity"]}.'
                    )
                    if return_to:
                        return redirect(return_to)
                    return redirect(url_for('orders.order_detail', order_id=order_id))

            # Снимок исходной позиции для аудита (до любых изменений).
            audit_source_info = {
                'plant_name': it.plant.name if it.plant else '',
                'size_name': it.size.name if it.size else '',
                'field_name': it.field.name if it.field else '',
                'year': it.year,
                'quantity': old_qty,
                'price': str(old_price),
            }

            # Применяем изменения транзакционно
            try:
                if orig_part is not None:
                    # обновляем исходную позицию
                    _record_order_item_history(
                        order_id=o.id,
                        item=it,
                        action_type='split_source',
                        before_qty=old_qty,
                        after_qty=orig_part['quantity'],
                    )
                    it.quantity = orig_part['quantity']
                    it.price = orig_part['price']
                    db.session.flush()
                else:
                    # удаляем исходную, сохраняя историю и логи
                    _record_order_item_history(
                        order_id=o.id,
                        item=it,
                        action_type='split_source',
                        before_qty=old_qty,
                        after_qty=0,
                    )
                    db.session.flush()
                    OrderItemHistory.query.filter_by(order_item_id=it.id).update(
                        {OrderItemHistory.order_item_id: None},
                        synchronize_session=False,
                    )
                    related_logs = DiggingLog.query.filter_by(order_item_id=it.id).all()
                    for log_entry in related_logs:
                        if log_entry.plant_id is None:
                            log_entry.plant_id = it.plant_id
                        if log_entry.size_id is None:
                            log_entry.size_id = it.size_id
                        if log_entry.field_id is None:
                            log_entry.field_id = it.field_id
                        if log_entry.year is None:
                            log_entry.year = it.year
                        log_entry.order_item_id = None
                    db.session.delete(it)
                    db.session.flush()

                # создаём остальные части
                for p in parts:
                    if (p['size_id'], p['field_id'], p['year']) == orig_key:
                        continue
                    new_item = OrderItem(
                        order_id=o.id,
                        plant_id=it.plant_id,
                        size_id=p['size_id'],
                        field_id=p['field_id'],
                        year=p['year'],
                        quantity=p['quantity'],
                        price=p['price'],
                    )
                    db.session.add(new_item)
                    db.session.flush()
                    _record_order_item_history(
                        order_id=o.id,
                        item=new_item,
                        action_type='split_dest',
                        before_qty=0,
                        after_qty=p['quantity'],
                    )

                # Аудит: фиксируем разбивку для админа
                try:
                    audit_parts = []
                    for p in parts:
                        size_obj = Size.query.get(p['size_id'])
                        field_obj = Field.query.get(p['field_id'])
                        audit_parts.append({
                            'size_name': size_obj.name if size_obj else f'#{p["size_id"]}',
                            'field_name': field_obj.name if field_obj else f'#{p["field_id"]}',
                            'year': p['year'],
                            'quantity': p['quantity'],
                            'price': str(p['price']),
                        })
                    _audit_paid_order_change(o, 'split_item', {
                        'item_id': source_id,
                        'source': audit_source_info,
                        'parts': audit_parts,
                    }, item_id=source_id)
                except Exception:
                    pass

                # Защита от гонки: финальный re-check относительно baseline'а.
                try:
                    db.session.flush()
                    from app.stock_helpers import assert_stock_for_order
                    assert_stock_for_order(o.id, baseline=stock_baseline)
                except ValueError as race_exc:
                    db.session.rollback()
                    flash(str(race_exc))
                    if return_to:
                        return redirect(return_to)
                    return redirect(url_for('orders.order_detail', order_id=order_id))

                db.session.commit()
                flash('Позиция разделена')
                log_action(f"Разделил позицию в заказе #{o.id}")
            except Exception as exc:
                db.session.rollback()
                flash(f'Ошибка при разделении позиции: {exc}')

        if return_to:
            return redirect(return_to)
        return redirect(url_for('orders.order_detail', order_id=order_id))
    
    # Получаем связанные документы (отгрузки)
    # --- ЛОГИКА СБОРА ДЕРЕВА ДОКУМЕНТОВ (СЧЕТ -> ЗАКАЗЫ -> ОПЛАТЫ/ОТГРУЗКИ) ---
    
    # 1. Определяем список заказов для отображения
    # Если есть номер счета - ищем все заказы этого клиента с таким счетом
    if o.invoice_number:
        related_orders = Order.query.filter_by(
            client_id=o.client_id, 
            invoice_number=o.invoice_number,
            is_deleted=False
        ).order_by(Order.date).all()
    else:
        # Если счета нет, показываем только текущий заказ
        related_orders = [o]

    # 2. Собираем данные (группы отгрузок) для КАЖДОГО заказа из списка
    orders_data = []
    
    for ord_obj in related_orders:
        ship_groups = []
        
        # А. Если Ghost
        if ord_obj.status == 'ghost':
            total_qty = sum(i.quantity for i in ord_obj.items)
            if total_qty > 0:
                ship_groups.append({
                    'date': ord_obj.date,
                    'type': 'ghost',
                    'doc_ids': [],
                    'total_qty': total_qty,
                    'comment': 'Историческая запись'
                })
        else:
            # Б. Если обычный заказ
            docs = Document.query.filter_by(order_id=ord_obj.id, doc_type='shipment').order_by(Document.date).all()
            groups_map = {}
            for d in docs:
                date_key = d.date.strftime('%Y-%m-%d')
                if date_key not in groups_map:
                    groups_map[date_key] = {
                        'date': d.date,
                        'type': 'real',
                        'doc_ids': [],
                        'total_qty': 0,
                        'comment': set()
                    }
                group = groups_map[date_key]
                group['doc_ids'].append(str(d.id))
                group['total_qty'] += sum(r.quantity for r in d.rows)
                if d.comment: group['comment'].add(d.comment)
            
            for val in groups_map.values():
                val['comment'] = ", ".join(val['comment'])
                ship_groups.append(val)
            ship_groups.sort(key=lambda x: x['date'])
            
        orders_data.append({
            'order': ord_obj,
            'is_current': (ord_obj.id == o.id),
            'shipments': ship_groups
        })

       # ... в конце функции ...
    active_projects = Project.query.filter_by(status='active').order_by(Project.name).all()

    # Синхронная проверка цен заказа: после любого POST на этой странице пользователь
    # будет редиректнут обратно на GET, и здесь мы обновим/закроем карточку «нет цены».
    try:
        from app.anomaly_engine import sync_price_anomaly_for_order
        sync_price_anomaly_for_order(o.id)
    except Exception:
        pass

    # Синхронная проверка совпадения года партии с фактом — карточки в дашборде
    # появляются/закрываются автоматически после любых правок состава заказа.
    try:
        from app.anomaly_engine import sync_year_mismatch_for_order
        sync_year_mismatch_for_order(o.id)
    except Exception:
        pass

    order_paid_total = _order_paid_total(o.id)
    order_locked_for_manager = _is_order_locked_for_manager(o, current_user)

    # История изменений позиций — только для админа. Запросом не утяжеляем
    # страницу для других ролей. Лимит 200 событий хватает с запасом, кнопка
    # «История .xlsx» уже есть рядом для полного экспорта.
    order_history_rows = []
    if current_user.is_authenticated and getattr(current_user, 'role', None) == 'admin':
        _hist = (
            OrderItemHistory.query
            .filter_by(order_id=o.id)
            .order_by(OrderItemHistory.created_at.desc(), OrderItemHistory.id.desc())
            .limit(200)
            .all()
        )
        for _h in _hist:
            try:
                _payload = json.loads(_h.snapshot_payload) if _h.snapshot_payload else {}
            except Exception:
                _payload = {}
            order_history_rows.append({
                'created_at': _h.created_at,
                'username': (_h.changed_by.username if _h.changed_by else None),
                'action_type': _h.action_type,
                'before_qty': _h.before_quantity,
                'after_qty': _h.after_quantity,
                'delta_qty': _h.delta_quantity or 0,
                'plant_name': _payload.get('plant_name')
                              or (_h.item.plant.name if _h.item and _h.item.plant else None),
                'size_name': _payload.get('size_name')
                             or (_h.item.size.name if _h.item and _h.item.size else None),
                'field_name': _payload.get('field_name')
                              or (_h.item.field.name if _h.item and _h.item.field else None),
                'year': _payload.get('year'),
            })

    return render_template('orders/order_detail.html', 
                           order=o, 
                           orders_data=orders_data, 
                           active_projects=active_projects, # <--- Передаем проекты
                           plants=sorted(Plant.query.all(), key=lambda x: x.name), 
                           sizes=Size.query.all(), 
                           fields=Field.query.all(), 
                           clients=sorted(Client.query.all(), key=lambda x: x.name),
                           current_year=msk_now().year,
                           order_paid_total=order_paid_total,
                           order_locked_for_manager=order_locked_for_manager,
                           order_history_rows=order_history_rows,
                           return_to=return_to)

@bp.route('/order/download_payment_file/<int:payment_id>')
@login_required
def download_payment_file(payment_id):
    p = Payment.query.get_or_404(payment_id)
    if p.file_path and os.path.exists(p.file_path):
        return send_file(p.file_path, as_attachment=True)
    flash("Файл не найден")
    return redirect(request.referrer)

@bp.route('/order/<int:order_id>/ship', methods=['POST'])
@login_required
def order_ship(order_id):
    o = Order.query.get_or_404(order_id)
    return_to = request.form.get('return_to') or request.args.get('return_to')
    if return_to == 'None': return_to = None
    item_id = int(request.form.get('item_id'))
    qty = int(request.form.get('quantity'))
    item = OrderItem.query.get(item_id)

    is_xhr = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    # Защита от дублей: если только что (в пределах 30 сек) уже была создана
    # точно такая же отгрузка по этой позиции (совпадают plant/size/field/year
    # /quantity и order), считаем это повторной отправкой (гонка hx-boost с
    # нашим fetch, сетевой retry, двойной клик на мобильном, два менеджера
    # одновременно нажали «Отгрузить») и возвращаем текущее состояние без
    # создания второго документа. Призрачных заказов это не касается: их
    # отгрузки делаются отдельным флоу и не проходят через этот endpoint.
    cutoff = msk_now() - timedelta(seconds=30)
    dup = (db.session.query(Document)
           .join(DocumentRow, DocumentRow.document_id == Document.id)
           .filter(Document.doc_type == 'shipment',
                   Document.order_id == o.id,
                   # ВАЖНО: проверяем только по заказу/позиции/количеству,
                   # без привязки к user_id. Если двое нажали «Отгрузить»
                   # на одну и ту же позицию за 30 сек — это всё равно дубль.
                   Document.date >= cutoff,
                   DocumentRow.plant_id == item.plant_id,
                   DocumentRow.size_id == item.size_id,
                   DocumentRow.field_from_id == item.field_id,
                   DocumentRow.year == item.year,
                   DocumentRow.quantity == qty)
           .first())
    if dup is not None:
        current_app.logger.info(
            "order_ship duplicate suppressed: order=%s item=%s qty=%s user=%s",
            o.id, item.id, qty, current_user.id,
        )
        if is_xhr:
            return jsonify({'status': 'ok', 'shipped': item.shipped_quantity,
                            'plan': item.quantity, 'duplicate': True})
        if return_to:
            return redirect(return_to)
        return redirect(url_for('orders.order_detail', order_id=order_id))

    if qty > (item.quantity - item.shipped_quantity):
        flash('Ошибка кол-ва')
    else:
        # Защита от гонки: перед изменением подтянем актуальное состояние
        # позиции из БД (мог уже отгрузить параллельный воркер) и ещё раз
        # проверим, что qty помещается в неотгруженный остаток.
        db.session.refresh(item)
        if qty > (item.quantity - item.shipped_quantity):
            db.session.rollback()
            flash(
                f'Кол-во к отгрузке ({qty}) больше, чем осталось '
                f'({item.quantity - item.shipped_quantity}). Возможно, '
                f'кто-то уже отгрузил эту позицию параллельно.'
            )
            if is_xhr:
                return jsonify({'status': 'error', 'message': 'race'}), 409
            if return_to:
                return redirect(return_to)
            return redirect(url_for('orders.order_detail', order_id=order_id))

        doc = Document(doc_type='shipment', user_id=current_user.id, order_id=o.id, comment=f"Отгрузка {o.id}", date=msk_now())
        db.session.add(doc)
        db.session.flush()

        db.session.add(DocumentRow(document_id=doc.id, plant_id=item.plant_id, size_id=item.size_id, field_from_id=item.field_id, year=item.year, quantity=qty))

        from app.utils import get_or_create_stock
        stock = get_or_create_stock(item.plant_id, item.size_id, item.field_id, item.year)
        stock.quantity -= qty

        item.shipped_quantity += qty

        if not o.project_id:
            from app.finance import resolve_project_id_for_yard_fields
            linked = resolve_project_id_for_yard_fields([item.field_id])
            if linked:
                o.project_id = linked

        if o.status == 'reserved': o.status = 'in_progress'
        if all(i.shipped_quantity >= i.quantity for i in o.items): o.status = 'shipped'

        db.session.commit()

        if is_xhr:
            return jsonify({'status': 'ok', 'shipped': item.shipped_quantity, 'plan': item.quantity})

        flash('Отгружено (остатки обновлены)')
        log_action(f"Отгрузил из заказа #{o.id}")

    if is_xhr:
        return jsonify({'status': 'error', 'message': 'Ошибка кол-ва'})

    if return_to:
        return redirect(return_to)
    return redirect(url_for('orders.order_detail', order_id=order_id))


@bp.route('/admin/shipment_duplicates')
@login_required
def admin_shipment_duplicates():
    """Диагностический отчёт: дубли частичных отгрузок.

    Только чтение. БД не меняет. Ищет пары (и более) Document(doc_type='shipment')
    у одного и того же пользователя и заказа, у которых совпадает DocumentRow
    по (plant_id, size_id, field_from_id, year, quantity) и которые созданы
    в пределах 120 сек друг от друга.

    Призрачные заказы (Order.status == 'ghost') и удалённые заказы исключены.
    Формат: по умолчанию HTML, ?format=json — JSON.
    """
    if current_user.role != 'admin':
        return jsonify({'error': 'forbidden'}), 403

    window_seconds = int(request.args.get('window', 120))

    rows = (db.session.query(
                Document.id.label('doc_id'),
                Document.order_id,
                Document.user_id,
                Document.date,
                DocumentRow.id.label('row_id'),
                DocumentRow.plant_id,
                DocumentRow.size_id,
                DocumentRow.field_from_id,
                DocumentRow.year,
                DocumentRow.quantity,
            )
            .join(DocumentRow, DocumentRow.document_id == Document.id)
            .join(Order, Order.id == Document.order_id)
            .filter(Document.doc_type == 'shipment',
                    Order.status != 'ghost',
                    Order.is_deleted == False)
            .order_by(Document.order_id, DocumentRow.plant_id,
                      DocumentRow.size_id, DocumentRow.field_from_id,
                      DocumentRow.year, DocumentRow.quantity, Document.date)
            .all())

    groups = {}
    for r in rows:
        key = (r.user_id, r.order_id, r.plant_id, r.size_id,
               r.field_from_id, r.year, r.quantity)
        groups.setdefault(key, []).append(r)

    duplicate_groups = []
    total_extra_qty = 0
    for key, items in groups.items():
        if len(items) < 2:
            continue
        items_sorted = sorted(items, key=lambda x: x.date)
        cluster = [items_sorted[0]]
        clusters = []
        for it in items_sorted[1:]:
            if (it.date - cluster[-1].date).total_seconds() <= window_seconds:
                cluster.append(it)
            else:
                if len(cluster) >= 2:
                    clusters.append(cluster)
                cluster = [it]
        if len(cluster) >= 2:
            clusters.append(cluster)
        for cl in clusters:
            extra = len(cl) - 1
            extra_qty = extra * cl[0].quantity
            total_extra_qty += extra_qty
            user = db.session.get(User, cl[0].user_id) if cl[0].user_id else None
            plant = db.session.get(Plant, cl[0].plant_id) if cl[0].plant_id else None
            size = db.session.get(Size, cl[0].size_id) if cl[0].size_id else None
            field = db.session.get(Field, cl[0].field_from_id) if cl[0].field_from_id else None
            order = db.session.get(Order, cl[0].order_id) if cl[0].order_id else None
            client = order.client.name if order and order.client else ''

            matching_item = (OrderItem.query
                             .filter_by(order_id=cl[0].order_id,
                                        plant_id=cl[0].plant_id,
                                        size_id=cl[0].size_id,
                                        field_id=cl[0].field_from_id,
                                        year=cl[0].year)
                             .first())

            duplicate_groups.append({
                'order_id': cl[0].order_id,
                'client': client,
                'user': user.username if user else None,
                'plant': plant.name if plant else None,
                'size': size.name if size else None,
                'field': field.name if field else None,
                'year': cl[0].year,
                'qty_each': cl[0].quantity,
                'count': len(cl),
                'extra_copies': extra,
                'extra_qty': extra_qty,
                'docs': [{
                    'doc_id': it.doc_id,
                    'row_id': it.row_id,
                    'date': it.date.strftime('%Y-%m-%d %H:%M:%S') if it.date else None,
                } for it in cl],
                'item_id': matching_item.id if matching_item else None,
                'item_shipped_quantity': matching_item.shipped_quantity if matching_item else None,
                'item_quantity': matching_item.quantity if matching_item else None,
                'order_status': order.status if order else None,
            })

    duplicate_groups.sort(key=lambda g: (-g['extra_qty'], g['order_id']))
    summary = {
        'window_seconds': window_seconds,
        'groups_found': len(duplicate_groups),
        'total_extra_qty': total_extra_qty,
    }

    if request.args.get('format') == 'json':
        return jsonify({'summary': summary, 'duplicates': duplicate_groups})

    return render_template('orders/shipment_duplicates.html',
                           summary=summary,
                           duplicates=duplicate_groups)


@bp.route('/admin/shipment_duplicates/cleanup', methods=['POST'])
@login_required
def admin_shipment_duplicates_cleanup():
    """Удаление лишних копий shipment-документов.

    На вход: doc_ids — список id документов, которые надо удалить
    (передаются как повторяющийся параметр формы doc_ids или JSON-массив).

    Для каждого doc_id выполняется ЖЁСТКАЯ проверка безопасности:
      1) Document.doc_type == 'shipment'
      2) Заказ не призрачный и не удалён
      3) Существует ДРУГОЙ Document(doc_type='shipment') с тем же
         (user_id, order_id) и идентичной DocumentRow (plant/size/field/year/qty).
         Если такого нет — значит это НЕ дубликат, а единственный документ, и мы
         его не трогаем.

    Если все проверки прошли:
      - удаляем DocumentRow и Document
      - возвращаем stock.quantity += qty (через get_or_create_stock)
      - пишем ActionLog

    OrderItem.shipped_quantity и Order.status НЕ трогаем: они — источник правды
    для отображения в редакторе заказа. Бывает, что задублировались только
    Document/списание склада, а shipped_quantity был изначально корректный
    (или уже поправлен вручную). Пусть редактор заказа остаётся как есть.

    Всё в одной транзакции. При любой ошибке — rollback.
    """
    if current_user.role != 'admin':
        return jsonify({'error': 'forbidden'}), 403

    data = request.get_json(silent=True) or {}
    raw_ids = data.get('doc_ids') if data else request.form.getlist('doc_ids')
    if not raw_ids:
        return jsonify({'error': 'no doc_ids provided'}), 400

    try:
        doc_ids = [int(x) for x in raw_ids]
    except (TypeError, ValueError):
        return jsonify({'error': 'invalid doc_ids'}), 400

    results = []
    from app.utils import get_or_create_stock

    try:
        for doc_id in doc_ids:
            doc = db.session.get(Document, doc_id)
            if not doc:
                results.append({'doc_id': doc_id, 'status': 'skip', 'reason': 'not_found'})
                continue
            if doc.doc_type != 'shipment':
                results.append({'doc_id': doc_id, 'status': 'skip', 'reason': 'wrong_doc_type'})
                continue

            order = db.session.get(Order, doc.order_id) if doc.order_id else None
            if not order:
                results.append({'doc_id': doc_id, 'status': 'skip', 'reason': 'no_order'})
                continue
            if order.status == 'ghost' or order.is_deleted:
                results.append({'doc_id': doc_id, 'status': 'skip', 'reason': 'ghost_or_deleted'})
                continue

            rows = list(doc.rows)
            if len(rows) != 1:
                results.append({'doc_id': doc_id, 'status': 'skip', 'reason': 'unexpected_row_count'})
                continue
            row = rows[0]

            twin_exists = (db.session.query(Document.id)
                           .join(DocumentRow, DocumentRow.document_id == Document.id)
                           .filter(Document.id != doc.id,
                                   Document.doc_type == 'shipment',
                                   Document.order_id == doc.order_id,
                                   Document.user_id == doc.user_id,
                                   DocumentRow.plant_id == row.plant_id,
                                   DocumentRow.size_id == row.size_id,
                                   DocumentRow.field_from_id == row.field_from_id,
                                   DocumentRow.year == row.year,
                                   DocumentRow.quantity == row.quantity)
                           .first())
            if not twin_exists:
                results.append({'doc_id': doc_id, 'status': 'skip', 'reason': 'no_twin_unique_document'})
                continue

            qty = int(row.quantity or 0)
            plant_id = row.plant_id
            size_id = row.size_id
            field_id = row.field_from_id
            year = row.year

            stock = get_or_create_stock(plant_id, size_id, field_id, year)
            stock.quantity = (stock.quantity or 0) + qty

            db.session.delete(doc)

            results.append({
                'doc_id': doc_id,
                'status': 'ok',
                'order_id': order.id,
                'qty_returned': qty,
                'plant_id': plant_id,
                'size_id': size_id,
                'field_id': field_id,
                'year': year,
            })

        db.session.commit()

        ok_results = [r for r in results if r['status'] == 'ok']
        for r in ok_results:
            log_action(
                f"Удалил дубль отгрузки Document #{r['doc_id']} "
                f"из заказа #{r['order_id']} ({r['qty_returned']} шт возвращено на склад)"
            )

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("shipment cleanup failed")
        return jsonify({'error': 'exception', 'message': str(e), 'partial': results}), 500

    summary = {
        'requested': len(doc_ids),
        'deleted': sum(1 for r in results if r['status'] == 'ok'),
        'skipped': sum(1 for r in results if r['status'] == 'skip'),
        'qty_returned_total': sum(r.get('qty_returned', 0) for r in results if r['status'] == 'ok'),
    }
    return jsonify({'summary': summary, 'results': results})


@bp.route('/order/<int:order_id>/send_shipment_report', methods=['POST'])
@login_required
def send_shipment_report(order_id):
    o = Order.query.get_or_404(order_id)
    return_to = request.form.get('return_to') or request.args.get('return_to')
    if return_to == 'None': return_to = None
    
    # --- ЗАЩИТА ОТ ДУБЛИКАТОВ (5 МИНУТ) ---
    cutoff_time = msk_now() - timedelta(minutes=5)
    recent_log = ActionLog.query.filter(
        ActionLog.user_id == current_user.id,
        ActionLog.action == f"Отправил отчет об отгрузке в ТГ по заказу #{o.id}",
        ActionLog.date >= cutoff_time
    ).first()
    
    if recent_log:
        flash('Отчет по этой отгрузке уже был отправлен недавно. Пожалуйста, подождите.', 'info')
        if return_to:
            return redirect(return_to)
        return redirect(url_for('orders.order_detail', order_id=order_id))
    # --------------------------------------
    
    comment = request.form.get('comment', '').strip()
    photos = request.files.getlist('photos')
    
    # 1. Сохранение фотографий
    saved_photos = []
    if photos and any(p.filename for p in photos):
        date_str = msk_now().strftime('%Y-%m-%d')
        upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'shipment', date_str)
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir, exist_ok=True)
        for p in photos:
            if p.filename:
                fname = secure_filename(p.filename)
                base, ext = os.path.splitext(fname)
                unique_fname = f"ord{o.id}_{int(time.time()*1000)}{ext}"
                fpath = os.path.join(upload_dir, unique_fname)
                p.save(fpath)
                saved_photos.append(fpath)
    
    # 2. Формирование текста отчета
    # Формат:
    #   🚚 Отгрузка по заказу #N
    #   👤 Клиент: ...
    #   💰 Сумма заказа: ...
    #   [💬 Комментарий: ...]
    #
    #   🚚 Отгружено DD.MM.YYYY:
    #       🌲 <Растение>
    #            ✅ <размер> ➔ N шт
    #   Отгружено всего DD.MM.YYYY: X шт
    #
    #   📦 Статус заказа:
    #       🌲 <Растение>
    #            ✅ <размер> ➔ shipped из plan шт.
    #   Отгружено всего по заказу: Y шт из Z шт
    today_msk = msk_now()
    today_date = today_msk.date()
    today_label = today_date.strftime('%d.%m.%Y')
    start_today = today_msk.replace(hour=0, minute=0, second=0, microsecond=0)
    next_day = start_today + timedelta(days=1)

    report_lines = []
    report_lines.append(f"🚚 <b>Отгрузка по заказу #{o.id}</b>")
    report_lines.append(f"👤 <b>Клиент:</b> {o.client.name}")

    total_sum = sum(item.sum for item in o.items)
    report_lines.append(f"💰 <b>Сумма заказа:</b> {total_sum:,.0f} ₽".replace(',', ' '))

    if comment:
        import html
        safe_comment = html.escape(comment)
        report_lines.append(f"💬 <b>Комментарий:</b> {safe_comment}")

    # --- Блок 1: что отгружено СЕГОДНЯ по этому заказу --------------------
    today_docs = (Document.query
                  .filter(Document.doc_type == 'shipment',
                          Document.order_id == o.id,
                          Document.date >= start_today,
                          Document.date < next_day)
                  .all())
    today_agg = {}  # (plant_id, size_id) -> {'plant': str, 'size': str, 'qty': int}
    for d in today_docs:
        for r in d.rows:
            key = (r.plant_id, r.size_id)
            slot = today_agg.setdefault(key, {
                'plant': r.plant.name if r.plant else '—',
                'size': r.size.name if r.size else '—',
                'qty': 0,
            })
            slot['qty'] += int(r.quantity or 0)

    today_total = sum(v['qty'] for v in today_agg.values())

    report_lines.append("")
    report_lines.append(f"🚚 <b>Отгружено {today_label}:</b>")
    if today_agg:
        # Группируем по растению, внутри сортируем размеры натурально
        today_by_plant = {}
        for (_, _), v in today_agg.items():
            today_by_plant.setdefault(v['plant'], []).append(v)
        for plant_name in sorted(today_by_plant.keys(), key=lambda s: s.lower()):
            sizes = sorted(today_by_plant[plant_name], key=lambda x: natural_key(x['size'] or ''))
            report_lines.append(f"       🌲 <b>{plant_name}</b>")
            for s in sizes:
                report_lines.append(f"              ✅  {s['size']} ➔ {s['qty']} шт")
        report_lines.append(f"<b>Отгружено всего {today_label}: {today_total} шт</b>")
    else:
        report_lines.append("<i>За сегодня по этому заказу отгрузок нет.</i>")

    # --- Блок 2: текущий статус заказа (план/факт по всем позициям) -------
    status_agg = {}  # (plant_id, size_id) -> {'plant', 'size', 'plan', 'fact'}
    for it in o.items:
        key = (it.plant_id, it.size_id)
        slot = status_agg.setdefault(key, {
            'plant': it.plant.name if it.plant else '—',
            'size': it.size.name if it.size else '—',
            'plan': 0,
            'fact': 0,
        })
        slot['plan'] += int(it.quantity or 0)
        slot['fact'] += int(it.shipped_quantity or 0)

    status_total_fact = sum(v['fact'] for v in status_agg.values())
    status_total_plan = sum(v['plan'] for v in status_agg.values())

    report_lines.append("")
    report_lines.append("📦 <b>Статус заказа:</b>")
    status_by_plant = {}
    for v in status_agg.values():
        status_by_plant.setdefault(v['plant'], []).append(v)
    for plant_name in sorted(status_by_plant.keys(), key=lambda s: s.lower()):
        sizes = sorted(status_by_plant[plant_name], key=lambda x: natural_key(x['size'] or ''))
        report_lines.append(f"       🌲 <b>{plant_name}</b>")
        for s in sizes:
            plan = s['plan']
            fact = s['fact']
            icon = "✅" if plan and fact >= plan else ("⏳" if fact > 0 else "❌")
            report_lines.append(f"              {icon}  {s['size']} ➔ {fact} из {plan} шт.")
    report_lines.append(
        f"<b>Отгружено всего по заказу: {status_total_fact} шт из {status_total_plan} шт</b>"
    )

    msg = "\n".join(report_lines)
    
    # 3. Отправка в ТГ
    # Сначала отправляем фото АЛЬБОМОМ
    if saved_photos:
        send_tg_photo_album_orders(saved_photos)
        time.sleep(1)
        
    # Затем отправляем текстовый отчет
    send_tg_message_orders(msg)
    
    flash('Отчет об отгрузке успешно отправлен в Telegram!', 'success')
    log_action(f"Отправил отчет об отгрузке в ТГ по заказу #{o.id}")
    
    if return_to:
        return redirect(return_to)
    return redirect(url_for('orders.order_detail', order_id=order_id))

@bp.route('/invoice/<int:client_id>/<path:invoice_number>')
@login_required
def invoice_detail(client_id, invoice_number):
    # 1. Ищем все заказы
    from urllib.parse import unquote
    invoice_number = unquote(invoice_number)
    
    orders = Order.query.filter(
        Order.client_id == client_id,
        Order.invoice_number == invoice_number,
        Order.is_deleted == False
    ).order_by(Order.date).all()
    
    if not orders:
        flash(f'Счет "{invoice_number}" не найден или заказы удалены')
        return redirect(url_for('orders.orders_list'))
        
    client = orders[0].client
    invoice_date = orders[0].invoice_date
    return_to = request.args.get('return_to') or request.referrer
    
    # 2. Считаем общую математику (используем Decimal из базы, не превращая в float)
    from decimal import Decimal
    
    stats = {
        'total_sum': Decimal(0),
        'paid_sum': Decimal(0),
        'shipped_sum': Decimal(0),
        'total_items_qty': 0,
        'orders_count': len(orders)
    }
    
    # 3. Сбор данных для "Товарной Матрицы"
    matrix = {}
    
    all_payments = []
    all_shipments = []
    
    for o in orders:
        stats['total_sum'] += o.total_sum
        stats['paid_sum'] += o.paid_sum
        
        # Собираем платежи
        for p in o.payments:
            all_payments.append({'date': p.date, 'amount': p.amount, 'comment': p.comment, 'order_id': o.id})
            
        # Собираем товары
        for item in o.items:
            stats['total_items_qty'] += item.quantity
            
            key = (item.plant.name, item.size.name)
            if key not in matrix:
                matrix[key] = {
                    'plant': item.plant.name,
                    'size': item.size.name,
                    'total_qty': 0,
                    'shipped_qty': 0,
                    'dug_qty': 0, # <--- Добавили поле
                    'avg_price': 0,
                    'total_money': 0,
                    'orders_links': []
                }
            
            row = matrix[key]
            row['total_qty'] += item.quantity
            row['shipped_qty'] += item.shipped_quantity
            row['dug_qty'] += (item.dug_quantity or 0) # <--- Суммируем выкопанное
            
            # ВАЖНО: item.price - это Decimal, умножаем напрямую
            # Если price вдруг None, заменяем на 0
            price = item.price if item.price is not None else Decimal(0)
            
            row['total_money'] += price * item.quantity
            
            if o.id not in row['orders_links']: row['orders_links'].append(o.id)
            
            # Считаем сумму отгруженного по факту (без float!)
            stats['shipped_sum'] += price * item.shipped_quantity

        # Собираем документы отгрузки
        docs = Document.query.filter_by(order_id=o.id, doc_type='shipment').all()
        for d in docs:
            qty = sum(r.quantity for r in d.rows)
            all_shipments.append({'date': d.date, 'id': d.id, 'qty': qty, 'order_id': o.id, 'comment': d.comment})

    # Досчитываем средние цены
    sorted_matrix = []
    for k, v in matrix.items():
        if v['total_qty'] > 0:
            v['avg_price'] = v['total_money'] / Decimal(v['total_qty'])
        sorted_matrix.append(v)
    
    # Сортировка
    sorted_matrix.sort(key=lambda x: x['plant'])
    all_payments.sort(key=lambda x: x['date'])
    all_shipments.sort(key=lambda x: x['date'])

    return render_template('orders/invoice_detail.html', 
                           invoice_number=invoice_number,
                           invoice_date=invoice_date,
                           client=client,
                           orders=orders,
                           stats=stats,
                           matrix=sorted_matrix,
                           payments=all_payments,
                           shipments=all_shipments,
                           return_to=return_to)

@bp.route('/orders/export')
@login_required
def export_orders():
    """Excel-выгрузка списка заказов.

    Учитывает все фильтры со страницы /orders (клиент, статус, даты, режим
    активные/скрытые) плюс multi-select «По номеру заказа» (`filter_ids`).
    Если пользователь отметил конкретные номера в фильтре — Excel выгрузит
    только их. Иначе — все заказы по текущим фильтрам.

    Дизайн копирует «красивый» отчёт `export_order_history`: для каждого
    заказа собственная зелёная шапка, светло-зелёная подшапка с клиентом,
    таблица позиций (без столбцов «Первоначально / Изменения» — мы хотим
    видеть только фактические числа на момент выгрузки) и блок ИТОГО /
    Оплачено / Остаток. Между заказами — пустая строка-разделитель.

    В конце листа — общий итог по выгруженной выборке (кол-во заказов,
    суммарная сумма, оплачено и остаток), чтобы при многозаказной выгрузке
    сразу был виден сводный показатель.
    """
    f_client = request.args.get('filter_client')
    f_status = request.args.get('filter_status')
    f_date_start = request.args.get('start_date')
    f_date_end = request.args.get('end_date')
    mode = request.args.get('mode', 'active')

    # Multi-select по номерам заказов из фильтра на /orders. Приходит как
    # ?filter_ids=1&filter_ids=5&filter_ids=12. Дополнительно поддерживаем
    # старый формат ?ids=1,5,12 на случай прямых ссылок.
    f_ids_raw = request.args.getlist('filter_ids')
    if not f_ids_raw and request.args.get('ids'):
        f_ids_raw = (request.args.get('ids') or '').split(',')
    selected_ids = []
    for v in f_ids_raw:
        v = (v or '').strip()
        if v.isdigit():
            selected_ids.append(int(v))

    # Применяем тот же набор фильтров, что и на странице /orders, плюс
    # фильтр по выбранным номерам. Это гарантирует, что Excel содержит
    # ровно то, что сейчас видно в списке.
    q = Order.query.filter_by(is_deleted=(mode == 'trash'))
    if f_client:
        q = q.filter(Order.client_id == int(f_client))
    if f_status:
        q = q.filter(Order.status == f_status)
    if not f_status:
        q = q.filter(Order.status != 'ghost')
    if f_date_start:
        q = q.filter(func.date(Order.date) >= f_date_start)
    if f_date_end:
        q = q.filter(func.date(Order.date) <= f_date_end)
    if selected_ids:
        q = q.filter(Order.id.in_(selected_ids))
    orders = q.order_by(Order.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"

    # --- Палитра/стили (как в export_order_history) -----------------------
    style_order_header = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    style_sub_header = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    style_table_header = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    style_total = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
    style_grand_total = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    font_order_header = Font(bold=True, color="FFFFFF", size=13)
    font_sub_header = Font(bold=True, color="1B5E20", size=11)
    font_table_header = Font(bold=True, color="000000")
    font_total = Font(bold=True)
    border_total = Border(top=Side(style='thick'))
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # 7 колонок (как в истории заказа), но без «Первоначально / Изменения».
    # «Поле» и «Год» нужны при многозаказной выгрузке: один и тот же размер
    # может встречаться у разных партий и без поля/года их не отличить.
    columns = ["Растение", "Размер", "Поле", "Год", "Цена", "Кол-во", "Сумма"]
    col_widths = [33, 14, 16, 8, 14, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_idx = 1
    grand_total_qty = 0
    grand_total_sum = Decimal('0')
    grand_total_paid = Decimal('0')
    exported_count = 0

    for o in orders:
        if not o.items:
            continue

        # Шапка заказа (зелёная) — №, дата, статус
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        cell = ws.cell(
            row=row_idx, column=1,
            value=f"Заказ №{o.id} от {o.date.strftime('%d.%m.%Y')}"
        )
        cell.fill = style_order_header
        cell.font = font_order_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 26
        row_idx += 1

        # Подшапка — клиент + счёт + статус
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        client_name = o.client.name if o.client else '—'
        sub_parts = [f"Клиент: {client_name}", f"Статус: {o.status}"]
        if o.invoice_number:
            inv_part = f"Счёт: {o.invoice_number}"
            if o.invoice_date:
                inv_part += f" от {o.invoice_date.strftime('%d.%m.%Y')}"
            sub_parts.append(inv_part)
        c_client = ws.cell(row=row_idx, column=1, value="    |    ".join(sub_parts))
        c_client.fill = style_sub_header
        c_client.font = font_sub_header
        c_client.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 22
        row_idx += 2

        # Заголовки таблицы позиций
        for col_num, col_name in enumerate(columns, 1):
            c = ws.cell(row=row_idx, column=col_num, value=col_name)
            c.fill = style_table_header
            c.font = font_table_header
            c.border = thin_border
            c.alignment = align_center
        row_idx += 1

        # Позиции — фактические числа на момент выгрузки.
        # Сортируем для удобочитаемости: растение → размер → поле → год.
        items_sorted = sorted(
            o.items,
            key=lambda it: (
                (it.plant.name if it.plant else '').lower(),
                natural_key(it.size.name if it.size else ''),
                (it.field.name if it.field else '').lower(),
                it.year or 0,
            ),
        )

        order_total_qty = 0
        order_total_sum = Decimal('0')

        for item in items_sorted:
            plant_name = item.plant.name if item.plant else '—'
            size_name = item.size.name if item.size else '—'
            field_name = item.field.name if item.field else '—'
            qty = int(item.quantity or 0)
            price = Decimal(str(item.price or 0))
            line_sum = price * Decimal(qty)

            ws.cell(row=row_idx, column=1, value=plant_name).border = thin_border
            ws.cell(row=row_idx, column=2, value=size_name).border = thin_border
            c_field = ws.cell(row=row_idx, column=3, value=field_name)
            c_field.border = thin_border
            c_field.alignment = align_center

            c_year = ws.cell(row=row_idx, column=4, value=item.year)
            c_year.border = thin_border
            c_year.alignment = align_center

            c_price = ws.cell(row=row_idx, column=5, value=float(price))
            c_price.border = thin_border
            c_price.number_format = '#,##0.00 "₽"'

            c_qty = ws.cell(row=row_idx, column=6, value=qty)
            c_qty.border = thin_border
            c_qty.alignment = align_center

            c_sum = ws.cell(row=row_idx, column=7, value=float(line_sum))
            c_sum.border = thin_border
            c_sum.number_format = '#,##0.00 "₽"'
            c_sum.font = Font(bold=True)

            order_total_qty += qty
            order_total_sum += line_sum
            row_idx += 1

        # ИТОГО по заказу
        c_label = ws.cell(row=row_idx, column=5, value="ИТОГО:")
        c_label.font = font_total
        c_label.alignment = align_right
        c_label.border = border_total
        c_label.fill = style_total
        c_t_qty = ws.cell(row=row_idx, column=6, value=order_total_qty)
        c_t_qty.font = font_total
        c_t_qty.alignment = align_center
        c_t_qty.border = border_total
        c_t_qty.fill = style_total
        c_t_sum = ws.cell(row=row_idx, column=7, value=float(order_total_sum))
        c_t_sum.font = font_total
        c_t_sum.number_format = '#,##0.00 "₽"'
        c_t_sum.border = border_total
        c_t_sum.fill = style_total
        # Закрасим пустые ячейки слева, чтобы тоновая полоса была сплошной
        for col_idx in range(1, 5):
            cc = ws.cell(row=row_idx, column=col_idx)
            cc.fill = style_total
            cc.border = border_total
        row_idx += 2

        # Блок «Сумма заказа / Оплачено / Остаток»
        total_sum = Decimal(str(o.total_sum or 0))
        paid_sum = Decimal(str(o.paid_sum or 0))
        debt = total_sum - paid_sum
        summary_titles = ["Сумма заказа", "Оплачено", "Остаток"]
        summary_values = [float(total_sum), float(paid_sum), float(debt)]
        summary_cols = [(1, 2), (3, 4), (5, 7)]

        for idx, title in enumerate(summary_titles):
            start_col, end_col = summary_cols[idx]
            ws.merge_cells(start_row=row_idx, start_column=start_col,
                           end_row=row_idx, end_column=end_col)
            t_cell = ws.cell(row=row_idx, column=start_col, value=title)
            t_cell.fill = style_sub_header
            t_cell.font = font_sub_header
            t_cell.alignment = align_center
            t_cell.border = thin_border
            for c in range(start_col + 1, end_col + 1):
                ws.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1

        for idx, value in enumerate(summary_values):
            start_col, end_col = summary_cols[idx]
            ws.merge_cells(start_row=row_idx, start_column=start_col,
                           end_row=row_idx, end_column=end_col)
            v_cell = ws.cell(row=row_idx, column=start_col, value=value)
            v_cell.font = Font(bold=True, size=12, color="1B5E20")
            v_cell.alignment = align_center
            v_cell.number_format = '#,##0.00 "₽"'
            v_cell.border = thin_border
            for c in range(start_col + 1, end_col + 1):
                ws.cell(row=row_idx, column=c).border = thin_border
        # Пустая строка-разделитель между заказами.
        row_idx += 3

        grand_total_qty += order_total_qty
        grand_total_sum += order_total_sum
        grand_total_paid += paid_sum
        exported_count += 1

    # --- Общий итог по выборке (только если выгружено больше 1 заказа) ----
    if exported_count > 1:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        gh = ws.cell(
            row=row_idx, column=1,
            value=f"ВСЕГО ПО ВЫГРУЗКЕ ({exported_count} заказов)",
        )
        gh.fill = style_grand_total
        gh.font = Font(bold=True, color="1B5E20", size=12)
        gh.alignment = Alignment(horizontal="center", vertical="center")
        gh.border = thin_border
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1

        grand_total_debt = grand_total_sum - grand_total_paid
        grand_titles = ["Сумма заказов", "Всего оплачено", "Всего остаток"]
        grand_values = [
            float(grand_total_sum),
            float(grand_total_paid),
            float(grand_total_debt),
        ]

        for idx, title in enumerate(grand_titles):
            start_col, end_col = ((1, 2), (3, 4), (5, 7))[idx]
            ws.merge_cells(start_row=row_idx, start_column=start_col,
                           end_row=row_idx, end_column=end_col)
            t_cell = ws.cell(row=row_idx, column=start_col, value=title)
            t_cell.fill = style_grand_total
            t_cell.font = font_sub_header
            t_cell.alignment = align_center
            t_cell.border = thin_border
            for c in range(start_col + 1, end_col + 1):
                ws.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1

        for idx, value in enumerate(grand_values):
            start_col, end_col = ((1, 2), (3, 4), (5, 7))[idx]
            ws.merge_cells(start_row=row_idx, start_column=start_col,
                           end_row=row_idx, end_column=end_col)
            v_cell = ws.cell(row=row_idx, column=start_col, value=value)
            v_cell.font = Font(bold=True, size=13, color="1B5E20")
            v_cell.alignment = align_center
            v_cell.number_format = '#,##0.00 "₽"'
            v_cell.border = thin_border
            for c in range(start_col + 1, end_col + 1):
                ws.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1

        c_qty = ws.cell(row=row_idx, column=6, value=grand_total_qty)
        c_qty.font = font_sub_header
        c_qty.alignment = align_center
        c_qty.border = thin_border
        c_qty.fill = style_grand_total
        c_qty_lbl = ws.cell(row=row_idx, column=5, value="Всего шт.:")
        c_qty_lbl.font = font_sub_header
        c_qty_lbl.alignment = align_right
        c_qty_lbl.border = thin_border
        c_qty_lbl.fill = style_grand_total

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'Коммерческое предложение {msk_now().strftime("%d.%m.%Y")}.xlsx'

    return send_file(
        buf,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/order/<int:order_id>/export_history')
@login_required
def export_order_history(order_id):
    if current_user.role == 'user2':
        flash('Доступ запрещен')
        return redirect(url_for('orders.order_detail', order_id=order_id))
    o = Order.query.get_or_404(order_id)
    history_rows = OrderItemHistory.query.filter_by(order_id=o.id).order_by(OrderItemHistory.created_at.asc(), OrderItemHistory.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Заказ_{o.id}"

    style_order_header = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    style_sub_header = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    style_table_header = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    style_total = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
    font_order_header = Font(bold=True, color="FFFFFF", size=13)
    font_sub_header = Font(bold=True, color="1B5E20", size=11)
    font_table_header = Font(bold=True, color="000000")
    font_total = Font(bold=True)
    border_total = Border(top=Side(style='thick'))
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    columns = ["Растение", "Размер", "Цена", "Первоначально", "Изменения", "Итог", "Сумма"]
    col_widths = [33, 16, 14, 16, 14, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    row_idx = 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    header_text = f"Статус заказа №{o.id} от {o.date.strftime('%d.%m.%Y')}"
    cell = ws.cell(row=row_idx, column=1, value=header_text)
    cell.fill = style_order_header
    cell.font = font_order_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_idx].height = 28
    row_idx += 1

    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    client_line = f"Клиент: {o.client.name}    |    Статус: {o.status}"
    c_client = ws.cell(row=row_idx, column=1, value=client_line)
    c_client.fill = style_sub_header
    c_client.font = font_sub_header
    c_client.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_idx].height = 22
    row_idx += 2

    for col_num, col_name in enumerate(columns, 1):
        c = ws.cell(row=row_idx, column=col_num, value=col_name)
        c.fill = style_table_header
        c.font = font_table_header
        c.border = thin_border
        c.alignment = align_center
    row_idx += 1

    # Сводим историю в агрегат по позиции (изменение количества).
    # Если у позиции нет событий, считаем её "без изменений": первоначально == итог.
    lines = {}

    def _line_key(payload):
        return (
            payload.get('plant_id'),
            payload.get('size_id'),
            payload.get('field_id'),
            payload.get('year'),
        )

    for h in history_rows:
        payload = {}
        if h.snapshot_payload:
            try:
                payload = json.loads(h.snapshot_payload)
            except Exception:
                payload = {}
        if not payload:
            continue

        key = _line_key(payload)
        if key not in lines:
            lines[key] = {
                'plant_name': payload.get('plant_name') or '-',
                'size_name': payload.get('size_name') or '-',
                'field_name': payload.get('field_name') or '-',
                'year': payload.get('year'),
                'price': Decimal(str(payload.get('price') or 0)),
                'initial_qty': None,
                'delta_qty': 0,
                'final_qty': None,
            }

        line = lines[key]

        # Первичная загрузка при создании заказа должна попадать в "Первоначально",
        # а не в "Изменения".
        if h.action_type == 'initial_item':
            if line['initial_qty'] is None:
                line['initial_qty'] = int(h.after_quantity if h.after_quantity is not None else 0)
            if h.after_quantity is not None:
                line['final_qty'] = int(h.after_quantity)
        else:
            if line['initial_qty'] is None:
                line['initial_qty'] = int(h.before_quantity if h.before_quantity is not None else 0)
            line['delta_qty'] += int(h.delta_quantity or 0)
            if h.after_quantity is not None:
                line['final_qty'] = int(h.after_quantity)

        if payload.get('price') is not None:
            line['price'] = Decimal(str(payload.get('price')))

    for item in o.items:
        key = (item.plant_id, item.size_id, item.field_id, item.year)
        if key not in lines:
            lines[key] = {
                'plant_name': item.plant.name if item.plant else '-',
                'size_name': item.size.name if item.size else '-',
                'field_name': item.field.name if item.field else '-',
                'year': item.year,
                'price': Decimal(str(item.price or 0)),
                'initial_qty': int(item.quantity),
                'delta_qty': 0,
                'final_qty': int(item.quantity),
            }
        else:
            lines[key]['final_qty'] = int(item.quantity)
            lines[key]['price'] = Decimal(str(item.price or 0))
            if lines[key]['initial_qty'] is None:
                lines[key]['initial_qty'] = int(item.quantity) - int(lines[key]['delta_qty'])

    for data in lines.values():
        if data['initial_qty'] is None:
            data['initial_qty'] = 0
        if data['final_qty'] is None:
            data['final_qty'] = data['initial_qty']

    total_final_qty = 0
    total_final_sum = Decimal('0')
    for data in sorted(lines.values(), key=lambda x: (x['plant_name'], x['size_name'], x['field_name'], x['year'] or 0)):
        final_sum = data['price'] * Decimal(data['final_qty'])
        ws.cell(row=row_idx, column=1, value=data['plant_name']).border = thin_border
        ws.cell(row=row_idx, column=2, value=data['size_name']).border = thin_border
        c_price = ws.cell(row=row_idx, column=3, value=float(data['price']))
        c_price.border = thin_border
        c_price.number_format = '#,##0.00 "₽"'

        c_initial = ws.cell(row=row_idx, column=4, value=int(data['initial_qty']))
        c_initial.border = thin_border
        c_initial.alignment = align_center

        c_delta = ws.cell(row=row_idx, column=5, value=int(data['delta_qty']))
        c_delta.border = thin_border
        c_delta.alignment = align_center

        c_final = ws.cell(row=row_idx, column=6, value=int(data['final_qty']))
        c_final.border = thin_border
        c_final.alignment = align_center

        c_sum = ws.cell(row=row_idx, column=7, value=float(final_sum))
        c_sum.border = thin_border
        c_sum.number_format = '#,##0.00 "₽"'
        c_sum.font = Font(bold=True)

        total_final_qty += int(data['final_qty'])
        total_final_sum += final_sum
        row_idx += 1

    c_label = ws.cell(row=row_idx, column=5, value="ИТОГО:")
    c_label.font = font_total
    c_label.alignment = align_right
    c_label.border = border_total
    c_label.fill = style_total

    c_t_qty = ws.cell(row=row_idx, column=6, value=total_final_qty)
    c_t_qty.font = font_total
    c_t_qty.alignment = align_center
    c_t_qty.border = border_total
    c_t_qty.fill = style_total

    c_t_sum = ws.cell(row=row_idx, column=7, value=float(total_final_sum))
    c_t_sum.font = font_total
    c_t_sum.number_format = '#,##0.00 "₽"'
    c_t_sum.border = border_total
    c_t_sum.fill = style_total

    row_idx += 2
    debt = Decimal(str(o.total_sum)) - Decimal(str(o.paid_sum))
    summary_titles = ["Итоговая сумма заказа", "Оплачено", "Остаток"]
    summary_values = [float(o.total_sum), float(o.paid_sum), float(debt)]
    summary_cols = [(1, 2), (3, 4), (5, 7)]

    for idx, title in enumerate(summary_titles):
        start_col, end_col = summary_cols[idx]
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
        t_cell = ws.cell(row=row_idx, column=start_col, value=title)
        t_cell.fill = style_sub_header
        t_cell.font = font_sub_header
        t_cell.alignment = align_center
        t_cell.border = thin_border
        for c in range(start_col + 1, end_col + 1):
            ws.cell(row=row_idx, column=c).border = thin_border
    row_idx += 1

    for idx, value in enumerate(summary_values):
        start_col, end_col = summary_cols[idx]
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
        v_cell = ws.cell(row=row_idx, column=start_col, value=value)
        v_cell.font = Font(bold=True, size=12, color="1B5E20")
        v_cell.alignment = align_center
        v_cell.number_format = '#,##0.00 "₽"'
        v_cell.border = thin_border
        for c in range(start_col + 1, end_col + 1):
            ws.cell(row=row_idx, column=c).border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'История заказа №{o.id} от {msk_now().strftime("%d.%m.%Y")}.xlsx'
    return send_file(
        buf,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/orders/export_client_catalog')
@login_required
def export_client_catalog():
    rows = _get_client_catalog_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс для заказа"

    ws.merge_cells("A1:E1")
    ws["A1"] = "ПРАЙС-ЛИСТ ДЛЯ ЗАКАЗА"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Дата выгрузки: {msk_now().strftime('%d.%m.%Y %H:%M')}  |  Внесите количество в столбец 'Заказать, шт'"
    ws["A2"].font = Font(size=11, color="1B5E20")
    ws["A2"].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:E3")
    ws["A3"] = "Показаны только доступные позиции. Размеры 'Нетов.' и 'Саженцы' исключены."
    ws["A3"].font = Font(size=10, color="455A64", italic=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 20

    header_row = 5
    headers = ["Наименование", "Размер", "Цена, ₽", "Свободный остаток, шт", "Заказать, шт"]
    for col, title in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=title)
        c.font = Font(bold=True, color="000000")
        c.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    ws.row_dimensions[header_row].height = 30

    widths = {1: 40, 2: 14, 3: 13, 4: 20, 5: 16}
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    data_start = header_row + 1
    row_idx = data_start

    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    zebra_fill = PatternFill(start_color="F9FBE7", end_color="F9FBE7", fill_type="solid")
    qty_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    for i, item in enumerate(rows):
        ws.cell(row=row_idx, column=1, value=item["plant_name"])
        ws.cell(row=row_idx, column=2, value=item["size_name"])
        c_price = ws.cell(row=row_idx, column=3, value=float(item["price"] or 0))
        c_free = ws.cell(row=row_idx, column=4, value=int(item["free_qty"] or 0))
        c_order = ws.cell(row=row_idx, column=5, value=None)

        c_price.number_format = '#,##0.00'
        c_free.number_format = '#,##0'
        c_order.number_format = '#,##0'

        for col in range(1, 6):
            c = ws.cell(row=row_idx, column=col)
            c.border = thin
            c.font = Font(size=12)
            if col in (3, 4, 5):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

        if i % 2 == 1:
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = zebra_fill

        c_order.fill = qty_fill
        row_idx += 1

    data_end = max(row_idx - 1, data_start)

    qty_validation = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    qty_validation.error = "Введите целое число 0 или больше."
    qty_validation.errorTitle = "Некорректное количество"
    ws.add_data_validation(qty_validation)
    if data_end >= data_start:
        qty_validation.add(f"E{data_start}:E{data_end}")
        ws.auto_filter.ref = f"A{header_row}:E{data_end}"

    ws.freeze_panes = f"A{data_start}"
    ws.sheet_view.zoomScale = 120

    footer_row = data_end + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=5)
    ws.cell(row=footer_row, column=1, value="Подсказка: в столбце 'Заказать, шт' клиент заполняет нужное количество.").font = Font(
        italic=True, color="1B5E20"
    )
    ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'Client_Catalog_{msk_now().strftime("%Y-%m-%d_%H-%M")}.xlsx'
    log_action("Выгрузил клиентский каталог (Excel)")
    return send_file(
        buf,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )