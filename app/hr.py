import calendar
import io
from datetime import datetime, date, timedelta
from decimal import Decimal
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from sqlalchemy import func, or_, and_, extract
from sqlalchemy.orm import joinedload
from app.models import (
    db, Employee, SalaryRate, TimeLog, Expense, EmployeePayment, AppSetting, BudgetItem,
    ForeignEmployeeProfile, ForeignEmployeeDocument, PatentPeriod, PatentPayment, ActionLog
)
from app.utils import msk_today, msk_now, MONTH_NAMES, RATE_TYPES_LABELS, EMPLOYEE_ROLES, natural_key, log_action
import os
import requests
import time

bp = Blueprint('hr', __name__)

from app.telegram import send_message as _tg_send_msg, send_photo as _tg_send_photo, send_photo_album as _tg_send_photo_album

def send_tg_message(text):
    _tg_send_msg(text, chat_type="hr")

def send_tg_photo(photo_path, caption=""):
    _tg_send_photo(photo_path, caption, chat_type="hr")

def send_tg_photo_album(photo_paths):
    _tg_send_photo_album(photo_paths, chat_type="hr")
# ---------------------------------


def _format_hours_value(value):
    hours = float(value or 0)
    if hours.is_integer():
        return str(int(hours))
    return f"{hours:.1f}".rstrip('0').rstrip('.')


def _build_hours_details(logs):
    lines =[]
    total_people = 0

    for l in logs:
        if getattr(l, 'is_day_off', False):
            total_people += 1
            lines.append(f"• {l.employee.name}: Выходной (В)")
            continue

        h_norm = l.hours_norm or 0
        h_norm_over = l.hours_norm_over or 0
        h_spec = l.hours_spec or 0
        h_spec_over = l.hours_spec_over or 0
        total_h = h_norm + h_norm_over + h_spec + h_spec_over

        if total_h <= 0:
            continue

        total_people += 1
        parts =[]
        if h_norm > 0:
            parts.append(f"{_format_hours_value(h_norm)} об")
        if h_norm_over > 0:
            parts.append(f"{_format_hours_value(h_norm_over)} пер")
        if h_spec > 0:
            parts.append(f"{_format_hours_value(h_spec)} коп")
        if h_spec_over > 0:
            parts.append(f"{_format_hours_value(h_spec_over)} пер.коп")

        detail = " + ".join(parts) if parts else "0 об"
        lines.append(f"• {l.employee.name}: {_format_hours_value(total_h)}ч ({detail})")

    return lines, total_people


def _ensure_hr_access():
    if current_user.role not in ['admin', 'user2', 'executive', 'brigadier']:
        return False
    return True


def _safe_rel_path(path_abs):
    root = os.path.abspath(current_app.config['UPLOAD_FOLDER'])
    target = os.path.abspath(path_abs)
    if not target.startswith(root):
        raise ValueError("Недопустимый путь файла")
    return os.path.relpath(target, root).replace("\\", "/")


def _employee_storage_dir(employee):
    safe_name = secure_filename(employee.name or f"employee_{employee.id}") or f"employee_{employee.id}"
    base = os.path.join(current_app.config['UPLOAD_FOLDER'], 'data-personel', f"{employee.id}_{safe_name}")
    os.makedirs(base, exist_ok=True)
    return base


def _add_months(src_date, months):
    month = src_date.month - 1 + int(months)
    year = src_date.year + month // 12
    month = month % 12 + 1
    day = min(src_date.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)

@bp.route('/personnel', methods=['GET', 'POST'])
@login_required
def personnel():
    # Впустили Руководителя
    if current_user.role not in['admin', 'user2', 'brigadier', 'executive']:
        return redirect(url_for('main.index'))
        
    active_tab = request.args.get('tab', 'summary')
    
    if request.method == 'POST':
        # Руководитель может только смотреть
        if current_user.role == 'executive':
            flash('У вас права только на просмотр')
            return redirect(url_for('hr.personnel', tab=active_tab))
            
        action = request.form.get('action')
    current_date = msk_today()
    today_str = current_date.strftime('%Y-%m-%d')
    try: 
        selected_year = int(request.args.get('year', current_date.year))
        selected_month = int(request.args.get('month', current_date.month))
    except (ValueError, TypeError): 
        selected_year = current_date.year
        selected_month = current_date.month
        
    start_date = date(selected_year, selected_month, 1)
    last_day = calendar.monthrange(selected_year, selected_month)[1]
    end_date = date(selected_year, selected_month, last_day)
    
    filter_emp_ids =[int(x) for x in request.args.getlist('employee_id') if x]
    
    def get_setting_int(key): 
        s = AppSetting.query.get(key)
        return int(s.value) if s and s.value.isdigit() else 0

    def get_setting_int_list(key):
        s = AppSetting.query.get(key)
        if not s or not s.value:
            return []
        out = []
        for part in str(s.value).split(','):
            part = part.strip()
            if part.isdigit():
                out.append(int(part))
        return out

    official_item_id = get_setting_int('hr_official_item')
    unofficial_item_id = get_setting_int('hr_unofficial_item')
    other_salary_item_ids = [
        bid for bid in get_setting_int_list('hr_other_salary_items')
        if bid and bid != official_item_id and bid != unofficial_item_id
    ]
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        # --- ИНДИВИДУАЛЬНАЯ ВЫПЛАТА ЗП ---
        if action == 'pay_individual' and current_user.role in ['admin', 'user2']:
            emp_id = int(request.form.get('employee_id'))
            amount = Decimal(request.form.get('amount') or 0)
            b_item_id = int(request.form.get('budget_item_id'))
            p_type = request.form.get('payment_type')
            p_date = datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d').date()
            
            emp = Employee.query.get(emp_id)
            
            if amount > 0:
                ex = Expense(
                    date=p_date,
                    budget_item_id=b_item_id,
                    description=f"ЗП {emp.name} (за {MONTH_NAMES[selected_month]} {selected_year})",
                    amount=amount,
                    payment_type=p_type,
                    employee_id=emp.id,
                    target_month=selected_month,
                    target_year=selected_year
                )
                db.session.add(ex)
                db.session.commit()
                flash(f'Выплата {amount:,.0f} руб. для {emp.name} успешно проведена'.replace(',', ' '))
                log_action(f"Выдал индивидуальную ЗП: {emp.name} ({amount} руб.)")
            else:
                flash('Сумма выплаты должна быть больше нуля', 'danger')
                
            return redirect(url_for('hr.personnel', year=selected_year, month=selected_month, tab=active_tab))

        # --- НОВАЯ КНОПКА: КОТЛОВОЙ МЕТОД ЗП БРИГАДОЙ ---
        if action == 'bulk_pay_brigade' and current_user.role == 'admin':
            total_amount = Decimal(request.form.get('total_amount') or 0)
            brigadier_id = int(request.form.get('brigadier_id'))
            b_item_id = int(request.form.get('budget_item_id'))
            p_type = request.form.get('payment_type')
            p_date = datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d').date()
            
            all_emps = Employee.query.filter_by(is_active=True).all()
            rates_db = SalaryRate.query.filter_by(year=selected_year).all()
            r_map = {} 
            for r in rates_db:
                if r.role not in r_map: r_map[r.role] = {}
                r_map[r.role][r.rate_type] = r.rate_value
                
            t_logs = TimeLog.query.filter(TimeLog.date >= start_date, TimeLog.date <= end_date).all()
            h_map = {}
            for l in t_logs:
                eid = l.employee_id
                if eid not in h_map: h_map[eid] = {'norm': 0, 'norm_over': 0, 'spec': 0, 'spec_over': 0, 'day_offs':[], 'worked_days': 0}
                if getattr(l, 'is_day_off', False):
                    h_map[eid]['day_offs'].append(l.date.day)
                else:
                    hn = float(l.hours_norm or 0)
                    hno = float(l.hours_norm_over or 0)
                    hs = float(l.hours_spec or 0)
                    hso = float(l.hours_spec_over or 0)
                    
                    h_map[eid]['norm'] += hn
                    h_map[eid]['norm_over'] += hno
                    h_map[eid]['spec'] += hs
                    h_map[eid]['spec_over'] += hso
                    
                    if (hn + hno + hs + hso) > 1:
                        h_map[eid]['worked_days'] += 1
                    
            exp_agg_b = db.session.query(
                Expense.employee_id, Expense.budget_item_id, func.sum(Expense.amount)
            ).filter(
                Expense.employee_id.isnot(None),
                or_(
                    and_(Expense.target_month == selected_month, Expense.target_year == selected_year),
                    and_(Expense.target_month.is_(None), extract('month', Expense.date) == selected_month, extract('year', Expense.date) == selected_year)
                )
            ).group_by(Expense.employee_id, Expense.budget_item_id).all()
            e_map = {}
            for eid, bid, total in exp_agg_b:
                if eid not in e_map: e_map[eid] = {}
                e_map[eid][bid] = total or 0
                
            adj_agg_b = db.session.query(
                EmployeePayment.employee_id, func.sum(EmployeePayment.amount)
            ).filter(EmployeePayment.date >= start_date, EmployeePayment.date <= end_date).group_by(EmployeePayment.employee_id).all()
            a_map = {r[0]: r[1] or 0 for r in adj_agg_b}
            
            remaining_amount = total_amount
            
            for emp in all_emps:
                if emp.id == brigadier_id: continue
                if emp.role != 'worker': continue # <--- КОТЕЛ ТЕПЕРЬ ТРОГАЕТ ТОЛЬКО РАБОЧИХ
                    
                h = h_map.get(emp.id, {'norm': 0, 'norm_over': 0, 'spec': 0, 'spec_over': 0, 'day_offs':[], 'worked_days': 0})
                paid_do = min(3, len(h['day_offs']))
                
                emp_exp = e_map.get(emp.id, {})
                emp_off_id = emp.official_budget_item_id or official_item_id
                emp_unoff_id = emp.unofficial_budget_item_id or unofficial_item_id
                p_off = emp_exp.get(emp_off_id, 0)
                p_unoff = emp_exp.get(emp_unoff_id, 0) if emp_unoff_id != emp_off_id else 0
                _ac = {emp_off_id, emp_unoff_id}
                p_other = Decimal(0)
                for bid in other_salary_item_ids:
                    if bid in _ac:
                        continue
                    p_other += Decimal(str(emp_exp.get(bid, 0) or 0))
                o_adj = Decimal(str(a_map.get(emp.id, 0) or 0))

                e_rates = r_map.get(emp.role, {})
                earned = Decimal(0)
                if emp.is_salary:
                    earned = emp.fixed_salary
                elif emp.role == 'mechanic':
                    earned = Decimal(str(h['worked_days'])) * e_rates.get('norm', Decimal(0))
                else:
                    earned += Decimal(str(h['norm'])) * e_rates.get('norm', Decimal(0))
                    earned += Decimal(str(h['norm_over'])) * e_rates.get('norm_over', Decimal(0))
                    earned += Decimal(str(h['spec'])) * e_rates.get('spec', Decimal(0))
                    earned += Decimal(str(h['spec_over'])) * e_rates.get('spec_over', Decimal(0))
                    earned += Decimal(str(paid_do * 10)) * e_rates.get('norm', Decimal(0))

                debt = earned - (Decimal(str(p_off)) + Decimal(str(p_unoff)) + p_other + o_adj)
                
                if debt > 0:
                    pay_val = min(debt, remaining_amount)
                    if pay_val > 0:
                        ex = Expense(
                            date=p_date,
                            budget_item_id=b_item_id,
                            description=f"ЗП {emp.name} (из котла {MONTH_NAMES[selected_month]} {selected_year})",
                            amount=pay_val,
                            payment_type=p_type,
                            employee_id=emp.id,
                            target_month=selected_month,
                            target_year=selected_year
                        )
                        db.session.add(ex)
                        remaining_amount -= pay_val
            
            if remaining_amount > 0:
                ex_brig = Expense(
                    date=p_date,
                    budget_item_id=b_item_id,
                    description=f"Остаток котла бригадиру (за {MONTH_NAMES[selected_month]} {selected_year})",
                    amount=remaining_amount,
                    payment_type=p_type,
                    employee_id=brigadier_id,
                    target_month=selected_month,
                    target_year=selected_year
                )
                db.session.add(ex_brig)
                
            db.session.commit()
            flash(f'Котловая выплата распределена! Остаток бригадиру: {remaining_amount:,.0f} руб.'.replace(',', ' '))
            log_action(f"Котловая выплата ЗП: {total_amount} (Ост. бригадиру: {remaining_amount})")
            return redirect(url_for('hr.personnel', year=selected_year, month=selected_month, tab=active_tab))
        
        # --- ОСТАЛЬНЫЕ ДЕЙСТВИЯ ---
        if action in['save_employee', 'save_rates', 'save_settings', 'delete_employee', 'add_adjustment'] and current_user.role in ['user2', 'brigadier']:
            flash('Доступ запрещен, только просмотр')
            return redirect(url_for('hr.personnel', year=selected_year, month=selected_month, tab=active_tab))
            
        if action == 'save_settings' and current_user.role == 'admin':
            s1 = AppSetting.query.get('hr_official_item') or AppSetting(key='hr_official_item')
            s1.value = request.form.get('official_item_id')
            db.session.add(s1)
            s2 = AppSetting.query.get('hr_unofficial_item') or AppSetting(key='hr_unofficial_item')
            s2.value = request.form.get('unofficial_item_id')
            db.session.add(s2)

            db.session.commit()
            flash('Настройки сохранены')

        elif action == 'save_employee_budget_items' and current_user.role == 'admin':
            # Массовое сохранение персональных статей ЗП для сотрудников
            changed = 0
            for emp in Employee.query.all():
                off_key = f'emp_off_{emp.id}'
                unoff_key = f'emp_unoff_{emp.id}'
                if off_key not in request.form and unoff_key not in request.form:
                    continue
                raw_off = (request.form.get(off_key) or '').strip()
                raw_unoff = (request.form.get(unoff_key) or '').strip()
                new_off = int(raw_off) if raw_off.isdigit() and int(raw_off) > 0 else None
                new_unoff = int(raw_unoff) if raw_unoff.isdigit() and int(raw_unoff) > 0 else None
                if emp.official_budget_item_id != new_off or emp.unofficial_budget_item_id != new_unoff:
                    emp.official_budget_item_id = new_off
                    emp.unofficial_budget_item_id = new_unoff
                    changed += 1
            if changed:
                db.session.commit()
            flash(f'Персональные статьи ЗП сохранены: изменено {changed} сотрудник(ов)')
            
        elif action == 'save_employee' and current_user.role == 'admin':
            e_id = request.form.get('id')
            role = request.form.get('role')
            if e_id: 
                emp = Employee.query.get(e_id)
                emp.name = request.form.get('name')
                emp.is_salary = bool(request.form.get('is_salary'))
                emp.fixed_salary = float(request.form.get('fixed_salary') or 0)
                emp.is_active = bool(request.form.get('is_active'))
                emp.role = role
            else: 
                db.session.add(Employee(name=request.form.get('name'), is_salary=bool(request.form.get('is_salary')), fixed_salary=float(request.form.get('fixed_salary') or 0), is_active=bool(request.form.get('is_active')), role=role))
            db.session.commit()
            flash('Сотрудник сохранен')

        elif action == 'delete_employee' and current_user.role == 'admin':
            emp_id = request.form.get('employee_id')
            emp = Employee.query.get(emp_id) if emp_id else None
            if emp:
                try:
                    TimeLog.query.filter_by(employee_id=emp.id).delete()
                    EmployeePayment.query.filter_by(employee_id=emp.id).delete()
                    Expense.query.filter_by(employee_id=emp.id).update({'employee_id': None})
                    db.session.delete(emp)
                    db.session.commit()
                    flash('Сотрудник удален')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Не удалось удалить сотрудника: {e}')
            else:
                flash('Сотрудник не найден')

        elif action == 'save_rates' and current_user.role == 'admin':
            year = int(request.form.get('year'))
            for role_code in EMPLOYEE_ROLES.keys():
                for r_type in RATE_TYPES_LABELS.keys():
                    rate_obj = SalaryRate.query.filter_by(year=year, role=role_code, rate_type=r_type).first()
                    if not rate_obj: 
                        rate_obj = SalaryRate(year=year, role=role_code, rate_type=r_type)
                        db.session.add(rate_obj)
                    rate_obj.rate_value = float(request.form.get(f'rate_{role_code}_{r_type}') or 0)
            db.session.commit()
            flash('Ставки обновлены')
            
        elif action == 'add_hours':
            emp_id = int(request.form.get('employee_id'))
            log_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
            log = TimeLog.query.filter_by(employee_id=emp_id, date=log_date).first()
            if not log: 
                log = TimeLog(employee_id=emp_id, date=log_date)
                db.session.add(log)
                
            is_day_off = request.form.get('is_day_off') == 'on'
            log.is_day_off = is_day_off
            
            if is_day_off:
                log.hours_norm = 0
                log.hours_norm_over = 0
                log.hours_spec = 0
                log.hours_spec_over = 0
            else:
                log.hours_norm = float(request.form.get('hours_norm') or 0)
                log.hours_norm_over = float(request.form.get('hours_norm_over') or 0)
                log.hours_spec = float(request.form.get('hours_spec') or 0)
                log.hours_spec_over = float(request.form.get('hours_spec_over') or 0)
                
            db.session.commit()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest': 
                return jsonify({'status':'ok'})
            flash('Часы внесены')
            
        elif action == 'add_adjustment':
            db.session.add(EmployeePayment(employee_id=int(request.form.get('employee_id')), date=datetime.strptime(request.form.get('date'), '%Y-%m-%d').date(), amount=float(request.form.get('amount') or 0), payment_type='other_manual', comment=request.form.get('comment')))
            db.session.commit()
            flash('Корректировка добавлена')
            
        elif action == 'send_daily_report':
            report_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
            
            # --- ЗАЩИТА ОТ ДУБЛИКАТОВ (5 минут) ---
            cutoff_time = msk_now() - timedelta(minutes=5)
            recent_log = ActionLog.query.filter(
                ActionLog.user_id == current_user.id,
                ActionLog.action == f"Отправил ежедневный отчет за {report_date.strftime('%Y-%m-%d')}",
                ActionLog.date >= cutoff_time
            ).first()
            
            if recent_log:
                # Если уже отправил недавно, просто перекидываем на страницу успеха без дублирования
                return redirect(url_for('hr.report_success'))
            # --------------------------------------
            
            daily_report_text = request.form.get('daily_report_text', '').strip()
            photos = request.files.getlist('photos')
            
            # Сохранение фотографий
            saved_photos =[]
            if photos and any(p.filename for p in photos):
                date_str = report_date.strftime('%Y-%m-%d')
                upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photo', date_str)
                if not os.path.exists(upload_dir):
                    os.makedirs(upload_dir, exist_ok=True)
                for p in photos:
                    if p.filename:
                        fname = secure_filename(p.filename)
                        base, ext = os.path.splitext(fname)
                        unique_fname = f"{base}_{int(time.time()*1000)}{ext}"
                        fpath = os.path.join(upload_dir, unique_fname)
                        p.save(fpath)
                        saved_photos.append(fpath)
            
            # Отправка в ТГ
            if saved_photos:
                send_tg_photo_album(saved_photos)
                time.sleep(1) 
                
            if daily_report_text:
                import html
                safe_text = html.escape(daily_report_text)
                msg_report = f"📝 <b>Отчет о работах за {report_date.strftime('%d.%m.%Y')}</b>\n👤 Внес(ла): {current_user.username}\n\n{safe_text}"
                send_tg_message(msg_report)
                time.sleep(1)

            logs = TimeLog.query.filter_by(date=report_date).join(Employee).order_by(Employee.name).all()
            lines, total_people = _build_hours_details(logs)
            
            if lines or daily_report_text or saved_photos:
                if lines:
                    msg = f"⏱ <b>Отчет по часам за {report_date.strftime('%d.%m.%Y')}</b>\n👤 Внес(ла): {current_user.username}\n\n"
                    msg += f"👥 <b>ДЕТАЛИЗАЦИЯ ({total_people} чел)</b>:\n"
                    msg += "\n".join(lines)
                    send_tg_message(msg)
                
                # Записываем действие в лог, чтобы сработала защита от дублей
                log_action(f"Отправил ежедневный отчет за {report_date.strftime('%Y-%m-%d')}")
                
                # Перенаправляем на зеленую страницу успеха
                return redirect(url_for('hr.report_success'))
            else:
                flash('Нет данных для отправки отчета.', 'warning')

        return redirect(url_for('hr.personnel', year=selected_year, month=selected_month, employee_id=filter_emp_ids, tab=active_tab))
        
    emp_query = Employee.query
    if filter_emp_ids: emp_query = emp_query.filter(Employee.id.in_(filter_emp_ids))
    
    # ПРЯЧЕМ МЕНЕДЖЕРОВ ОТ ВСЕХ, КРОМЕ АДМИНА И РУКОВОДИТЕЛЯ
    if current_user.role not in ['admin', 'executive']:
        emp_query = emp_query.filter(Employee.role != 'manager')
        
    employees = emp_query.order_by(Employee.name).all()
    
    # Скрываем их и из выпадающего списка фильтра
    if current_user.role not in ['admin', 'executive']:
        all_employees_list = Employee.query.filter(Employee.role != 'manager').order_by(Employee.name).all()
    else:
        all_employees_list = Employee.query.order_by(Employee.name).all()
    
    rates_db = SalaryRate.query.filter_by(year=selected_year).all()
    rates_map = {} 
    for r in rates_db:
        if r.role not in rates_map: rates_map[r.role] = {}
        rates_map[r.role][r.rate_type] = r.rate_value
        
    # ЕДИНЫЙ ЗАПРОС ВМЕСТО АГРЕГАЦИИ В БД: Обрабатываем все логи за месяц в Python
    all_logs = TimeLog.query.filter(TimeLog.date >= start_date, TimeLog.date <= end_date).all()
    log_map = {}
    hours_map = {}
    
    for l in all_logs:
        eid = l.employee_id
        if eid not in log_map: log_map[eid] = {}
        if eid not in hours_map: hours_map[eid] = {'norm': 0.0, 'norm_over': 0.0, 'spec': 0.0, 'spec_over': 0.0, 'day_offs':[], 'worked_days': 0}
        
        log_map[eid][l.date.day] = l
        
        if getattr(l, 'is_day_off', False):
            hours_map[eid]['day_offs'].append(l.date.day)
        else:
            hn = float(l.hours_norm or 0)
            hno = float(l.hours_norm_over or 0)
            hs = float(l.hours_spec or 0)
            hso = float(l.hours_spec_over or 0)
            hours_map[eid]['norm'] += hn
            hours_map[eid]['norm_over'] += hno
            hours_map[eid]['spec'] += hs
            hours_map[eid]['spec_over'] += hso
            if (hn + hno + hs + hso) > 1:
                hours_map[eid]['worked_days'] += 1
            
    for eid in hours_map:
        hours_map[eid]['day_offs'].sort()

    exp_agg_b = db.session.query(
        Expense.employee_id, Expense.budget_item_id, func.sum(Expense.amount)
    ).filter(
        Expense.employee_id.isnot(None),
        or_(
            and_(Expense.target_month == selected_month, Expense.target_year == selected_year),
            and_(Expense.target_month.is_(None), extract('month', Expense.date) == selected_month, extract('year', Expense.date) == selected_year)
        )
    ).group_by(Expense.employee_id, Expense.budget_item_id).all()
    exp_map = {}
    for eid, bid, total in exp_agg_b:
        if eid not in exp_map: exp_map[eid] = {}
        exp_map[eid][bid] = total or 0

    adj_agg = db.session.query(
        EmployeePayment.employee_id, func.sum(EmployeePayment.amount)
    ).filter(EmployeePayment.date >= start_date, EmployeePayment.date <= end_date).group_by(EmployeePayment.employee_id).all()
    adj_map = {r[0]: r[1] or 0 for r in adj_agg}

    def _calc_payroll_row(emp):
        h = hours_map.get(emp.id, {'norm': 0.0, 'norm_over': 0.0, 'spec': 0.0, 'spec_over': 0.0, 'day_offs':[], 'worked_days': 0})
        paid_day_offs = min(3, len(h['day_offs']))
            
        emp_exp = exp_map.get(emp.id, {})

        # Персональные статьи ЗП (если не заданы — используется глобальная настройка)
        emp_off_id = emp.official_budget_item_id or official_item_id
        emp_unoff_id = emp.unofficial_budget_item_id or unofficial_item_id

        paid_official = emp_exp.get(emp_off_id, 0)
        paid_unofficial = emp_exp.get(emp_unoff_id, 0) if emp_unoff_id != emp_off_id else 0

        # Прочие выплаты сотруднику по дополнительным статьям (глобальный CSV, редкие случаи)
        _already_counted = {emp_off_id, emp_unoff_id}
        paid_other_exp = Decimal(0)
        for bid in other_salary_item_ids:
            if bid in _already_counted:
                continue
            paid_other_exp += Decimal(str(emp_exp.get(bid, 0) or 0))

        # Ручные корректировки/выплаты (EmployeePayment) трактуем как выплату:
        # + => выдали на руки (уменьшает долг), - => вернули/удержали (увеличивает долг).
        manual_adjust = Decimal(str(adj_map.get(emp.id, 0) or 0))
        other_adjust = paid_other_exp + manual_adjust

        emp_rates = rates_map.get(emp.role, {})
        earned_base = Decimal(0)
        if emp.is_salary: 
            earned_base = emp.fixed_salary
        elif emp.role == 'mechanic':
            earned_base = Decimal(str(h['worked_days'])) * emp_rates.get('norm', Decimal(0))
        else:
            earned_base += Decimal(str(h['norm'])) * emp_rates.get('norm', Decimal(0))
            earned_base += Decimal(str(h['norm_over'])) * emp_rates.get('norm_over', Decimal(0))
            earned_base += Decimal(str(h['spec'])) * emp_rates.get('spec', Decimal(0))
            earned_base += Decimal(str(h['spec_over'])) * emp_rates.get('spec_over', Decimal(0))
            earned_base += Decimal(str(paid_day_offs * 10)) * emp_rates.get('norm', Decimal(0))

        total_earned = Decimal(str(earned_base))
        paid_official = Decimal(str(paid_official or 0))
        paid_unofficial = Decimal(str(paid_unofficial or 0))
        balance = total_earned - (paid_official + paid_unofficial + other_adjust)

        return {
            'emp': emp, 'hours': h, 'paid_day_offs': paid_day_offs,
            'earned': total_earned,
            'paid_official': paid_official,
            'paid_unofficial': paid_unofficial,
            'paid_other_exp': paid_other_exp,
            'manual_adjust': manual_adjust,
            'other_adjust': other_adjust,
            'balance': balance,
        }

    data = []
    total_summary = {'earned': Decimal(0), 'official': Decimal(0), 'unofficial': Decimal(0), 'other': Decimal(0), 'paid_total': Decimal(0), 'balance': Decimal(0)}
    for emp in employees:
        row = _calc_payroll_row(emp)
        data.append(row)
        total_summary['earned'] += row['earned']
        total_summary['official'] += row['paid_official']
        total_summary['unofficial'] += row['paid_unofficial']
        total_summary['other'] += row['other_adjust']
        total_summary['paid_total'] += (row['paid_official'] + row['paid_unofficial'] + row['other_adjust'])
        total_summary['balance'] += row['balance']

    # Итоги для инфо-плашек считаются по всем сотрудникам без исключений
    all_summary = {'earned': Decimal(0), 'official': Decimal(0), 'unofficial': Decimal(0), 'other': Decimal(0), 'paid_total': Decimal(0), 'balance': Decimal(0)}
    all_summary_employees = Employee.query.order_by(Employee.name).all()
    for emp in all_summary_employees:
        row = _calc_payroll_row(emp)
        all_summary['earned'] += row['earned']
        all_summary['official'] += row['paid_official']
        all_summary['unofficial'] += row['paid_unofficial']
        all_summary['other'] += row['other_adjust']
        all_summary['paid_total'] += (row['paid_official'] + row['paid_unofficial'] + row['other_adjust'])
        all_summary['balance'] += row['balance']

    timesheet_data =[]; days_in_month = list(range(1, last_day + 1))
        
    for emp in employees:
        row = {'emp': emp, 'days': {}}
        h = hours_map.get(emp.id, {'day_offs': []})
        paid_days = h['day_offs'][:3]
        
        for d in days_in_month:
            l = log_map.get(emp.id, {}).get(d)
            is_day_off = getattr(l, 'is_day_off', False) if l else False
            is_paid = is_day_off and (d in paid_days)
            total_h = (l.hours_norm + l.hours_norm_over + l.hours_spec + l.hours_spec_over) if l else 0
            row['days'][d] = {'total': total_h, 'obj': l, 'is_day_off': is_day_off, 'is_paid': is_paid}
        timesheet_data.append(row)
            
    if request.args.get('export') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = f"Кадры {MONTH_NAMES[selected_month]} {selected_year}"
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_font = Font(bold=True)
        align_center = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        title_cell = ws.cell(row=1, column=1, value=f"Кадровый отчет за {MONTH_NAMES[selected_month]} {selected_year}")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        headers =["Сотрудник", "Роль", "Обыч. ч", "О.Пер. ч", "Спец. ч", "С.Пер. ч", "Вых. опл.", "Заработано", "Выплачено офиц.", "Выплачено неофиц.", "Прочее (выпл.+корр.)", "Остаток"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border
            ws.column_dimensions[chr(64 + col)].width = 15

        row_idx = 3
        for r in data:
            ws.cell(row=row_idx, column=1, value=r['emp'].name).border = border
            ws.cell(row=row_idx, column=2, value=EMPLOYEE_ROLES.get(r['emp'].role, r['emp'].role)).border = border
            ws.cell(row=row_idx, column=3, value=float(r['hours']['norm'])).border = border
            ws.cell(row=row_idx, column=4, value=float(r['hours']['norm_over'])).border = border
            ws.cell(row=row_idx, column=5, value=float(r['hours']['spec'])).border = border
            ws.cell(row=row_idx, column=6, value=float(r['hours']['spec_over'])).border = border
            ws.cell(row=row_idx, column=7, value=int(r['paid_day_offs'])).border = border
            ws.cell(row=row_idx, column=8, value=float(r['earned'])).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=8).border = border
            ws.cell(row=row_idx, column=9, value=float(r['paid_official'])).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=9).border = border
            ws.cell(row=row_idx, column=10, value=float(r['paid_unofficial'])).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=10).border = border
            ws.cell(row=row_idx, column=11, value=float(r['other_adjust'])).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=11).border = border
            ws.cell(row=row_idx, column=12, value=float(r['balance'])).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=12).border = border
            row_idx += 1

        ws.cell(row=row_idx, column=7, value="Итого:").font = Font(bold=True)
        ws.cell(row=row_idx, column=8, value=float(total_summary['earned'])).font = Font(bold=True)
        ws.cell(row=row_idx, column=9, value=float(total_summary['official'])).font = Font(bold=True)
        ws.cell(row=row_idx, column=10, value=float(total_summary['unofficial'])).font = Font(bold=True)
        ws.cell(row=row_idx, column=11, value=float(total_summary['other'])).font = Font(bold=True)
        ws.cell(row=row_idx, column=12, value=float(total_summary['balance'])).font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, download_name=f'hr_report_{selected_year}_{selected_month}.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('hr/personnel.html', 
                           active_tab=active_tab, 
                           data=data, 
                           summary=total_summary, 
                           summary_all=all_summary,
                           timesheet_data=timesheet_data, 
                           days_in_month=days_in_month, 
                           year=selected_year, month=selected_month, 
                           employees=all_employees_list, 
                           filter_emp_ids=filter_emp_ids, 
                           rates_map=rates_map, 
                           month_names=MONTH_NAMES, 
                           rate_labels=RATE_TYPES_LABELS, 
                           roles=EMPLOYEE_ROLES, 
                           budget_items=BudgetItem.query.all(), 
                           settings={'off': official_item_id, 'unoff': unofficial_item_id, 'others': other_salary_item_ids},
                           today_str=today_str)

@bp.route('/personnel/export')
@login_required
def personnel_export(): return redirect(url_for('hr.personnel'))


@bp.route('/personnel/foreign', methods=['GET'])
@login_required
def foreign_employees():
    if not _ensure_hr_access():
        return redirect(url_for('main.index'))

    employees = Employee.query.options(
        joinedload(Employee.foreign_profile)
    ).filter_by(is_active=True).order_by(Employee.name).all()

    # Pre-load current patent periods for all employees in one query
    current_periods = PatentPeriod.query.filter_by(is_current=True).order_by(PatentPeriod.end_date.desc()).all()
    period_map = {}
    for p in current_periods:
        if p.employee_id not in period_map:
            period_map[p.employee_id] = p

    rows = []
    today = msk_today()
    for emp in employees:
        profile = emp.foreign_profile
        period = period_map.get(emp.id)
        days_left = None
        if period and period.end_date:
            days_left = (period.end_date - today).days
        rows.append({
            'employee': emp,
            'profile': profile,
            'period': period,
            'days_left': days_left
        })
    rows.sort(key=lambda x: natural_key(x['employee'].name))
    return render_template('hr/foreign_employees.html', rows=rows, today=today)


@bp.route('/personnel/foreign/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def foreign_employee_card(employee_id):
    if not _ensure_hr_access():
        return redirect(url_for('main.index'))

    employee = Employee.query.get_or_404(employee_id)
    profile = ForeignEmployeeProfile.query.filter_by(employee_id=employee.id).first()

    if request.method == 'POST':
        action = request.form.get('action')
        try:
            if action == 'save_profile':
                if not profile:
                    profile = ForeignEmployeeProfile(employee_id=employee.id, full_name=employee.name)
                    db.session.add(profile)
                profile.full_name = request.form.get('full_name', '').strip() or employee.name
                profile.phone = request.form.get('phone', '').strip()
                profile.citizenship = request.form.get('citizenship', '').strip()
                dob_raw = request.form.get('date_of_birth')
                profile.date_of_birth = datetime.strptime(dob_raw, '%Y-%m-%d').date() if dob_raw else None
                profile.passport_number = request.form.get('passport_number', '').strip()
                profile.passport_issued_by = request.form.get('passport_issued_by', '').strip()
                profile.migration_card_number = request.form.get('migration_card_number', '').strip()
                profile.registration_address = request.form.get('registration_address', '').strip()
                profile.inn = request.form.get('inn', '').strip()
                profile.snils = request.form.get('snils', '').strip()
                profile.notes = request.form.get('notes', '').strip()
                db.session.commit()
                flash('Карточка сотрудника сохранена')

            elif action == 'upload_document':
                category = (request.form.get('category') or 'other').strip()
                title = (request.form.get('title') or '').strip()
                file = request.files.get('document_file')
                if not file or not file.filename:
                    raise ValueError('Выберите файл для загрузки')
                allowed_ext = {'pdf', 'jpg', 'jpeg', 'png', 'webp', 'doc', 'docx'}
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                if ext not in allowed_ext:
                    raise ValueError('Недопустимый формат файла')
                base_dir = _employee_storage_dir(employee)
                docs_dir = os.path.join(base_dir, 'documents', category)
                os.makedirs(docs_dir, exist_ok=True)
                safe_name = secure_filename(file.filename)
                unique_name = f"{int(msk_now().timestamp())}_{safe_name}"
                abs_path = os.path.join(docs_dir, unique_name)
                file.save(abs_path)
                rel_path = _safe_rel_path(abs_path)
                db.session.add(ForeignEmployeeDocument(
                    employee_id=employee.id,
                    category=category,
                    title=title,
                    original_name=file.filename,
                    stored_name=unique_name,
                    file_rel_path=rel_path,
                    uploaded_by_user_id=current_user.id,
                    size_bytes=os.path.getsize(abs_path)
                ))
                db.session.commit()
                flash('Документ загружен')

            elif action == 'delete_document':
                doc_id = int(request.form.get('document_id'))
                doc = ForeignEmployeeDocument.query.get(doc_id)
                if not doc or doc.employee_id != employee.id:
                    raise ValueError('Документ не найден')
                abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], doc.file_rel_path)
                if os.path.exists(abs_path):
                    os.remove(abs_path)
                db.session.delete(doc)
                db.session.commit()
                flash('Документ удален')

            elif action == 'create_patent_period':
                start_raw = request.form.get('patent_start_date')
                end_raw = request.form.get('patent_end_date')
                months_mode = int(request.form.get('months_mode') or 0)
                if not start_raw:
                    raise ValueError('Укажите дату начала патента')
                start_date = datetime.strptime(start_raw, '%Y-%m-%d').date()
                if months_mode in [1, 2, 3]:
                    end_date = _add_months(start_date, months_mode)
                else:
                    if not end_raw:
                        raise ValueError('Укажите дату окончания патента')
                    end_date = datetime.strptime(end_raw, '%Y-%m-%d').date()
                if end_date < start_date:
                    raise ValueError('Дата окончания не может быть раньше даты начала')

                PatentPeriod.query.filter_by(employee_id=employee.id, is_current=True).update(
                    {PatentPeriod.is_current: False, PatentPeriod.status: 'archived'},
                    synchronize_session=False
                )
                db.session.add(PatentPeriod(
                    employee_id=employee.id,
                    start_date=start_date,
                    end_date=end_date,
                    status='active',
                    is_current=True,
                    created_by_user_id=current_user.id
                ))
                db.session.commit()
                flash('Период патента создан')

            elif action == 'add_patent_payment':
                period_id = int(request.form.get('period_id'))
                period = PatentPeriod.query.get(period_id)
                if not period or period.employee_id != employee.id:
                    raise ValueError('Период патента не найден')
                payment_date_raw = request.form.get('payment_date')
                if not payment_date_raw:
                    raise ValueError('Укажите дату оплаты')
                payment_date = datetime.strptime(payment_date_raw, '%Y-%m-%d').date()
                months_paid = int(request.form.get('months_paid') or 1)
                if months_paid not in [1, 2, 3]:
                    raise ValueError('Период оплаты может быть только 1, 2 или 3 месяца')
                amount = float(request.form.get('amount') or 0)
                if amount <= 0:
                    raise ValueError('Сумма оплаты должна быть больше 0')

                next_base_date = period.end_date if period.end_date and period.end_date >= payment_date else payment_date
                new_end_date = _add_months(next_base_date, months_paid)

                check_rel_path = None
                check_original_name = None
                check_file = request.files.get('check_file')
                if check_file and check_file.filename:
                    allowed_ext = {'pdf', 'jpg', 'jpeg', 'png', 'webp'}
                    ext = check_file.filename.rsplit('.', 1)[1].lower() if '.' in check_file.filename else ''
                    if ext not in allowed_ext:
                        raise ValueError('Недопустимый формат чека')
                    base_dir = _employee_storage_dir(employee)
                    checks_dir = os.path.join(base_dir, 'patent_checks')
                    os.makedirs(checks_dir, exist_ok=True)
                    safe_name = secure_filename(check_file.filename)
                    unique_name = f"check_{int(msk_now().timestamp())}_{safe_name}"
                    check_abs = os.path.join(checks_dir, unique_name)
                    check_file.save(check_abs)
                    check_rel_path = _safe_rel_path(check_abs)
                    check_original_name = check_file.filename

                payment = PatentPayment(
                    employee_id=employee.id,
                    patent_period_id=period.id,
                    payment_date=payment_date,
                    months_paid=months_paid,
                    amount=amount,
                    period_end_after_payment=new_end_date,
                    check_file_rel_path=check_rel_path,
                    check_original_name=check_original_name,
                    comment=(request.form.get('comment') or '').strip(),
                    created_by_user_id=current_user.id
                )
                period.end_date = new_end_date
                db.session.add(payment)
                db.session.commit()
                flash('Оплата патента сохранена')
            else:
                flash('Неизвестное действие')
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка: {e}')
        return redirect(url_for('hr.foreign_employee_card', employee_id=employee.id))

    profile = ForeignEmployeeProfile.query.filter_by(employee_id=employee.id).first()
    docs = ForeignEmployeeDocument.query.filter_by(employee_id=employee.id).order_by(ForeignEmployeeDocument.uploaded_at.desc()).all()
    periods = PatentPeriod.query.filter_by(employee_id=employee.id).order_by(PatentPeriod.start_date.desc()).all()
    current_period = next((p for p in periods if p.is_current), periods[0] if periods else None)
    payments = []
    if current_period:
        payments = PatentPayment.query.filter_by(patent_period_id=current_period.id).order_by(PatentPayment.payment_date.desc(), PatentPayment.id.desc()).all()
    return render_template(
        'hr/foreign_employee_card.html',
        employee=employee,
        profile=profile,
        docs=docs,
        periods=periods,
        current_period=current_period,
        payments=payments
    )


@bp.route('/personnel/foreign/<int:employee_id>/document/<int:doc_id>/download', methods=['GET'])
@login_required
def foreign_employee_document_download(employee_id, doc_id):
    if not _ensure_hr_access():
        return redirect(url_for('main.index'))
    doc = ForeignEmployeeDocument.query.get_or_404(doc_id)
    if doc.employee_id != employee_id:
        return redirect(url_for('main.index'))
    abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], doc.file_rel_path)
    if not os.path.exists(abs_path):
        flash('Файл не найден')
        return redirect(url_for('hr.foreign_employee_card', employee_id=employee_id))
    return send_file(abs_path, as_attachment=True, download_name=doc.original_name)


@bp.route('/personnel/foreign/<int:employee_id>/payment/<int:payment_id>/check', methods=['GET'])
@login_required
def foreign_employee_payment_check_download(employee_id, payment_id):
    if not _ensure_hr_access():
        return redirect(url_for('main.index'))
    payment = PatentPayment.query.get_or_404(payment_id)
    if payment.employee_id != employee_id or not payment.check_file_rel_path:
        flash('Чек не найден')
        return redirect(url_for('hr.foreign_employee_card', employee_id=employee_id))
    abs_path = os.path.join(current_app.config['UPLOAD_FOLDER'], payment.check_file_rel_path)
    if not os.path.exists(abs_path):
        flash('Файл чека не найден')
        return redirect(url_for('hr.foreign_employee_card', employee_id=employee_id))
    filename = payment.check_original_name or os.path.basename(abs_path)
    return send_file(abs_path, as_attachment=True, download_name=filename)


@bp.route('/patents/reminders/run', methods=['GET', 'POST'])
def run_patent_reminders():
    token = request.headers.get('X-Reminder-Token') or request.args.get('token')
    expected = os.environ.get('PATENT_REMINDER_TOKEN')
    if not expected or token != expected:
        return jsonify({'status': 'forbidden'}), 403
    from app.patent_reminders import run_patent_reminders_job
    sent_count, message = run_patent_reminders_job()
    return jsonify({'status': 'ok', 'sent': sent_count, 'message': message})

@bp.route('/personnel/report_success')
@login_required
def report_success():
    """Страница успешной отправки отчета (предотвращает повторную отправку при перезагрузке)"""
    if current_user.role not in['admin', 'user2', 'brigadier', 'executive']:
        return redirect(url_for('main.index'))
    return render_template('hr/report_success.html')