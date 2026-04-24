import io
import re
import calendar
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app
from flask_login import login_required, current_user
from sqlalchemy import func, and_, extract
from sqlalchemy.orm import joinedload
from openpyxl import load_workbook, Workbook
from app.models import db, Order, OrderItem, Client, CompetitorSnapshot, CompetitorRow, Plant, Size, StockBalance, Field
from app.utils import msk_now, MONTH_NAMES, log_action
from app.services import calculate_cost_data

bp = Blueprint('crm', __name__)

RUSSIAN_REGIONS_LIST = [
    "Все регионы РФ", "Москва и Московская обл.", "Санкт-Петербург и Ленинградская обл.",
    "Тульская обл.", "Воронежская обл.", "Калужская обл.", "Тверская обл.", "Рязанская обл.",
    "Ярославская обл.", "Владимирская обл.", "Ивановская обл.", "Костромская обл.", 
    "Смоленская обл.", "Брянская обл.", "Орловская обл.", "Липецкая обл.", "Тамбовская обл.",
    "Курская обл.", "Белгородская обл.", "Краснодарский край", "Ростовская обл.",
    "Нижегородская обл.", "Республика Татарстан", "Свердловская обл."
]

@bp.route('/crm/client_analytics')
@login_required
def crm_client_analytics():
    if current_user.role not in ['admin', 'executive']:
        return redirect(url_for('main.index'))

    # Фильтры
    client_id = request.args.get('client_id', type=int)
    year_start = request.args.get('year_start', type=int, default=msk_now().year)
    year_end = request.args.get('year_end', type=int, default=msk_now().year)
    
    start_date = datetime(year_start, 1, 1)
    end_date = datetime(year_end, 12, 31, 23, 59, 59)

    clients = Client.query.order_by(Client.name).all()
    
    stats = {
        'total_qty': 0, 'total_sum': 0, 'avg_check': 0,
        'plants_breakdown': [], # Для таблицы (иерархия)
        'top_item': {'name': '-', 'size': '-', 'sum': 0}, # Топ позиция
        'status_funnel': {'shipped_pct': 0, 'canceled_pct': 0, 'reserved_pct': 0}, # Проценты
        'charts': {'labels': [], 'qty': [], 'sum': [], 'avg': [], 'colors': []} # Данные графиков
    }
    
    if client_id:
        # 1. Запрос заказов
        orders = Order.query.options(
            joinedload(Order.items).joinedload(OrderItem.plant),
            joinedload(Order.items).joinedload(OrderItem.size),
        ).filter(
            Order.client_id == client_id,
            Order.date >= start_date,
            Order.date <= end_date,
            Order.is_deleted == False 
        ).order_by(Order.date).all()

        agg_plants = {} 
        monthly_stats = {} 
        funnel_sum = {'total_potential': 0, 'shipped': 0, 'canceled': 0, 'reserved': 0}
        active_orders_count = 0
        flat_items_stats = {}

        for order in orders:
            order_potential = 0
            for item in order.items:
                order_potential += item.quantity * item.price
            
            funnel_sum['total_potential'] += order_potential

            if order.status == 'canceled':
                funnel_sum['canceled'] += order_potential
                continue 
            
            active_orders_count += 1
            
            m_key = (order.date.year, order.date.month)
            if m_key not in monthly_stats: monthly_stats[m_key] = {'qty': 0, 'sum': 0, 'count': 0}
            monthly_stats[m_key]['count'] += 1 

            order_fact_sum = 0 

            for item in order.items:
                fact_qty = item.shipped_quantity
                fact_sum = fact_qty * item.price
                
                reserved_qty = item.quantity - item.shipped_quantity
                funnel_sum['reserved'] += reserved_qty * item.price
                
                funnel_sum['shipped'] += fact_sum
                order_fact_sum += fact_sum

                if fact_qty > 0:
                    stats['total_qty'] += fact_qty
                    stats['total_sum'] += fact_sum
                    
                    monthly_stats[m_key]['qty'] += fact_qty
                    monthly_stats[m_key]['sum'] += fact_sum
                    
                    p_name = item.plant.name
                    s_name = item.size.name
                    
                    if p_name not in agg_plants:
                        agg_plants[p_name] = {'name': p_name, 'sum': 0, 'qty': 0, 'sizes': {}}
                    
                    agg_plants[p_name]['sum'] += fact_sum
                    agg_plants[p_name]['qty'] += fact_qty
                    
                    if s_name not in agg_plants[p_name]['sizes']:
                        agg_plants[p_name]['sizes'][s_name] = {'name': s_name, 'sum': 0, 'qty': 0}
                    
                    agg_plants[p_name]['sizes'][s_name]['sum'] += fact_sum
                    agg_plants[p_name]['sizes'][s_name]['qty'] += fact_qty

                    combo_key = (p_name, s_name)
                    flat_items_stats[combo_key] = flat_items_stats.get(combo_key, 0) + fact_sum

        # 2. Итоги
        if active_orders_count > 0:
            stats['avg_check'] = stats['total_sum'] / active_orders_count

        if funnel_sum['total_potential'] > 0:
            stats['status_funnel']['shipped_pct'] = (funnel_sum['shipped'] / funnel_sum['total_potential']) * 100
            stats['status_funnel']['canceled_pct'] = (funnel_sum['canceled'] / funnel_sum['total_potential']) * 100
            stats['status_funnel']['reserved_pct'] = (funnel_sum['reserved'] / funnel_sum['total_potential']) * 100

        if flat_items_stats:
            best_combo = max(flat_items_stats.items(), key=lambda x: x[1])
            stats['top_item'] = {'name': best_combo[0][0], 'size': best_combo[0][1], 'sum': best_combo[1]}

        sorted_plants = sorted(agg_plants.values(), key=lambda x: x['sum'], reverse=True)
        for p in sorted_plants:
            p['sizes_list'] = sorted(p['sizes'].values(), key=lambda x: x['sum'], reverse=True)
        stats['plants_breakdown'] = sorted_plants

        sorted_keys = sorted(monthly_stats.keys())
        
        season_spring = 'rgba(76, 175, 80, 0.6)'
        season_autumn = 'rgba(255, 152, 0, 0.6)'
        season_none = 'rgba(200, 200, 200, 0.3)'

        for (y, m) in sorted_keys:
            d = monthly_stats[(y, m)]
            label = f"{MONTH_NAMES[m]} {y}"
            stats['charts']['labels'].append(label)
            stats['charts']['qty'].append(d['qty'])
            stats['charts']['sum'].append(d['sum'])
            
            avg = d['sum'] / d['count'] if d['count'] > 0 else 0
            stats['charts']['avg'].append(avg)

            if 3 <= m <= 7: color = season_spring
            elif 8 <= m <= 12: color = season_autumn
            else: color = season_none
            stats['charts']['colors'].append(color)

    return render_template(
        'crm/crm_analytics.html',
        clients=clients,
        stats=stats,
        filters={'client_id': client_id, 'year_start': year_start, 'year_end': year_end},
        years_range=range(2017, msk_now().year + 2)
    )

@bp.route('/crm/abc')
@login_required
def crm_abc():
    if current_user.role not in ['admin', 'executive']: return redirect(url_for('main.index'))
    
    start_year = request.args.get('year_start', 2017, type=int)
    end_year = request.args.get('year_end', msk_now().year, type=int)
    
    # 1. Получаем данные в разбивке: Растение + Размер
    # Исключаем удаленные и отмененные
    results = db.session.query(
        OrderItem.plant_id,
        OrderItem.size_id,
        func.sum(OrderItem.shipped_quantity * OrderItem.price).label('rev'),
        func.sum(OrderItem.shipped_quantity).label('qty'),
        func.count(func.distinct(func.strftime('%Y-%m', Order.date))).label('months')
    ).join(Order).filter(
        Order.is_deleted == False,
        Order.status != 'canceled',
        func.extract('year', Order.date) >= start_year,
        func.extract('year', Order.date) <= end_year,
        OrderItem.shipped_quantity > 0
    ).group_by(OrderItem.plant_id, OrderItem.size_id).all()

    # 2. Подготовка справочников
    from app.models import Plant, Size
    plants_map = {p.id: p.name for p in Plant.query.all()}
    sizes_map = {s.id: s.name for s in Size.query.all()}

    # 3. Собираем плоский список для расчета долей
    all_items = []
    total_revenue = 0
    
    for r in results:
        rev = float(r.rev)
        total_revenue += rev
        all_items.append({
            'plant_id': r.plant_id,
            'size_id': r.size_id,
            'plant_name': plants_map.get(r.plant_id, 'Unknown'),
            'size_name': sizes_map.get(r.size_id, 'Unknown'),
            'revenue': rev,
            'qty': r.qty,
            'months': r.months
        })

    # 4. Сортируем по выручке для расчета ABC
    all_items.sort(key=lambda x: x['revenue'], reverse=True)
    
    # 5. Расчет классов для каждой позиции (SKU)
    current_accum = 0
    grouped_data = {} # plant_id -> { summary, items: [] }

    for item in all_items:
        current_accum += item['revenue']
        share_accum = (current_accum / total_revenue * 100) if total_revenue > 0 else 0
        
        # ABC
        if share_accum <= 80: abc = 'A'
        elif share_accum <= 95: abc = 'B'
        else: abc = 'C'
        
        # XYZ (Стабильность)
        if item['months'] >= 8: xyz = 'X'
        elif item['months'] >= 4: xyz = 'Y'
        else: xyz = 'Z'
        
        item['abc'] = abc
        item['xyz'] = xyz
        item['matrix'] = f"{abc}{xyz}"
        item['share_accum'] = share_accum

        # Группировка по растению
        pid = item['plant_id']
        if pid not in grouped_data:
            pid = item['plant_id']
        if pid not in grouped_data:
            grouped_data[pid] = {
                'name': item['plant_name'],
                'total_revenue': 0,
                'total_qty': 0,
                'variants': [], # <--- ПЕРЕИМЕНОВАЛИ В variants
                'months_max': 0
            }
        
        grouped_data[pid]['variants'].append(item)
        grouped_data[pid]['total_revenue'] += item['revenue']
        grouped_data[pid]['total_qty'] += item['qty']
        if item['months'] > grouped_data[pid]['months_max']:
            grouped_data[pid]['months_max'] = item['months']

    # 6. Расчет ABC для ГРУПП (Растений целиком)
    # Преобразуем словарь в список и сортируем по общей выручке растения
    final_list = sorted(grouped_data.values(), key=lambda x: x['total_revenue'], reverse=True)
    
    plant_accum = 0
    for plant in final_list:
        plant_accum += plant['total_revenue']
        p_share = (plant_accum / total_revenue * 100) if total_revenue > 0 else 0
        
        if p_share <= 80: p_abc = 'A'
        elif p_share <= 95: p_abc = 'B'
        else: p_abc = 'C'
        
        # XYZ для растения берем по максимуму его размеров
        if plant['months_max'] >= 8: p_xyz = 'X'
        elif plant['months_max'] >= 4: p_xyz = 'Y'
        else: p_xyz = 'Z'
        
        plant['abc'] = p_abc
        plant['xyz'] = p_xyz
        plant['matrix'] = f"{p_abc}{p_xyz}"
        plant['share_accum'] = p_share
        
        # Сортируем размеры внутри растения (A -> C)
        plant['variants'].sort(key=lambda x: x['revenue'], reverse=True)

    return render_template('crm/crm_abc.html', 
                           data=final_list, 
                           filters={'start': start_year, 'end': end_year},
                           years=range(2017, msk_now().year + 2))

@bp.route('/crm/yoy')
@login_required
def crm_yoy():
    if current_user.role not in ['admin', 'executive']: return redirect(url_for('main.index'))
    
    f_clients = [int(x) for x in request.args.getlist('client_id')]
    start_year = request.args.get('year_start', 2017, type=int)
    end_year = request.args.get('year_end', msk_now().year, type=int)
    
    q = db.session.query(
        func.extract('year', Order.date).label('yr'),
        func.extract('month', Order.date).label('mth'),
        func.sum(OrderItem.shipped_quantity * OrderItem.price)
    ).join(OrderItem).filter(
        Order.is_deleted == False,
        Order.status != 'canceled',
        func.extract('year', Order.date) >= start_year,
        func.extract('year', Order.date) <= end_year,
        OrderItem.shipped_quantity > 0
    )
    
    if f_clients:
        q = q.filter(Order.client_id.in_(f_clients))
        
    results = q.group_by('yr', 'mth').all()
    
    years_data = {}
    for yr in range(start_year, end_year + 1):
        years_data[yr] = [0] * 12
        
    for r in results:
        y, m, val = int(r[0]), int(r[1]), float(r[2])
        if y in years_data:
            years_data[y][m-1] = val
            
    colors = ['#4CAF50', '#2196F3', '#FF9800', '#F44336', '#9C27B0', '#795548', '#607D8B', '#00BCD4']
    chart_datasets = []
    i = 0
    for yr, data in years_data.items():
        if sum(data) > 0:
            chart_datasets.append({
                'label': str(yr),
                'data': data,
                'borderColor': colors[i % len(colors)],
                'backgroundColor': 'transparent',
                'tension': 0.3
            })
            i += 1

    return render_template('crm/crm_yoy.html', 
                           datasets=chart_datasets,
                           all_clients=Client.query.order_by(Client.name).all(),
                           selected_clients=f_clients,
                           filters={'start': start_year, 'end': end_year},
                           years=range(2017, msk_now().year + 2))

@bp.route('/crm/seasonality')
@login_required
def crm_seasonality():
    if current_user.role not in ['admin', 'executive']: return redirect(url_for('main.index'))
    
    year = request.args.get('year', msk_now().year, type=int)
    
    results = db.session.query(
        Client.name,
        func.extract('month', Order.date).label('mth'),
        func.sum(OrderItem.shipped_quantity * OrderItem.price)
    ).join(Order, Client.id == Order.client_id)\
     .join(OrderItem, Order.id == OrderItem.order_id)\
     .filter(
        Order.is_deleted == False,
        Order.status != 'canceled',
        func.extract('year', Order.date) == year,
        OrderItem.shipped_quantity > 0
    ).group_by(Client.name, 'mth').all()
    
    matrix = {}
    max_val = 0
    
    for r in results:
        c_name, m, val = r[0], int(r[1]), float(r[2])
        if c_name not in matrix: matrix[c_name] = {'data': [0]*12, 'total': 0}
        matrix[c_name]['data'][m-1] = val
        matrix[c_name]['total'] += val
        if val > max_val: max_val = val
        
    sorted_matrix = sorted(matrix.items(), key=lambda x: x[1]['total'], reverse=True)
    
    return render_template('crm/crm_seasonality.html', 
                           matrix=sorted_matrix, 
                           max_val=max_val,
                           target_year=year,
                           years=range(2017, msk_now().year + 2),
                           month_names=MONTH_NAMES)

@bp.route('/crm/price_calculator', methods=['GET', 'POST'])
@login_required
def crm_price_calculator():
    if current_user.role not in ['admin', 'executive']:
        return redirect(url_for('main.index'))

    # ... (код обработки POST/загрузки файла оставляем без изменений, он выше) ...
    # Если ты используешь старый код, убедись, что блок if request.method == 'POST': остался на месте.
    # Я привожу код начиная с блока GET (отображения), где меняется логика сортировки.
    
    # Обработка загрузки файла (Только Админ)
    if request.method == 'POST':
        if current_user.role != 'admin':
            flash('Только администратор может загружать цены конкурентов')
            return redirect(url_for('crm.crm_price_calculator'))
            
        file = request.files.get('file')
        snap_name = request.form.get('name', f'Загрузка от {msk_now().strftime("%d.%m.%Y")}')
        
        if file:
            try:
                # 1. Кэш себестоимости и цен
                calc_year = msk_now().year - 1
                cost_data = calculate_cost_data(calc_year)
                base_cost = cost_data['cumulative_cost']
                
                our_prices_cache = {}
                stocks = db.session.query(Plant.name, Size.name, func.max(StockBalance.price))\
                    .join(Plant).join(Size)\
                    .group_by(Plant.name, Size.name).all()
                for p_name, s_name, price in stocks:
                    our_prices_cache[(p_name.lower().strip(), s_name.lower().strip())] = float(price or 0)

                # 2. Читаем Excel
                wb = load_workbook(file)
                ws = wb.active
                
                # --- ЛОГИКА ОБЪЕДИНЕНИЯ ПО ДАТЕ ---
                now = msk_now()
                # Определяем начало и конец сегодняшнего дня
                start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
                end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)

                # Ищем, есть ли уже отчет за сегодня
                snapshot = CompetitorSnapshot.query.filter(
                    CompetitorSnapshot.date >= start_of_day,
                    CompetitorSnapshot.date <= end_of_day
                ).first()

                if snapshot:
                    # Если нашли - обновляем имя (чтобы было видно название последней загрузки)
                    snapshot.name = snap_name
                    # Мы НЕ создаем новый объект, а используем найденный
                    flash(f'Данные добавлены в существующий отчет за сегодня.')
                else:
                    # Если отчета за сегодня нет - создаем новый
                    snapshot = CompetitorSnapshot(name=snap_name, user_id=current_user.id, date=now)
                    db.session.add(snapshot)
                
                db.session.flush() # Получаем ID (старого или нового отчета)
                # ----------------------------------
                
                from app.crm_validator import validate_competitor_row, format_reason_badges
                import json as _json

                count = 0
                rejected_count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0] or (len(row) < 3 or not row[2]):
                        continue

                    p_name = str(row[0]).strip()
                    s_name = str(row[1]).strip() if len(row) > 1 and row[1] else "-"
                    raw_price = str(row[2]) if row[2] else "0"
                    clean_price = raw_price.replace(' ', '').replace('\xa0', '').replace('руб', '').replace('р', '').replace(',', '.')
                    try:
                        c_price = float(clean_price)
                    except ValueError:
                        c_price = 0.0
                    comp_name = str(row[3]).strip() if len(row) > 3 and row[3] else "Неизвестно"
                    link_val = str(row[4]).strip() if len(row) > 4 and row[4] else None

                    # Новые поля (обратно совместимо: если колонок нет — None/'')
                    pack_type = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                    form_v = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                    excerpt = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                    try:
                        conf_v = float(row[8]) if len(row) > 8 and row[8] not in (None, '') else None
                    except (ValueError, TypeError):
                        conf_v = None

                    key = (p_name.lower(), s_name.lower())
                    my_price = our_prices_cache.get(key, 0)

                    # Валидация: помечаем подозрительные, но всё равно сохраняем с флагом.
                    vresult = validate_competitor_row({
                        'plant_name': p_name,
                        'size': s_name,
                        'price': c_price,
                        'pack_type': pack_type,
                        'form': form_v,
                        'source_excerpt': excerpt,
                        'url': link_val or '',
                    }, our_item={
                        'plant_name': p_name,
                        'size': s_name,
                        'haircut': '',  # при ручной загрузке форму из INPUT не знаем
                        'our_price': my_price,
                    })
                    is_rej = not vresult['ok']
                    if is_rej:
                        rejected_count += 1

                    item = CompetitorRow(
                        snapshot_id=snapshot.id,
                        plant_name=p_name,
                        size_name=s_name,
                        competitor_name=comp_name,
                        competitor_price=c_price,
                        source_link=link_val,
                        our_price_at_moment=my_price,
                        our_cost_at_moment=base_cost,
                        pack_type=pack_type or None,
                        form=form_v or None,
                        source_excerpt=excerpt or None,
                        confidence=conf_v,
                        is_rejected=is_rej,
                        reject_reasons=_json.dumps(vresult['reasons'], ensure_ascii=False) if is_rej else None,
                    )
                    db.session.add(item)
                    count += 1

                db.session.commit()
                log_action(f"Загрузил цены конкурентов: {snap_name} ({count} поз., отклонено валидатором: {rejected_count})")
                if rejected_count:
                    flash(f'Загружено {count} позиций. Из них {rejected_count} отправлены во вкладку «Отклонено» — проверьте.')
                else:
                    flash(f'Успешно загружено {count} позиций.')
                return redirect(url_for('crm.crm_price_calculator', snapshot_id=snapshot.id))
                
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка обработки файла: {e}')
                
        return redirect(url_for('crm.crm_price_calculator'))

    # Отображение (GET)
    snapshot_id = request.args.get('snapshot_id', type=int)
    snapshots = CompetitorSnapshot.query.order_by(CompetitorSnapshot.date.desc()).all()
    
    current_snapshot = None
    rows_data = []
    
    if snapshot_id:
        current_snapshot = CompetitorSnapshot.query.get(snapshot_id)
    elif snapshots:
        current_snapshot = snapshots[0]
        
    rejected_rows = []
    if current_snapshot:
        from app.crm_validator import format_reason_badges
        # Подготовка данных
        for r in current_snapshot.rows:
            our_p = float(r.our_price_at_moment or 0)
            comp_p = float(r.competitor_price or 0)
            cost = float(r.our_cost_at_moment or 0)

            diff_rub = our_p - comp_p
            diff_pct = ((our_p - comp_p) / comp_p * 100) if comp_p > 0 else 0

            our_margin = our_p - cost
            comp_margin_scenario = comp_p - cost

            entry = {
                'id': r.id,
                'snapshot_id': r.snapshot_id,
                'plant': r.plant_name,
                'size': r.size_name,
                'competitor': r.competitor_name,
                'link': r.source_link,
                'comp_price': comp_p,
                'our_price': our_p,
                'diff_rub': diff_rub,
                'cost': cost,
                'our_margin': our_margin,
                'comp_margin_scenario': comp_margin_scenario,
                'pack_type': getattr(r, 'pack_type', None),
                'form': getattr(r, 'form', None),
                'excerpt': getattr(r, 'source_excerpt', None),
                'confidence': float(getattr(r, 'confidence', None) or 0),
                'is_rejected': bool(getattr(r, 'is_rejected', False)),
                'reject_badges': format_reason_badges(getattr(r, 'reject_reasons', None)),
            }
            if entry['is_rejected']:
                rejected_rows.append(entry)
            else:
                rows_data.append(entry)

        # --- ЛОГИКА СОРТИРОВКИ ---
        import re
        def size_sort_key(s_str):
            match = re.search(r'\d+', s_str or '')
            if match:
                return int(match.group())
            return 0

        rows_data.sort(key=lambda x: (x['plant'] or '', -size_sort_key(x['size'] or '')))
        rejected_rows.sort(key=lambda x: (x['plant'] or '', -size_sort_key(x['size'] or '')))

    # Список товаров для графика
    history_products = db.session.query(
        CompetitorRow.plant_name, 
        CompetitorRow.size_name
    ).distinct().order_by(CompetitorRow.plant_name, CompetitorRow.size_name).all()
    product_options = [f"{p} | {s}" for p, s in history_products]

     # Импортируем список регионов (если он в том же файле, просто берем переменную)
    from app.crm import RUSSIAN_REGIONS_LIST 

    # Есть ли ключ Groq — для UI (включать/отключать кнопку авто-анализа).
    import os as _os
    ai_enabled = bool(_os.environ.get('GROQ_API_KEY'))
    groq_model_name = _os.environ.get('GROQ_MODEL', 'llama-3.3-70b-versatile')

    return render_template('crm/crm_price_calc.html',
                           snapshots=snapshots,
                           current=current_snapshot,
                           rows=rows_data,
                           rejected_rows=rejected_rows,
                           product_options=product_options,
                           regions_list=RUSSIAN_REGIONS_LIST,
                           ai_enabled=ai_enabled,
                           groq_model_name=groq_model_name)

@bp.route('/crm/price_calculator/template')
@login_required
def crm_price_calculator_template():
    if current_user.role not in ['admin', 'executive']:
        return redirect(url_for('main.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон цен"

    # 9 колонок: 5 старых + 4 новых для качественной фильтрации.
    # Старые 5-колоночные файлы тоже принимаются (поля просто остаются пустыми).
    headers = [
        "Наименование (как в базе)", "Размер", "Цена конкурента", "Конкурент", "Ссылка",
        "Тип упаковки (RB/WRB/grunt/C2/C3/C5/C7/C10/C15/C20/P9/other)",
        "Форма (free/ball/niwaki/topiary/pompon/stamm/other)",
        "Цитата с сайта (до 150 симв.)",
        "Уверенность (0..1)",
    ]
    ws.append(headers)

    # Примеры
    ws.append([
        "Туя западная Smaragd", "100-120", 2900, "Питомник А", "https://example.com/tuia-smaragd",
        "RB", "free", "Туя западная Smaragd, ком 100-120", 0.9,
    ])
    ws.append([
        "Спирея серая", "80-100", 1800, "Садовый Центр Б", "https://example.com/spirea-grey",
        "RB", "ball", "Спирея серая, стриженый шар, ком", 0.8,
    ])

    # Ширина колонок
    widths = [30, 12, 14, 20, 30, 18, 18, 40, 12]
    for i, w in enumerate(widths, start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(buf, download_name='competitor_prices_template.xlsx', as_attachment=True)

@bp.route('/crm/price_calculator/delete/<int:snapshot_id>', methods=['POST'])
@login_required
def crm_delete_snapshot(snapshot_id):
    if current_user.role != 'admin':
        flash('Только администратор может удалять снимки')
        return redirect(url_for('crm.crm_price_calculator'))
        
    try:
        snapshot = CompetitorSnapshot.query.get_or_404(snapshot_id)
        db.session.delete(snapshot)
        db.session.commit()
        log_action(f"Удалил сравнение цен конкурентов: {snapshot.name}")
        flash('Сравнение цен удалено')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении: {e}')
        
    return redirect(url_for('crm.crm_price_calculator'))


@bp.route('/crm/price_calculator/delete_row/<int:row_id>', methods=['POST'])
@login_required
def crm_delete_row(row_id):
    if current_user.role != 'admin':
        flash('Только администратор может удалять строки')
        return redirect(url_for('crm.crm_price_calculator'))

    # Причина удаления нужна для обучения промта (Фаза 5).
    reason_code = (request.form.get('reason_code') or '').strip().lower() or 'other'
    allowed = {'pot', 'niwaki', 'price', 'wrong_plant', 'wrong_region', 'other'}
    if reason_code not in allowed:
        reason_code = 'other'

    try:
        row = CompetitorRow.query.get_or_404(row_id)
        snapshot_id = row.snapshot_id
        # Перед физическим удалением сохраняем «посмертную» причину в ActionLog,
        # чтобы можно было строить статистику/негативные примеры.
        preview = f"{row.plant_name} / {row.size_name} у {row.competitor_name} за {row.competitor_price} руб"
        db.session.delete(row)
        db.session.commit()
        log_action(
            f"Удалил строку сравнения: {preview} [reason={reason_code}]"
        )
        flash('Строка удалена')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении строки: {e}')
        snapshot_id = request.args.get('snapshot_id')

    return redirect(url_for('crm.crm_price_calculator', snapshot_id=snapshot_id))


@bp.route('/crm/price_calculator/restore_row/<int:row_id>', methods=['POST'])
@login_required
def crm_restore_row(row_id):
    """Вернуть строку из «Отклонено» в основной список (сбросить is_rejected)."""
    if current_user.role not in ['admin', 'executive']:
        flash('Только администратор или руководитель может восстанавливать строки')
        return redirect(url_for('crm.crm_price_calculator'))

    try:
        row = CompetitorRow.query.get_or_404(row_id)
        snapshot_id = row.snapshot_id
        row.is_rejected = False
        row.reject_reasons = None
        db.session.commit()
        log_action(
            f"Вернул из отклонённых: {row.plant_name} / {row.size_name} (конкурент {row.competitor_name})"
        )
        flash('Строка возвращена в основной список')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка восстановления строки: {e}')
        snapshot_id = request.args.get('snapshot_id')

    return redirect(url_for('crm.crm_price_calculator', snapshot_id=snapshot_id))

@bp.route('/crm/api/price_history', methods=['POST'])
@login_required
def crm_api_price_history():
    # Получаем "Растение | Размер"
    selection = request.json.get('selection', '')
    if not selection or ' | ' not in selection:
        return jsonify({'error': 'Invalid selection'})
    
    plant, size = selection.split(' | ', 1)
    
    # Ищем все записи по этому товару, сортируем по дате загрузки
    rows = db.session.query(CompetitorRow, CompetitorSnapshot.date)\
        .join(CompetitorSnapshot)\
        .filter(CompetitorRow.plant_name == plant, CompetitorRow.size_name == size)\
        .order_by(CompetitorSnapshot.date).all()
    
    if not rows:
        return jsonify({'error': 'No data'})

    # Собираем данные
    dates = []
    datasets = {} # { 'Наш питомник': [100, 120...], 'Конкурент А': [90, 95...] }
    
    # 1. Собираем все уникальные даты
    sorted_rows = sorted(rows, key=lambda x: x[1])
    unique_dates = sorted(list(set([r[1].strftime('%d.%m.%Y') for r in sorted_rows])))
    
    for d in unique_dates:
        dates.append(d)
    
    # Инициализируем массивы нулями/None
    datasets['Наш Питомник'] = [None] * len(dates)
    
    # 2. Заполняем данными
    for row_obj, snap_date in rows:
        d_str = snap_date.strftime('%d.%m.%Y')
        idx = dates.index(d_str)
        
        # Наша цена (она одна)
        if row_obj.our_price_at_moment:
            datasets['Наш Питомник'][idx] = float(row_obj.our_price_at_moment)
            
        # Цена конкурента
        c_name = row_obj.competitor_name
        if c_name not in datasets:
            datasets[c_name] = [None] * len(dates)
        
        if row_obj.competitor_price:
            datasets[c_name][idx] = float(row_obj.competitor_price)

    return jsonify({'labels': dates, 'datasets': datasets})

@bp.route('/crm/download_prompt_input_template')
@login_required
def crm_download_prompt_input_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Список растений"
    # Заголовки для пользователя
    ws.append(["Наименование", "Размер", "Стрижка (шар/стрижка/пусто)"])
    # Примеры
    ws.append(["Туя западная Smaragd", "100-120", ""])
    ws.append(["Спирея серая", "80-100", "шар"])

    # Ширина колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='prompt_plant_list_template.xlsx', as_attachment=True)


# ------------------------------------------------------------------
# Подготовка данных для промта (используется и при скачивании .txt,
# и при автоматическом запуске Groq).
# ------------------------------------------------------------------

def _parse_plant_list_xlsx(file_obj):
    """Читает xlsx-список позиций для анализа. Возвращает список dict:
    [{'plant': ..., 'size': ..., 'haircut': ...}, ...]. Пустые имена пропускаются."""
    wb = load_workbook(file_obj)
    ws = wb.active
    seen = set()
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        p_name = str(row[0]).strip()
        s_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        haircut = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        key = (p_name.lower(), s_name.lower(), haircut.lower())
        if key in seen:
            continue
        seen.add(key)
        items.append({'plant': p_name, 'size': s_name, 'haircut': haircut})
    return items


def _enrich_items_with_our_prices(items):
    """По каждой позиции подтягивает нашу цену и среднюю себестоимость.
    Не мутирует items, возвращает новый список dict с добавленными полями
    our_price и our_cost. Цена — максимум по StockBalance для (plant, size);
    себестоимость — кумулятивная за прошлый год (как при импорте)."""
    try:
        calc_year = msk_now().year - 1
        cost_data = calculate_cost_data(calc_year)
        base_cost = float(cost_data.get('cumulative_cost') or 0)
    except Exception:
        base_cost = 0.0
    our_prices_cache = {}
    try:
        stocks = db.session.query(Plant.name, Size.name, func.max(StockBalance.price)) \
            .join(Plant).join(Size).group_by(Plant.name, Size.name).all()
        for p_name, s_name, price in stocks:
            if p_name and s_name:
                our_prices_cache[(p_name.lower().strip(), s_name.lower().strip())] = float(price or 0)
    except Exception:
        pass
    out = []
    for it in items:
        key = (it['plant'].lower().strip(), it['size'].lower().strip())
        price = our_prices_cache.get(key, 0.0)
        out.append({**it, 'our_price': price, 'our_cost': base_cost})
    return out


def _collect_negative_examples(limit=8):
    """Подтягивает последние негативные примеры для подмешивания в промт.

    Источники:
    1. CompetitorRow с is_rejected=True + reject_reasons — то, что отбраковал
       валидатор (самый достоверный сигнал).
    2. ActionLog с префиксом «Удалил строку сравнения» и суффиксом
       [reason=...] — ручные удаления админом (человеческая разметка).
    """
    import json as _json
    import re as _re
    from app.models import ActionLog

    result = []
    seen_tags = set()

    # 1. Валидатор.
    try:
        rows = (
            db.session.query(CompetitorRow)
            .filter(CompetitorRow.is_rejected == True)  # noqa: E712
            .filter(CompetitorRow.reject_reasons.isnot(None))
            .order_by(CompetitorRow.id.desc())
            .limit(limit * 3)
            .all()
        )
    except Exception:
        rows = []

    for r in rows:
        reasons = getattr(r, 'reject_reasons', None)
        if not reasons:
            continue
        try:
            tags = _json.loads(reasons) if isinstance(reasons, str) else reasons
            if not isinstance(tags, list):
                continue
        except Exception:
            continue
        key_tag = ','.join(sorted(tags))
        if key_tag in seen_tags:
            continue
        seen_tags.add(key_tag)
        excerpt = getattr(r, 'source_excerpt', '') or ''
        result.append({
            'plant': r.plant_name or '',
            'size': r.size_name or '',
            'competitor': r.competitor_name or '',
            'price': float(r.competitor_price or 0),
            'reasons': tags,
            'excerpt': excerpt[:140],
        })
        if len(result) >= limit:
            return result

    # 2. Ручные удаления из ActionLog.
    # Формат: "Удалил строку сравнения: {plant} / {size} у {comp} за {price} руб [reason={code}]"
    try:
        logs = (
            db.session.query(ActionLog)
            .filter(ActionLog.action.like('Удалил строку сравнения:%'))
            .order_by(ActionLog.id.desc())
            .limit(limit * 4)
            .all()
        )
    except Exception:
        logs = []
    rx = _re.compile(
        r"Удалил строку сравнения:\s*(?P<plant>.+?)\s*/\s*(?P<size>.+?)\s+у\s+(?P<comp>.+?)\s+за\s+(?P<price>[\d.]+)\s*руб\s*\[reason=(?P<reason>\w+)\]",
        _re.IGNORECASE,
    )
    for log in logs:
        m = rx.search(log.action or '')
        if not m:
            continue
        reason_code = m.group('reason').lower()
        if reason_code in ('other',):
            continue  # не учим тегом 'other'
        tag = f"user:{reason_code}"
        if tag in seen_tags:
            continue
        seen_tags.add(tag)
        try:
            price_v = float(m.group('price'))
        except (ValueError, TypeError):
            price_v = 0.0
        result.append({
            'plant': m.group('plant').strip(),
            'size': m.group('size').strip(),
            'competitor': m.group('comp').strip(),
            'price': price_v,
            'reasons': [tag],
            'excerpt': '',
        })
        if len(result) >= limit:
            break

    return result


def _build_market_analysis_prompt(items_enriched, regions_str, negatives=None):
    """Собирает финальный текст промта по списку позиций. items_enriched —
    элементы с полями plant/size/haircut/our_price/our_cost.
    negatives — опциональный список свежих негативных примеров."""
    import json as _json

    # Входной список для модели — JSON, чтобы не ошибиться с парсингом.
    input_list = []
    for it in items_enriched:
        input_list.append({
            'plant': it.get('plant', ''),
            'size': it.get('size', ''),
            'form_required': 'ball' if (it.get('haircut', '').lower() in ('шар', 'ball', 'стрижка', 'trim')) else 'free',
            'our_price_rub': round(float(it.get('our_price') or 0), 0),
            'price_corridor_min': round(float(it.get('our_price') or 0) * 0.3, 0),
            'price_corridor_max': round(float(it.get('our_price') or 0) * 3.0, 0),
        })
    input_json = _json.dumps(input_list, ensure_ascii=False, indent=2)

    negatives = negatives or []
    negatives_block = ""
    if negatives:
        lines = []
        for n in negatives:
            lines.append(
                f"- REJECT: '{n['plant']} {n['size']} у {n['competitor']} за {int(n['price'])} руб' — "
                f"причина: {', '.join(n['reasons'])}. Цитата: \"{n['excerpt']}\""
            )
        negatives_block = (
            "\n\nДополнительные НЕГАТИВНЫЕ примеры (эти позиции были отклонены ранее — НЕ повторяй такие ошибки):\n"
            + "\n".join(lines)
        )

    prompt = f"""Роль: Ты профессиональный аналитик рынка посадочного материала
(декоративные деревья и кустарники открытого грунта) РФ со стажем 20+ лет.
Ты отвечаешь только проверенными фактами, у тебя есть доступ к актуальным
каталогам ПИТОМНИКОВ (например: lkpitomnik.ru, knp.moscow, vashsad.ua,
pitomnik-plants.ru, poselki.net, landshaftni.ru, sazhency-online.ru,
posadi-derevo.ru, landscape-mix.ru, zelsad.ru, gardentools.ru — и прочих
СПЕЦИАЛИЗИРОВАННЫХ питомников и садовых центров РФ).

ЗАДАЧА
Для каждой позиции из INPUT найди предложения конкурентов-ПИТОМНИКОВ
в указанных регионах и верни СТРОГО JSON по схеме OUTPUT.

РЕГИОНЫ ПОИСКА: {regions_str}

============================================================
АНТИ-ГАЛЛЮЦИНАЦИОННЫЕ ПРАВИЛА — САМОЕ ВАЖНОЕ
============================================================
1. Если у тебя НЕТ доступа к живому поиску/браузеру в ходе этого запроса —
   ВЕРНИ {{"results": []}}. НЕ ВЫДУМЫВАЙ цены, URL или товары, которых
   не видел своими глазами на конкретной карточке товара.
2. Каждая строка в results должна быть привязана к одной РЕАЛЬНОЙ карточке
   товара на сайте питомника. URL должен вести на страницу товара, а не
   на главную, категорию или поиск.
3. В поле source_excerpt — БУКВАЛЬНАЯ цитата с этой карточки (как на сайте,
   40–180 символов). Цитата ОБЯЗАТЕЛЬНО должна содержать либо цену, либо
   размер, либо артикул (иначе строка будет отклонена валидатором).
4. Если на карточке нет видимой цены (написано «цену уточняйте», «под заказ»,
   «свяжитесь с менеджером») — НЕ включай строку в results.
5. Если цена на карточке — в рублях за 1 шт. Если за комплект/опт с ценником
   «от N шт» — НЕ включай.

ЗАПРЕЩЁННЫЕ ИСТОЧНИКИ (строго НЕ брать цены оттуда — это не питомники):
- Маркетплейсы: ozon.ru, wildberries.ru, market.yandex, avito.ru, youla.ru.
- Магазины инструмента и DIY: leroymerlin, castorama, obi, petrovich, maxidom.
- Производители инструмента/оборудования: stihl, husqvarna, makita, gardena, karcher.
- Соцсети/блоги/агрегаторы: vk.com, t.me, dzen.ru, pinterest, wikipedia.
Если URL ведёт на любой из этих сайтов — НЕ включай строку.

ГЛОССАРИЙ УПАКОВКИ (pack_type)
- RB / WRB / grunt = растение с комом земли (в сетке/мешковине) или
  выкопанное из грунта. ТОЛЬКО ТАКИЕ позиции нам нужны.
- C2 / C3 / C5 / C7 / C10 / C15 / C20 = пластиковый контейнер 2–20 л. ЗАПРЕЩЕНО.
- P9 = технологический горшок 9×9 см. ЗАПРЕЩЕНО.
- other = неизвестно.
Признаки горшка в описании: «C2», «C3», «C10», «P9», «контейнер», «горшок»,
«N литров», «pot», «литраж», «объём контейнера». Если видишь эти слова в
цитате или URL — строка ОТБРАК, не включаем.

ГЛОССАРИЙ ФОРМЫ (form)
- free    = свободнорастущее природной формы (обычный куст/дерево). По умолчанию.
- ball    = стриженый шар (полусфера/сфера).
- niwaki  = садовый бонсай, многоярусная стрижка ветвей. ДОРОГО.
- topiary = топиар (спираль, куб, пирамида).
- pompon  = шары на штамбе (помпоны).
- stamm   = штамбовая форма.
- other   = иное.

КАК СОПОСТАВЛЯТЬ FORM
- Если form_required = 'free' — допустим ТОЛЬКО form=free.
  Ниваки / шары / помпоны / стамбы / топиары — НЕ конкуренты, НЕ включать.
- Если form_required = 'ball' — допустимы form ∈ {{ball, topiary, pompon}}.
  Свободнорастущие (free) — НЕ включать. Ниваки НЕ включать (слишком дорого).

ПРАВИЛА ЦЕН
- Для каждой позиции в INPUT указаны price_corridor_min и price_corridor_max
  (0.3× и 3× от нашей цены).
- Если цена конкурента вне коридора — всё равно можешь вернуть строку,
  НО в warnings добавь "price_outlier" и перепроверь цитату.
- НЕ ВЫДУМЫВАЙ цены — бери только те, что видишь на карточке товара.

ФОРМАТ ОТВЕТА (OUTPUT)
Верни ОДИН JSON-объект:
{{
  "results": [
    {{
      "plant_name": "строка ровно из INPUT.plant",
      "size": "строка ровно из INPUT.size",
      "competitor": "название питомника (бренд/домен без www и без .ru/.com)",
      "price": 1234,
      "url": "полный https URL карточки товара",
      "pack_type": "RB|WRB|grunt|other",
      "form": "free|ball|niwaki|topiary|pompon|stamm|other",
      "source_excerpt": "БУКВАЛЬНАЯ цитата с карточки, 40–180 симв., с цифрами (цена/размер/артикул)",
      "confidence": 0.0,
      "warnings": ["price_outlier", "size_mismatch"]
    }}
  ]
}}

ЖЁСТКИЕ ПРАВИЛА
1. plant_name и size — БУКВАЛЬНО из INPUT, не перефразируй.
2. Одна позиция INPUT → несколько строк в results, если нашёл несколько
   питомников. Минимум 3 источника желательно.
3. Никаких строк без URL. URL должен быть ПОЛНЫЙ и валидный, https.
4. Если по позиции не нашёл ничего достойного — просто пропусти её.
5. НЕ группируй результаты; каждая цена — отдельная строка.
6. Ответ — чистый JSON без markdown (```), без пояснений до или после.

НЕГАТИВНЫЕ FEW-SHOT (что НЕ включать):
- "Туя Smaragd 100-120 в контейнере C10 за 1290 руб" — pack_type=C10.
- "Сосна горная Mughus 80-100, ниваки за 45 000 руб" — form=niwaki, а нужна free.
- "Спирея серая 80-100, свободный куст 650 руб", если form_required=ball.
- Любой URL на ozon/wildberries/avito/leroymerlin/vk/dzen/wikipedia/t.me.
- source_excerpt типа "декоративное растение" — без цифр, подозрительно.{negatives_block}

INPUT (позиции для анализа):
{input_json}
"""
    return prompt


@bp.route('/crm/download_ai_prompt', methods=['GET', 'POST'])
@login_required
def crm_download_ai_prompt():
    if request.method != 'POST':
        return redirect(url_for('crm.crm_price_calculator'))

    regions = request.form.getlist('regions')
    if not regions:
        regions_str = "ЦФО (Москва и МО, Тула, Воронеж, Калуга, Тверь и др.)"
    else:
        regions_str = ", ".join(regions)

    file = request.files.get('plant_file')
    if not file:
        flash('Список растений не загружен')
        return redirect(url_for('crm.crm_price_calculator'))

    try:
        items = _parse_plant_list_xlsx(file)
    except Exception as e:
        flash(f"Ошибка чтения файла: {e}")
        return redirect(url_for('crm.crm_price_calculator'))

    if not items:
        flash('В файле не найдено позиций')
        return redirect(url_for('crm.crm_price_calculator'))

    enriched = _enrich_items_with_our_prices(items)
    negatives = _collect_negative_examples(limit=8)
    prompt_text = _build_market_analysis_prompt(enriched, regions_str, negatives=negatives)

    buf = io.BytesIO()
    buf.write(prompt_text.encode('utf-8'))
    buf.seek(0)
    return send_file(buf, download_name='market_analysis_prompt.txt', mimetype='text/plain', as_attachment=True)


# ------------------------------------------------------------------
# АВТО-АНАЛИЗ через Groq: формируем промт, зовём модель, парсим JSON,
# прогоняем через валидатор, создаём CompetitorSnapshot + CompetitorRow.
# ------------------------------------------------------------------

def _groq_json_request(prompt_text, model=None, temperature=0.1, timeout_sec=None):
    """Отправляет один запрос в Groq, требует JSON-объект в ответе. Возвращает
    (dict, raw_content), либо бросает исключение.

    Модель берётся из env GROQ_MODEL (можно переключить на 'groq/compound-beta'
    — у неё есть встроенный web_search и она НЕ галлюцинирует). По умолчанию —
    'llama-3.3-70b-versatile' (быстрая, но БЕЗ браузера: будет придумывать сайты
    и цены, если не подсунуть реальные данные в промт).

    timeout_sec — жёсткий HTTP-таймаут на ответ модели, чтобы один зависший
    батч не съедал весь gunicorn timeout. Дефолт — из env GROQ_TIMEOUT_SEC (45).
    """
    import os as _os
    import json as _json
    from groq import Groq

    api_key = _os.environ.get('GROQ_API_KEY')
    if not api_key:
        raise RuntimeError('GROQ_API_KEY не задан в окружении')

    if model is None:
        model = _os.environ.get('GROQ_MODEL', 'llama-3.3-70b-versatile')
    if timeout_sec is None:
        try:
            timeout_sec = float(_os.environ.get('GROQ_TIMEOUT_SEC', '45'))
        except ValueError:
            timeout_sec = 45.0

    client = Groq(api_key=api_key, timeout=timeout_sec)
    # compound-beta и подобные grouded-модели принимают response_format через
    # tools иначе, поэтому для них не принуждаем json_object (сами извлекаем).
    kwargs = {
        'messages': [
            {"role": "system", "content": "Ты аккуратный аналитик. Отвечаешь только валидным JSON без пояснений."},
            {"role": "user", "content": prompt_text},
        ],
        'model': model,
        'temperature': temperature,
    }
    if 'compound' not in (model or '').lower():
        kwargs['response_format'] = {"type": "json_object"}
    resp = client.chat.completions.create(**kwargs)
    content = resp.choices[0].message.content or '{}'
    # На случай, когда compound-beta вернула обрамление ```json ... ``` — вытащим:
    stripped = content.strip()
    if stripped.startswith('```'):
        # срезаем первые ```(lang)?  и последние ```
        stripped = re.sub(r'^```[a-zA-Z0-9_-]*\s*\n?', '', stripped)
        stripped = re.sub(r'\n?```\s*$', '', stripped)
    try:
        data = _json.loads(stripped)
    except _json.JSONDecodeError:
        # Иногда модель отдаёт ведущий текст — попробуем выдернуть первый JSON-объект.
        m = re.search(r'\{.*\}', stripped, re.DOTALL)
        if not m:
            raise
        data = _json.loads(m.group(0))
    return data, content


@bp.route('/crm/price_calculator/ai_run', methods=['POST'])
@login_required
def crm_ai_run():
    """Полный авто-цикл: xlsx -> промт -> Groq -> JSON -> валидатор -> Snapshot."""
    if current_user.role not in ['admin', 'executive']:
        flash('Нет прав на авто-анализ')
        return redirect(url_for('crm.crm_price_calculator'))

    import os as _os
    import json as _json
    from app.crm_validator import validate_competitor_row

    if not _os.environ.get('GROQ_API_KEY'):
        flash('GROQ_API_KEY не задан в окружении — авто-анализ недоступен')
        return redirect(url_for('crm.crm_price_calculator'))

    regions = request.form.getlist('regions')
    regions_str = ", ".join(regions) if regions else "ЦФО (Москва и МО, Тула, Воронеж, Калуга, Тверь и др.)"

    file = request.files.get('plant_file')
    if not file:
        flash('Список растений не загружен')
        return redirect(url_for('crm.crm_price_calculator'))

    try:
        items = _parse_plant_list_xlsx(file)
    except Exception as e:
        flash(f'Ошибка чтения файла: {e}')
        return redirect(url_for('crm.crm_price_calculator'))
    if not items:
        flash('В файле не найдено позиций')
        return redirect(url_for('crm.crm_price_calculator'))

    enriched = _enrich_items_with_our_prices(items)
    negatives = _collect_negative_examples(limit=8)

    # Бьём на батчи, чтобы уложиться в контекст и не терять на таймауте.
    # Держим batch небольшим, чтобы уложиться в gunicorn timeout (120с по умолчанию).
    BATCH_SIZE = 10
    batches = [enriched[i:i + BATCH_SIZE] for i in range(0, len(enriched), BATCH_SIZE)]
    groq_model = _os.environ.get('GROQ_MODEL', 'llama-3.3-70b-versatile')
    current_app.logger.info(
        'crm_ai_run: user=%s items=%s batches=%s model=%s regions=%r',
        current_user.id, len(enriched), len(batches), groq_model, regions_str[:120]
    )

    all_results = []
    raw_chunks = []
    errors = []
    for bi, batch in enumerate(batches, 1):
        prompt_text = _build_market_analysis_prompt(batch, regions_str, negatives=negatives)
        try:
            data, raw = _groq_json_request(prompt_text)
            raw_chunks.append(f"# BATCH {bi}/{len(batches)}\n{raw}")
            results = data.get('results') if isinstance(data, dict) else None
            if not isinstance(results, list):
                errors.append(f'batch #{bi}: поле results отсутствует или не list')
                current_app.logger.warning('crm_ai_run batch %s: no results list, keys=%r', bi, list(data.keys()) if isinstance(data, dict) else type(data).__name__)
                continue
            all_results.extend(results)
            current_app.logger.info('crm_ai_run batch %s/%s: got %s rows', bi, len(batches), len(results))
        except Exception as e:
            errors.append(f'batch #{bi}: {e}')
            current_app.logger.warning('crm_ai_run batch %s failed: %s', bi, e)

    if not all_results:
        details = ('; '.join(errors)) if errors else 'Модель не вернула ни одной строки'
        current_app.logger.warning('crm_ai_run: no results. details=%s', details)
        flash(f'Авто-анализ не дал результатов. {details}')
        return redirect(url_for('crm.crm_price_calculator'))

    # Себестоимость/наши цены — те же, что и в ручной ветке импорта.
    try:
        calc_year = msk_now().year - 1
        base_cost = float(calculate_cost_data(calc_year).get('cumulative_cost') or 0)
    except Exception:
        base_cost = 0.0
    # Быстрый индекс по enriched: для поиска нашей цены и haircut.
    our_idx = {(it['plant'].lower().strip(), it['size'].lower().strip()): it for it in enriched}

    # Создаём снапшот (обновляем имя, если уже есть за сегодня — как в ручной ветке).
    now = msk_now()
    start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    snap_name = f"Авто-анализ {now.strftime('%d.%m.%Y %H:%M')}"
    snapshot = CompetitorSnapshot.query.filter(
        CompetitorSnapshot.date >= start_of_day,
        CompetitorSnapshot.date <= end_of_day,
        CompetitorSnapshot.name.like('Авто-анализ %'),
    ).first()
    if snapshot:
        snapshot.name = snap_name
    else:
        snapshot = CompetitorSnapshot(name=snap_name, user_id=current_user.id, date=now)
        db.session.add(snapshot)
    db.session.flush()

    # Сохраняем сырой ответ (обрезаем до 200 КБ, чтобы не раздувать БД).
    raw_joined = "\n\n".join(raw_chunks)
    if len(raw_joined) > 200_000:
        raw_joined = raw_joined[:200_000] + '\n...[truncated]'
    snapshot.raw_ai_response = raw_joined

    clean_cnt = 0
    rejected_cnt = 0
    for item in all_results:
        if not isinstance(item, dict):
            continue
        p_name = str(item.get('plant_name') or '').strip()
        s_name = str(item.get('size') or '').strip()
        if not p_name:
            continue
        try:
            price = float(item.get('price') or 0)
        except (ValueError, TypeError):
            price = 0.0
        comp_name = str(item.get('competitor') or 'Неизвестно').strip()
        url_v = str(item.get('url') or '').strip() or None
        pack = str(item.get('pack_type') or '').strip()
        form_v = str(item.get('form') or '').strip()
        excerpt = str(item.get('source_excerpt') or '').strip()[:500]
        try:
            conf = float(item.get('confidence')) if item.get('confidence') is not None else None
        except (ValueError, TypeError):
            conf = None

        our_item = our_idx.get((p_name.lower().strip(), s_name.lower().strip())) or {
            'plant_name': p_name, 'size': s_name, 'haircut': '', 'our_price': 0,
        }

        vresult = validate_competitor_row({
            'plant_name': p_name,
            'size': s_name,
            'price': price,
            'pack_type': pack,
            'form': form_v,
            'source_excerpt': excerpt,
            'url': url_v or '',
        }, our_item={
            'plant_name': our_item.get('plant', p_name),
            'size': our_item.get('size', s_name),
            'haircut': our_item.get('haircut', ''),
            'our_price': our_item.get('our_price', 0),
        })
        is_rej = not vresult['ok']
        if is_rej:
            rejected_cnt += 1
        else:
            clean_cnt += 1

        db.session.add(CompetitorRow(
            snapshot_id=snapshot.id,
            plant_name=p_name,
            size_name=s_name or '-',
            competitor_name=comp_name,
            competitor_price=price,
            source_link=url_v,
            our_price_at_moment=float(our_item.get('our_price') or 0),
            our_cost_at_moment=base_cost,
            pack_type=pack or None,
            form=form_v or None,
            source_excerpt=excerpt or None,
            confidence=conf,
            is_rejected=is_rej,
            reject_reasons=_json.dumps(vresult['reasons'], ensure_ascii=False) if is_rej else None,
        ))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка сохранения результата: {e}')
        return redirect(url_for('crm.crm_price_calculator'))

    log_action(
        f"Авто-анализ рынка (Groq): чисто {clean_cnt}, отклонено {rejected_cnt}, батчей {len(batches)}"
    )
    err_suffix = f' Ошибки в {len(errors)} батчах, см. логи.' if errors else ''
    flash(
        f'Готово. Добавлено {clean_cnt + rejected_cnt} строк '
        f'(в «Отклонено» — {rejected_cnt}).{err_suffix}'
    )
    return redirect(url_for('crm.crm_price_calculator', snapshot_id=snapshot.id))


# ------------------------------------------------------------------
# ПОИСК ПО AVITO: своя выдача + валидатор, сохраняется как Snapshot.
# ------------------------------------------------------------------

@bp.route('/crm/price_calculator/avito_scan', methods=['POST'])
@login_required
def crm_avito_scan():
    """Сканирует Avito по позициям из загруженного XLSX и сохраняет
    найденное в CompetitorSnapshot. Работает без Groq — чистый парсинг."""
    if current_user.role not in ['admin', 'executive']:
        flash('Нет прав на поиск по Avito')
        return redirect(url_for('crm.crm_price_calculator'))

    import os as _os
    import json as _json
    import time as _time
    import requests as _requests
    from app.crm_validator import validate_competitor_row
    from app import crm_avito

    regions = request.form.getlist('regions')
    region_slug = crm_avito.pick_region_slug(regions)

    file = request.files.get('plant_file')
    if not file:
        flash('Список растений не загружен (используйте форму «Создать Промт»).')
        return redirect(url_for('crm.crm_price_calculator'))

    try:
        items = _parse_plant_list_xlsx(file)
    except Exception as e:
        flash(f'Ошибка чтения файла: {e}')
        return redirect(url_for('crm.crm_price_calculator'))
    if not items:
        flash('В файле не найдено позиций')
        return redirect(url_for('crm.crm_price_calculator'))

    # Ограничимся 40 позициями, чтобы не получить бан за массовый парсинг.
    MAX_ITEMS = 40
    if len(items) > MAX_ITEMS:
        flash(f'В файле {len(items)} позиций — отсканируем первые {MAX_ITEMS}, '
              f'чтобы не нарваться на блокировку Avito.')
        items = items[:MAX_ITEMS]

    enriched = _enrich_items_with_our_prices(items)
    try:
        calc_year = msk_now().year - 1
        base_cost = float(calculate_cost_data(calc_year).get('cumulative_cost') or 0)
    except Exception:
        base_cost = 0.0

    now = msk_now()
    snap_name = f"Avito-поиск {now.strftime('%d.%m.%Y %H:%M')}"
    snapshot = CompetitorSnapshot(name=snap_name, user_id=current_user.id, date=now)
    db.session.add(snapshot)
    db.session.flush()

    session = _requests.Session()
    clean_cnt = 0
    rejected_cnt = 0
    no_result_cnt = 0
    blocked = False
    errors = []

    current_app.logger.info(
        'crm_avito_scan: user=%s items=%s region_slug=%s',
        current_user.id, len(items), region_slug
    )

    for idx, it in enumerate(enriched):
        query = crm_avito.build_query_for_item(it.get('plant', ''), it.get('size', ''))
        if not query:
            no_result_cnt += 1
            continue
        try:
            listings = crm_avito.search(query, region_slug=region_slug, session=session)
        except crm_avito.AvitoBlockedError as e:
            current_app.logger.warning('crm_avito_scan: blocked on item %s (%r): %s', idx, query, e)
            errors.append(str(e))
            blocked = True
            break
        except Exception as e:
            current_app.logger.warning('crm_avito_scan: unexpected error on %r: %s', query, e)
            errors.append(f'{query}: {e}')
            continue

        # По каждой позиции — максимум 10 первых карточек, чтобы не раздувать базу
        # и держать результат релевантным (Avito сортирует по дате).
        for lst in listings[:10]:
            our_item = {
                'plant_name': it.get('plant', ''),
                'size': it.get('size', ''),
                'haircut': it.get('haircut', ''),
                'our_price': it.get('our_price', 0),
            }
            raw = {
                'plant_name': it.get('plant', ''),
                'size': it.get('size', ''),
                'price': lst.price,
                'pack_type': '',
                'form': '',
                'source_excerpt': lst.excerpt,
                'url': lst.url,
            }
            vresult = validate_competitor_row(raw, our_item=our_item)
            is_rej = not vresult['ok']
            if is_rej:
                rejected_cnt += 1
            else:
                clean_cnt += 1
            db.session.add(CompetitorRow(
                snapshot_id=snapshot.id,
                plant_name=it.get('plant', ''),
                size_name=it.get('size', '') or '-',
                competitor_name=f'Avito ({lst.location})' if lst.location else 'Avito',
                competitor_price=lst.price,
                source_link=lst.url,
                our_price_at_moment=float(it.get('our_price') or 0),
                our_cost_at_moment=base_cost,
                pack_type=None,
                form=None,
                source_excerpt=lst.excerpt[:500] if lst.excerpt else None,
                confidence=None,
                is_rejected=is_rej,
                reject_reasons=_json.dumps(vresult['reasons'], ensure_ascii=False) if is_rej else None,
            ))

        if not listings:
            no_result_cnt += 1

        # Вежливая пауза, чтобы не получить rate-limit.
        _time.sleep(1.0)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.warning('crm_avito_scan commit failed: %s', e)
        flash(f'Ошибка сохранения результата: {e}')
        return redirect(url_for('crm.crm_price_calculator'))

    total = clean_cnt + rejected_cnt
    if blocked:
        flash(
            f'Avito временно заблокировал запросы. Успели собрать {total} строк '
            f'(в «Отклонено» — {rejected_cnt}). Попробуйте позже или запустите '
            f'локально с российским IP.'
        )
    elif total == 0:
        flash('Avito не вернул ни одной подходящей карточки. Возможно, блок или изменилась вёрстка — см. логи.')
    else:
        flash(
            f'Готово (Avito). Добавлено {total} строк '
            f'(в «Отклонено» — {rejected_cnt}; позиций без результатов: {no_result_cnt}).'
        )
    log_action(f'Поиск Avito: чисто {clean_cnt}, отклонено {rejected_cnt}, позиций {len(items)}')
    return redirect(url_for('crm.crm_price_calculator', snapshot_id=snapshot.id))