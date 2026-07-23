"""Блюпринт раздела «Учет ВиУМ» (вспомогательные материалы).

Раздел доступен только пользователям с ролями `admin` и `executive`.
Маршруты:

* `GET  /vium`                           — отчёт остатков
* `GET  /vium/operations`                — журнал операций (с фильтрами)
* `GET  /vium/material/<id>`             — карточка материала (партии + история)
* `POST /vium/material/new`              — создать материал
* `POST /vium/material/<id>/edit`        — редактировать материал
* `POST /vium/material/<id>/archive`     — архив / разархив
* `GET  /vium/operation/new`             — форма создания операции
* `POST /vium/operation/new`             — сохранить операцию
* `GET  /vium/operation/<id>`            — карточка операции
* `GET  /vium/export`                    — Excel остатков
* `GET  /vium/inbox`                     — очередь оцифровки PDF-счетов
* `GET  /vium/inbox/<id>`                — карточка очереди (фазы 2/3)
* `POST /vium/inbox/<id>/parse`          — запустить парсер (фаза 3)
* `POST /vium/inbox/<id>/commit`         — провести поступление (фаза 3)
* `POST /vium/inbox/<id>/skip`           — пометить «не оцифровывать»
"""
from __future__ import annotations

import io
import os
import json
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash,
    send_file, abort, current_app, send_from_directory,
)
from flask_login import login_required, current_user
from sqlalchemy import or_, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import (
    db,
    ViumMaterial, ViumLot, ViumOperation, ViumOperationLine,
    ViumInvoiceQueue, ViumMaterialAlias,
    ViumTechCardLine, ViumPlannedConsume,
    PaymentInvoice,
    Plant, Size,
)
from app import vium_service, vium_techcard, vium_fot
from app.utils import msk_now

bp = Blueprint('vium', __name__, url_prefix='/vium')


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

ALLOWED_ROLES = ('admin', 'executive')

UNIT_CHOICES = ['шт', 'кг', 'м', 'м2', 'рулон', 'упак', 'л']
KIND_CHOICES = [
    ('intake',   'Поступление'),
    ('consume',  'Расход'),
    ('writeoff', 'Списание'),
    ('adjust',   'Корректировка'),
]


def _check_access():
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))
    if current_user.role not in ALLOWED_ROLES:
        flash('Раздел «Учет ВиУМ» доступен только администратору и руководителю.')
        return redirect(url_for('directory.directory'))
    return None


def _parse_decimal(raw, default=None):
    if raw in (None, ''):
        return default
    try:
        return Decimal(str(raw).replace(',', '.').strip())
    except (InvalidOperation, ValueError):
        return default


def _parse_date(raw):
    if not raw:
        return msk_now().date()
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except Exception:
        return msk_now().date()


# ---------------------------------------------------------------------------
# Отчёт остатков
# ---------------------------------------------------------------------------

@bp.route('')
@login_required
def vium_report():
    deny = _check_access()
    if deny is not None:
        return deny

    show_archived = request.args.get('show_archived') == '1'

    q = ViumMaterial.query
    if not show_archived:
        q = q.filter(ViumMaterial.is_archived == False)  # noqa: E712
    materials = q.order_by(ViumMaterial.name.asc()).all()

    overview = vium_service.materials_overview()

    rows = []
    total_value = Decimal('0')
    for m in materials:
        info = overview.get(m.id, {
            'qty': Decimal('0'),
            'value': Decimal('0'),
            'avg_price': Decimal('0'),
            'lots': 0,
        })
        rows.append({
            'material': m,
            'qty': info['qty'],
            'value': info['value'],
            'avg_price': info['avg_price'],
            'lots': info['lots'],
        })
        total_value += info['value']

    return render_template(
        'vium/vium_report.html',
        rows=rows,
        total_value=total_value,
        show_archived=show_archived,
    )


# ---------------------------------------------------------------------------
# Журнал операций
# ---------------------------------------------------------------------------

@bp.route('/operations')
@login_required
def vium_operations():
    deny = _check_access()
    if deny is not None:
        return deny

    f_kind = request.args.get('kind') or ''
    f_material = request.args.get('material_id') or ''
    f_start = request.args.get('start_date') or ''
    f_end = request.args.get('end_date') or ''

    q = ViumOperation.query

    if f_kind:
        q = q.filter(ViumOperation.kind == f_kind)
    if f_material:
        try:
            mid = int(f_material)
            q = q.join(ViumOperationLine, ViumOperationLine.operation_id == ViumOperation.id).filter(
                ViumOperationLine.material_id == mid
            ).distinct()
        except ValueError:
            pass
    if f_start:
        try:
            d = datetime.strptime(f_start, '%Y-%m-%d').date()
            q = q.filter(ViumOperation.date >= d)
        except Exception:
            pass
    if f_end:
        try:
            d = datetime.strptime(f_end, '%Y-%m-%d').date()
            q = q.filter(ViumOperation.date <= d)
        except Exception:
            pass

    operations = q.order_by(ViumOperation.date.desc(), ViumOperation.id.desc()).limit(500).all()

    materials = ViumMaterial.query.order_by(ViumMaterial.name.asc()).all()

    return render_template(
        'vium/vium_operations.html',
        operations=operations,
        materials=materials,
        kind_choices=KIND_CHOICES,
        f_kind=f_kind,
        f_material=f_material,
        f_start=f_start,
        f_end=f_end,
    )


@bp.route('/operation/<int:op_id>')
@login_required
def vium_operation_view(op_id: int):
    deny = _check_access()
    if deny is not None:
        return deny
    op = ViumOperation.query.get_or_404(op_id)

    # Расширяем lot_consumption — для UI красиво показать партии.
    lines_view = []
    for line in op.lines:
        lines_view.append({
            'line': line,
            'consumption': vium_service.parse_lot_consumption(line.lot_consumption),
        })

    return render_template(
        'vium/vium_operation_view.html',
        op=op,
        lines_view=lines_view,
        kind_label=vium_service.op_kind_label(op.kind),
    )


# ---------------------------------------------------------------------------
# Карточка материала
# ---------------------------------------------------------------------------

@bp.route('/material/<int:material_id>')
@login_required
def vium_material_view(material_id: int):
    deny = _check_access()
    if deny is not None:
        return deny
    m = ViumMaterial.query.get_or_404(material_id)

    bal = vium_service.current_balance(material_id)
    lots = vium_service.list_active_lots(material_id)

    history_lines = (
        ViumOperationLine.query
        .filter(ViumOperationLine.material_id == material_id)
        .join(ViumOperation, ViumOperation.id == ViumOperationLine.operation_id)
        .order_by(ViumOperation.date.desc(), ViumOperation.id.desc())
        .limit(200)
        .all()
    )

    return render_template(
        'vium/vium_material.html',
        m=m,
        balance=bal,
        lots=lots,
        history_lines=history_lines,
        unit_choices=UNIT_CHOICES,
    )


# ---------------------------------------------------------------------------
# Справочник: создать / редактировать
# ---------------------------------------------------------------------------

@bp.route('/material/new', methods=['POST'])
@login_required
def vium_material_create():
    deny = _check_access()
    if deny is not None:
        return deny

    name = (request.form.get('name') or '').strip()
    unit = (request.form.get('unit') or 'шт').strip() or 'шт'
    desc = (request.form.get('description') or '').strip() or None

    if not name:
        flash('Название материала не может быть пустым.')
        return redirect(request.referrer or url_for('vium.vium_report'))

    if ViumMaterial.query.filter(func.lower(ViumMaterial.name) == name.lower()).first():
        flash(f'Материал «{name}» уже есть в справочнике.')
        return redirect(request.referrer or url_for('vium.vium_report'))

    m = ViumMaterial(name=name, unit=unit, description=desc)
    db.session.add(m)
    db.session.commit()
    flash(f'Материал «{name}» добавлен.')
    return redirect(url_for('vium.vium_material_view', material_id=m.id))


@bp.route('/material/<int:material_id>/edit', methods=['POST'])
@login_required
def vium_material_edit(material_id: int):
    deny = _check_access()
    if deny is not None:
        return deny

    m = ViumMaterial.query.get_or_404(material_id)
    name = (request.form.get('name') or '').strip()
    unit = (request.form.get('unit') or '').strip()
    desc = (request.form.get('description') or '').strip() or None

    if not name:
        flash('Название материала не может быть пустым.')
        return redirect(url_for('vium.vium_material_view', material_id=m.id))

    dup = ViumMaterial.query.filter(
        func.lower(ViumMaterial.name) == name.lower(),
        ViumMaterial.id != m.id,
    ).first()
    if dup:
        flash(f'Материал «{name}» уже есть.')
        return redirect(url_for('vium.vium_material_view', material_id=m.id))

    m.name = name
    if unit:
        m.unit = unit
    m.description = desc
    db.session.commit()
    flash('Материал обновлён.')
    return redirect(url_for('vium.vium_material_view', material_id=m.id))


@bp.route('/material/<int:material_id>/archive', methods=['POST'])
@login_required
def vium_material_archive(material_id: int):
    deny = _check_access()
    if deny is not None:
        return deny
    m = ViumMaterial.query.get_or_404(material_id)
    m.is_archived = not bool(m.is_archived)
    db.session.commit()
    flash('Материал {}.'.format('в архиве' if m.is_archived else 'возвращён в работу'))
    return redirect(request.referrer or url_for('vium.vium_report'))


# ---------------------------------------------------------------------------
# Создание операции (приход / расход / списание / корректировка)
# ---------------------------------------------------------------------------

@bp.route('/operation/new', methods=['GET', 'POST'])
@login_required
def vium_operation_new():
    deny = _check_access()
    if deny is not None:
        return deny

    materials = (
        ViumMaterial.query
        .filter(ViumMaterial.is_archived == False)  # noqa: E712
        .order_by(ViumMaterial.name.asc())
        .all()
    )

    if request.method == 'GET':
        prefill_kind = request.args.get('kind') or 'intake'
        return render_template(
            'vium/vium_operation_form.html',
            materials=materials,
            kind_choices=KIND_CHOICES,
            prefill_kind=prefill_kind,
            today=msk_now().date(),
        )

    kind = (request.form.get('kind') or 'intake').strip()
    if kind not in {k for k, _ in KIND_CHOICES}:
        flash('Неизвестный тип операции.')
        return redirect(url_for('vium.vium_operation_new'))

    op_date = _parse_date(request.form.get('date'))
    comment = (request.form.get('comment') or '').strip() or None

    raw_material_ids = request.form.getlist('material_id[]')
    raw_qtys = request.form.getlist('qty[]')
    raw_prices = request.form.getlist('unit_price[]')
    raw_notes = request.form.getlist('note[]')

    op = ViumOperation(
        kind=kind,
        date=op_date,
        comment=comment,
        created_by_user_id=getattr(current_user, 'id', None),
    )
    db.session.add(op)
    db.session.flush()

    line_count = 0
    for idx, mid_raw in enumerate(raw_material_ids):
        mid_raw = (mid_raw or '').strip()
        if not mid_raw:
            continue
        try:
            mid = int(mid_raw)
        except ValueError:
            continue
        qty = _parse_decimal(raw_qtys[idx] if idx < len(raw_qtys) else None, default=Decimal('0'))
        if not qty or qty == 0:
            continue
        if kind == 'consume' or kind == 'writeoff':
            if qty < 0:
                qty = -qty
        price = _parse_decimal(raw_prices[idx] if idx < len(raw_prices) else None)
        note = (raw_notes[idx] if idx < len(raw_notes) else '') or None
        if note:
            note = note.strip() or None

        if kind == 'intake' and (price is None or price < 0):
            price = Decimal('0')

        line = ViumOperationLine(
            operation_id=op.id,
            material_id=mid,
            qty=qty,
            unit_price=price,
            note=note,
        )
        db.session.add(line)
        line_count += 1

    if line_count == 0:
        db.session.rollback()
        flash('Нужно добавить хотя бы одну строку.')
        return redirect(url_for('vium.vium_operation_new'))

    try:
        db.session.flush()
        vium_service.apply_operation(op)
        db.session.commit()
    except ValueError as e:
        db.session.rollback()
        flash(f'Ошибка проведения операции: {e}')
        return redirect(url_for('vium.vium_operation_new'))
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('vium operation save failed')
        flash(f'Не удалось сохранить операцию: {e}')
        return redirect(url_for('vium.vium_operation_new'))

    flash('Операция «{}» проведена.'.format(vium_service.op_kind_label(op.kind)))
    return redirect(url_for('vium.vium_operation_view', op_id=op.id))


# ---------------------------------------------------------------------------
# Excel-выгрузка остатков
# ---------------------------------------------------------------------------

@bp.route('/export')
@login_required
def vium_export():
    deny = _check_access()
    if deny is not None:
        return deny

    show_archived = request.args.get('show_archived') == '1'

    q = ViumMaterial.query
    if not show_archived:
        q = q.filter(ViumMaterial.is_archived == False)  # noqa: E712
    materials = q.order_by(ViumMaterial.name.asc()).all()

    overview = vium_service.materials_overview()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Остатки ВиУМ'

    head_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    head_font = Font(bold=True, color='FFFFFF', size=11)
    total_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    total_font = Font(bold=True, color='1B5E20')
    thin = Side(style='thin', color='BDBDBD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centered = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    headers = [
        'Материал', 'Ед.изм.', 'Остаток', 'Партий', 'Ср. цена, ₽',
        'Стоимость, ₽', 'Описание',
    ]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = centered
        cell.border = border

    total_value = Decimal('0')
    row = 2
    for m in materials:
        info = overview.get(m.id, {
            'qty': Decimal('0'),
            'value': Decimal('0'),
            'avg_price': Decimal('0'),
            'lots': 0,
        })
        ws.cell(row=row, column=1, value=m.name).alignment = left
        ws.cell(row=row, column=2, value=m.unit).alignment = centered
        ws.cell(row=row, column=3, value=float(info['qty'])).alignment = right
        ws.cell(row=row, column=4, value=int(info['lots'])).alignment = centered
        ws.cell(row=row, column=5, value=float(info['avg_price'])).alignment = right
        ws.cell(row=row, column=6, value=float(info['value'])).alignment = right
        ws.cell(row=row, column=7, value=m.description or '').alignment = left
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = border
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=3).number_format = '#,##0.000'
        if m.is_archived:
            italic = Font(italic=True, color='9E9E9E')
            for c in range(1, 8):
                ws.cell(row=row, column=c).font = italic
        total_value += info['value']
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value='ИТОГО')
    ws.cell(row=total_row, column=6, value=float(total_value))
    ws.cell(row=total_row, column=6).number_format = '#,##0.00'
    for c in range(1, 8):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        if c == 6:
            cell.alignment = right
        elif c == 1:
            cell.alignment = left
        else:
            cell.alignment = centered

    widths = [38, 10, 12, 10, 14, 16, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f'Остатки ВиУМ {msk_now().strftime("%d.%m.%Y")}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=fn,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ---------------------------------------------------------------------------
# Инбокс оцифровки PDF (фаза 2 — заглушка, фаза 3 — реальный парсер)
# ---------------------------------------------------------------------------

@bp.route('/inbox')
@login_required
def vium_inbox():
    deny = _check_access()
    if deny is not None:
        return deny

    f_status = request.args.get('status') or 'open'
    q = ViumInvoiceQueue.query.join(
        PaymentInvoice, PaymentInvoice.id == ViumInvoiceQueue.invoice_id
    )
    if f_status == 'open':
        q = q.filter(ViumInvoiceQueue.status.in_(('new', 'parsing', 'ready', 'error')))
    elif f_status == 'done':
        q = q.filter(ViumInvoiceQueue.status == 'done')
    elif f_status == 'skipped':
        q = q.filter(ViumInvoiceQueue.status == 'skipped')
    items = q.order_by(ViumInvoiceQueue.created_at.desc()).limit(300).all()

    return render_template('vium/vium_inbox.html', items=items, f_status=f_status)


@bp.route('/inbox/<int:item_id>')
@login_required
def vium_inbox_view(item_id: int):
    deny = _check_access()
    if deny is not None:
        return deny

    item = ViumInvoiceQueue.query.get_or_404(item_id)
    materials = (
        ViumMaterial.query
        .filter(ViumMaterial.is_archived == False)  # noqa: E712
        .order_by(ViumMaterial.name.asc())
        .all()
    )

    parsed_lines = []
    payload = item.parsed_payload
    if payload:
        try:
            parsed_lines = json.loads(payload)
            if not isinstance(parsed_lines, list):
                parsed_lines = []
        except Exception:
            parsed_lines = []

    # Подготовим suggest для каждой строки (фаза 3)
    suggested = []
    try:
        from app.vium_inbox import suggest_material  # фаза 3
    except Exception:
        suggest_material = None
    for ln in parsed_lines:
        desc = (ln.get('description') if isinstance(ln, dict) else '') or ''
        if suggest_material is not None and desc:
            try:
                suggested.append(suggest_material(desc))
            except Exception:
                suggested.append(None)
        else:
            suggested.append(None)

    return render_template(
        'vium/vium_inbox_card.html',
        item=item,
        materials=materials,
        parsed_lines=parsed_lines,
        suggested=suggested,
    )


@bp.route('/inbox/<int:item_id>/parse', methods=['POST'])
@login_required
def vium_inbox_parse(item_id: int):
    deny = _check_access()
    if deny is not None:
        return deny

    item = ViumInvoiceQueue.query.get_or_404(item_id)
    try:
        from app.vium_inbox import parse_queue_item  # фаза 3
    except ImportError:
        flash('Парсер PDF пока не подключён (фаза 3).')
        return redirect(url_for('vium.vium_inbox_view', item_id=item.id))

    try:
        parse_queue_item(item)
        db.session.commit()
        flash('Счёт распознан, проверьте позиции.')
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('vium parse failed')
        flash(f'Не удалось распарсить счёт: {e}')
    return redirect(url_for('vium.vium_inbox_view', item_id=item.id))


@bp.route('/inbox/<int:item_id>/commit', methods=['POST'])
@login_required
def vium_inbox_commit(item_id: int):
    deny = _check_access()
    if deny is not None:
        return deny

    item = ViumInvoiceQueue.query.get_or_404(item_id)

    raw_materials = request.form.getlist('material_id[]')
    raw_qtys = request.form.getlist('qty[]')
    raw_prices = request.form.getlist('unit_price[]')
    raw_descs = request.form.getlist('description[]')
    raw_skip = set(request.form.getlist('skip[]'))

    lines_payload = []
    for idx, mid_raw in enumerate(raw_materials):
        if str(idx) in raw_skip:
            continue
        mid_raw = (mid_raw or '').strip()
        if not mid_raw:
            continue
        try:
            mid = int(mid_raw)
        except ValueError:
            continue
        qty = _parse_decimal(raw_qtys[idx] if idx < len(raw_qtys) else None, default=Decimal('0'))
        if not qty or qty == 0:
            continue
        price = _parse_decimal(raw_prices[idx] if idx < len(raw_prices) else None, default=Decimal('0'))
        desc = (raw_descs[idx] if idx < len(raw_descs) else '') or ''
        lines_payload.append({
            'material_id': mid,
            'qty': qty,
            'unit_price': price or Decimal('0'),
            'description': desc.strip(),
        })

    if not lines_payload:
        flash('Нужно подтвердить хотя бы одну позицию.')
        return redirect(url_for('vium.vium_inbox_view', item_id=item.id))

    try:
        from app.vium_inbox import commit_intake  # фаза 3
    except ImportError:
        # Фаза 2: запасной путь — простое создание ViumOperation/Lot вручную.
        commit_intake = None

    try:
        if commit_intake is not None:
            op = commit_intake(item, lines_payload, user_id=getattr(current_user, 'id', None))
        else:
            op = _commit_intake_fallback(item, lines_payload)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('vium commit failed')
        flash(f'Не удалось провести поступление: {e}')
        return redirect(url_for('vium.vium_inbox_view', item_id=item.id))

    flash(f'Поступление #{op.id} проведено.')
    return redirect(url_for('vium.vium_operation_view', op_id=op.id))


def _commit_intake_fallback(item: ViumInvoiceQueue, lines_payload: list[dict]) -> ViumOperation:
    """Простое проведение из инбокса без обучающих алиасов (фаза 2)."""
    op = ViumOperation(
        kind='intake',
        date=msk_now().date(),
        comment=(item.invoice.original_name if item.invoice else None),
        invoice_id=item.invoice_id,
        created_by_user_id=getattr(current_user, 'id', None),
    )
    db.session.add(op)
    db.session.flush()
    for ln in lines_payload:
        line = ViumOperationLine(
            operation_id=op.id,
            material_id=int(ln['material_id']),
            qty=Decimal(str(ln['qty'])),
            unit_price=Decimal(str(ln.get('unit_price') or 0)),
            note=(ln.get('description') or None),
        )
        db.session.add(line)
    db.session.flush()
    vium_service.apply_intake(op)

    item.status = 'done'
    item.operation_id = op.id
    item.processed_at = datetime.now()
    return op


@bp.route('/inbox/<int:item_id>/skip', methods=['POST'])
@login_required
def vium_inbox_skip(item_id: int):
    deny = _check_access()
    if deny is not None:
        return deny
    item = ViumInvoiceQueue.query.get_or_404(item_id)
    item.status = 'skipped'
    item.processed_at = datetime.now()
    db.session.commit()
    flash('Счёт помечен как «не оцифровывать».')
    return redirect(url_for('vium.vium_inbox'))


# ---------------------------------------------------------------------------
# Тех.карты ВиУМ (фаза 4)
# ---------------------------------------------------------------------------

@bp.route('/techcards')
@login_required
def vium_techcards():
    deny = _check_access()
    if deny is not None:
        return deny

    pairs = vium_techcard.all_pairs_with_cards()
    plants = Plant.query.order_by(Plant.name.asc()).all()
    sizes = Size.query.order_by(Size.name.asc()).all()
    return render_template(
        'vium/techcards.html',
        pairs=pairs, plants=plants, sizes=sizes,
    )


@bp.route('/techcards/<int:plant_id>/<int:size_id>')
@login_required
def vium_techcard_edit(plant_id: int, size_id: int):
    deny = _check_access()
    if deny is not None:
        return deny

    plant = Plant.query.get_or_404(plant_id)
    size = Size.query.get_or_404(size_id)
    lines = vium_techcard.card_for(plant_id, size_id)
    materials = ViumMaterial.query.filter_by(is_archived=False).order_by(ViumMaterial.name.asc()).all()
    return render_template(
        'vium/techcard_edit.html',
        plant=plant, size=size, lines=lines, materials=materials,
    )


@bp.route('/techcards/<int:plant_id>/<int:size_id>/save', methods=['POST'])
@login_required
def vium_techcard_save(plant_id: int, size_id: int):
    deny = _check_access()
    if deny is not None:
        return deny

    plant = Plant.query.get_or_404(plant_id)
    size = Size.query.get_or_404(size_id)

    # Удалить отмеченные строки
    delete_ids = set()
    for key in request.form.keys():
        if key.startswith('delete_line_'):
            try:
                delete_ids.add(int(key.split('_')[-1]))
            except (TypeError, ValueError):
                pass
    if delete_ids:
        for ln in ViumTechCardLine.query.filter(ViumTechCardLine.id.in_(delete_ids)).all():
            db.session.delete(ln)

    # Существующие строки — обновить qty/note (или удалить если qty=0)
    existing = {
        ln.id: ln for ln in vium_techcard.card_for(plant_id, size_id)
        if ln.id not in delete_ids
    }
    for line_id, ln in existing.items():
        qty_raw = request.form.get(f'qty_{line_id}')
        note_raw = request.form.get(f'note_{line_id}', '')
        mat_raw = request.form.get(f'mat_{line_id}')
        if qty_raw is None:
            continue
        qty = _parse_decimal(qty_raw)
        if qty is None or qty <= 0:
            db.session.delete(ln)
            continue
        ln.qty_per_bush = qty
        ln.note = (note_raw or '').strip() or None
        if mat_raw:
            try:
                ln.material_id = int(mat_raw)
            except (TypeError, ValueError):
                pass

    # Новые строки (форма new_*)
    new_ids = set()
    for key in request.form.keys():
        if key.startswith('new_mat_'):
            try:
                new_ids.add(int(key.split('_')[-1]))
            except (TypeError, ValueError):
                pass
    for nid in new_ids:
        mat_raw = request.form.get(f'new_mat_{nid}')
        qty_raw = request.form.get(f'new_qty_{nid}')
        note_raw = request.form.get(f'new_note_{nid}', '')
        if not mat_raw:
            continue
        try:
            mat_id = int(mat_raw)
        except (TypeError, ValueError):
            continue
        qty = _parse_decimal(qty_raw)
        if qty is None or qty <= 0:
            continue
        # Если такая пара+материал уже была в БД (включая неактивную) — обновим.
        existing_any = ViumTechCardLine.query.filter_by(
            plant_id=plant_id, size_id=size_id, material_id=mat_id
        ).first()
        if existing_any:
            existing_any.qty_per_bush = qty
            existing_any.note = (note_raw or '').strip() or None
            existing_any.is_active = True
        else:
            db.session.add(ViumTechCardLine(
                plant_id=plant_id, size_id=size_id, material_id=mat_id,
                qty_per_bush=qty, note=(note_raw or '').strip() or None,
                is_active=True,
            ))

    try:
        db.session.commit()
        flash(f'Тех.карта сохранена: {plant.name} · {size.name}', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('vium_techcard_save failed')
        flash(f'Не удалось сохранить тех.карту: {e}', 'danger')

    return redirect(url_for('vium.vium_techcard_edit', plant_id=plant_id, size_id=size_id))


# ---------------------------------------------------------------------------
# Плановый отчёт по ВиУМ (фаза 4)
# ---------------------------------------------------------------------------

def _resolve_plan_period() -> tuple[date, date]:
    today = msk_now().date()
    s = _parse_date(request.args.get('start'))
    e = _parse_date(request.args.get('end'))
    if not s:
        s = today.replace(day=1)
    if not e:
        if s.month == 12:
            e = date(s.year, 12, 31)
        else:
            e = date(s.year, s.month + 1, 1) - timedelta(days=1)
    if e < s:
        s, e = e, s
    return s, e


@bp.route('/plan-report')
@login_required
def vium_plan_report():
    deny = _check_access()
    if deny is not None:
        return deny

    start, end = _resolve_plan_period()
    materials = vium_techcard.materials_plan_overview(start, end)
    fot_per_unit = vium_fot.period_fot_per_unit(start, end)
    fot_amount = vium_fot.period_fot_amount(start, end)
    dug_qty = vium_fot.period_dug_qty(start, end)
    missing_dates = vium_fot.missing_hours_dates(start, end)
    pairs = vium_techcard.pairs_plan_overview(start, end, fot_per_unit=fot_per_unit)

    return render_template(
        'vium/plan_report.html',
        start=start, end=end,
        materials=materials, pairs=pairs,
        fot_per_unit=fot_per_unit,
        fot_amount=fot_amount,
        dug_qty=dug_qty,
        missing_dates=missing_dates,
    )


@bp.route('/plan-report/export')
@login_required
def vium_plan_report_export():
    deny = _check_access()
    if deny is not None:
        return deny

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    start, end = _resolve_plan_period()
    materials = vium_techcard.materials_plan_overview(start, end)
    fot_per_unit = vium_fot.period_fot_per_unit(start, end)
    pairs = vium_techcard.pairs_plan_overview(start, end, fot_per_unit=fot_per_unit)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Материалы'
    headers = [
        'Материал', 'Ед.',
        'Приход за период', 'Факт-расход за период',
        'План-расход за период',
        'План-остаток (накопл.)', 'Факт-остаток (партии)',
        'Ср. цена, ₽',
    ]
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='FFE5E5E5')
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

    row = 2
    for m in materials:
        ws.cell(row=row, column=1, value=m['material'].name)
        ws.cell(row=row, column=2, value=m['material'].unit)
        ws.cell(row=row, column=3, value=float(m['intake_period'] or 0))
        ws.cell(row=row, column=4, value=float(m['real_consume_period'] or 0))
        ws.cell(row=row, column=5, value=float(m['planned_consume_period'] or 0))
        ws.cell(row=row, column=6, value=float(m['plan_balance'] or 0))
        ws.cell(row=row, column=7, value=float(m['fact_balance'] or 0))
        ws.cell(row=row, column=8, value=float(m['avg_price'] or 0))
        row += 1

    ws2 = wb.create_sheet('Пары растение-размер')
    headers2 = [
        'Растение', 'Размер', 'Выкопано за период',
        'Сумма по тех.карте, ₽',
        'ФОТ ₽/шт', 'ФОТ за выкопку, ₽',
        'Итого с ФОТ, ₽',
    ]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

    row = 2
    for p in pairs:
        ws2.cell(row=row, column=1, value=p['plant_name'])
        ws2.cell(row=row, column=2, value=p['size_name'])
        ws2.cell(row=row, column=3, value=int(p['dug_qty'] or 0))
        ws2.cell(row=row, column=4, value=float(p['total_value'] or 0))
        ws2.cell(row=row, column=5, value=float(p.get('fot_unit') or 0))
        ws2.cell(row=row, column=6, value=float(p.get('fot_value') or 0))
        ws2.cell(row=row, column=7, value=float(p.get('total_with_fot') or 0))
        row += 1

    for sheet in (ws, ws2):
        for col in sheet.columns:
            try:
                length = max(len(str(c.value or '')) for c in col)
            except ValueError:
                length = 12
            sheet.column_dimensions[col[0].column_letter].width = max(12, length + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f'vium_plan_report_{start:%Y%m%d}_{end:%Y%m%d}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=fn,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
