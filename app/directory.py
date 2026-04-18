import io
import os
import time
from werkzeug.utils import secure_filename
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, current_app, send_from_directory, jsonify
from flask_login import login_required, current_user
from openpyxl import Workbook, load_workbook
from app.models import db, Plant, Size, Field, Client, Supplier, Document, DocumentRow, StockBalance, AppSetting, FileArchive
from app.utils import log_action, get_or_create_stock, msk_now, natural_key

bp = Blueprint('directory', __name__)

@bp.route('/directory', methods=['GET', 'POST'])
@login_required
def directory():
    if request.method == 'POST':
        if current_user.role != 'admin': 
            return redirect(url_for('directory.directory'))
        
        action = request.form.get('action')
        type_ = request.form.get('type')
        model = {'plant': Plant, 'size': Size, 'field': Field, 'client': Client, 'supplier': Supplier}.get(type_)
        
        if not model: 
            return redirect(url_for('directory.directory'))
        
        try:
            if action == 'add':
                name = request.form.get('name')
                if type_ == 'field': 
                    db.session.add(model(
                        name=name, 
                        investor_id=(int(request.form.get('investor_id')) if request.form.get('investor_id') else None), 
                        planting_year=(int(request.form.get('planting_year')) if request.form.get('planting_year') else 2017)
                    ))
                elif type_ == 'client':
                    fixed_balance = request.form.get('fixed_balance')
                    fixed_balance_date = request.form.get('fixed_balance_date')
                    from datetime import datetime
                    fb_val = float(fixed_balance) if fixed_balance else None
                    fbd_val = datetime.strptime(fixed_balance_date, '%Y-%m-%d').date() if fixed_balance_date else None
                    db.session.add(model(name=name, fixed_balance=fb_val, fixed_balance_date=fbd_val))
                else: 
                    db.session.add(model(name=name))
                db.session.commit()
                flash('Запись добавлена')
                log_action(f"Добавил в справочник {type_}: {name}")
                
            elif action == 'edit':
                item = model.query.get(request.form.get('id'))
                if item:
                    old_name = item.name
                    item.name = request.form.get('name')
                    if type_ == 'field': 
                        item.investor_id = int(request.form.get('investor_id')) if request.form.get('investor_id') else None
                        item.planting_year = int(request.form.get('planting_year')) if request.form.get('planting_year') else 2017
                    elif type_ == 'client':
                        fixed_balance = request.form.get('fixed_balance')
                        fixed_balance_date = request.form.get('fixed_balance_date')
                        from datetime import datetime
                        item.fixed_balance = float(fixed_balance) if fixed_balance else None
                        item.fixed_balance_date = datetime.strptime(fixed_balance_date, '%Y-%m-%d').date() if fixed_balance_date else None
                    db.session.commit()
                    flash('Обновлено')
                    log_action(f"Изменил {type_}: {old_name} -> {item.name}")
                    
            elif action == 'delete':
                item = model.query.get(request.form.get('id'))
                if item: 
                    name = item.name
                    db.session.delete(item)
                    db.session.commit()
                    flash('Удалено')
                    log_action(f"Удалил из {type_}: {name}")

            elif action == 'bulk_delete':
                ids = request.form.getlist('ids[]')
                success_count = 0
                error_count = 0
                for item_id in ids:
                    obj = model.query.get(item_id)
                    if obj:
                        try:
                            db.session.delete(obj)
                            db.session.flush()
                            success_count += 1
                        except Exception:
                            db.session.rollback()
                            error_count += 1
                if success_count:
                    db.session.commit()
                
                msg = f'Удалено: {success_count}.'
                if error_count > 0:
                    msg += f' Не удалено (заняты): {error_count}.'
                flash(msg)
                log_action(f"Массовое удаление из {type_}: {success_count} шт.")
        except Exception as e: 
            db.session.rollback()
            flash(f'Ошибка: {e}')
        return redirect(url_for('directory.directory'))
        
    return render_template('directory/directory.html', 
                           plants=sorted(Plant.query.all(), key=lambda x: x.name), 
                           sizes=sorted(Size.query.all(), key=natural_key),
                           fields=sorted(Field.query.all(), key=natural_key), 
                           clients=sorted(Client.query.all(), key=lambda x: x.name), 
                           suppliers=sorted(Supplier.query.all(), key=lambda x: x.name))


@bp.route('/directory/download_stock_template')
@login_required
def download_stock_template():
    if current_user.role != 'admin': return redirect(url_for('directory.directory'))
    wb = Workbook()
    ws = wb.active
    ws.append(["Растение", "Размер", "Поле", "Год (Партия)", "Кол-во (шт)", "Цена закупки", "Цена продажи (База)"])
    ws.append(["Сосна горная", "C3", "Поле 1", msk_now().year, 100, 500, 1500])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='stock_import_template.xlsx', as_attachment=True)

@bp.route('/directory/stock_import', methods=['POST'])
@login_required
def stock_import():
    if current_user.role != 'admin': return redirect(url_for('directory.directory'))
    file = request.files.get('file')
    if not file: return redirect(url_for('directory.directory'))
    try:
        wb = load_workbook(file)
        ws = wb.active
        plants_map = {p.name.lower().strip(): p.id for p in Plant.query.all()}
        sizes_map = {s.name.lower().strip(): s.id for s in Size.query.all()}
        fields_map = {f.name.lower().strip(): f.id for f in Field.query.all()}
        
        data_by_year = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            y = int(row[3]) if row[3] else msk_now().year
            if y not in data_by_year: data_by_year[y] = []
            data_by_year[y].append(row)

        from datetime import datetime
        for year, rows in data_by_year.items():
            doc_date = datetime(year, 1, 1)
            doc = Document(doc_type='correction', user_id=current_user.id, date=doc_date, comment=f'Импорт остатков (Партия {year})')
            db.session.add(doc)
            db.session.flush()

            for r in rows:
                pid = plants_map.get(str(r[0]).lower().strip())
                sid = sizes_map.get(str(r[1]).lower().strip())
                fid = fields_map.get(str(r[2]).lower().strip())
                if pid and sid and fid:
                    new_qty = int(r[4] or 0)
                    price_buy = float(r[5] or 0)
                    price_sell = float(r[6] or 0)
                    
                    st = get_or_create_stock(pid, sid, fid, year)
                    
                    # ИСПРАВЛЕНО: Вычисляем дельту для истории, чтобы отчеты не двоились
                    delta = new_qty - st.quantity
                    st.quantity = new_qty # Устанавливаем абсолютное значение
                    
                    st.purchase_price = price_buy
                    st.price = price_sell
                    
                    # Записываем в документ только разницу, если она есть
                    if delta != 0:
                        db.session.add(DocumentRow(document_id=doc.id, plant_id=pid, size_id=sid, field_to_id=fid, year=year, quantity=delta))

        db.session.commit()
        flash("Остатки обновлены (коррекция записана)")
        log_action("Импорт остатков с коррекцией")
    except Exception as e: 
        db.session.rollback()
        flash(f"Ошибка: {e}")
    return redirect(url_for('directory.directory'))

@bp.route('/directory/download_template')
@login_required
def download_directory_template():
    if current_user.role != 'admin': return redirect(url_for('directory.directory'))
    wb = Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    ws1 = wb.create_sheet("Plant"); ws1.append(["Наименование", "Характеристика"])
    ws2 = wb.create_sheet("Field"); ws2.append(["Наименование", "Партнер (Имя Клиента)", "Год посадки (число)"])
    for name in ['Size', 'Client', 'Supplier']: wb.create_sheet(name).append(["Наименование"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='import_template.xlsx', as_attachment=True)

@bp.route('/directory/import', methods=['POST'])
@login_required
def import_directory():
    if current_user.role != 'admin': return redirect(url_for('directory.directory'))
    file = request.files.get('file')
    if not file: return redirect(url_for('directory.directory'))
    try:
        wb = load_workbook(file)
        if 'Plant' in wb.sheetnames:
            for row in wb['Plant'].iter_rows(min_row=2, values_only=True):
                if row and row[0] and not Plant.query.filter_by(name=str(row[0]).strip()).first():
                    db.session.add(Plant(name=str(row[0]).strip(), characteristic=str(row[1]) if len(row)>1 else None))
        if 'Field' in wb.sheetnames:
            for row in wb['Field'].iter_rows(min_row=2, values_only=True):
                if row and row[0] and not Field.query.filter_by(name=str(row[0]).strip()).first():
                    inv = Client.query.filter_by(name=str(row[1])).first() if len(row)>1 and row[1] else None
                    db.session.add(Field(name=str(row[0]).strip(), investor_id=inv.id if inv else None, planting_year=int(row[2]) if len(row)>2 and row[2] else 2017))
        for m_name, model in {'Size':Size, 'Client':Client, 'Supplier':Supplier}.items():
            if m_name in wb.sheetnames:
                for row in wb[m_name].iter_rows(min_row=2, values_only=True):
                    if row and row[0] and not model.query.filter_by(name=str(row[0]).strip()).first():
                        db.session.add(model(name=str(row[0]).strip()))
        db.session.commit()
        flash('Справочники импортированы')
        log_action("Импорт справочников Excel")
    except Exception as e: 
        db.session.rollback()
        flash(f'Ошибка: {e}')
    return redirect(url_for('directory.directory'))

@bp.route('/archive', methods=['GET', 'POST'])
@login_required
def archive():
    if current_user.role != 'admin':
        flash("Доступ только для администратора")
        return redirect(url_for('main.index'))

    if request.method == 'POST':
        file = request.files.get('file')
        category = request.form.get('category')
        comment = request.form.get('comment')

        if file and file.filename:
            try:
                now = msk_now()
                year_str = str(now.year)
                
                # Безопасное имя файла (латиница)
                filename = secure_filename(file.filename)
                # Добавляем таймштамп, чтобы имена не повторялись
                save_name = f"{int(now.timestamp())}_{filename}"
                
                # Относительный путь: archive/YYYY/name.ext
                rel_path = f"archive/{year_str}/{save_name}"
                
                # Полный путь к папке
                upload_path = current_app.config['UPLOAD_FOLDER']
                target_dir = os.path.join(upload_path, 'archive', year_str)
                os.makedirs(target_dir, exist_ok=True)
                
                full_path = os.path.join(target_dir, save_name)
                
                # Сохраняем на диск
                file.save(full_path)
                file_size = os.path.getsize(full_path)

                # Сохраняем в БД (в filename пишем относительный путь, чтобы старые файлы тоже работали)
                new_file = FileArchive(
                    filename=rel_path,
                    original_name=file.filename,
                    category=category,
                    comment=comment,
                    size_bytes=file_size
                )
                db.session.add(new_file)
                db.session.commit()
                
                log_action(f"Загрузил в архив: {file.filename}")
                flash('Файл успешно загружен в архив')
            except Exception as e:
                flash(f'Ошибка загрузки: {e}')
        else:
            flash('Файл не выбран')
        return redirect(url_for('directory.archive'))

    files = FileArchive.query.order_by(FileArchive.uploaded_at.desc()).all()
    return render_template('directory/archive.html', files=files)

@bp.route('/archive/download/<int:file_id>')
@login_required
def archive_download(file_id):
    if current_user.role != 'admin':
        return redirect(url_for('main.index'))
    
    file_record = FileArchive.query.get_or_404(file_id)
    try:
        from flask import send_file
        import os
        full_path = os.path.join(current_app.config['UPLOAD_FOLDER'], file_record.filename)
        return send_file(full_path, as_attachment=True, download_name=file_record.original_name)
    except Exception as e:
        flash(f"Файл не найден на диске: {e}")
        return redirect(url_for('directory.archive'))

@bp.route('/archive/delete/<int:file_id>', methods=['POST'])
@login_required
def archive_delete(file_id):
    if current_user.role != 'admin':
        return redirect(url_for('main.index'))
        
    file_record = FileArchive.query.get_or_404(file_id)
    try:
        # Удаляем с диска
        full_path = os.path.join(current_app.config['UPLOAD_FOLDER'], file_record.filename)
        if os.path.exists(full_path):
            os.remove(full_path)
        
        # Удаляем из БД
        db.session.delete(file_record)
        db.session.commit()
        
        log_action(f"Удалил из архива: {file_record.original_name}")
        flash('Файл удален')
    except Exception as e:
        flash(f'Ошибка удаления: {e}')
        
    return redirect(url_for('directory.archive'))


# ==========================================================
# ФОТОГРАФИИ РАСТЕНИЙ ДЛЯ КЛИЕНТСКОГО КАТАЛОГА
# ==========================================================
from app.photo_storage import (
    IMAGE_EXTENSIONS,
    get_legacy_photo_rel_dir,
    get_primary_photo_rel_dir,
    resolve_photo_source,
)


def _sanitize_photo_basename(name: str) -> str:
    # Сохраняем читаемое исходное имя (включая кириллицу), убирая только опасные символы.
    cleaned = os.path.basename((name or "").strip())
    cleaned = cleaned.replace("/", "_").replace("\\", "_").replace("\x00", "")
    cleaned = "".join(ch for ch in cleaned if ch not in '<>:"|?*' and ord(ch) >= 32)
    return cleaned.rstrip(". ").strip()

@bp.route('/directory/plant/<int:plant_id>/photos', methods=['GET', 'POST'])
@login_required
def manage_plant_photos(plant_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Доступ запрещен'}), 403

    plant = Plant.query.get_or_404(plant_id)
    root = current_app.config['UPLOAD_FOLDER']
    rel_dir = get_primary_photo_rel_dir(plant.name)
    upload_dir = os.path.join(root, *rel_dir.split('/'))
    
    if request.method == 'POST':
        return _handle_plant_photo_upload(plant, upload_dir)

    # GET: Возвращаем список фото для модального окна
    photos = []
    source_rel_dir, filenames = resolve_photo_source(root, plant.id, plant.name)
    for filename in filenames:
        photos.append({
            'name': filename,
            'url': url_for('main.serve_uploaded_file', filename=f"{source_rel_dir}/{filename}")
        })
    
    return jsonify({'plant_name': plant.name, 'photos': photos})


def _handle_plant_photo_upload(plant, upload_dir):
    if current_user.role != 'admin':
        return jsonify({'error': 'Доступ запрещен'}), 403

    if request.method != 'POST':
        return redirect(url_for('directory.directory'))

    # Загрузка фото
    try:
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir, exist_ok=True)
    except Exception as exc:
        current_app.logger.exception("Не удалось создать каталог фото: %s", upload_dir)
        flash(f'Ошибка каталога загрузки: {exc}')
        return redirect(url_for('directory.directory'))

    files = request.files.getlist('photos')
    if not files:
        flash('Файлы не получены. Повторите загрузку и проверьте, что выбраны изображения.')
        return redirect(url_for('directory.directory'))

    saved = 0
    rejected_reasons = []
    for file in files:
        if file and file.filename:
            original_name = (file.filename or "").strip()
            base_name, ext = os.path.splitext(original_name)
            ext = (ext or "").lower().strip()
            if ext not in IMAGE_EXTENSIONS:
                rejected_reasons.append(f"{original_name or '(без имени)'}: неподдерживаемое расширение '{ext or 'нет'}'")
                continue

            display_base = _sanitize_photo_basename(base_name)
            if not display_base:
                display_base = f"photo_{int(time.time()*1000)}"

            unique_name = f"{display_base}{ext}"
            suffix = 2
            while os.path.exists(os.path.join(upload_dir, unique_name)):
                unique_name = f"{display_base}_{suffix}{ext}"
                suffix += 1
            try:
                file.save(os.path.join(upload_dir, unique_name))
                saved += 1
            except Exception as exc:
                current_app.logger.exception(
                    "Ошибка сохранения фото '%s' в '%s' для plant_id=%s",
                    original_name,
                    upload_dir,
                    plant.id,
                )
                rejected_reasons.append(f"{original_name}: ошибка сохранения ({exc})")

    log_action(f"Загрузил {saved} фото для растения: {plant.name} (пропущено {len(rejected_reasons)})")
    if rejected_reasons:
        details = "; ".join(rejected_reasons[:3])
        more = f" (+{len(rejected_reasons) - 3} еще)" if len(rejected_reasons) > 3 else ""
        flash(
            f"Загружено {saved} фото. Пропущено {len(rejected_reasons)}: {details}{more}. "
            f"Поддерживаются: {', '.join(IMAGE_EXTENSIONS)}"
        )
    else:
        flash(f'Успешно загружено {saved} фото')
    return redirect(url_for('directory.directory'))


@bp.route('/directory/plant/photos', methods=['POST'])
@login_required
def upload_plant_photos():
    if current_user.role != 'admin':
        return jsonify({'error': 'Доступ запрещен'}), 403

    plant_id = request.form.get('plant_id', type=int)
    if not plant_id:
        flash('Не выбран объект растения для загрузки фото.')
        return redirect(url_for('directory.directory'))

    plant = Plant.query.get_or_404(plant_id)
    root = current_app.config['UPLOAD_FOLDER']
    rel_dir = get_primary_photo_rel_dir(plant.name)
    upload_dir = os.path.join(root, *rel_dir.split('/'))
    return _handle_plant_photo_upload(plant, upload_dir)

@bp.route('/directory/plant/<int:plant_id>/photos/<filename>/delete', methods=['POST'])
@login_required
def delete_plant_photo(plant_id, filename):
    if current_user.role != 'admin':
        return jsonify({'success': False}), 403
        
    plant = Plant.query.get_or_404(plant_id)
    # Удаляем только файлы из текущего каталога фото без путевых сегментов.
    if not filename or os.path.basename(filename) != filename:
        return jsonify({'success': False, 'error': 'Некорректное имя файла'}), 400

    root = current_app.config['UPLOAD_FOLDER']
    source_rel_dir, _ = resolve_photo_source(root, plant.id, plant.name)
    primary_rel_dir = get_primary_photo_rel_dir(plant.name)
    legacy_rel_dir = get_legacy_photo_rel_dir(plant.id, plant.name)

    for rel_dir in (source_rel_dir, primary_rel_dir, legacy_rel_dir):
        file_path = os.path.join(root, *f"{rel_dir}/{filename}".split('/'))
        if os.path.exists(file_path):
            os.remove(file_path)
            log_action(f"Удалил фото {filename} у растения {plant.name}")
            return jsonify({'success': True})

    return jsonify({'success': False, 'error': 'Файл не найден'})